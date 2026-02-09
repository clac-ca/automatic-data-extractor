"""
ADE hook: ``on_sheet_end``.

This hook runs once per worksheet after ADE has finished writing all normalized
output tables into the *output* workbook/worksheet. Use it for sheet-level
formatting, summaries, or post-processing that needs the final written state.

Contract
--------
- Called once per output worksheet.
- Must return ``None`` (any other return value is an error).

Guidance
--------
- Prefer deterministic, idempotent formatting operations.
- Avoid heavy scans; reuse facts captured earlier in ``state`` or ``tables``.
- Keep data edits earlier in the pipeline (mapped/transformed/validated).
"""

from __future__ import annotations

from collections.abc import Mapping, MutableMapping, Sequence
from typing import TYPE_CHECKING, Any

from ade_engine.models import TableResult

if TYPE_CHECKING:
    import openpyxl
    import openpyxl.worksheet.worksheet
    from ade_engine.extensions.registry import Registry
    from ade_engine.infrastructure.observability.logger import RunLogger
    from ade_engine.infrastructure.settings import Settings


def register(registry: Registry) -> None:
    """Register hook(s) with the ADE registry."""
    registry.register_hook(on_sheet_end, hook="on_sheet_end", priority=0)

    # ---------------------------------------------------------------------
    # Examples (uncomment to enable, then customize as needed)
    # ---------------------------------------------------------------------
    # registry.register_hook(on_sheet_end_example_1_record_output_bounds, hook="on_sheet_end", priority=10)
    # registry.register_hook(on_sheet_end_example_insert_original_headers_row, hook="on_sheet_end", priority=20)


def on_sheet_end(
    *,
    output_sheet: openpyxl.worksheet.worksheet.Worksheet,  # Output worksheet (openpyxl Worksheet)
    output_workbook: openpyxl.Workbook,  # Output workbook (openpyxl Workbook)
    tables: Sequence[TableResult],  # TableResult objects for this sheet, in write order
    input_file_name: str,  # Input filename (basename)
    settings: Settings,  # Engine Settings
    metadata: Mapping[str, Any],  # Run/sheet metadata (filenames, sheet_index, etc.)
    state: MutableMapping[str, Any],  # Mutable dict shared across the run
    logger: RunLogger,  # RunLogger (structured events + text logs)
) -> None:  # noqa: ARG001
    """Default implementation is a placeholder (no-op)."""
    return None


# ----------------------------
# Examples (uncomment in register() to enable)
# ----------------------------


def on_sheet_end_example_1_record_output_bounds(
    *,
    output_sheet: openpyxl.worksheet.worksheet.Worksheet,
    output_workbook: openpyxl.Workbook,
    tables: Sequence[TableResult],
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example: record output bounds for the sheet and log a short summary."""
    sheet_name = str(getattr(output_sheet, "title", ""))
    bounds = {
        "max_row": int(output_sheet.max_row or 0),
        "max_col": int(output_sheet.max_column or 0),
        "dimension": output_sheet.calculate_dimension()
        if hasattr(output_sheet, "calculate_dimension")
        else None,
    }

    stats = state.get("sheet_output_bounds")
    if not isinstance(stats, dict):
        stats = {}
        state["sheet_output_bounds"] = stats
    stats[sheet_name] = bounds

    if logger:
        logger.info(
            "Sheet output bounds recorded",
            extra={"data": {"sheet_name": sheet_name, **bounds}},
        )


def on_sheet_end_example_insert_original_headers_row(
    *,
    output_sheet: openpyxl.worksheet.worksheet.Worksheet,
    output_workbook: openpyxl.Workbook,
    tables: Sequence[TableResult],
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example: insert a row with original headers beneath each output header row."""
    for table_result in reversed(list(tables)):
        if table_result.output_region is None:
            continue

        original_by_field = {col.field_name: col.header for col in table_result.mapped_columns}
        header_row = table_result.output_region.min_row
        insert_row = header_row + 1
        output_sheet.insert_rows(insert_row, amount=1)

        # If you add Excel structured tables elsewhere, update their `ref` after row inserts.
        for col in range(
            table_result.output_region.min_col, table_result.output_region.max_col + 1
        ):
            field_name = output_sheet.cell(row=header_row, column=col).value
            output_sheet.cell(row=insert_row, column=col).value = original_by_field.get(field_name)
