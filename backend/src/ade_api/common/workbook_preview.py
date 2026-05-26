from __future__ import annotations

import csv
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX
from pydantic import Field

from ade_api.common.schema import BaseSchema

DEFAULT_PREVIEW_ROWS = 10_000
DEFAULT_PREVIEW_COLUMNS = 10_000
MAX_PREVIEW_ROWS = 10_000
MAX_PREVIEW_COLUMNS = 10_000


class WorkbookCellFormat(BaseSchema):
    """Display formatting for one preview cell."""

    row: int = Field(ge=0)
    column: int = Field(ge=0)
    bg_color: str | None = Field(default=None, alias="bgColor")
    text_color: str | None = Field(default=None, alias="textColor")
    bold: bool | None = None
    italic: bool | None = None
    horizontal_align: str | None = Field(default=None, alias="horizontalAlign")
    wrap_text: bool | None = Field(default=None, alias="wrapText")


class WorkbookSheetPreview(BaseSchema):
    """Preview for a single workbook sheet."""

    name: str
    index: int = Field(ge=0)
    rows: list[list[str]]
    total_rows: int = Field(alias="totalRows")
    total_columns: int = Field(alias="totalColumns")
    truncated_rows: bool = Field(alias="truncatedRows")
    truncated_columns: bool = Field(alias="truncatedColumns")
    hidden_rows: list[int] = Field(default_factory=list, alias="hiddenRows")
    hidden_columns: list[int] = Field(default_factory=list, alias="hiddenColumns")
    cell_formats: list[WorkbookCellFormat] = Field(default_factory=list, alias="cellFormats")


def get_xlsx_hidden_columns(
    path: Path,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> list[int]:
    """Extract 0-based hidden column indices without loading the whole workbook."""

    try:
        with zipfile.ZipFile(path, "r") as archive:
            rel_path = _resolve_xlsx_sheet_path(archive, sheet_name, sheet_index)
            if not rel_path:
                return []

            ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            sheet_xml = archive.read(rel_path)
            sheet_root = ET.fromstring(sheet_xml)

            hidden_cols = []
            cols_el = sheet_root.find(".//ns:cols", ns)
            if cols_el is not None:
                for col in cols_el.findall("ns:col", ns):
                    if col.attrib.get("hidden") in ("1", "true"):
                        min_attr = col.attrib.get("min")
                        max_attr = col.attrib.get("max")
                        if min_attr is None or max_attr is None:
                            continue
                        c_min = int(min_attr)
                        c_max = int(max_attr)
                        # min and max are 1-based, convert to 0-based for list indices
                        for col_idx in range(c_min, c_max + 1):
                            hidden_cols.append(col_idx - 1)
            return sorted(list(set(hidden_cols)))
    except Exception:
        # Fallback gracefully if ZIP/XML parsing fails for any reason
        return []


def _resolve_xlsx_sheet_path(
    archive: zipfile.ZipFile,
    sheet_name: str | None,
    sheet_index: int | None,
) -> str | None:
    # Step 1: Parse workbook to map sheet to relationship ID
    wb_content = archive.read("xl/workbook.xml")
    wb_root = ET.fromstring(wb_content)
    ns = {
        "ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    sheets = []
    for sheet_el in wb_root.findall(".//ns:sheet", ns):
        name = sheet_el.attrib.get("name")
        r_id = sheet_el.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        sheets.append({"name": name, "r_id": r_id})

    # Step 2: Determine which sheet we need
    target_sheet = None
    if sheet_name is not None:
        for s in sheets:
            if s["name"] == sheet_name:
                target_sheet = s
                break
    else:
        idx = sheet_index if sheet_index is not None else 0
        if 0 <= idx < len(sheets):
            target_sheet = sheets[idx]

    if not target_sheet:
        return None

    # Step 3: Find relationship targets to get sheet file path
    rel_content = archive.read("xl/_rels/workbook.xml.rels")
    rel_root = ET.fromstring(rel_content)
    rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    rel_map = {}
    for rel in rel_root.findall(".//rel:Relationship", rel_ns):
        rel_map[rel.attrib.get("Id")] = rel.attrib.get("Target")

    rel_path = rel_map.get(target_sheet["r_id"])
    if not rel_path:
        return None

    rel_path = rel_path.lstrip("/")
    if not rel_path.startswith("xl/"):
        rel_path = "xl/" + rel_path

    return rel_path


def get_xlsx_hidden_rows(
    path: Path,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> list[int]:
    """Extract 0-based hidden row indices without loading the whole workbook."""

    try:
        with zipfile.ZipFile(path, "r") as archive:
            rel_path = _resolve_xlsx_sheet_path(archive, sheet_name, sheet_index)
            if not rel_path:
                return []

            ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            sheet_xml = archive.read(rel_path)
            sheet_root = ET.fromstring(sheet_xml)

            hidden_rows = []
            for row in sheet_root.findall(".//ns:sheetData/ns:row", ns):
                if row.attrib.get("hidden") not in ("1", "true"):
                    continue
                row_number = row.attrib.get("r")
                if row_number is None:
                    continue
                hidden_rows.append(int(row_number) - 1)
            return sorted(list(set(hidden_rows)))
    except Exception:
        return []


def build_workbook_preview_from_xlsx(
    path: Path,
    *,
    max_rows: int = DEFAULT_PREVIEW_ROWS,
    max_columns: int = DEFAULT_PREVIEW_COLUMNS,
    trim_empty_columns: bool = False,
    trim_empty_rows: bool = False,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> WorkbookSheetPreview:
    with path.open("rb") as handle:
        workbook = openpyxl.load_workbook(
            handle,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            index, sheet = _select_xlsx_sheet(workbook, sheet_name, sheet_index)
            preview = _preview_xlsx_sheet(
                sheet,
                index=index,
                max_rows=max_rows,
                max_columns=max_columns,
                trim_empty_columns=trim_empty_columns,
                trim_empty_rows=trim_empty_rows,
            )
            preview.hidden_rows = get_xlsx_hidden_rows(path, sheet_name=preview.name)
            preview.hidden_columns = get_xlsx_hidden_columns(path, sheet_name=preview.name)
            return preview
        finally:
            workbook.close()


def build_workbook_preview_from_csv(
    path: Path,
    *,
    max_rows: int = DEFAULT_PREVIEW_ROWS,
    max_columns: int = DEFAULT_PREVIEW_COLUMNS,
    trim_empty_columns: bool = False,
    trim_empty_rows: bool = False,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> WorkbookSheetPreview:
    name = path.stem or "Sheet1"
    if sheet_name and sheet_name != name:
        raise KeyError(f"Sheet {sheet_name!r} not found")
    if sheet_index not in (None, 0):
        raise IndexError("sheet_index out of range")

    rows: list[list[str]] = []
    total_rows = 0
    total_columns = 0

    with path.open("r", newline="", encoding="utf-8", errors="replace") as handle:
        reader = csv.reader(handle)
        for row in reader:
            total_rows += 1
            total_columns = max(total_columns, len(row))
            if len(rows) < max_rows:
                rows.append([_normalize_cell(cell) for cell in row])

    return _preview_sheet_from_rows(
        name=name,
        index=0,
        rows=rows,
        total_rows=total_rows,
        total_columns=total_columns,
        max_rows=max_rows,
        max_columns=max_columns,
        trim_empty_columns=trim_empty_columns,
        trim_empty_rows=trim_empty_rows,
    )


def _select_xlsx_sheet(
    workbook: Any,
    sheet_name: str | None,
    sheet_index: int | None,
) -> tuple[int, Any]:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found")
        index = workbook.sheetnames.index(sheet_name)
        return index, workbook[sheet_name]
    effective_index = sheet_index if sheet_index is not None else 0
    if effective_index < 0 or effective_index >= len(workbook.sheetnames):
        raise IndexError("sheet_index out of range")
    return effective_index, workbook[workbook.sheetnames[effective_index]]


def _preview_xlsx_sheet(
    sheet: Any,
    *,
    index: int,
    max_rows: int,
    max_columns: int,
    trim_empty_columns: bool,
    trim_empty_rows: bool,
) -> WorkbookSheetPreview:
    rows: list[list[str]] = []
    cell_formats: list[WorkbookCellFormat] = []
    for row_index, row in enumerate(sheet.iter_rows(
        min_row=1,
        max_row=max_rows,
        max_col=max_columns,
        values_only=False,
    )):
        rows.append([_normalize_cell(cell.value) for cell in row])
        for column_index, cell in enumerate(row):
            cell_format = _extract_cell_format(cell, row_index, column_index)
            if cell_format:
                cell_formats.append(cell_format)

    total_rows = sheet.max_row or 0
    total_columns = sheet.max_column or 0
    preview = _preview_sheet_from_rows(
        name=sheet.title,
        index=index,
        rows=rows,
        total_rows=total_rows,
        total_columns=total_columns,
        max_rows=max_rows,
        max_columns=max_columns,
        trim_empty_columns=trim_empty_columns,
        trim_empty_rows=trim_empty_rows,
    )
    preview.cell_formats = cell_formats
    return preview


def _preview_sheet_from_rows(
    *,
    name: str,
    index: int,
    rows: Sequence[Sequence[str]],
    total_rows: int,
    total_columns: int,
    max_rows: int,
    max_columns: int,
    trim_empty_columns: bool,
    trim_empty_rows: bool,
) -> WorkbookSheetPreview:
    max_row_length = max((len(row) for row in rows), default=0)
    preview_columns = min(max_columns, total_columns or max_row_length)
    normalized_rows = [_normalize_row(list(row), preview_columns) for row in rows]
    if trim_empty_rows:
        normalized_rows = _trim_empty_rows(normalized_rows)
    if trim_empty_columns:
        normalized_rows = _trim_empty_columns(normalized_rows)

    return WorkbookSheetPreview(
        name=name,
        index=index,
        rows=normalized_rows,
        total_rows=total_rows,
        total_columns=total_columns,
        truncated_rows=total_rows > max_rows,
        truncated_columns=total_columns > max_columns,
    )


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _normalize_row(row: Sequence[str], length: int) -> list[str]:
    return [row[index] if index < len(row) else "" for index in range(length)]


def _trim_empty_columns(rows: Sequence[Sequence[str]]) -> list[list[str]]:
    if not rows:
        return [list(row) for row in rows]

    column_count = max((len(row) for row in rows), default=0)
    if column_count == 0:
        return [list(row) for row in rows]

    keep_indices = [
        index for index in range(column_count) if any(row[index].strip() for row in rows)
    ]

    if not keep_indices:
        return [[] for _ in rows]

    return [[row[index] for index in keep_indices] for row in rows]


def _trim_empty_rows(rows: Sequence[Sequence[str]]) -> list[list[str]]:
    return [list(row) for row in rows if any(cell.strip() for cell in row)]


def _extract_cell_format(cell: Any, row_index: int, column_index: int) -> WorkbookCellFormat | None:
    fill = getattr(cell, "fill", None)
    font = getattr(cell, "font", None)
    alignment = getattr(cell, "alignment", None)

    fill_type = getattr(fill, "fill_type", None)
    bg_color = None
    if fill_type and fill_type != "none":
        bg_color = _css_color_from_openpyxl_color(getattr(fill, "fgColor", None))

    text_color = _css_color_from_openpyxl_color(getattr(font, "color", None))
    bold = True if getattr(font, "bold", False) else None
    italic = True if getattr(font, "italic", False) else None
    horizontal_align = getattr(alignment, "horizontal", None)
    wrap_text = True if getattr(alignment, "wrap_text", False) else None

    if not any([bg_color, text_color, bold, italic, horizontal_align, wrap_text]):
        return None

    return WorkbookCellFormat(
        row=row_index,
        column=column_index,
        bg_color=bg_color,
        text_color=text_color,
        bold=bold,
        italic=italic,
        horizontal_align=horizontal_align,
        wrap_text=wrap_text,
    )


def _css_color_from_openpyxl_color(color: Any) -> str | None:
    color_type = getattr(color, "type", None)
    raw: str | None = None

    if color_type == "rgb":
        raw = getattr(color, "rgb", None)
    elif color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if isinstance(indexed, int) and 0 <= indexed < len(COLOR_INDEX):
            raw = COLOR_INDEX[indexed]

    if not raw or not isinstance(raw, str):
        return None

    raw = raw.upper()
    if len(raw) == 8:
        rgb = raw[2:]
        return f"#{rgb}"
    if len(raw) == 6:
        return f"#{raw}"
    return None


__all__ = [
    "DEFAULT_PREVIEW_COLUMNS",
    "DEFAULT_PREVIEW_ROWS",
    "MAX_PREVIEW_COLUMNS",
    "MAX_PREVIEW_ROWS",
    "WorkbookCellFormat",
    "WorkbookSheetPreview",
    "build_workbook_preview_from_csv",
    "build_workbook_preview_from_xlsx",
    "get_xlsx_hidden_rows",
]
