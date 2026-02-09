"""Run orchestration service coordinating DB state and queueing."""

from __future__ import annotations

import asyncio
import json
import logging
import tempfile
import time
import unicodedata
from collections.abc import AsyncIterator, Iterator, Sequence
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import TimeoutError as FuturesTimeout
from contextlib import contextmanager
from datetime import UTC, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import openpyxl
from fastapi import UploadFile
from sqlalchemy import case, func, insert, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ade_api.common.cursor_listing import ResolvedCursorSort
from ade_api.common.list_filters import FilterItem, FilterJoinOperator
from ade_api.common.logging import log_context
from ade_api.common.sse import sse_text
from ade_api.common.time import utc_now
from ade_api.common.workbook_preview import (
    WorkbookSheetPreview,
    build_workbook_preview_from_csv,
    build_workbook_preview_from_xlsx,
)
from ade_api.features.admin_settings.service import RuntimeSettingsService
from ade_api.features.configs.deps import compute_dependency_digest, has_engine_dependency
from ade_api.features.configs.exceptions import (
    ConfigEngineDependencyMissingError,
    ConfigStateError,
    ConfigStorageNotFoundError,
    ConfigurationNotFoundError,
)
from ade_api.features.configs.repository import ConfigurationsRepository
from ade_api.features.configs.storage import ConfigStorage
from ade_api.features.documents.repository import DocumentsRepository
from ade_api.features.documents.upload_metadata import parse_upload_run_options
from ade_api.features.workspaces.repository import WorkspacesRepository
from ade_api.features.workspaces.settings import read_processing_paused
from ade_api.settings import Settings
from ade_db.models import (
    Configuration,
    ConfigurationStatus,
    File,
    FileKind,
    FileVersion,
    FileVersionOrigin,
    Run,
    RunField,
    RunMetrics,
    RunOperation,
    RunStatus,
    RunTableColumn,
)
from ade_storage import StorageAdapter

from .exceptions import (
    RunDocumentMissingError,
    RunInputDocumentRequiredForProcessError,
    RunInputMissingError,
    RunLogsFileMissingError,
    RunNotCancellableError,
    RunNotFoundError,
    RunOutputMissingError,
    RunOutputNotReadyError,
    RunOutputPreviewParseError,
    RunOutputPreviewSheetNotFoundError,
    RunOutputPreviewUnsupportedError,
    RunOutputSheetParseError,
    RunOutputSheetUnsupportedError,
    RunSafeModeEnabledError,
)
from .filters import RunColumnFilters
from .repository import RunsRepository
from .schemas import (
    RunBatchCreateOptions,
    RunCreateOptions,
    RunCreateOptionsBase,
    RunInput,
    RunLinks,
    RunOutput,
    RunOutputSheet,
    RunPage,
    RunResource,
)

__all__ = [
    "RunInputMissingError",
    "RunDocumentMissingError",
    "RunLogsFileMissingError",
    "RunNotFoundError",
    "RunOutputNotReadyError",
    "RunOutputMissingError",
    "RunsService",
]

logger = logging.getLogger(__name__)

_DEPS_DIGEST_CACHE_SIZE = 256
_OUTPUT_FALLBACK_FILENAME = "output"
_MAX_FILENAME_LENGTH = 255
_CONFIG_DIGEST_SNAPSHOT_KEY = "__config_digest_snapshot"
_RUN_EVENTS_STREAM_POLL_SECONDS = 0.25
_RUN_EVENTS_STREAM_KEEPALIVE_SECONDS = 15.0
_RUN_EVENTS_STREAM_TERMINAL_EOF_GRACE_SECONDS = 3.0


@lru_cache(maxsize=_DEPS_DIGEST_CACHE_SIZE)
def _cached_deps_digest(config_path: str, config_cache_key: str) -> str:
    return compute_dependency_digest(Path(config_path))


def _deps_digest_cache_key(configuration: Configuration) -> str:
    if configuration.published_digest:
        return configuration.published_digest
    if configuration.updated_at:
        return configuration.updated_at.isoformat()
    return "unknown"



# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #


def _run_with_timeout(func, *, timeout: float, **kwargs):
    """Run a callable with a timeout to avoid hanging on large workbook operations."""
    if timeout <= 0:
        return func(**kwargs)
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(func, **kwargs)
        return future.result(timeout=timeout)


def _normalize_run_operation(value: RunOperation | str) -> RunOperation:
    if isinstance(value, RunOperation):
        return value
    return RunOperation(str(value))



# --------------------------------------------------------------------------- #
# Main service
# --------------------------------------------------------------------------- #


class RunsService:
    """Coordinate run persistence, queueing, and serialization for the API.

    Responsibilities:
    - create and persist Run rows
    - enforce batch creation rules
    - read artifacts/logs written by ade-worker
    - serialize run resources for API responses
    """

    def __init__(
        self,
        *,
        session: Session,
        settings: Settings,
        storage: ConfigStorage | None = None,
        blob_storage: StorageAdapter,
    ) -> None:
        self._session = session
        self._settings = settings
        self._configs = ConfigurationsRepository(session)
        self._runs = RunsRepository(session)
        self._documents = DocumentsRepository(session)
        self._workspaces = WorkspacesRepository(session)
        self._storage = storage or ConfigStorage(
            settings=settings,
        )
        self._blob_storage = blob_storage
        self._runtime_settings = RuntimeSettingsService(
            session=session,
        )

        default_max = Run.__table__.c.max_attempts.default
        self._run_max_attempts = int(default_max.arg) if default_max is not None else 3

    # --------------------------------------------------------------------- #
    # Run lifecycle: creation and queueing
    # --------------------------------------------------------------------- #

    def prepare_run(
        self,
        *,
        configuration_id: UUID,
        options: RunCreateOptions,
    ) -> Run:
        """Create the queued run row and enqueue execution."""
        operation = _normalize_run_operation(options.operation)

        logger.debug(
            "run.prepare.start",
            extra=log_context(
                configuration_id=configuration_id,
                operation=operation.value,
                dry_run=options.dry_run,
                input_document_id=options.input_document_id,
            ),
        )

        safe_mode = self._runtime_settings.get_effective_values().safe_mode
        if safe_mode.enabled:
            raise RunSafeModeEnabledError(safe_mode.detail)

        configuration = self._resolve_configuration(configuration_id)
        if operation is RunOperation.PUBLISH:
            return self.prepare_publish_run(
                workspace_id=configuration.workspace_id,
                configuration_id=configuration.id,
            )

        run_options_payload = options.model_dump(mode="json", exclude_none=True)
        run_options_payload["operation"] = operation.value
        deps_digest = self._resolve_deps_digest(configuration)
        input_document_id = options.input_document_id
        if operation is RunOperation.PROCESS:
            if not input_document_id:
                raise RunInputDocumentRequiredForProcessError(
                    "input_document_required_for_process"
                )
            self._require_document(
                workspace_id=configuration.workspace_id,
                document_id=input_document_id,
            )

            selected_sheet_names = self._select_input_sheet_names(options)
            self._insert_runs_for_documents(
                configuration=configuration,
                document_ids=[input_document_id],
                deps_digest=deps_digest,
                input_sheet_names_by_document_id={
                    input_document_id: selected_sheet_names or None,
                },
                run_options_by_document_id={
                    input_document_id: run_options_payload,
                },
                existing_statuses=[RunStatus.QUEUED, RunStatus.RUNNING],
                operation=RunOperation.PROCESS,
            )
            run = self._require_active_run(
                configuration_id=configuration.id,
                document_id=input_document_id,
            )
        else:
            config_digest_snapshot = self._resolve_config_digest_snapshot(configuration)
            run_options_payload[_CONFIG_DIGEST_SNAPSHOT_KEY] = config_digest_snapshot
            run = self._insert_validate_run(
                configuration=configuration,
                deps_digest=deps_digest,
                run_options=run_options_payload,
            )

        # Touch configuration usage timestamp.
        configuration.last_used_at = utc_now()  # type: ignore[attr-defined]
        self._session.flush()

        logger.info(
            "run.prepare.success",
            extra=log_context(
                workspace_id=run.workspace_id,
                configuration_id=run.configuration_id,
                run_id=run.id,
                operation=run.operation.value,
                input_document_id=input_document_id,
                dry_run=options.dry_run,
            ),
        )
        return run

    def prepare_publish_run(
        self,
        *,
        workspace_id: UUID,
        configuration_id: UUID,
    ) -> Run:
        """Create (or reuse) a queued publish run for a draft configuration."""

        logger.debug(
            "run.prepare.publish.start",
            extra=log_context(
                workspace_id=workspace_id,
                configuration_id=configuration_id,
            ),
        )

        configuration = self._resolve_configuration(configuration_id)
        if configuration.workspace_id != workspace_id:
            raise ConfigurationNotFoundError("configuration_not_found")
        if configuration.status is not ConfigurationStatus.DRAFT:
            raise ConfigStateError("Configuration must be a draft before publishing")

        existing = self._runs.get_active_publish_run(configuration_id=configuration.id)
        if existing is not None:
            logger.info(
                "run.prepare.publish.reuse",
                extra=log_context(
                    workspace_id=workspace_id,
                    configuration_id=configuration_id,
                    run_id=existing.id,
                ),
            )
            return existing

        try:
            deps_digest = self._resolve_deps_digest(configuration)
            config_digest_snapshot = self._resolve_config_digest_snapshot(configuration)
        except RunInputMissingError as exc:
            raise ConfigStorageNotFoundError("configuration_storage_missing") from exc
        run_options_payload: dict[str, Any] = {
            "operation": RunOperation.PUBLISH.value,
            _CONFIG_DIGEST_SNAPSHOT_KEY: config_digest_snapshot,
        }

        try:
            with self._session.begin_nested():
                run = self._insert_publish_run(
                    configuration=configuration,
                    deps_digest=deps_digest,
                    run_options=run_options_payload,
                )
        except IntegrityError:
            existing = self._runs.get_active_publish_run(configuration_id=configuration.id)
            if existing is not None:
                logger.info(
                    "run.prepare.publish.reuse_after_conflict",
                    extra=log_context(
                        workspace_id=workspace_id,
                        configuration_id=configuration_id,
                        run_id=existing.id,
                    ),
                )
                return existing
            raise

        configuration.last_used_at = utc_now()  # type: ignore[attr-defined]
        self._session.flush()

        logger.info(
            "run.prepare.publish.success",
            extra=log_context(
                workspace_id=workspace_id,
                configuration_id=configuration_id,
                run_id=run.id,
            ),
        )
        return run

    def prepare_run_for_workspace(
        self,
        *,
        workspace_id: UUID,
        input_document_id: UUID | None,
        configuration_id: UUID | None,
        options: RunCreateOptionsBase | None = None,
    ) -> Run:
        """Create a run for the workspace, resolving the active configuration if needed."""

        configuration = self._resolve_workspace_configuration(
            workspace_id=workspace_id,
            configuration_id=configuration_id,
        )
        run_options = self._merge_run_options(
            input_document_id=input_document_id,
            options=options,
        )
        return self.prepare_run(
            configuration_id=configuration.id,
            options=run_options,
        )

    def prepare_runs_batch(
        self,
        *,
        configuration_id: UUID,
        document_ids: Sequence[UUID],
        options: RunBatchCreateOptions,
        input_sheet_names_by_document_id: dict[UUID, list[str]] | None = None,
        active_sheet_only_by_document_id: dict[UUID, bool] | None = None,
        skip_existing_check: bool = False,
    ) -> list[Run]:
        """Create queued runs for each document id, enforcing all-or-nothing semantics."""
        operation = _normalize_run_operation(options.operation)

        logger.debug(
            "run.prepare.batch.start",
            extra=log_context(
                configuration_id=configuration_id,
                document_count=len(document_ids),
                operation=operation.value,
                dry_run=options.dry_run,
            ),
        )

        safe_mode = self._runtime_settings.get_effective_values().safe_mode
        if safe_mode.enabled:
            raise RunSafeModeEnabledError(safe_mode.detail)

        if not document_ids:
            return []
        if operation is not RunOperation.PROCESS:
            raise RunInputDocumentRequiredForProcessError(
                "input_document_required_for_process"
            )

        configuration = self._resolve_configuration(configuration_id)
        self._require_documents(
            workspace_id=configuration.workspace_id,
            document_ids=document_ids,
        )

        batch_active_sheet_only = bool(getattr(options, "active_sheet_only", False))
        normalized_sheet_names: dict[UUID, list[str] | None] = {}
        run_options_by_document_id: dict[UUID, RunCreateOptions] = {}

        if skip_existing_check:
            new_document_ids = list(document_ids)
        else:
            existing = self._runs.list_active_for_documents(
                configuration_id=configuration.id,
                document_ids=list(document_ids),
            )
            existing_version_ids = [
                run.input_file_version_id
                for run in existing
                if run.input_file_version_id is not None
            ]
            if existing_version_ids:
                rows = self._session.execute(
                    select(FileVersion.file_id).where(FileVersion.id.in_(existing_version_ids))
                )
                existing_ids = {row[0] for row in rows}
            else:
                existing_ids = set()
            new_document_ids = [doc_id for doc_id in document_ids if doc_id not in existing_ids]

        if new_document_ids:
            for document_id in new_document_ids:
                input_sheet_names = None
                if (
                    input_sheet_names_by_document_id
                    and document_id in input_sheet_names_by_document_id
                ):
                    raw_names = input_sheet_names_by_document_id.get(document_id) or []
                    normalized = self._select_input_sheet_names(
                        RunCreateOptions(
                            input_document_id=document_id,
                            input_sheet_names=raw_names,
                        ),
                    )
                    input_sheet_names = normalized or None
                active_sheet_only = batch_active_sheet_only
                if (
                    active_sheet_only_by_document_id
                    and document_id in active_sheet_only_by_document_id
                ):
                    active_sheet_only = bool(active_sheet_only_by_document_id.get(document_id))
                if input_sheet_names:
                    active_sheet_only = False
                normalized_sheet_names[document_id] = input_sheet_names
                run_options_by_document_id[document_id] = RunCreateOptions(
                    operation=RunOperation.PROCESS,
                    dry_run=options.dry_run,
                    log_level=options.log_level,
                    input_document_id=document_id,
                    input_sheet_names=input_sheet_names,
                    active_sheet_only=active_sheet_only,
                    metadata=options.metadata,
                )

            deps_digest = self._resolve_deps_digest(configuration)
            self._insert_runs_for_documents(
                configuration=configuration,
                document_ids=new_document_ids,
                deps_digest=deps_digest,
                input_sheet_names_by_document_id=normalized_sheet_names,
                run_options_by_document_id={
                    doc_id: run_options.model_dump(mode="json", exclude_none=True)
                    for doc_id, run_options in run_options_by_document_id.items()
                },
                existing_statuses=[RunStatus.QUEUED, RunStatus.RUNNING],
                operation=RunOperation.PROCESS,
            )

        configuration.last_used_at = utc_now()  # type: ignore[attr-defined]
        self._session.flush()

        runs = self._runs.list_active_for_documents(
            configuration_id=configuration.id,
            document_ids=list(document_ids),
        )

        if runs:
            logger.info(
                "run.prepare.batch.success",
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                    count=len(runs),
                ),
            )
        return runs

    def prepare_runs_batch_for_workspace(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
        configuration_id: UUID | None,
        options: RunBatchCreateOptions,
    ) -> list[Run]:
        """Create batch runs for the workspace, resolving the active configuration if needed."""

        configuration = self._resolve_workspace_configuration(
            workspace_id=workspace_id,
            configuration_id=configuration_id,
        )
        return self.prepare_runs_batch(
            configuration_id=configuration.id,
            document_ids=document_ids,
            options=options,
        )

    def load_run_options(self, run: Run) -> RunCreateOptions:
        """Rehydrate run options from the run row."""

        payload: dict[str, Any] = dict(run.run_options or {})
        if "operation" not in payload:
            payload["operation"] = (
                run.operation.value
                if isinstance(run.operation, RunOperation)
                else str(run.operation or RunOperation.PROCESS.value)
            )

        if "input_document_id" not in payload and run.input_file_version_id:
            document_id = self._resolve_document_id_for_version(run.input_file_version_id)
            if document_id is not None:
                payload["input_document_id"] = str(document_id)
        if "input_sheet_names" not in payload and run.input_sheet_names:
            payload["input_sheet_names"] = list(run.input_sheet_names)

        try:
            return RunCreateOptions(**payload)
        except Exception:
            logger.warning(
                "run.options.load.failed",
                extra=log_context(run_id=run.id),
            )
            fallback_operation = (
                run.operation
                if isinstance(run.operation, RunOperation)
                else RunOperation.PROCESS
            )
            fallback_document_id = (
                self._resolve_document_id_for_version(run.input_file_version_id)
                if run.input_file_version_id
                else None
            )
            fallback_value = (
                str(fallback_document_id)
                if fallback_document_id is not None
                else (
                    str(run.input_file_version_id)
                    if fallback_operation is RunOperation.PROCESS and run.input_file_version_id
                    else None
                )
            )
            return RunCreateOptions(
                operation=fallback_operation,
                input_document_id=fallback_value,
                input_sheet_names=list(run.input_sheet_names or []),
            )

    def enqueue_pending_runs_for_configuration(
        self,
        *,
        configuration_id: UUID,
        batch_size: int | None = None,
    ) -> int:
        """Queue runs for uploaded documents without runs using the active configuration."""

        try:
            configuration = self._resolve_configuration(configuration_id)
        except ConfigurationNotFoundError:
            return 0
        if self._processing_paused(configuration.workspace_id):
            logger.info(
                "run.pending.enqueue.skip.processing_paused",
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                ),
            )
            return 0
        if configuration.status is not ConfigurationStatus.ACTIVE:
            logger.info(
                "run.pending.enqueue.skip.inactive_config",
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                    status=configuration.status.value,
                ),
            )
            return 0

        batch_size = batch_size or 100
        if batch_size <= 0:
            return 0

        total = 0
        while True:
            limit = batch_size
            documents = self._pending_documents(
                workspace_id=configuration.workspace_id,
                configuration_id=configuration.id,
                limit=limit,
            )
            if not documents:
                break
            logger.debug(
                "run.pending.enqueue.batch batch_size=%s",
                limit,
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                    batch_size=limit,
                ),
            )
            document_ids = [doc.id for doc in documents]
            sheet_names_by_document_id: dict[UUID, list[str]] = {}
            active_sheet_only_by_document_id: dict[UUID, bool] = {}
            for document in documents:
                run_options = parse_upload_run_options(document.attributes)
                if run_options and run_options.input_sheet_names is not None:
                    sheet_names_by_document_id[document.id] = list(run_options.input_sheet_names)
                if run_options and run_options.active_sheet_only:
                    active_sheet_only_by_document_id[document.id] = True
            runs = self.prepare_runs_batch(
                configuration_id=configuration.id,
                document_ids=document_ids,
                options=RunBatchCreateOptions(),
                input_sheet_names_by_document_id=sheet_names_by_document_id or None,
                active_sheet_only_by_document_id=active_sheet_only_by_document_id or None,
                skip_existing_check=True,
            )
            total += len(runs)
            if len(document_ids) < limit:
                break

        if total:
            logger.info(
                "run.pending.enqueue.completed",
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                    count=total,
                ),
            )
        return total

    def is_processing_paused(self, *, workspace_id: UUID) -> bool:
        return self._processing_paused(workspace_id)

    def _pending_documents(
        self,
        *,
        workspace_id: UUID,
        configuration_id: UUID,
        limit: int,
    ) -> list[File]:
        if limit <= 0:
            return []
        pending_run_exists = (
            select(Run.id)
            .where(
                Run.input_file_version_id == FileVersion.id,
                FileVersion.file_id == File.id,
                Run.configuration_id == configuration_id,
            )
            .limit(1)
            .exists()
        )
        stmt = (
            select(File)
            .where(
                File.workspace_id == workspace_id,
                File.kind == FileKind.INPUT,
                File.last_run_id.is_(None),
                File.deleted_at.is_(None),
                ~pending_run_exists,
            )
            .order_by(File.created_at.asc())
            .limit(limit)
        )
        result = self._session.execute(stmt)
        return list(result.scalars().all())

    def _require_active_run(
        self,
        *,
        configuration_id: UUID,
        document_id: UUID,
    ) -> Run:
        runs = self._runs.list_active_for_documents(
            configuration_id=configuration_id,
            document_ids=[document_id],
        )
        if not runs:
            raise RuntimeError(f"No active run found for document {document_id}")
        return runs[0]

    def _resolve_deps_digest(self, configuration: Configuration) -> str:
        try:
            config_path = self._storage.ensure_config_path(
                configuration.workspace_id,
                configuration.id,
            )
        except ConfigStorageNotFoundError as exc:  # pragma: no cover - surface as run error
            raise RunInputMissingError("Configuration files are missing") from exc
        if not has_engine_dependency(config_path):
            raise ConfigEngineDependencyMissingError(
                "Configuration must declare ade-engine in dependency manifests."
            )
        cache_key = _deps_digest_cache_key(configuration)
        return _cached_deps_digest(str(config_path), cache_key)

    def _resolve_config_digest_snapshot(self, configuration: Configuration) -> str:
        try:
            config_path = self._storage.ensure_config_path(
                configuration.workspace_id,
                configuration.id,
            )
        except ConfigStorageNotFoundError as exc:  # pragma: no cover - surface as run error
            raise RunInputMissingError("Configuration files are missing") from exc
        return self._storage.compute_config_digest(config_path)

    def _insert_validate_run(
        self,
        *,
        configuration: Configuration,
        deps_digest: str,
        run_options: dict[str, Any],
    ) -> Run:
        run = Run(
            id=uuid4(),
            configuration_id=configuration.id,
            workspace_id=configuration.workspace_id,
            input_file_version_id=None,
            input_sheet_names=None,
            run_options=run_options,
            deps_digest=deps_digest,
            operation=RunOperation.VALIDATE,
            status=RunStatus.QUEUED,
            available_at=utc_now(),
            attempt_count=0,
            max_attempts=self._run_max_attempts,
        )
        self._session.add(run)
        self._session.flush()
        self._session.refresh(run)
        return run

    def _insert_publish_run(
        self,
        *,
        configuration: Configuration,
        deps_digest: str,
        run_options: dict[str, Any],
    ) -> Run:
        run = Run(
            id=uuid4(),
            configuration_id=configuration.id,
            workspace_id=configuration.workspace_id,
            input_file_version_id=None,
            input_sheet_names=None,
            run_options=run_options,
            deps_digest=deps_digest,
            operation=RunOperation.PUBLISH,
            status=RunStatus.QUEUED,
            available_at=utc_now(),
            attempt_count=0,
            max_attempts=self._run_max_attempts,
        )
        self._session.add(run)
        self._session.flush()
        self._session.refresh(run)
        return run

    def _insert_runs_for_documents(
        self,
        *,
        configuration: Configuration,
        document_ids: Sequence[UUID],
        deps_digest: str,
        input_sheet_names_by_document_id: dict[UUID, list[str] | None] | None,
        run_options_by_document_id: dict[UUID, dict[str, Any] | None] | None,
        existing_statuses: Sequence[RunStatus] | None,
        operation: RunOperation = RunOperation.PROCESS,
    ) -> None:
        if not document_ids:
            return

        base_stmt = (
            select(File.id, File.current_version_id)
            .where(File.workspace_id == configuration.workspace_id)
            .where(File.kind == FileKind.INPUT)
            .where(File.deleted_at.is_(None))
            .where(File.id.in_(document_ids))
        )
        result = self._session.execute(base_stmt)
        version_by_doc: dict[UUID, UUID] = {
            doc_id: version_id
            for doc_id, version_id in result.all()
            if version_id is not None
        }
        version_to_doc = {version_id: doc_id for doc_id, version_id in version_by_doc.items()}
        eligible_ids = list(version_by_doc.keys())
        if not eligible_ids:
            return

        if existing_statuses:
            existing_stmt = select(Run.input_file_version_id).where(
                Run.configuration_id == configuration.id,
                Run.input_file_version_id.in_(list(version_by_doc.values())),
                Run.status.in_(list(existing_statuses)),
            )
            existing_result = self._session.execute(existing_stmt)
            existing_versions = {version_id for (version_id,) in existing_result.all()}
            eligible_ids = [
                doc_id
                for doc_id, version_id in version_by_doc.items()
                if version_id not in existing_versions
            ]
            if not eligible_ids:
                return

        def build_rows(ids: Sequence[UUID]) -> list[dict[str, Any]]:
            now = utc_now()
            return [
                {
                    "id": uuid4(),
                    "configuration_id": configuration.id,
                    "workspace_id": configuration.workspace_id,
                    "input_file_version_id": version_by_doc[doc_id],
                    "input_sheet_names": (
                        input_sheet_names_by_document_id.get(doc_id)
                        if input_sheet_names_by_document_id
                        else None
                    ),
                    "run_options": (
                        run_options_by_document_id.get(doc_id)
                        if run_options_by_document_id
                        else None
                    ),
                    "deps_digest": deps_digest,
                    "operation": operation,
                    "status": RunStatus.QUEUED,
                    "available_at": now,
                    "attempt_count": 0,
                    "max_attempts": self._run_max_attempts,
                }
                for doc_id in ids
            ]

        def update_last_run_ids(rows: Sequence[dict[str, Any]]) -> None:
            if not rows:
                return
            doc_to_run = {
                version_to_doc[row["input_file_version_id"]]: row["id"]
                for row in rows
                if row.get("input_file_version_id") in version_to_doc
            }
            self._session.execute(
                update(File)
                .where(File.id.in_(list(doc_to_run.keys())))
                .values(last_run_id=case(doc_to_run, value=File.id))
            )

        rows = build_rows(eligible_ids)
        for attempt in range(2):
            try:
                with self._session.begin_nested():
                    self._session.execute(insert(Run), rows)
                    update_last_run_ids(rows)
                return
            except IntegrityError:
                if attempt:
                    raise
                if existing_statuses:
                    existing_stmt = select(Run.input_file_version_id).where(
                        Run.configuration_id == configuration.id,
                        Run.input_file_version_id.in_(list(version_by_doc.values())),
                        Run.status.in_(list(existing_statuses)),
                    )
                    existing_result = self._session.execute(existing_stmt)
                    existing_versions = {version_id for (version_id,) in existing_result.all()}
                    remaining = [
                        doc_id
                        for doc_id, version_id in version_by_doc.items()
                        if version_id not in existing_versions
                    ]
                    if not remaining:
                        return
                    rows = build_rows(remaining)

    # --------------------------------------------------------------------- #
    # Public read APIs (runs, summaries, outputs)
    # --------------------------------------------------------------------- #

    def get_run(self, run_id: UUID) -> Run | None:
        """Return the run instance for ``run_id`` if it exists."""

        logger.debug(
            "run.get.start",
            extra=log_context(run_id=run_id),
        )
        run = self._runs.get(run_id)
        if run is None:
            logger.debug(
                "run.get.miss",
                extra=log_context(run_id=run_id),
            )
        else:
            logger.debug(
                "run.get.hit",
                extra=log_context(
                    workspace_id=run.workspace_id,
                    configuration_id=run.configuration_id,
                    run_id=run.id,
                    status=run.status.value,
                ),
            )
        return run

    def cancel_run(self, *, run_id: UUID) -> Run:
        """Cancel a queued/running run and return the updated row.

        Cancellation is idempotent for runs already marked as cancelled.
        """

        run = self._require_run(run_id)
        if run.status is RunStatus.CANCELLED:
            return run
        if run.status not in {RunStatus.QUEUED, RunStatus.RUNNING}:
            raise RunNotCancellableError(
                f"Run {run_id} cannot be cancelled from status {run.status.value}."
            )

        now = utc_now()
        run.status = RunStatus.CANCELLED
        run.completed_at = now
        run.claimed_by = None
        run.claim_expires_at = None
        run.error_message = "Run cancelled by user"
        self._session.flush()
        self._session.refresh(run)

        logger.info(
            "run.cancelled",
            extra=log_context(
                workspace_id=run.workspace_id,
                configuration_id=run.configuration_id,
                run_id=run.id,
                operation=run.operation.value,
            ),
        )
        return run

    def get_run_metrics(self, *, run_id: UUID) -> RunMetrics | None:
        """Return persisted run metrics for ``run_id`` when available."""

        run = self._require_run(run_id)
        return self._runs.get_metrics(run.id)

    def list_run_fields(self, *, run_id: UUID) -> list[RunField]:
        """Return field summaries for ``run_id``."""

        run = self._require_run(run_id)
        return self._runs.list_fields(run.id)

    def list_run_columns(
        self,
        *,
        run_id: UUID,
        filters: RunColumnFilters,
    ) -> list[RunTableColumn]:
        """Return detected columns for ``run_id`` with optional filters."""

        run = self._require_run(run_id)
        return self._runs.list_columns(run_id=run.id, filters=filters)

    def list_runs(
        self,
        *,
        workspace_id: UUID,
        configuration_id: UUID | None = None,
        filters: list[FilterItem],
        join_operator: FilterJoinOperator,
        q: str | None,
        resolved_sort: ResolvedCursorSort[Run],
        limit: int,
        cursor: str | None,
        include_total: bool,
    ) -> RunPage:
        """Return paginated runs for ``workspace_id`` with optional filters."""

        logger.debug(
            "run.list.start",
            extra=log_context(
                workspace_id=workspace_id,
                configuration_id=configuration_id,
                filters=[item.model_dump() for item in filters],
                order_by=str(resolved_sort.order_by),
                limit=limit,
                cursor=cursor,
                q=q,
            ),
        )

        page_result = self._runs.list_by_workspace(
            workspace_id=workspace_id,
            configuration_id=configuration_id,
            filters=filters,
            join_operator=join_operator,
            q=q,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
        )
        resources = [self.to_resource(run, resolve_paths=False) for run in page_result.items]
        response = RunPage(
            items=resources,
            meta=page_result.meta,
            facets=page_result.facets,
        )

        logger.info(
            "run.list.success",
            extra=log_context(
                workspace_id=workspace_id,
                configuration_id=configuration_id,
                limit=response.meta.limit,
                count=len(response.items),
                total=response.meta.total_count,
            ),
        )
        return response

    def list_runs_for_configuration(
        self,
        *,
        configuration_id: UUID,
        filters: list[FilterItem],
        join_operator: FilterJoinOperator,
        q: str | None,
        resolved_sort: ResolvedCursorSort[Run],
        limit: int,
        cursor: str | None,
        include_total: bool,
    ) -> RunPage:
        """Return paginated runs for ``configuration_id`` scoped to its workspace."""

        configuration = self._resolve_configuration(configuration_id)
        return self.list_runs(
            workspace_id=configuration.workspace_id,
            configuration_id=configuration.id,
            filters=filters,
            join_operator=join_operator,
            q=q,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
        )

    def to_resource(self, run: Run, *, resolve_paths: bool = True) -> RunResource:
        """Convert ``run`` into its API representation."""

        # Timing and failure info
        started_at = self._ensure_utc(run.started_at)
        completed_at = self._ensure_utc(run.completed_at)
        duration_seconds = (
            (completed_at - started_at).total_seconds() if started_at and completed_at else None
        )

        failure_code = None
        failure_stage = None
        failure_message = run.error_message

        input_meta = self._build_input_metadata(
            run=run,
            files_counts={},
            sheets_counts={},
        )
        output_meta = self._build_output_metadata(
            run=run,
        )
        links = self._links(
            workspace_id=run.workspace_id,
            run_id=run.id,
        )

        return RunResource(
            id=run.id,
            workspace_id=run.workspace_id,
            configuration_id=run.configuration_id,
            operation=run.operation,
            status=run.status,
            failure_code=failure_code,
            failure_stage=failure_stage,
            failure_message=failure_message,
            created_at=self._ensure_utc(run.created_at) or utc_now(),
            started_at=started_at,
            completed_at=completed_at,
            duration_seconds=duration_seconds,
            exit_code=run.exit_code,
            input=input_meta,
            output=output_meta,
            links=links,
        )

    def _build_input_metadata(
        self,
        *,
        run: Run,
        files_counts: dict[str, Any],
        sheets_counts: dict[str, Any],
    ) -> RunInput:
        document_id = None
        file_version_id = None
        version_no = None
        filename: str | None = None
        content_type: str | None = None
        size_bytes: int | None = None
        download_url: str | None = None

        if run.input_file_version_id:
            file_version_id = str(run.input_file_version_id)
            download_url = (
                f"/api/v1/workspaces/{run.workspace_id}/runs/{run.id}/input/download"
            )
            try:
                document, version = self._require_document_with_version(
                    workspace_id=run.workspace_id,
                    file_version_id=run.input_file_version_id,
                )
                document_id = str(document.id)
                version_no = version.version_no
                filename = version.filename_at_upload or document.name
                content_type = version.content_type
                size_bytes = version.byte_size
            except RunDocumentMissingError:
                logger.warning(
                    "run.input.metadata.missing_document",
                    extra=log_context(
                        workspace_id=run.workspace_id,
                        configuration_id=run.configuration_id,
                        run_id=run.id,
                        document_id=document_id or file_version_id,
                    ),
                )

        return RunInput(
            document_id=document_id,
            file_version_id=file_version_id,
            version_no=version_no,
            filename=filename,
            content_type=content_type,
            size_bytes=size_bytes,
            download_url=download_url,
            input_sheet_names=run.input_sheet_names,
            input_file_count=files_counts.get("total"),
            input_sheet_count=sheets_counts.get("total"),
        )

    def _build_output_metadata(
        self,
        *,
        run: Run,
    ) -> RunOutput:
        output_version_id = run.output_file_version_id
        output_file: File | None = None
        output_version: FileVersion | None = None
        if output_version_id is not None:
            output_version = self._session.get(FileVersion, output_version_id)
            if output_version is not None:
                output_file = self._session.get(File, output_version.file_id)

        ready = output_version is not None

        filename: str | None = None
        size_bytes: int | None = None
        content_type: str | None = None
        version_no: int | None = None

        if output_version is not None:
            version_no = output_version.version_no
            size_bytes = output_version.byte_size
            content_type = output_version.content_type or "application/octet-stream"
            filename = output_version.filename_at_upload or (
                output_file.name if output_file is not None else None
            )

        return RunOutput(
            ready=ready,
            download_url=(
                f"/api/v1/workspaces/{run.workspace_id}/runs/{run.id}/output/download"
            ),
            filename=filename,
            content_type=content_type,
            size_bytes=size_bytes,
            has_output=ready,
            file_version_id=str(output_version_id) if output_version_id else None,
            version_no=version_no,
        )

    @staticmethod
    def _links(*, workspace_id: UUID, run_id: UUID) -> RunLinks:
        base = f"/api/v1/workspaces/{workspace_id}/runs/{run_id}"
        events_stream = f"{base}/events/stream"
        events_download = f"{base}/events/download"
        output_metadata = f"{base}/output"
        output_download = f"{output_metadata}/download"
        input_download = f"{base}/input/download"

        return RunLinks(
            self=base,
            events_stream=events_stream,
            events_download=events_download,
            input_download=input_download,
            output_download=output_download,
            output_metadata=output_metadata,
        )

    def get_run_input_metadata(
        self,
        *,
        run_id: UUID,
    ) -> RunInput:
        run = self._require_run(run_id)
        resource = self.to_resource(run)
        if resource.input.document_id is None:
            raise RunInputMissingError("Run input is unavailable")
        if resource.input.filename is None:
            raise RunDocumentMissingError("Run input file is unavailable")
        return resource.input

    def stream_run_input(
        self,
        *,
        run_id: UUID,
    ) -> tuple[Run, File, FileVersion, Iterator[bytes]]:
        run = self._require_run(run_id)
        if not run.input_file_version_id:
            raise RunInputMissingError("Run input is unavailable")
        document, version = self._require_document_with_version(
            workspace_id=run.workspace_id,
            file_version_id=run.input_file_version_id,
        )

        stream = self._blob_storage.stream(
            document.blob_name,
            version_id=version.storage_version_id,
            chunk_size=self._settings.blob_download_chunk_size_bytes,
        )

        def _guarded() -> Iterator[bytes]:
            try:
                yield from stream
            except FileNotFoundError as exc:
                logger.warning(
                    "run.input.stream.file_lost",
                    extra=log_context(
                        workspace_id=run.workspace_id,
                        configuration_id=run.configuration_id,
                        run_id=run.id,
                        document_id=document.id,
                        stored_uri=document.blob_name,
                    ),
                )
                raise RunDocumentMissingError("Run input file is unavailable") from exc

        return run, document, version, _guarded()

    def get_run_output_metadata(
        self,
        *,
        run_id: UUID,
    ) -> RunOutput:
        run = self._require_run(run_id)
        resource = self.to_resource(run)
        return resource.output

    def upload_manual_output(
        self,
        *,
        run_id: UUID,
        upload: UploadFile,
        actor_id: UUID | None = None,
    ) -> RunOutput:
        """Upload a manual output file for a completed run."""

        run = self._require_run(run_id)
        if run.status in {RunStatus.QUEUED, RunStatus.RUNNING}:
            raise RunOutputNotReadyError("Run output cannot be uploaded until the run completes.")

        document, _ = self._require_document_with_version(
            workspace_id=run.workspace_id,
            file_version_id=run.input_file_version_id,
        )

        upload_name = self._normalise_filename(upload.filename)
        content_type = self._normalise_content_type(upload.content_type)

        output_name = f"{document.name or upload_name} (Output)"
        output_name_key = f"output:{document.id}"
        output_file = self._ensure_output_file(
            workspace_id=run.workspace_id,
            document_id=document.id,
            name=output_name,
            name_key=output_name_key,
            actor_id=actor_id,
        )

        if upload.file is None:  # pragma: no cover - UploadFile always provides a file
            raise RuntimeError("Upload stream is not available")

        stored = self._blob_storage.write(
            output_file.blob_name,
            upload.file,
            max_bytes=self._settings.storage_upload_max_bytes,
        )

        storage_version_id = stored.version_id
        now = datetime.now(tz=UTC)

        version_no = self._next_version_no(file_id=output_file.id)
        file_version = FileVersion(
            file_id=output_file.id,
            version_no=version_no,
            origin=FileVersionOrigin.MANUAL,
            run_id=run.id,
            created_by_user_id=actor_id,
            sha256=stored.sha256,
            byte_size=stored.byte_size,
            content_type=content_type,
            filename_at_upload=upload_name,
            storage_version_id=storage_version_id,
        )
        self._session.add(file_version)
        self._session.flush()

        output_file.current_version_id = file_version.id
        output_file.updated_at = now
        run.output_file_version_id = file_version.id
        self._session.flush()

        return self._build_output_metadata(run=run)

    def resolve_output_for_download(
        self,
        *,
        run_id: UUID,
    ) -> tuple[Run, File, FileVersion]:
        run = self._require_run(run_id)
        if run.status in {
            RunStatus.QUEUED,
            RunStatus.RUNNING,
        }:
            raise RunOutputNotReadyError("Run output is not available until the run completes.")
        try:
            output_file, output_version = self._require_output_version(run=run)
        except RunOutputMissingError as err:
            if run.status is RunStatus.FAILED:
                raise RunOutputMissingError(
                    "Run failed and no output is available",
                ) from err
            if run.status is RunStatus.CANCELLED:
                raise RunOutputMissingError(
                    "Run was cancelled and no output is available",
                ) from err
            raise
        return run, output_file, output_version

    def stream_run_output(
        self,
        *,
        run_id: UUID,
    ) -> tuple[Run, File, FileVersion, Iterator[bytes]]:
        run, output_file, output_version = self.resolve_output_for_download(run_id=run_id)
        stream = self._blob_storage.stream(
            output_file.blob_name,
            version_id=output_version.storage_version_id,
            chunk_size=self._settings.blob_download_chunk_size_bytes,
        )

        def _guarded() -> Iterator[bytes]:
            try:
                yield from stream
            except FileNotFoundError as exc:
                logger.warning(
                    "run.output.stream.missing",
                    extra=log_context(
                        workspace_id=run.workspace_id,
                        configuration_id=run.configuration_id,
                        run_id=run.id,
                        blob_name=output_file.blob_name,
                    ),
                )
                raise RunOutputMissingError("Run output is unavailable") from exc

        return run, output_file, output_version, _guarded()

    def get_run_output_preview(
        self,
        *,
        run_id: UUID,
        max_rows: int,
        max_columns: int,
        trim_empty_columns: bool = False,
        trim_empty_rows: bool = False,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
    ) -> WorkbookSheetPreview:
        """Return a table-ready preview for a run output workbook."""

        effective_sheet_index = sheet_index
        if sheet_name is None and sheet_index is None:
            effective_sheet_index = 0

        logger.debug(
            "run.output.preview.start",
            extra=log_context(
                run_id=run_id,
                sheet_name=sheet_name,
                sheet_index=effective_sheet_index,
            ),
        )

        run, output_file, output_version = self.resolve_output_for_download(run_id=run_id)
        output_name = output_version.filename_at_upload or output_file.name
        suffix = Path(output_name).suffix.lower()
        timeout = self._settings.preview_timeout_seconds

        try:
            with self._download_blob_to_tempfile(
                blob_name=output_file.blob_name,
                version_id=output_version.storage_version_id,
                suffix=suffix,
            ) as path:
                if suffix == ".xlsx":
                    preview = _run_with_timeout(
                        build_workbook_preview_from_xlsx,
                        timeout=timeout,
                        path=path,
                        max_rows=max_rows,
                        max_columns=max_columns,
                        trim_empty_columns=trim_empty_columns,
                        trim_empty_rows=trim_empty_rows,
                        sheet_name=sheet_name,
                        sheet_index=effective_sheet_index,
                    )
                elif suffix == ".csv":
                    preview = _run_with_timeout(
                        build_workbook_preview_from_csv,
                        timeout=timeout,
                        path=path,
                        max_rows=max_rows,
                        max_columns=max_columns,
                        trim_empty_columns=trim_empty_columns,
                        trim_empty_rows=trim_empty_rows,
                        sheet_name=sheet_name,
                        sheet_index=effective_sheet_index,
                    )
                else:
                    raise RunOutputPreviewUnsupportedError(
                        f"Preview is not supported for output file type {suffix!r}."
                    )
        except FileNotFoundError as exc:
            raise RunOutputMissingError("Run output is unavailable") from exc
        except (KeyError, IndexError) as exc:
            requested = sheet_name if sheet_name is not None else str(effective_sheet_index)
            raise RunOutputPreviewSheetNotFoundError(
                f"Sheet {requested!r} was not found in run {run_id!r} output."
            ) from exc
        except FuturesTimeout as exc:
            raise RunOutputPreviewParseError(
                f"Preview timed out after {timeout:g}s for run {run_id!r} output."
            ) from exc
        except RunOutputPreviewUnsupportedError:
            raise
        except Exception as exc:  # pragma: no cover - defensive fallback
            raise RunOutputPreviewParseError(
                f"Preview generation failed for run {run_id!r} output ({type(exc).__name__})."
            ) from exc

        logger.info(
            "run.output.preview.success",
            extra=log_context(
                run_id=run.id,
                workspace_id=run.workspace_id,
                configuration_id=run.configuration_id,
                sheet_name=preview.name,
                sheet_index=preview.index,
            ),
        )
        return preview

    def list_run_output_sheets(
        self,
        *,
        run_id: UUID,
    ) -> list[RunOutputSheet]:
        """Return worksheet metadata for a run output workbook."""

        logger.debug(
            "run.output.sheets.list.start",
            extra=log_context(run_id=run_id),
        )

        run, output_file, output_version = self.resolve_output_for_download(run_id=run_id)
        output_name = output_version.filename_at_upload or output_file.name
        suffix = Path(output_name).suffix.lower()
        timeout = self._settings.preview_timeout_seconds

        if suffix == ".xlsx":
            try:
                with self._download_blob_to_tempfile(
                    blob_name=output_file.blob_name,
                    version_id=output_version.storage_version_id,
                    suffix=suffix,
                ) as path:
                    sheets = _run_with_timeout(self._inspect_workbook, timeout=timeout, path=path)
            except FuturesTimeout as exc:
                raise RunOutputSheetParseError(
                    f"Worksheet inspection timed out after {timeout:g}s for run {run_id!r} output."
                ) from exc
            except FileNotFoundError as exc:
                raise RunOutputMissingError("Run output is unavailable") from exc
            except Exception as exc:  # pragma: no cover - defensive fallback
                raise RunOutputSheetParseError(
                    f"Worksheet inspection failed for run {run_id!r} output ({type(exc).__name__})."
                ) from exc
            logger.info(
                "run.output.sheets.list.success",
                extra=log_context(
                    run_id=run.id,
                    workspace_id=run.workspace_id,
                    configuration_id=run.configuration_id,
                    sheet_count=len(sheets),
                    kind="workbook",
                ),
            )
            return sheets

        if suffix == ".csv":
            name = self._default_sheet_name(output_name)
            sheets = [RunOutputSheet(name=name, index=0, kind="file", is_active=True)]
            logger.info(
                "run.output.sheets.list.success",
                extra=log_context(
                    run_id=run.id,
                    workspace_id=run.workspace_id,
                    configuration_id=run.configuration_id,
                    sheet_count=len(sheets),
                    kind="file",
                ),
            )
            return sheets

        raise RunOutputSheetUnsupportedError(
            f"Sheets are not supported for output file type {suffix!r}."
        )

    def stream_run_logs(self, *, run_id: UUID) -> Iterator[bytes]:
        """Stream NDJSON run logs from storage."""

        run = self._require_run(run_id)
        blob_name = self._run_log_blob_name(workspace_id=run.workspace_id, run_id=run.id)
        stream = self._blob_storage.stream(
            blob_name,
            chunk_size=self._settings.blob_download_chunk_size_bytes,
        )

        def _guarded() -> Iterator[bytes]:
            try:
                yield from stream
            except FileNotFoundError as exc:
                logger.warning(
                    "run.logs.missing",
                    extra=log_context(
                        workspace_id=run.workspace_id,
                        configuration_id=run.configuration_id,
                        run_id=run.id,
                        blob_name=blob_name,
                    ),
                )
                raise RunLogsFileMissingError("Run log stream is unavailable") from exc

        return _guarded()

    async def stream_run_events(
        self,
        *,
        run_id: UUID,
        cursor: int = 0,
    ) -> AsyncIterator[dict[str, str]]:
        """Tail run NDJSON logs from blob storage and yield SSE event messages."""

        snapshot = self._run_stream_snapshot(run_id=run_id)
        blob_name = self._run_log_blob_name(workspace_id=snapshot["workspace_id"], run_id=run_id)
        offset = max(0, int(cursor))
        buffer = b""
        last_keepalive_at = time.monotonic()
        terminal_observed_at: float | None = None
        terminal_statuses = {
            RunStatus.SUCCEEDED,
            RunStatus.FAILED,
            RunStatus.CANCELLED,
        }

        yield sse_text(
            "ready",
            json.dumps(
                {
                    "runId": str(run_id),
                    "status": snapshot["status"].value,
                    "cursor": offset,
                },
                separators=(",", ":"),
            ),
            event_id=offset,
        )

        while True:
            saw_bytes = False
            try:
                stream = self._blob_storage.stream_range(
                    blob_name,
                    start_offset=offset,
                    chunk_size=self._settings.blob_download_chunk_size_bytes,
                )
                for chunk in stream:
                    if not chunk:
                        continue
                    saw_bytes = True
                    buffer += chunk
                    lines = buffer.splitlines(keepends=True)
                    buffer = b""
                    for line in lines:
                        if not line.endswith(b"\n"):
                            buffer = line
                            break
                        offset += len(line)
                        raw = line.decode("utf-8", errors="replace").rstrip("\r\n")
                        event_name = self._extract_event_name(raw)
                        if not event_name:
                            continue
                        yield sse_text("message", raw, event_id=offset)
                        if event_name == "run.complete":
                            snapshot = self._run_stream_snapshot(run_id=run_id)
                            yield sse_text(
                                "end",
                                json.dumps(
                                    {
                                        "runId": str(run_id),
                                        "status": snapshot["status"].value,
                                        "cursor": offset,
                                        "reason": "run_complete",
                                    },
                                    separators=(",", ":"),
                                ),
                                event_id=offset,
                            )
                            return
            except FileNotFoundError:
                # A stream client may connect before the worker writes the first log chunk.
                pass

            now = time.monotonic()
            if now - last_keepalive_at >= _RUN_EVENTS_STREAM_KEEPALIVE_SECONDS:
                yield sse_text("keepalive", "{}")
                last_keepalive_at = now

            snapshot = self._run_stream_snapshot(run_id=run_id)
            if snapshot["status"] in terminal_statuses:
                if terminal_observed_at is None:
                    terminal_observed_at = now
                if saw_bytes or buffer:
                    terminal_observed_at = now
                grace_elapsed = (
                    now - terminal_observed_at
                    >= _RUN_EVENTS_STREAM_TERMINAL_EOF_GRACE_SECONDS
                )
                if grace_elapsed and not saw_bytes and not buffer:
                    yield sse_text(
                        "end",
                        json.dumps(
                            {
                                "runId": str(run_id),
                                "status": snapshot["status"].value,
                                "cursor": offset,
                                "reason": "terminal_status",
                            },
                            separators=(",", ":"),
                        ),
                        event_id=offset,
                    )
                    return
            else:
                terminal_observed_at = None

            await asyncio.sleep(_RUN_EVENTS_STREAM_POLL_SECONDS)

    def _run_stream_snapshot(
        self,
        *,
        run_id: UUID,
    ) -> dict[str, Any]:
        row = self._session.execute(
            select(
                Run.workspace_id,
                Run.status,
            ).where(Run.id == run_id)
        ).one_or_none()
        if row is None:
            raise RunNotFoundError("Run not found")
        workspace_id, status = row
        return {
            "workspace_id": workspace_id,
            "status": status if isinstance(status, RunStatus) else RunStatus(str(status)),
        }

    @staticmethod
    def _run_log_blob_name(*, workspace_id: UUID, run_id: UUID) -> str:
        return f"{workspace_id}/runs/{run_id}/logs/events.ndjson"

    @staticmethod
    def _extract_event_name(raw: str) -> str | None:
        if not raw.strip():
            return None
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            return None
        if not isinstance(payload, dict):
            return None
        event_name = payload.get("event")
        if not isinstance(event_name, str) or not event_name.strip():
            return None
        return event_name

    @staticmethod
    def _ensure_utc(dt: datetime | None) -> datetime | None:
        if dt is None:
            return None
        if dt.tzinfo is None:
            return dt.replace(tzinfo=UTC)
        return dt

    @staticmethod
    def _coerce_datetime(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value if value.tzinfo else value.replace(tzinfo=UTC)
        if isinstance(value, str):
            try:
                parsed = datetime.fromisoformat(value)
            except ValueError:
                return None
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
        return None

    def _require_run(self, run_id: UUID) -> Run:
        run = self._runs.get(run_id)
        if run is None:
            logger.warning(
                "run.require_run.not_found",
                extra=log_context(run_id=run_id),
            )
            raise RunNotFoundError(run_id)
        return run

    def _resolve_document_id_for_version(self, file_version_id: UUID) -> UUID | None:
        stmt = (
            select(File.id)
            .join(FileVersion, FileVersion.file_id == File.id)
            .where(
                FileVersion.id == file_version_id,
                File.kind == FileKind.INPUT,
                File.deleted_at.is_(None),
            )
        )
        result = self._session.execute(stmt).scalar_one_or_none()
        return result if isinstance(result, UUID) else (UUID(str(result)) if result else None)

    def _require_document_with_version(
        self,
        *,
        workspace_id: UUID,
        file_version_id: UUID,
    ) -> tuple[File, FileVersion]:
        version = self._session.get(FileVersion, file_version_id)
        if version is None:
            raise RunDocumentMissingError(f"Document version {file_version_id} not found")
        document = self._session.get(File, version.file_id)
        if (
            document is None
            or document.workspace_id != workspace_id
            or document.kind != FileKind.INPUT
            or document.deleted_at is not None
        ):
            raise RunDocumentMissingError(f"Document {version.file_id} not found")
        return document, version

    def _require_output_version(self, *, run: Run) -> tuple[File, FileVersion]:
        output_version_id = run.output_file_version_id
        if not output_version_id:
            raise RunOutputMissingError("Run output is unavailable")
        version = self._session.get(FileVersion, output_version_id)
        if version is None:
            raise RunOutputMissingError("Run output is unavailable")
        output_file = self._session.get(File, version.file_id)
        if (
            output_file is None
            or output_file.workspace_id != run.workspace_id
            or output_file.kind != FileKind.OUTPUT
        ):
            raise RunOutputMissingError("Run output is unavailable")
        return output_file, version

    def _ensure_output_file(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        name: str,
        name_key: str,
        actor_id: UUID | None,
    ) -> File:
        stmt = select(File).where(
            File.workspace_id == workspace_id,
            File.kind == FileKind.OUTPUT,
            File.name_key == name_key,
            File.deleted_at.is_(None),
        )
        existing = self._session.execute(stmt).scalar_one_or_none()
        if existing is not None:
            return existing

        file_id = uuid4()
        output_file = File(
            id=file_id,
            workspace_id=workspace_id,
            kind=FileKind.OUTPUT,
            name=name,
            name_key=name_key,
            blob_name=self._file_blob_name(workspace_id=workspace_id, file_id=file_id),
            source_file_id=document_id,
            attributes={},
            uploaded_by_user_id=actor_id,
            comment_count=0,
        )
        self._session.add(output_file)
        try:
            self._session.flush()
        except IntegrityError:
            self._session.rollback()
            existing = self._session.execute(stmt).scalar_one_or_none()
            if existing is not None:
                return existing
            raise
        return output_file

    def _next_version_no(self, *, file_id: UUID) -> int:
        stmt = select(func.max(FileVersion.version_no)).where(FileVersion.file_id == file_id)
        current = self._session.execute(stmt).scalar_one()
        return int(current or 0) + 1

    @staticmethod
    def _normalise_filename(name: str | None) -> str:
        if name is None:
            return _OUTPUT_FALLBACK_FILENAME

        candidate = name.strip()
        if not candidate:
            return _OUTPUT_FALLBACK_FILENAME

        filtered = "".join(ch for ch in candidate if unicodedata.category(ch)[0] != "C").strip()
        if not filtered:
            return _OUTPUT_FALLBACK_FILENAME

        if len(filtered) > _MAX_FILENAME_LENGTH:
            filtered = filtered[:_MAX_FILENAME_LENGTH].rstrip()

        return filtered or _OUTPUT_FALLBACK_FILENAME

    @staticmethod
    def _normalise_content_type(content_type: str | None) -> str | None:
        if content_type is None:
            return None
        candidate = content_type.strip()
        return candidate or None

    @staticmethod
    def _file_blob_name(*, workspace_id: UUID, file_id: UUID) -> str:
        return f"{workspace_id}/files/{file_id}"

    def _require_document(self, *, workspace_id: UUID, document_id: UUID) -> File:
        document = self._documents.get_document(
            workspace_id=workspace_id,
            document_id=document_id,
        )
        if document is None:
            logger.warning(
                "run.require_document.not_found",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                ),
            )
            raise RunDocumentMissingError(f"Document {document_id} not found")
        return document

    def _require_documents(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
    ) -> list[File]:
        if not document_ids:
            return []

        stmt = select(File).where(
            File.workspace_id == workspace_id,
            File.kind == FileKind.INPUT,
            File.deleted_at.is_(None),
            File.id.in_(document_ids),
        )
        result = self._session.execute(stmt)
        documents = list(result.scalars())
        found = {doc.id for doc in documents}
        missing = [document_id for document_id in document_ids if document_id not in found]
        if missing:
            logger.warning(
                "run.require_documents.not_found",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=str(missing[0]),
                    missing_count=len(missing),
                ),
            )
            raise RunDocumentMissingError(f"Document {missing[0]} not found")
        return documents

    def _resolve_configuration(self, configuration_id: UUID) -> Configuration:
        configuration = self._configs.get_by_id(configuration_id)
        if configuration is None:
            logger.warning(
                "run.config.resolve.not_found",
                extra=log_context(configuration_id=configuration_id),
            )
            raise ConfigurationNotFoundError(configuration_id)
        if configuration.status == ConfigurationStatus.ARCHIVED:
            logger.warning(
                "run.config.resolve.archived",
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                    status=configuration.status.value,
                ),
            )
        return configuration

    def _processing_paused(self, workspace_id: UUID) -> bool:
        workspace = self._workspaces.get_workspace(workspace_id)
        if workspace is None:
            return False
        return read_processing_paused(workspace.settings)

    def _resolve_workspace_configuration(
        self,
        *,
        workspace_id: UUID,
        configuration_id: UUID | None,
    ) -> Configuration:
        if configuration_id is None:
            configuration = self._configs.get_active(workspace_id)
            if configuration is None:
                logger.warning(
                    "run.config.resolve.active_missing",
                    extra=log_context(workspace_id=workspace_id),
                )
                raise ConfigurationNotFoundError("active_configuration_not_found")
        else:
            configuration = self._configs.get(
                workspace_id=workspace_id,
                configuration_id=configuration_id,
            )
            if configuration is None:
                logger.warning(
                    "run.config.resolve.not_found",
                    extra=log_context(
                        workspace_id=workspace_id,
                        configuration_id=configuration_id,
                    ),
                )
                raise ConfigurationNotFoundError(configuration_id)

        if configuration.status == ConfigurationStatus.ARCHIVED:
            logger.warning(
                "run.config.resolve.archived",
                extra=log_context(
                    workspace_id=configuration.workspace_id,
                    configuration_id=configuration.id,
                    status=configuration.status.value,
                ),
            )
        return configuration

    @staticmethod
    def _inspect_workbook(path: Path) -> list[RunOutputSheet]:
        with path.open("rb") as raw:
            workbook = openpyxl.load_workbook(
                raw,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            try:
                sheetnames = workbook.sheetnames
                active = workbook.active.title if sheetnames else None
                return [
                    RunOutputSheet(
                        name=title,
                        index=index,
                        kind="worksheet",
                        is_active=title == active,
                    )
                    for index, title in enumerate(sheetnames)
                ]
            finally:
                workbook.close()

    @staticmethod
    def _default_sheet_name(name: str | None) -> str:
        stem = Path(name or "Sheet").stem.strip() or "Sheet"
        return stem

    @contextmanager
    def _download_blob_to_tempfile(
        self,
        *,
        blob_name: str,
        version_id: str | None,
        suffix: str | None = None,
    ) -> Iterator[Path]:
        suffix = suffix or ""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / f"run-output{suffix}"
            with path.open("wb") as handle:
                for chunk in self._blob_storage.stream(
                    blob_name,
                    version_id=version_id,
                    chunk_size=self._settings.blob_download_chunk_size_bytes,
                ):
                    handle.write(chunk)
            yield path

    @staticmethod
    def _epoch_seconds(dt: datetime | None) -> int | None:
        if dt is None:
            return None
        return int(dt.timestamp())

    @staticmethod
    def _merge_run_options(
        *,
        input_document_id: UUID | None,
        options: RunCreateOptionsBase | None,
    ) -> RunCreateOptions:
        payload = options.model_dump(exclude_none=True) if options else {}
        return RunCreateOptions(input_document_id=input_document_id, **payload)

    @staticmethod
    def _select_input_sheet_names(options: RunCreateOptions) -> list[str]:
        """Normalize requested sheet names into a unique, ordered list."""

        if getattr(options, "active_sheet_only", False):
            return []
        names: list[str] = []
        for name in getattr(options, "input_sheet_names", None) or []:
            if not isinstance(name, str):
                continue
            cleaned = name.strip()
            if cleaned and cleaned not in names:
                names.append(cleaned)
        return names

    @staticmethod
    def _select_input_sheet_name(options: RunCreateOptions) -> str | None:
        """Resolve the first selected sheet name from the run options, if any."""

        if getattr(options, "active_sheet_only", False):
            return None
        normalized = RunsService._select_input_sheet_names(options)
        return normalized[0] if normalized else None

    # ------------------------------------------------------------------ #
    # Internal helpers
    # ------------------------------------------------------------------ #

    def settings(self) -> Settings:
        """Expose the bound settings instance for caller reuse."""

        return self._settings
