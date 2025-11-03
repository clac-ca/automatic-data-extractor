"""Manifest-driven pipeline execution for Jobs worker."""

from __future__ import annotations

import inspect
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .loader import (
    ConfigPackageLoader,
    LoadedColumnModule,
    LoadedHookModule,
    LoadedRowModule,
)


@dataclass(slots=True)
class SheetData:
    name: str
    rows: list[list[Any]]


@dataclass(slots=True)
class TableData:
    id: str
    sheet_name: str
    source_file: str
    header_row: int
    header_values: list[Any]
    data_rows: list[list[Any]]


@dataclass(slots=True)
class ColumnPlugin:
    field_name: str
    path: str
    detectors: list[Callable[..., Any]]
    transform: Callable[..., Any] | None
    validate: Callable[..., Any] | None
    field_meta: dict[str, Any]


@dataclass(slots=True)
class RowPlugin:
    path: str
    detectors: list[Callable[..., Any]]


@dataclass(slots=True)
class ColumnAssignment:
    column_id: str
    column_index: int
    header: str | None
    raw_values: list[Any]
    target_field: str | None
    confidence: float
    contributors: list[dict[str, Any]]
    transformed_values: list[Any]
    warnings: list[str]
    issues: list[dict[str, Any]]


@dataclass(slots=True)
class PipelineResult:
    table: TableData
    tables_summary: list[dict[str, Any]]
    assignments: list[ColumnAssignment]
    diagnostics: list[dict[str, Any]]
    artifact_snapshot: dict[str, Any]
    sheet_title: str


class PipelineRunner:
    """Execute manifest-driven pipeline inside the worker subprocess."""

    _ISSUE_CODE_ALIASES: dict[str, str] = {
        "missing": "required_missing",
        "required": "required_missing",
        "blank": "required_missing",
        "pattern": "pattern_mismatch",
        "regex_mismatch": "pattern_mismatch",
        "format": "invalid_format",
        "invalid": "invalid_format",
    }

    def __init__(
        self,
        *,
        config_dir: Path,
        manifest: dict[str, Any],
        job_context: dict[str, Any],
        input_paths: list[str],
    ) -> None:
        self._config_dir = config_dir.resolve()
        self._manifest = manifest
        self._job_context = job_context
        self._env = manifest.get("env") or {}
        self._loader = ConfigPackageLoader(self._config_dir)
        self._diagnostics: list[dict[str, Any]] = []
        self._hooks: dict[str, list[LoadedHookModule]] = {}
        self._row_plugins: list[RowPlugin] = []
        self._column_plugins: list[ColumnPlugin] = []
        self._column_plugins_by_field: dict[str, ColumnPlugin] = {}
        self._input_paths = [Path(path).resolve() for path in input_paths]
        engine_section = manifest.get("engine") if isinstance(manifest.get("engine"), dict) else {}
        writer_defaults = (engine_section.get("writer") or {}) if isinstance(engine_section, dict) else {}
        defaults_section = (engine_section.get("defaults") or {}) if isinstance(engine_section, dict) else {}
        self._artifact: dict[str, Any] = {
            "artifact_version": "1.1",
            "engine": self._build_engine_snapshot(writer_defaults, defaults_section),
            "rules": {},
            "sheets": [],
            "pass_history": [],
            "annotations": [],
        }
        self._sheet_title = str(writer_defaults.get("output_sheet") or "Normalized")
        self._sample_size = 8

    def execute(self) -> PipelineResult:
        """Run row detection, mapping, transform, and validation passes."""

        self._prime_modules()
        self._artifact["rules"] = self._rules_snapshot()

        self._run_hook_group("on_job_start", artifact=self._artifact, stage="on_job_start")

        sheets, source_file = self._load_sheet_data()
        tables_summary, table_data, sheet_artifacts = self._detect_tables(sheets, source_file)
        self._artifact["sheets"] = sheet_artifacts
        structure_stats = {
            "tables": len(tables_summary),
            "rows": sum(table.get("row_count", 0) for table in tables_summary),
            "columns": sum(table.get("column_count", 0) for table in tables_summary),
        }
        self._record_pass(name="structure", stats=structure_stats)
        self._run_hook_group("on_after_extract", artifact=self._artifact, stage="on_after_extract")

        assignments = self._map_columns(table_data)
        mapping_stats = {
            "mapped": sum(1 for assignment in assignments if assignment.target_field),
            "unmapped": sum(1 for assignment in assignments if assignment.target_field is None),
        }
        self._record_pass(name="mapping", stats=mapping_stats)
        self._run_hook_group("after_mapping", artifact=self._artifact, stage="after_mapping")

        transform_summary = self._transform_columns(assignments, table_data)
        changed_cells = sum(
            1
            for assignment in assignments
            if assignment.target_field
            for original, transformed in zip(assignment.raw_values, assignment.transformed_values)
            if original != transformed
        )
        fields_with_warnings = sum(
            1
            for assignment in assignments
            if assignment.target_field and assignment.warnings
        )
        transform_stats = {
            "changed_cells": changed_cells,
            "fields_with_warnings": fields_with_warnings,
        }
        self._record_pass(name="transform", stats=transform_stats)
        self._run_hook_group("after_transform", artifact=self._artifact, stage="after_transform")

        validate_summary = self._validate_columns(assignments, table_data)
        validation_summary = validate_summary.get("summary", {})
        validate_stats = {
            "errors": int(validation_summary.get("total_errors", 0)),
            "warnings": int(validation_summary.get("total_warnings", 0)),
        }
        self._record_pass(name="validate", stats=validate_stats)
        self._run_hook_group("after_validate", artifact=self._artifact, stage="after_validate")

        return PipelineResult(
            table=table_data,
            tables_summary=tables_summary,
            assignments=assignments,
            diagnostics=self._diagnostics,
            artifact_snapshot=deepcopy(self._artifact),
            sheet_title=self._sheet_title,
        )

    def run_job_end(self, artifact: dict[str, Any]) -> None:
        """Execute job end hooks once the worker has built the artifact."""

        self._run_hook_group("on_job_end", artifact=artifact, stage="on_job_end")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _prime_modules(self) -> None:
        try:
            loaded_columns = self._loader.load_column_modules(self._manifest)
            hooks = self._loader.load_hook_modules(self._manifest)
            rows = self._loader.load_row_type_modules()
        except Exception as exc:  # pragma: no cover - defensive
            self._add_diagnostic(
                level="error",
                code="runtime.module_import_error",
                message=str(exc),
            )
            raise

        columns_meta = (self._manifest.get("columns") or {}).get("meta", {}) or {}
        self._column_plugins = []
        for item in loaded_columns:
            detectors = [
                function
                for name, function in inspect.getmembers(item.module, inspect.isfunction)
                if name.startswith("detect_")
            ]
            transform = getattr(item.module, "transform", None)
            validate = getattr(item.module, "validate", None)
            plugin = ColumnPlugin(
                field_name=item.field_name,
                path=item.path,
                detectors=detectors,
                transform=transform,
                validate=validate,
                field_meta=columns_meta.get(item.field_name, {}),
            )
            self._column_plugins.append(plugin)
            self._column_plugins_by_field[item.field_name] = plugin

        self._row_plugins = []
        for row_module in rows:
            detectors = [
                function
                for name, function in inspect.getmembers(row_module.module, inspect.isfunction)
                if name.startswith("detect_")
            ]
            if detectors:
                self._row_plugins.append(RowPlugin(path=row_module.path, detectors=detectors))

        self._hooks = hooks

    def _load_sheet_data(self) -> tuple[list[SheetData], str]:
        if self._input_paths:
            primary = self._input_paths[0]
            if not primary.exists():
                raise FileNotFoundError(f"Input spreadsheet not found at {primary}")
            workbook = load_workbook(primary, data_only=True, read_only=True)
            sheets: list[SheetData] = []
            for worksheet in workbook.worksheets:
                rows: list[list[Any]] = []
                for row in worksheet.iter_rows(values_only=True):
                    rows.append([cell for cell in row])
                sheets.append(SheetData(name=worksheet.title, rows=rows))
            workbook.close()
            return sheets or [SheetData(name="Sheet1", rows=[])], primary.name

        # Fallback: synthesize a single sheet using manifest column labels.
        columns = self._manifest.get("columns") or {}
        order = list(columns.get("order") or [])
        meta = columns.get("meta") or {}
        header = [str(meta.get(field, {}).get("label") or field) for field in order]
        return [SheetData(name="Sheet1", rows=[header])], "generated"

    def _detect_tables(
        self, sheets: list[SheetData], source_file: str
    ) -> tuple[list[dict[str, Any]], TableData, list[dict[str, Any]]]:
        tables_summary: list[dict[str, Any]] = []
        sheet_artifacts: list[dict[str, Any]] = []
        primary_table: TableData | None = None

        for sheet in sheets:
            if not sheet.rows:
                sheet.rows.append([])

            header_row = self._find_header_row(sheet, source_file)
            header_values = sheet.rows[header_row - 1] if 0 <= header_row - 1 < len(sheet.rows) else []
            data_rows = sheet.rows[header_row:] if header_row < len(sheet.rows) else []
            table_id = f"{sheet.name or 'Sheet'}-table-1"

            table = TableData(
                id=table_id,
                sheet_name=sheet.name,
                source_file=source_file,
                header_row=header_row,
                header_values=list(header_values),
                data_rows=[list(row) for row in data_rows],
            )
            if primary_table is None:
                primary_table = table

            row_classification: list[dict[str, Any]] = []
            for row_index, row_values in enumerate(sheet.rows, start=1):
                if self._is_empty_row(row_values):
                    continue
                scores, contributions = self._evaluate_row_detectors(
                    row_index=row_index,
                    row_values=row_values,
                    sheet_name=sheet.name,
                    source_file=source_file,
                )
                if not scores and not contributions:
                    continue
                label, confidence = self._select_row_label(scores)
                row_classification.append(
                    {
                        "row_index": row_index,
                        "label": label,
                        "confidence": confidence,
                        "scores_by_type": dict(scores),
                        "rule_traces": contributions,
                    }
                )

            table_artifact = {
                "id": table_id,
                "range": self._table_range(table),
                "data_range": self._table_data_range(table),
                "header": {
                    "kind": "raw",
                    "row_index": header_row,
                    "source_header": [self._stringify(value) for value in table.header_values],
                },
                "columns": [
                    {
                        "column_id": self._column_identifier(table_id, index),
                        "source_header": self._stringify(header_value),
                        "order": index,
                    }
                    for index, header_value in enumerate(table.header_values, start=1)
                ],
                "mapping": [],
                "transforms": [],
                "validation": {"issues": [], "summary_by_field": {}},
            }

            sheet_artifacts.append(
                {
                    "id": f"sheet_{len(sheet_artifacts) + 1}",
                    "name": sheet.name,
                    "row_classification": row_classification,
                    "tables": [table_artifact],
                }
            )

            tables_summary.append(
                {
                    "id": table_id,
                    "sheet_name": sheet.name,
                    "source_file": source_file,
                    "header_row": header_row,
                    "data_row_start": header_row + 1,
                    "row_count": len(table.data_rows),
                    "column_count": len(table.header_values),
                    "header": [self._stringify(value) for value in table.header_values],
                }
            )

        if primary_table is None:
            # Guaranteed to exist because we populate fallback rows, but guard defensively.
            primary_table = TableData(
                id="Sheet1-table-1",
                sheet_name="Sheet1",
                source_file=source_file,
                header_row=1,
                header_values=[],
                data_rows=[],
            )

        return tables_summary, primary_table, sheet_artifacts

    def _find_header_row(self, sheet: SheetData, source_file: str) -> int:
        best_index = 1
        best_score = float("-inf")

        for row_index, row_values in enumerate(sheet.rows, start=1):
            if self._is_empty_row(row_values):
                continue
            scores, _ = self._evaluate_row_detectors(
                row_index=row_index,
                row_values=row_values,
                sheet_name=sheet.name,
                source_file=source_file,
            )
            header_score = scores.get("header", 0.0)
            if header_score > best_score:
                best_score = header_score
                best_index = row_index

        return max(best_index, 1)

    def _evaluate_row_detectors(
        self,
        *,
        row_index: int,
        row_values: list[Any],
        sheet_name: str,
        source_file: str,
    ) -> tuple[dict[str, float], list[dict[str, Any]]]:
        if not self._row_plugins:
            return {}, []

        scores: dict[str, float] = {}
        contributions: list[dict[str, Any]] = []
        context = {
            "job_id": self._job_context.get("job_id"),
            "source_file": source_file,
            "sheet_name": sheet_name,
            "row_index": row_index,
            "row_values_sample": row_values[: self._sample_size],
            "manifest": self._manifest,
            "env": self._env,
            "artifact": self._artifact,
        }

        for plugin in self._row_plugins:
            for detector in plugin.detectors:
                try:
                    result = self._invoke(detector, **context)
                except Exception as exc:  # pragma: no cover - defensive
                    self._add_diagnostic(
                        level="error",
                        code="runtime.row_detector.exception",
                        message=str(exc),
                        path=f"{plugin.path}:{detector.__name__}",
                        stage="detect_tables",
                    )
                    continue

                module_id = self._module_identifier(plugin.path)
                for label, delta in (result or {}).get("scores", {}).items():
                    try:
                        value = float(delta)
                    except (TypeError, ValueError):
                        continue
                    scores[label] = scores.get(label, 0.0) + value
                    contributions.append(
                        {
                            "rule": f"{module_id}:{detector.__name__}",
                            "label": label,
                            "delta": value,
                        }
                    )

        return scores, contributions

    def _map_columns(self, table: TableData) -> list[ColumnAssignment]:
        columns_section = self._manifest.get("columns") or {}
        order: list[str] = list(columns_section.get("order") or [])
        meta: dict[str, dict[str, Any]] = columns_section.get("meta") or {}
        score_threshold = float(
            self._manifest.get("engine", {})
            .get("defaults", {})
            .get("mapping_score_threshold", 0.0)
        )

        column_values: list[ColumnAssignment] = []
        header_values = table.header_values or []
        row_count = len(table.data_rows)
        for index, header_value in enumerate(header_values, start=1):
            values: list[Any] = []
            for row in table.data_rows:
                value = row[index - 1] if index - 1 < len(row) else None
                values.append(value)

            column_id = self._column_identifier(table.id, index)
            assignment = ColumnAssignment(
                column_id=column_id,
                column_index=index,
                header=self._stringify(header_value),
                raw_values=values,
                target_field=None,
                confidence=0.0,
                contributors=[],
                transformed_values=list(values),
                warnings=[],
                issues=[],
            )

            scores: dict[str, float] = {}
            contributors: list[dict[str, Any]] = []
            column_context = {
                "job_id": self._job_context.get("job_id"),
                "source_file": table.source_file,
                "sheet_name": table.sheet_name,
                "table": self._table_context(table),
                "column_index": index,
                "header": assignment.header,
                "values_sample": assignment.raw_values[: self._sample_size],
                "manifest": self._manifest,
                "env": self._env,
                "job_context": self._job_context,
            }

            for plugin in self._column_plugins:
                module_id = self._module_identifier(plugin.path)
                column_context["field_name"] = plugin.field_name
                column_context["field_meta"] = plugin.field_meta
                for detector in plugin.detectors:
                    try:
                        result = self._invoke(detector, **column_context)
                    except Exception as exc:  # pragma: no cover - defensive
                        self._add_diagnostic(
                            level="error",
                            code="runtime.column_detector.exception",
                            message=str(exc),
                            path=f"{plugin.path}:{detector.__name__}",
                            stage="map_columns",
                        )
                        continue
                    for target, delta in (result or {}).get("scores", {}).items():
                        try:
                            value = float(delta)
                        except (TypeError, ValueError):
                            continue
                        scores[target] = scores.get(target, 0.0) + value
                        contributors.append(
                            {
                                "target": target,
                                "delta": value,
                                "rule": f"{module_id}:{detector.__name__}",
                            }
                        )

            if scores:
                best_score = max(scores.values())
                candidates = [field for field, score in scores.items() if score == best_score]
                selected_field = self._select_best_field(candidates, assignment.header, meta, order)
                if selected_field and best_score >= score_threshold:
                    assignment.target_field = selected_field
                    assignment.confidence = best_score
                    assignment.contributors = [
                        contributor
                        for contributor in contributors
                        if contributor["target"] == selected_field
                    ]

            column_values.append(assignment)

        table_artifact = self._find_table_artifact(table.id)
        if table_artifact is not None:
            table_artifact["mapping"] = [
                {
                    "raw": {"column": assignment.column_id, "header": assignment.header},
                    "target_field": assignment.target_field,
                    "score": assignment.confidence,
                    "contributors": [
                        {
                            "rule": contributor.get("rule"),
                            "delta": contributor.get("delta"),
                        }
                        for contributor in assignment.contributors
                    ],
                }
                for assignment in column_values
            ]

        return column_values

    def _transform_columns(self, assignments: list[ColumnAssignment], table: TableData) -> dict[str, Any]:
        table_artifact = self._find_table_artifact(table.id)
        transform_entries: list[dict[str, Any]] = []
        for assignment in assignments:
            if not assignment.target_field:
                continue
            plugin = self._column_plugins_by_field.get(assignment.target_field)
            if plugin is None or plugin.transform is None:
                continue

            module_id = self._module_identifier(plugin.path)
            context = {
                "job_id": self._job_context.get("job_id"),
                "source_file": table.source_file,
                "sheet_name": table.sheet_name,
                "table": self._table_context(table),
                "column_index": assignment.column_index,
                "header": assignment.header,
                "values": list(assignment.transformed_values),
                "field_name": assignment.target_field,
                "field_meta": plugin.field_meta,
                "manifest": self._manifest,
                "env": self._env,
                "job_context": self._job_context,
            }
            try:
                result = self._invoke(plugin.transform, **context) or {}
            except Exception as exc:  # pragma: no cover - defensive
                self._add_diagnostic(
                    level="error",
                    code="runtime.transform.exception",
                    message=str(exc),
                    path=f"{plugin.path}:transform",
                    stage="transform_values",
                )
                continue

            values = list(result.get("values", assignment.transformed_values))
            assignment.transformed_values = values

            warnings = result.get("warnings") or []
            if isinstance(warnings, (str, bytes)):
                warnings = [warnings]
            assignment.warnings = [self._stringify(item) for item in warnings if item]

            changed = sum(
                1
                for original, transformed in zip(assignment.raw_values, assignment.transformed_values)
                if original != transformed
            )
            entry = {
                "target_field": assignment.target_field,
                "transform": f"{module_id}:transform",
                "changed": changed,
                "total": len(assignment.transformed_values),
            }
            if assignment.warnings:
                entry["warnings"] = list(assignment.warnings)
                entry["notes"] = "; ".join(assignment.warnings)
            transform_entries.append(entry)

        summary = {
            "name": "transform_values",
            "status": "succeeded",
            "summary": {
                "fields": [
                    {
                        "field": assignment.target_field,
                        "warnings": assignment.warnings,
                    }
                    for assignment in assignments
                    if assignment.target_field
                ]
            },
        }
        if table_artifact is not None:
            table_artifact["transforms"] = transform_entries
        return summary

    def _validate_columns(self, assignments: list[ColumnAssignment], table: TableData) -> dict[str, Any]:
        total_errors = 0
        total_warnings = 0
        aggregated_issues: list[dict[str, Any]] = []
        summary_by_field: dict[str, dict[str, int]] = {}
        table_artifact = self._find_table_artifact(table.id)
        for assignment in assignments:
            if not assignment.target_field:
                continue
            plugin = self._column_plugins_by_field.get(assignment.target_field)
            if plugin is None or plugin.validate is None:
                continue

            module_id = self._module_identifier(plugin.path)
            context = {
                "job_id": self._job_context.get("job_id"),
                "source_file": table.source_file,
                "sheet_name": table.sheet_name,
                "table": self._table_context(table),
                "column_index": assignment.column_index,
                "header": assignment.header,
                "values": list(assignment.transformed_values),
                "field_name": assignment.target_field,
                "field_meta": plugin.field_meta,
                "manifest": self._manifest,
                "env": self._env,
                "job_context": self._job_context,
            }
            try:
                result = self._invoke(plugin.validate, **context) or {}
            except Exception as exc:  # pragma: no cover - defensive
                self._add_diagnostic(
                    level="error",
                    code="runtime.validate.exception",
                    message=str(exc),
                    path=f"{plugin.path}:validate",
                    stage="validate_values",
                )
                continue

            issues = result.get("issues") or []
            normalized_issues: list[dict[str, Any]] = []
            for issue in issues:
                if isinstance(issue, dict):
                    entry = issue.copy()
                else:
                    entry = {"message": self._stringify(issue)}
                entry.setdefault("severity", "error")
                entry.setdefault("target_field", assignment.target_field)
                entry.setdefault("column", assignment.column_id)
                if plugin and plugin.validate is not None:
                    entry.setdefault("rule", f"{module_id}:validate")
                canonical_code = self._canonical_issue_code(entry.get("code"))
                if canonical_code:
                    entry["code"] = canonical_code
                elif entry.get("code") is not None:
                    entry["code"] = str(entry.get("code", "")).strip()
                row_index_value = entry.get("row_index")
                if (
                    isinstance(row_index_value, int)
                    and row_index_value >= 1
                    and not entry.get("a1")
                ):
                    entry["a1"] = self._issue_a1(
                        table=table,
                        column_index=assignment.column_index,
                        row_index=row_index_value,
                    )
                normalized_issues.append(entry)
                if entry["severity"] == "warning":
                    total_warnings += 1
                else:
                    total_errors += 1
            assignment.issues = normalized_issues
            summary_entry = summary_by_field.setdefault(
                assignment.target_field,
                {"errors": 0, "warnings": 0, "missing": 0},
            )
            for issue_entry in normalized_issues:
                aggregated_issues.append(issue_entry)
                if issue_entry.get("severity") == "warning":
                    summary_entry["warnings"] += 1
                else:
                    summary_entry["errors"] += 1
                if issue_entry.get("code") == "required_missing":
                    summary_entry["missing"] += 1

        summary = {
            "name": "validate_values",
            "status": "succeeded",
            "summary": {
                "total_errors": total_errors,
                "total_warnings": total_warnings,
                "fields": [
                    {
                        "field": assignment.target_field,
                        "issue_count": len(assignment.issues),
                    }
                    for assignment in assignments
                    if assignment.target_field
                ],
            },
        }
        if table_artifact is not None:
            table_artifact["validation"] = {
                "issues": aggregated_issues,
                "summary_by_field": summary_by_field,
            }
        return summary

    def _run_hook_group(self, name: str, *, artifact: dict[str, Any], stage: str) -> None:
        for hook in self._hooks.get(name, []):
            run_fn = getattr(hook.module, "run", None)
            if not callable(run_fn):
                self._add_diagnostic(
                    level="warning",
                    code="runtime.hook.missing_run",
                    message=f"Hook {hook.path} does not define run()",
                    path=hook.path,
                    stage=stage,
                )
                continue
            context = {
                "job_context": self._job_context,
                "manifest": self._manifest,
                "env": self._env,
                "artifact": deepcopy(artifact),
            }
            try:
                result = self._invoke(run_fn, **context)
            except Exception as exc:  # pragma: no cover - defensive
                self._add_diagnostic(
                    level="error",
                    code="runtime.hook.exception",
                    message=str(exc),
                    path=f"{hook.path}:run",
                    stage=stage,
                )
                continue

            if isinstance(result, dict) and result:
                annotation = {
                    "stage": stage,
                    "hook": hook.path,
                    "annotated_at": datetime.now(timezone.utc).isoformat(),
                }
                annotation.update(result)
                artifact.setdefault("annotations", []).append(annotation)

    # ------------------------------------------------------------------
    # Utility helpers
    # ------------------------------------------------------------------

    def _invoke(self, func: Callable[..., Any], **context: Any) -> Any:
        signature = None
        try:
            signature = inspect.signature(func)
        except (TypeError, ValueError):
            return func(**context)

        accepts_kwargs = any(
            parameter.kind == inspect.Parameter.VAR_KEYWORD for parameter in signature.parameters.values()
        )
        bound_arguments: dict[str, Any] = {}
        for name, parameter in signature.parameters.items():
            if name == "self":
                continue
            if parameter.kind in (inspect.Parameter.VAR_KEYWORD, inspect.Parameter.VAR_POSITIONAL):
                continue
            if name in context:
                bound_arguments[name] = context[name]
            elif parameter.default is inspect._empty and parameter.kind in (
                inspect.Parameter.POSITIONAL_OR_KEYWORD,
                inspect.Parameter.KEYWORD_ONLY,
            ):
                raise TypeError(f"Missing argument {name} for function {func.__name__}")
        if accepts_kwargs:
            for key, value in context.items():
                if key not in bound_arguments:
                    bound_arguments[key] = value
        return func(**bound_arguments)

    def _table_context(self, table: TableData) -> dict[str, Any]:
        return {
            "id": table.id,
            "sheet_name": table.sheet_name,
            "header_row": table.header_row,
            "data_row_start": table.header_row + 1,
            "row_count": len(table.data_rows),
            "column_count": len(table.header_values),
        }

    def _select_best_field(
        self,
        candidates: list[str],
        header: str | None,
        meta: dict[str, dict[str, Any]],
        order: list[str],
    ) -> str | None:
        valid_candidates = [field for field in candidates if field in order]
        if not valid_candidates:
            return None

        header_normalized = (header or "").strip().lower()
        if header_normalized:
            for field in valid_candidates:
                label = str(meta.get(field, {}).get("label") or field).strip().lower()
                if label == header_normalized:
                    return field
            for field in valid_candidates:
                synonyms = [str(value).strip().lower() for value in meta.get(field, {}).get("synonyms", [])]
                if header_normalized in synonyms:
                    return field

        # Fall back to the candidate that appears first in manifest order.
        for field in order:
            if field in valid_candidates:
                return field

        return valid_candidates[0]

    def _select_row_label(self, scores: dict[str, float]) -> tuple[str | None, float]:
        if not scores:
            return None, 0.0
        label = max(scores, key=scores.get)
        return label, float(scores.get(label, 0.0))

    def _column_identifier(self, table_id: str, index: int) -> str:
        return f"{table_id}.col.{index}"

    def _find_table_artifact(self, table_id: str) -> dict[str, Any] | None:
        for sheet in self._artifact.get("sheets", []):
            for table in sheet.get("tables", []):
                if table.get("id") == table_id:
                    return table
        return None

    def _column_letter(self, index: int) -> str:
        if index <= 0:
            return "A"
        letters: list[str] = []
        current = index
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            letters.append(chr(65 + remainder))
        return "".join(reversed(letters))

    def _issue_a1(self, *, table: TableData, column_index: int, row_index: int) -> str:
        row_number = max(table.header_row + row_index, 1)
        return f"{self._column_letter(column_index)}{row_number}"

    def _canonical_issue_code(self, code: Any) -> str | None:
        if code is None:
            return None
        text = str(code).strip()
        if not text:
            return None
        return self._ISSUE_CODE_ALIASES.get(text.lower(), text)

    def _table_range(self, table: TableData) -> str:
        column_count = max(len(table.header_values), 1)
        start_col = self._column_letter(1)
        end_col = self._column_letter(column_count)
        start_row = max(table.header_row, 1)
        end_row = max(table.header_row + len(table.data_rows), start_row)
        return f"{start_col}{start_row}:{end_col}{end_row}"

    def _table_data_range(self, table: TableData) -> str | None:
        if not table.data_rows:
            return None
        column_count = max(len(table.header_values), 1)
        start_col = self._column_letter(1)
        end_col = self._column_letter(column_count)
        start_row = max(table.header_row + 1, 1)
        end_row = max(table.header_row + len(table.data_rows), start_row)
        return f"{start_col}{start_row}:{end_col}{end_row}"

    def _build_engine_snapshot(
        self,
        writer_defaults: dict[str, Any],
        engine_defaults: dict[str, Any],
    ) -> dict[str, Any]:
        writer_section = {
            "mode": writer_defaults.get("mode", "row_streaming"),
            "append_unmapped_columns": writer_defaults.get("append_unmapped_columns", True),
            "unmapped_prefix": writer_defaults.get("unmapped_prefix", "raw_"),
            "output_sheet": writer_defaults.get("output_sheet"),
        }
        defaults_section = {
            "timeout_ms": engine_defaults.get("timeout_ms"),
            "memory_mb": engine_defaults.get("memory_mb"),
            "runtime_network_access": engine_defaults.get("runtime_network_access"),
            "mapping_score_threshold": engine_defaults.get("mapping_score_threshold"),
        }
        defaults_section = {key: value for key, value in defaults_section.items() if value is not None}

        snapshot: dict[str, Any] = {"writer": writer_section}
        if defaults_section:
            snapshot["defaults"] = defaults_section
        return snapshot

    def _rules_snapshot(self) -> dict[str, Any]:
        row_rules: dict[str, dict[str, Any]] = {}
        for plugin in self._row_plugins:
            module_id = self._module_identifier(plugin.path)
            for detector in plugin.detectors:
                rule_id = f"{module_id}.{detector.__name__}"
                row_rules[rule_id] = {
                    "impl": f"{module_id}:{detector.__name__}",
                }

        column_detect_rules: dict[str, dict[str, Any]] = {}
        transform_rules: dict[str, dict[str, Any]] = {}
        validate_rules: dict[str, dict[str, Any]] = {}
        for plugin in self._column_plugins:
            module_id = self._module_identifier(plugin.path)
            for detector in plugin.detectors:
                rule_id = f"{module_id}.{detector.__name__}"
                column_detect_rules[rule_id] = {
                    "impl": f"{module_id}:{detector.__name__}",
                    "field": plugin.field_name,
                }
            if plugin.transform is not None:
                transform_rules[f"{module_id}.transform"] = {
                    "impl": f"{module_id}:transform",
                    "field": plugin.field_name,
                }
            if plugin.validate is not None:
                validate_rules[f"{module_id}.validate"] = {
                    "impl": f"{module_id}:validate",
                    "field": plugin.field_name,
                }

        return {
            "row_types": row_rules,
            "column_detect": column_detect_rules,
            "transform": transform_rules,
            "validate": validate_rules,
        }

    def _record_pass(self, name: str, *, stats: dict[str, Any] | None = None) -> None:
        history = self._artifact.setdefault("pass_history", [])
        entry = {
            "pass": len(history) + 1,
            "name": name,
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }
        if stats:
            entry["stats"] = stats
        history.append(entry)

    def _module_identifier(self, script_path: str) -> str:
        normalized = script_path.replace("\\", "/")
        if normalized.endswith(".py"):
            normalized = normalized[:-3]
        return normalized.replace("/", ".")

    @staticmethod
    def _stringify(value: Any) -> str | None:
        if value is None:
            return None
        return str(value)

    @staticmethod
    def _is_empty_row(row: list[Any]) -> bool:
        return all(value in (None, "") for value in row)

    def _add_diagnostic(
        self,
        *,
        level: str,
        code: str,
        message: str,
        path: str | None = None,
        stage: str | None = None,
    ) -> None:
        diagnostic = {
            "level": level,
            "code": code,
            "message": message,
        }
        if path:
            diagnostic["path"] = path
        if stage:
            diagnostic["stage"] = stage
        self._diagnostics.append(diagnostic)


__all__ = ["PipelineResult", "PipelineRunner", "ColumnAssignment"]
