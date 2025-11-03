"""Sandboxed job runner executed as a subprocess."""

from __future__ import annotations

import json
import os
import resource
import socket
import sys
import traceback
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

CURRENT_FILE = Path(__file__).resolve()
PROJECT_ROOT = CURRENT_FILE.parents[4]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from openpyxl import load_workbook, Workbook
except ModuleNotFoundError as exc:  # pragma: no cover - dependency should exist
    raise SystemExit(f"openpyxl is required for job execution: {exc}")

from backend.app.features.jobs.runtime import PipelineRunner


def main() -> None:
    try:
        request = _read_request()
        apply_resource_limits()
        result = run_job(request)
    except Exception as exc:  # pragma: no cover - defensive path
        diagnostics = [
            {
                "level": "error",
                "code": "exception",
                "message": f"Unhandled error: {exc}",
            }
        ]
        result = {
            "schema": "ade.run_result/v1",
            "status": "failed",
            "diagnostics": diagnostics,
            "error_message": str(exc),
        }
        traceback.print_exc()
    sys.stdout.write(json.dumps(result))
    sys.stdout.flush()


def _read_request() -> dict[str, Any]:
    raw = sys.stdin.read()
    if not raw:
        raise ValueError("Empty run request")
    payload = json.loads(raw)
    if payload.get("schema") != "ade.run_request/v1":
        raise ValueError("Unsupported run request schema")
    return payload


def apply_resource_limits() -> None:
    cpu_seconds = int(os.environ.get("ADE_WORKER_CPU_SECONDS", "60"))
    mem_limit = int(os.environ.get("ADE_WORKER_MEM_MB", "512")) * 1024 * 1024
    try:
        resource.setrlimit(resource.RLIMIT_CPU, (cpu_seconds, cpu_seconds))
    except Exception:  # pragma: no cover - may not be supported
        pass
    try:
        resource.setrlimit(resource.RLIMIT_AS, (mem_limit, mem_limit))
    except Exception:  # pragma: no cover
        pass


def disable_network(*, allow: bool) -> None:
    if allow:
        return

    def _blocked(*_args: Any, **_kwargs: Any) -> None:
        raise ConnectionError("Networking is disabled for this job")

    socket.socket = lambda *a, **k: (_blocked(*a, **k))  # type: ignore[assignment]
    socket.create_connection = lambda *a, **k: (_blocked(*a, **k))  # type: ignore[assignment]


def run_job(request: dict[str, Any]) -> dict[str, Any]:
    manifest_path = Path(request["manifest_path"]).resolve()
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest not found at {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    engine_section = manifest.get("engine", {}) if isinstance(manifest, dict) else {}
    defaults_section = (
        engine_section.get("defaults", {}) if isinstance(engine_section, dict) else {}
    )
    manifest_allows_network = bool(defaults_section.get("runtime_network_access", False))
    env_override = os.environ.get("ADE_RUNTIME_NETWORK_ACCESS")
    if env_override is None:
        env_override = os.environ.get("ADE_ALLOW_NETWORK")
    if env_override is not None:
        allow_network = env_override.lower() in {"1", "true", "yes", "on"}
    else:
        allow_network = manifest_allows_network

    disable_network(allow=allow_network)

    input_paths = [str(path) for path in request.get("input_paths", [])]

    job_context = {
        "job_id": request.get("job_id"),
        "config_version_id": request.get("config_version_id"),
        "trace_id": os.environ.get("ADE_TRACE_ID"),
        "input_paths": input_paths,
    }

    pipeline = PipelineRunner(
        config_dir=manifest_path.parent,
        manifest=manifest,
        job_context=job_context,
        input_paths=input_paths,
    )
    pipeline_result = pipeline.execute()

    sheet_title = pipeline_result.sheet_title

    work_dir = Path(request["work_dir"]) if request.get("work_dir") else manifest_path.parent
    artifact_path = Path(os.environ.get("ADE_ARTIFACT_PATH", work_dir / "artifact.json"))
    output_path = Path(os.environ.get("ADE_OUTPUT_PATH", work_dir / "normalized.xlsx"))

    started_at = datetime.now(timezone.utc)

    writer_config = engine_section.get("writer", {}) if isinstance(engine_section, dict) else {}
    append_unmapped = bool(writer_config.get("append_unmapped_columns", True))
    unmapped_prefix = str(writer_config.get("unmapped_prefix", "raw_"))

    columns_section = manifest.get("columns", {}) if isinstance(manifest, dict) else {}
    manifest_order = list(columns_section.get("order", []) or [])
    manifest_meta = columns_section.get("meta", {}) or {}
    enabled_fields = [field for field in manifest_order if manifest_meta.get(field, {}).get("enabled", True)]
    header_labels = [str(manifest_meta.get(field, {}).get("label") or field) for field in enabled_fields]

    target_plan = [
        {
            "field": field,
            "output_header": header_labels[index],
            "order": index + 1,
        }
        for index, field in enumerate(enabled_fields)
    ]

    field_assignments = {
        assignment.target_field: assignment
        for assignment in pipeline_result.assignments
        if assignment.target_field
    }

    unmapped_assignments = [
        assignment for assignment in pipeline_result.assignments if assignment.target_field is None
    ]

    row_count = 0
    for assignment in field_assignments.values():
        row_count = max(row_count, len(assignment.transformed_values))
    for assignment in unmapped_assignments:
        row_count = max(row_count, len(assignment.raw_values))

    appended_plan: list[dict[str, Any]] = []
    appended_headers: list[str] = []
    if append_unmapped and unmapped_assignments:
        used_headers = {label for label in header_labels}

        def build_output_header(assignment: Any) -> None:
            base = assignment.header if assignment.header else f"column_{assignment.column_index}"
            base = " ".join(str(base).split())
            sanitized = "".join(char if char.isalnum() else "_" for char in base).strip("_")
            if not sanitized:
                sanitized = f"col_{assignment.column_index}"
            candidate = f"{unmapped_prefix}{sanitized}"

            counter = 1
            final_candidate = candidate
            while final_candidate in used_headers:
                suffix = f"_{counter}"
                final_candidate = f"{candidate}{suffix}"
                counter += 1
            used_headers.add(final_candidate)
            appended_plan.append(
                {
                    "source_header": assignment.header,
                    "output_header": final_candidate,
                    "order": len(appended_plan) + 1,
                    "column": assignment.column_id,
                }
            )
            appended_headers.append(final_candidate)

        for assignment in unmapped_assignments:
            build_output_header(assignment)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(header_labels + appended_headers)

    for row_index in range(row_count):
        output_row: list[Any] = []
        for field in enabled_fields:
            assignment = field_assignments.get(field)
            value = None
            if assignment and row_index < len(assignment.transformed_values):
                value = assignment.transformed_values[row_index]
            output_row.append(value)
        if append_unmapped and unmapped_assignments:
            for assignment in unmapped_assignments:
                value = None
                if row_index < len(assignment.raw_values):
                    value = assignment.raw_values[row_index]
                output_row.append(value)
        sheet.append(output_row)

    workbook.save(output_path)
    workbook.close()

    completed_at = datetime.now(timezone.utc)
    artifact = deepcopy(pipeline_result.artifact_snapshot)
    artifact.setdefault("annotations", [])
    pass_history = artifact.setdefault("pass_history", [])
    pass_history.append(
        {
            "pass": len(pass_history) + 1,
            "name": "generate",
            "completed_at": completed_at.isoformat(),
            "stats": {
                "rows_written": row_count,
                "columns_written": len(target_plan) + len(appended_plan),
            },
        }
    )

    issues_found = sum(len(assignment.issues) for assignment in pipeline_result.assignments)
    artifact.update(
        {
            "schema": "ade.artifact/v1",
            "job": {
                "job_id": request.get("job_id"),
                "config_version_id": request.get("config_version_id"),
                "trace_id": os.environ.get("ADE_TRACE_ID"),
                "status": "succeeded",
                "started_at": started_at.isoformat(),
                "completed_at": completed_at.isoformat(),
                "source_file": pipeline_result.table.source_file,
            },
            "config": {
                "config_version_id": request.get("config_version_id"),
                "manifest_version": (manifest.get("info") or {}).get("version"),
            },
            "output": {
                "format": "xlsx",
                "sheet": sheet_title,
                "path": str(output_path),
                "column_plan": {
                    "target": target_plan,
                    "appended_unmapped": appended_plan,
                },
            },
            "summary": {
                "rows_written": row_count,
                "columns_written": len(target_plan) + len(appended_plan),
                "issues_found": issues_found,
            },
        }
    )

    pipeline.run_job_end(artifact)

    artifact_path.write_text(json.dumps(artifact, indent=2), encoding="utf-8")

    return {
        "schema": "ade.run_result/v1",
        "status": "succeeded",
        "artifact_path": str(artifact_path),
        "output_path": str(output_path),
        "diagnostics": pipeline_result.diagnostics,
    }

if __name__ == "__main__":
    main()
