from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from ade_api.common.workbook_preview import build_workbook_preview_from_xlsx


def test_xlsx_preview_reports_hidden_rows_and_columns(tmp_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["visible", "hidden column", "visible"])
    sheet.append(["hidden row", "hidden row", "hidden row"])
    sheet.append(["visible", "visible", "visible"])
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions["B"].hidden = True
    sheet["A1"].fill = PatternFill("solid", fgColor="FF008000")
    sheet["A1"].font = Font(color="FFFFFFFF", bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", wrap_text=True)

    path = tmp_path / "hidden.xlsx"
    workbook.save(path)

    preview = build_workbook_preview_from_xlsx(path)

    assert preview.hidden_rows == [1]
    assert preview.hidden_columns == [1]
    assert preview.cell_formats[0].row == 0
    assert preview.cell_formats[0].column == 0
    assert preview.cell_formats[0].bg_color == "#008000"
    assert preview.cell_formats[0].text_color == "#FFFFFF"
    assert preview.cell_formats[0].bold is True
    assert preview.cell_formats[0].horizontal_align == "center"
    assert preview.cell_formats[0].wrap_text is True


def test_xlsx_preview_treats_excel_argb_fills_as_opaque_rgb(tmp_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "member_id"
    sheet["A1"].fill = PatternFill("solid", fgColor="0033CCCC")

    path = tmp_path / "argb-fill.xlsx"
    workbook.save(path)

    preview = build_workbook_preview_from_xlsx(path)

    assert preview.cell_formats[0].bg_color == "#33CCCC"
