import asyncio
import io
import json
from datetime import timedelta
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import pytest
from fastapi import FastAPI
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import func, select

from backend.app.features.jobs.models import Job, JobStatus
from backend.app.features.jobs.repository import JobsRepository
from backend.app.shared.core.config import get_settings
from backend.app.shared.core.time import utc_now
from backend.app.shared.db.session import get_sessionmaker
from backend.tests.utils import login

pytestmark = pytest.mark.asyncio


async def _auth(async_client: AsyncClient, identity: dict[str, Any]) -> tuple[dict[str, str], str]:
    owner = identity["workspace_owner"]
    token, _ = await login(
        async_client,
        email=owner["email"],  # type: ignore[index]
        password=owner["password"],  # type: ignore[index]
    )
    return {"Authorization": f"Bearer {token}"}, identity["workspace_id"]  # type: ignore[return-value]


def _csrf_headers(client: AsyncClient) -> dict[str, str]:
    settings = get_settings()
    token = client.cookies.get(settings.session_csrf_cookie_name)
    assert token, "CSRF cookie missing"
    return {"X-CSRF-Token": token}


def _manifest(title: str, fields: list[str]) -> dict[str, Any]:
    return {
        "config_script_api_version": "1",
        "info": {
            "schema": "ade.manifest/v1.0",
            "title": title,
            "version": "1.0.0",
        },
        "engine": {
            "defaults": {
                "timeout_ms": 120000,
                "memory_mb": 256,
                "runtime_network_access": False,
                "mapping_score_threshold": 0.0,
            },
            "writer": {
                "mode": "row_streaming",
                "append_unmapped_columns": True,
                "unmapped_prefix": "raw_",
                "output_sheet": "Normalized",
            },
        },
        "env": {},
        "hooks": {
            "on_activate": [],
            "on_job_start": [],
            "on_after_extract": [],
            "on_job_end": [],
        },
        "columns": {
            "order": fields,
            "meta": {
                field: {
                    "label": field.title(),
                    "required": True,
                    "enabled": True,
                    "script": f"columns/{field}.py",
                }
                for field in fields
            },
        },
    }


def _package(
    manifest: dict[str, Any],
    name: str,
    *,
    slow_iterations: int | None = None,
) -> bytes:
    buffer = io.BytesIO()
    with ZipFile(buffer, "w") as archive:
        archive.writestr("manifest.json", json.dumps(manifest, indent=2))
        archive.writestr("columns/__init__.py", "")
        for field in manifest["columns"]["order"]:
            slow_body = (
                f"    for _ in range({slow_iterations}):\n        pass\n"
                if slow_iterations
                else ""
            )
            archive.writestr(
                f"columns/{field}.py",
                (
                    "def detect_default(*, header, values_sample, column_index, table, job_context, env):\n"
                    + slow_body
                    + f"    return {{'scores': {{'{field}': 1.0}}}}\n\n"
                    + "def transform(*, header, values, column_index, table, job_context, env):\n"
                    + "    return {'values': list(values), 'warnings': []}\n"
                ),
            )
    return buffer.getvalue()


async def _create_config(
    async_client: AsyncClient,
    headers: dict[str, str],
    workspace_id: str,
    *,
    slow_iterations: int | None = None,
) -> tuple[str, dict[str, Any], str]:
    base = f"/api/v1/workspaces/{workspace_id}/configs"
    manifest = _manifest("Job Config", ["member_id", "name"])
    package_bytes = _package(manifest, "job-config", slow_iterations=slow_iterations)
    created = await async_client.post(
        base,
        headers=headers,
        data={
            "slug": "job-config",
            "title": "Job Config",
            "manifest_json": json.dumps(manifest),
        },
        files={"package": ("job-config.zip", package_bytes, "application/zip")},
    )
    created.raise_for_status()
    detail = created.json()
    version = detail["versions"][0]
    return version["config_version_id"], manifest, detail["config_id"]


async def _upload_document(
    async_client: AsyncClient,
    headers: dict[str, str],
    workspace_id: str,
    workbook: Workbook,
    filename: str = "input.xlsx",
) -> str:
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = await async_client.post(
        f"/api/v1/workspaces/{workspace_id}/documents",
        headers=headers,
        files={
            "file": (
                filename,
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    response.raise_for_status()
    return response.json()["document_id"]


async def _wait_for_job(
    async_client: AsyncClient,
    headers: dict[str, str],
    workspace_id: str,
    job_id: str,
    *,
    timeout: float = 10.0,
) -> dict[str, Any]:
    base = f"/api/v1/workspaces/{workspace_id}/jobs/{job_id}"
    deadline = asyncio.get_event_loop().time() + timeout
    while True:
        detail = await async_client.get(base, headers=headers)
        detail.raise_for_status()
        payload = detail.json()
        if payload["status"] in {"succeeded", "failed"}:
            return payload
        if asyncio.get_event_loop().time() >= deadline:
            raise AssertionError(f"Job {job_id} did not complete in time")
        await asyncio.sleep(0.1)


async def test_job_submission_produces_artifacts(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, manifest, _ = await _create_config(async_client, headers, workspace_id)

    base = f"/api/v1/workspaces/{workspace_id}/jobs"
    submitted = await async_client.post(
        base,
        headers=headers,
        json={"config_version_id": version_id},
    )
    assert submitted.status_code == 202, submitted.text
    job_payload = submitted.json()
    job_id = job_payload["job_id"]
    assert submitted.headers["Location"].endswith(f"/jobs/{job_id}")

    detail_payload = await _wait_for_job(async_client, headers, workspace_id, job_id)
    assert detail_payload["status"] == "succeeded"
    assert detail_payload["artifact_uri"].endswith("artifact.json")
    assert detail_payload["output_uri"].endswith("normalized.xlsx")

    assert detail_payload["config_version"]["config_version_id"] == version_id

    artifact = await async_client.get(f"{base}/{job_id}/artifact", headers=headers)
    artifact.raise_for_status()
    artifact_payload = artifact.json()
    assert artifact_payload["config"]["config_version_id"] == version_id

    output = await async_client.get(f"{base}/{job_id}/output", headers=headers)
    output.raise_for_status()
    assert output.content
    assert output.headers["content-type"] in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }


async def test_run_request_records_activation_python(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, config_id = await _create_config(async_client, headers, workspace_id)

    activation_headers = {**headers, **_csrf_headers(async_client)}
    activate = await async_client.post(
        f"/api/v1/workspaces/{workspace_id}/configs/{config_id}/versions/{version_id}/activate",
        headers=activation_headers,
    )
    assert activate.status_code == 200, activate.text

    submitted = await async_client.post(
        f"/api/v1/workspaces/{workspace_id}/jobs",
        headers=headers,
        json={"config_version_id": version_id},
    )
    assert submitted.status_code == 202, submitted.text
    job_id = submitted.json()["job_id"]

    detail_payload = await _wait_for_job(async_client, headers, workspace_id, job_id)
    activation = detail_payload["config_version"].get("activation")
    assert activation is not None
    assert activation["status"] == "succeeded"

    run_request = json.loads(Path(detail_payload["run_request_uri"]).read_text(encoding="utf-8"))
    assert run_request["python_executable"] == activation["python_executable"]
    assert activation["python_executable"] and Path(activation["python_executable"]).exists()


async def test_job_submission_with_document_input(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, _ = await _create_config(async_client, headers, workspace_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Member ID", "First Name"])
    sheet.append(["1001", "Alice"])
    document_id = await _upload_document(async_client, headers, workspace_id, workbook)

    jobs_endpoint = f"/api/v1/workspaces/{workspace_id}/jobs"
    payload = {"config_version_id": version_id, "document_ids": [document_id]}
    submitted = await async_client.post(jobs_endpoint, headers=headers, json=payload)
    assert submitted.status_code == 202, submitted.text
    job_payload = submitted.json()
    job_id = job_payload["job_id"]
    completed = await _wait_for_job(async_client, headers, workspace_id, job_id)
    assert completed["status"] == "succeeded"
    assert completed["input_hash"]

    # Resubmitting with the same document should reuse the existing job based on the derived hash.
    resubmitted = await async_client.post(jobs_endpoint, headers=headers, json=payload)
    assert resubmitted.status_code == 202, resubmitted.text
    resubmitted_payload = resubmitted.json()
    assert resubmitted_payload["job_id"] == job_id
    assert resubmitted_payload["status"] == "succeeded"

    run_request = json.loads(Path(completed["run_request_uri"]).read_text(encoding="utf-8"))
    assert run_request["input_paths"]
    assert run_request["input_documents"]
    first_document = run_request["input_documents"][0]
    assert first_document["document_id"] == document_id
    assert first_document["filename"].startswith("input") or first_document["filename"].endswith(".xlsx")


async def test_job_submission_missing_document_id_returns_error(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, _ = await _create_config(async_client, headers, workspace_id)

    response = await async_client.post(
        f"/api/v1/workspaces/{workspace_id}/jobs",
        headers=headers,
        json={"config_version_id": version_id, "document_ids": ["nonexistent"]},
    )
    assert response.status_code == 400
    assert "Document" in response.text


async def test_job_submission_accepts_multiple_documents(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, _ = await _create_config(async_client, headers, workspace_id)

    workbook_a = Workbook()
    sheet_a = workbook_a.active
    sheet_a.append(["Member ID", "First Name"])
    sheet_a.append(["2001", "Bruno"])

    workbook_b = Workbook()
    sheet_b = workbook_b.active
    sheet_b.append(["Member ID", "First Name"])
    sheet_b.append(["2002", "Casey"])

    doc_a = await _upload_document(async_client, headers, workspace_id, workbook_a, filename="input-a.xlsx")
    doc_b = await _upload_document(async_client, headers, workspace_id, workbook_b, filename="input-b.xlsx")

    payload = {"config_version_id": version_id, "document_ids": [doc_a, doc_b]}
    response = await async_client.post(
        f"/api/v1/workspaces/{workspace_id}/jobs",
        headers=headers,
        json=payload,
    )
    assert response.status_code == 202, response.text
    job_detail = await _wait_for_job(async_client, headers, workspace_id, response.json()["job_id"])
    run_request = json.loads(Path(job_detail["run_request_uri"]).read_text(encoding="utf-8"))
    assert len(run_request["input_paths"]) == 2
    document_ids = [doc["document_id"] for doc in run_request["input_documents"]]
    assert document_ids == [doc_a, doc_b]


async def test_concurrent_job_submissions_share_existing_job(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, _ = await _create_config(async_client, headers, workspace_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Member ID", "First Name"])
    sheet.append(["1001", "Alex"])
    document_id = await _upload_document(async_client, headers, workspace_id, workbook)

    payload = {"config_version_id": version_id, "document_ids": [document_id]}
    base = f"/api/v1/workspaces/{workspace_id}/jobs"

    first = await async_client.post(base, headers=headers, json=payload)
    second = await async_client.post(base, headers=headers, json=payload)

    statuses = {first.status_code, second.status_code}
    assert statuses == {202}

    job_ids = {first.json()["job_id"], second.json()["job_id"]}
    assert len(job_ids) == 1
    job_id = job_ids.pop()

    detail = await _wait_for_job(async_client, headers, workspace_id, job_id)
    assert detail["status"] == "succeeded"
    assert detail["attempt"] == 1


async def test_queue_returns_429_without_persisting_job(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    app: FastAPI,
) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, _ = await _create_config(
        async_client, headers, workspace_id, slow_iterations=500_000
    )

    queue = app.state.job_queue
    assert queue is not None
    await queue.stop()
    original_settings = queue._settings
    original_queue = queue._queue
    queue._settings = original_settings.model_copy(update={"queue_max_size": 1, "queue_max_concurrency": 1})
    queue._queue = asyncio.Queue(maxsize=queue._settings.queue_max_size)
    queue._inflight.clear()

    try:
        await queue.start()

        base = f"/api/v1/workspaces/{workspace_id}/jobs"
        first = await async_client.post(base, headers=headers, json={"config_version_id": version_id})
        second = await async_client.post(base, headers=headers, json={"config_version_id": version_id})
        third = await async_client.post(base, headers=headers, json={"config_version_id": version_id})

        assert first.status_code == 202
        assert second.status_code == 202
        assert third.status_code == 429

        detail = third.json().get("detail", {})
        assert detail.get("queue_size") == 1
        assert detail.get("max_size") == 1
        assert detail.get("max_concurrency") == 1

        first_id = first.json()["job_id"]
        second_id = second.json()["job_id"]

        await _wait_for_job(async_client, headers, workspace_id, first_id)
        await _wait_for_job(async_client, headers, workspace_id, second_id)
    finally:
        await queue.stop()
        queue._settings = original_settings
        queue._queue = original_queue
        queue._inflight.clear()
        await queue.start()

    session_factory = get_sessionmaker(settings=get_settings())
    async with session_factory() as session:
        total = await session.scalar(
            select(func.count()).select_from(Job).where(Job.config_version_id == version_id)
        )
    assert total == 2


async def test_rehydration_requeues_stale_running_job(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    app: FastAPI,
) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, config_id = await _create_config(async_client, headers, workspace_id)

    settings = get_settings()
    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        repo = JobsRepository(session)
        job = await repo.create_job(
            workspace_id=workspace_id,
            config_id=config_id,
            config_version_id=version_id,
            actor_id=None,
            input_hash="rehydrate-test",
            trace_id="rehydrate-test",
            document_ids=[],
        )
        job.status = JobStatus.RUNNING.value
        job.started_at = utc_now() - timedelta(seconds=300)
        job.last_heartbeat = None
        await session.flush()
        await session.commit()
        job_id = job.id

    queue = app.state.job_queue
    assert queue is not None
    await queue.stop()
    await queue.start()

    before_retry = await async_client.get(
        f"/api/v1/workspaces/{workspace_id}/jobs/{job_id}", headers=headers
    )
    assert before_retry.status_code == 200
    assert before_retry.json()["status"] in {"queued", "running"}

    detail = await _wait_for_job(async_client, headers, workspace_id, job_id, timeout=20.0)
    assert detail["status"] == "succeeded"

    log_lines = Path(detail["logs_uri"]).read_text(encoding="utf-8").splitlines()
    assert any(json.loads(line)["event"] == "retry" for line in log_lines)


async def test_retry_endpoint_enqueues_new_attempt(async_client: AsyncClient, seed_identity: dict[str, Any]) -> None:
    headers, workspace_id = await _auth(async_client, seed_identity)
    version_id, _, config_id = await _create_config(async_client, headers, workspace_id)

    base = f"/api/v1/workspaces/{workspace_id}/jobs"
    submitted = await async_client.post(base, headers=headers, json={"config_version_id": version_id})
    assert submitted.status_code == 202
    job_id = submitted.json()["job_id"]

    original = await _wait_for_job(async_client, headers, workspace_id, job_id)
    assert original["status"] == "succeeded"
    assert original["attempt"] == 1

    retry = await async_client.post(f"{base}/{job_id}/retry", headers=headers)
    assert retry.status_code == 202, retry.text
    retry_payload = retry.json()
    assert retry_payload["job_id"] != job_id
    assert retry_payload["retry_of_job_id"] == job_id
    assert retry_payload["attempt"] == 2

    location = retry.headers.get("Location")
    assert location and location.endswith(f"/jobs/{retry_payload['job_id']}")

    retried = await _wait_for_job(async_client, headers, workspace_id, retry_payload["job_id"])
    assert retried["status"] == "succeeded"
    assert retried["attempt"] == 2
