"""Integration tests for config management and job orchestration."""

from __future__ import annotations

import asyncio
import io
import json
import textwrap
import zipfile
from hashlib import sha256
from pathlib import Path
from typing import Any
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from backend.app.features.configs.storage import ConfigStorage
from backend.app.features.configs.spec import ManifestLoader
from backend.app.shared.core.config import get_settings
from backend.tests.utils import login

pytestmark = pytest.mark.asyncio


def _csrf_headers(client: AsyncClient) -> dict[str, str]:
    settings = get_settings()
    token = client.cookies.get(settings.session_csrf_cookie_name)
    assert token, "CSRF cookie not present"
    return {"X-CSRF-Token": token}


def _column_script(field: str, *, include_transform: bool = True) -> str:
    body = f"""
    def detect_{field}(*, header, values_sample, column_index, table, job_context, env):
        return {{"scores": {{"{field}": 1.0}}}}
    """
    if include_transform:
        body += """

    def transform(*, header, values, column_index, table, job_context, env):
        return {"values": list(values), "warnings": []}
    """
    return textwrap.dedent(body).strip() + "\n"


def _column_script_with_required_validator(field: str) -> str:
    body = f"""
    def detect_{field}(*, header, values_sample, column_index, table, job_context, env):
        return {{"scores": {{"{field}": 1.0}}}}

    def transform(*, values, **_):
        return {{"values": list(values), "warnings": []}}

    def validate(*, values, **_):
        issues = []
        if not values:
            issues.append({{
                "row_index": 1,
                "code": "missing",
                "severity": "error",
                "message": "Missing value",
            }})
        return {{"issues": issues}}
    """
    return textwrap.dedent(body).strip() + "\n"


def _build_manifest(sheet_title: str = "Members") -> tuple[dict[str, Any], dict[str, str]]:
    manifest: dict[str, Any] = {
        "config_script_api_version": "1",
        "info": {"schema": "ade.manifest/v1.0", "title": "QA Config", "version": "1.0.0"},
        "engine": {"writer": {"output_sheet": sheet_title}},
        "columns": {
            "order": ["member_id", "first_name", "legacy_code"],
            "meta": {
                "member_id": {
                    "label": "Member ID",
                    "required": True,
                    "enabled": True,
                    "script": "columns/member_id.py",
                },
                "first_name": {
                    "label": "First Name",
                    "required": False,
                    "enabled": True,
                    "script": "columns/first_name.py",
                },
                "legacy_code": {
                    "label": "Legacy Code",
                    "required": False,
                    "enabled": False,
                    "script": "columns/legacy_code.py",
                },
            },
        },
    }
    scripts = {
        "columns/member_id.py": _column_script("member_id"),
        "columns/first_name.py": _column_script("first_name"),
    }
    return manifest, scripts


def _build_package(
    manifest: dict[str, Any], scripts: dict[str, str], *, extra_files: dict[str, str] | None = None
) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest, indent=2))
        archive.writestr("columns/__init__.py", "")
        for path, content in scripts.items():
            archive.writestr(path, content)
        for path, content in (extra_files or {}).items():
            archive.writestr(path, content)
    return buffer.getvalue()


def _config_endpoint(workspace_id: str) -> str:
    return f"/api/v1/workspaces/{workspace_id}/configs"


def _jobs_endpoint(workspace_id: str) -> str:
    return f"/api/v1/workspaces/{workspace_id}/jobs"


async def _wait_for_job(
    async_client: AsyncClient,
    *,
    workspace_id: str,
    job_id: str,
    headers: dict[str, str] | None = None,
    timeout: float = 15.0,
) -> dict[str, Any]:
    """Poll the job detail endpoint until the job completes."""

    headers = headers or {}
    detail_url = f"{_jobs_endpoint(workspace_id)}/{job_id}"
    loop = asyncio.get_event_loop()
    deadline = loop.time() + timeout
    while True:
        response = await async_client.get(detail_url, headers=headers)
        response.raise_for_status()
        payload = response.json()
        if payload.get("status") in {"succeeded", "failed"}:
            return payload
        if loop.time() >= deadline:
            pytest.fail(f"job {job_id} did not complete in time: {payload}")
        await asyncio.sleep(0.1)


async def test_create_config_allows_column_without_transform(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest()
    scripts["columns/member_id.py"] = _column_script("member_id", include_transform=False)
    package_bytes = _build_package(manifest, scripts)

    slug = f"invalid-{uuid4().hex[:8]}"
    headers = _csrf_headers(async_client)
    response = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": slug,
            "title": "Invalid Config",
            "manifest_json": json.dumps(manifest),
        },
        files={
            "package": ("package.zip", package_bytes, "application/zip"),
        },
        headers=headers,
    )

    assert response.status_code == 201
    payload = response.json()
    versions = payload.get("versions", [])
    assert versions, "Config version was not created"
    assert payload.get("active_version") is None


async def test_create_config_stores_canonical_manifest_and_package_hashes(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest(sheet_title="Audited")
    package_bytes = _build_package(manifest, scripts)

    slug = f"config-{uuid4().hex[:8]}"
    headers = _csrf_headers(async_client)
    response = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": slug,
            "title": "Canonical Config",
            "description": "Smoke test",
            "manifest_json": json.dumps(manifest),
        },
        files={
            "package": ("package.zip", package_bytes, "application/zip"),
        },
        headers=headers,
    )
    assert response.status_code == 201, response.text

    payload = response.json()
    assert payload.get("active_version") is None
    versions = payload.get("versions", [])
    assert len(versions) == 1
    version = versions[0]

    loader = ManifestLoader()
    canonical_manifest = loader.load(manifest).model_dump(mode="json")
    canonical_json = json.dumps(canonical_manifest, sort_keys=True, separators=(",", ":"))
    expected_manifest_hash = sha256(canonical_json.encode("utf-8")).hexdigest()
    assert version["manifest_sha256"] == expected_manifest_hash
    assert version["config_script_api_version"] == "1"

    package_path = Path(version["package_path"])
    assert package_path.exists()
    assert (package_path / "manifest.json").exists()

    sequence = int(version["sequence"])
    archive_path = package_path.parent / f"v{sequence:04d}.zip"
    assert archive_path.exists()

    storage = ConfigStorage(get_settings())
    expected_package_hash = storage.compute_package_hash(archive_path)
    assert version["package_sha256"] == expected_package_hash


async def test_activate_version_builds_environment_and_runs_hooks(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest()
    manifest["hooks"] = {
        "on_activate": [
            {
                "script": "hooks/on_activate.py",
                "enabled": True,
            }
        ]
    }
    scripts["hooks/__init__.py"] = ""
    scripts["hooks/on_activate.py"] = textwrap.dedent(
        """
        def run(*, manifest, env, artifact, job_context):
            return {"note": "activated"}
        """
    ).strip()
    package_bytes = _build_package(manifest, scripts)

    headers = _csrf_headers(async_client)
    create = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": "activation-demo",
            "title": "Activation Demo",
            "manifest_json": json.dumps(manifest),
        },
        files={"package": ("package.zip", package_bytes, "application/zip")},
        headers=headers,
    )
    create.raise_for_status()
    created = create.json()
    config_id = created["config_id"]
    version = created["versions"][0]
    version_id = version["config_version_id"]

    activate = await async_client.post(
        f"{_config_endpoint(workspace_id)}/{config_id}/versions/{version_id}/activate",
        headers=headers,
    )
    assert activate.status_code == 200, activate.text
    activated = activate.json()["active_version"]
    activation = activated.get("activation")
    assert activation is not None, activated
    assert activation["status"] == "succeeded"
    assert activation["annotations"] and activation["annotations"][0]["note"] == "activated"

    python_exec = activation["python_executable"]
    assert python_exec and Path(python_exec).exists()
    packages_uri = activation["packages_uri"]
    assert packages_uri and Path(packages_uri).exists()
    hooks_uri = activation["hooks_uri"]
    assert hooks_uri and Path(hooks_uri).exists()

    storage = ConfigStorage(get_settings())
    activation_dir = storage.activation_dir(config_id, version["sequence"])
    result_path = activation_dir / "result.json"
    assert result_path.exists()
    result_payload = json.loads(result_path.read_text(encoding="utf-8"))
    assert result_payload["status"] == "succeeded"


async def test_activate_version_fails_on_invalid_requirements(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest()
    package_bytes = _build_package(
        manifest,
        scripts,
        extra_files={"requirements.txt": "not_a_valid_requirement!!!\n"},
    )

    headers = _csrf_headers(async_client)
    create = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": "activation-error",
            "title": "Activation Error",
            "manifest_json": json.dumps(manifest),
        },
        files={"package": ("package.zip", package_bytes, "application/zip")},
        headers=headers,
    )
    create.raise_for_status()
    created = create.json()
    config_id = created["config_id"]
    version = created["versions"][0]
    version_id = version["config_version_id"]

    activate = await async_client.post(
        f"{_config_endpoint(workspace_id)}/{config_id}/versions/{version_id}/activate",
        headers=headers,
    )
    assert activate.status_code == 400, activate.text
    detail = activate.json()
    diagnostics = detail.get("detail", {}).get("diagnostics")
    assert diagnostics

    refreshed = await async_client.get(f"{_config_endpoint(workspace_id)}/{config_id}", headers=headers)
    refreshed.raise_for_status()
    assert refreshed.json().get("active_version") is None

    storage = ConfigStorage(get_settings())
    activation_dir = storage.activation_dir(config_id, version["sequence"])
    result_path = activation_dir / "result.json"
    assert result_path.exists()
    result_payload = json.loads(result_path.read_text(encoding="utf-8"))
    assert result_payload["status"] == "failed"


async def test_job_artifact_reflects_manifest_columns(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest(sheet_title="Members")
    manifest.setdefault("hooks", {}).setdefault("after_mapping", []).append(
        {"script": "hooks/after_mapping.py"}
    )
    scripts["hooks/__init__.py"] = ""
    scripts["hooks/after_mapping.py"] = (
        textwrap.dedent(
            """
            def run(*, artifact, **_):
                return {"notes": "mapping complete"}
            """
        ).strip()
        + "\n"
    )
    package_bytes = _build_package(manifest, scripts)

    slug = f"job-{uuid4().hex[:8]}"
    headers = _csrf_headers(async_client)
    create_response = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": slug,
            "title": "Job Config",
            "manifest_json": json.dumps(manifest),
        },
        files={
            "package": ("package.zip", package_bytes, "application/zip"),
        },
        headers=headers,
    )
    assert create_response.status_code == 201, create_response.text
    config_payload = create_response.json()
    version_id = config_payload["versions"][0]["config_version_id"]

    job_response = await async_client.post(
        _jobs_endpoint(workspace_id),
        json={"config_version_id": version_id},
        headers=headers,
    )
    assert job_response.status_code == 202, job_response.text
    job_id = job_response.json()["job_id"]
    job_payload = await _wait_for_job(
        async_client,
        workspace_id=workspace_id,
        job_id=job_id,
        headers=headers,
    )
    if job_payload["status"] != "succeeded":
        pytest.fail(f"job failed: {job_payload}")
    assert job_payload["attempt"] == 1
    assert job_payload["logs_uri"]
    assert job_payload["run_request_uri"]
    trace_id = job_payload["trace_id"]

    artifact_response = await async_client.get(
        f"{_jobs_endpoint(workspace_id)}/{job_id}/artifact"
    )
    assert artifact_response.status_code == 200
    artifact = artifact_response.json()
    assert artifact["job"]["trace_id"] == trace_id
    pass_names = [p["name"] for p in artifact["pass_history"]]
    assert pass_names == [
        "structure",
        "mapping",
        "transform",
        "validate",
        "generate",
    ]
    assert all("stats" in entry for entry in artifact["pass_history"])
    mapping_pass = next(entry for entry in artifact["pass_history"] if entry["name"] == "mapping")
    assert set(mapping_pass["stats"]) == {"mapped", "unmapped"}
    assert artifact["annotations"]
    assert any(
        annotation["stage"] == "after_mapping"
        and annotation["hook"] == "hooks/after_mapping.py"
        and annotation.get("notes") == "mapping complete"
        for annotation in artifact["annotations"]
    )
    for annotation in artifact["annotations"]:
        assert "annotated_at" in annotation

    assert artifact["summary"]["rows_written"] >= 0
    assert artifact["output"]["sheet"] == "Members"
    assert artifact["output"]["column_plan"]["target"] == [
        {"field": "member_id", "output_header": "Member ID", "order": 1},
        {"field": "first_name", "output_header": "First Name", "order": 2},
    ]

    sheet_entry = artifact["sheets"][0]
    table_entry = sheet_entry["tables"][0]
    mapping_entries = table_entry["mapping"]
    assert any(entry["target_field"] == "member_id" for entry in mapping_entries)
    assert any(entry["raw"]["header"] == "Legacy Code" for entry in mapping_entries)

    transform_targets = {entry["target_field"] for entry in table_entry["transforms"]}
    assert {"member_id", "first_name"}.issubset(transform_targets)

    validation_section = table_entry["validation"]
    assert isinstance(validation_section["issues"], list)
    assert isinstance(validation_section["summary_by_field"], dict)

    output_response = await async_client.get(
        f"{_jobs_endpoint(workspace_id)}/{job_id}/output"
    )
    assert (
        output_response.headers.get("content-type")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(output_response.content))
    sheet = workbook.active
    assert sheet.title == "Members"
    rows = list(sheet.iter_rows(values_only=True))
    assert rows == [("Member ID", "First Name")]

    # Inspect run request and logs
    run_request = json.loads(Path(job_payload["run_request_uri"]).read_text(encoding="utf-8"))
    assert run_request["schema"] == "ade.run_request/v1"
    assert run_request["job_id"] == job_id

    log_lines = [
        json.loads(line)
        for line in Path(job_payload["logs_uri"]).read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    assert any(entry.get("event") == "worker.spawn" for entry in log_lines)
    exit_event = next(entry for entry in log_lines if entry.get("event") == "worker.exit")
    assert exit_event.get("state") == "succeeded"
    assert exit_event.get("duration_ms", 0) >= 0
    exit_detail = exit_event.get("detail", {})
    assert exit_detail.get("exit_code") == 0


async def test_validation_issues_include_a1_and_canonical_codes(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest(sheet_title="Audit")
    scripts["columns/member_id.py"] = _column_script_with_required_validator("member_id")
    package_bytes = _build_package(manifest, scripts)

    slug = f"job-{uuid4().hex[:8]}"
    headers = _csrf_headers(async_client)
    create_response = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": slug,
            "title": "Validation Config",
            "manifest_json": json.dumps(manifest),
        },
        files={"package": ("package.zip", package_bytes, "application/zip")},
        headers=headers,
    )
    assert create_response.status_code == 201, create_response.text
    config_payload = create_response.json()
    version_id = config_payload["versions"][0]["config_version_id"]

    job_response = await async_client.post(
        _jobs_endpoint(workspace_id),
        json={"config_version_id": version_id},
        headers=headers,
    )
    assert job_response.status_code == 202, job_response.text
    job_id = job_response.json()["job_id"]
    job_payload = await _wait_for_job(
        async_client,
        workspace_id=workspace_id,
        job_id=job_id,
        headers=headers,
    )
    assert job_payload["status"] == "succeeded"
    artifact_response = await async_client.get(
        f"{_jobs_endpoint(workspace_id)}/{job_id}/artifact"
    )
    assert artifact_response.status_code == 200
    artifact = artifact_response.json()

    sheet_entry = artifact["sheets"][0]
    table_entry = sheet_entry["tables"][0]
    issues = table_entry["validation"]["issues"]
    assert issues, "Expected validation issues to be recorded"

    member_issue = next(
        issue for issue in issues if issue.get("target_field") == "member_id"
    )
    assert member_issue["code"] == "required_missing"
    assert member_issue["row_index"] == 1
    assert member_issue["a1"] == "A2"
    assert member_issue["severity"] == "error"


async def test_job_appends_unmapped_columns_when_enabled(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest(sheet_title="Members")
    writer_section = manifest.setdefault("engine", {}).setdefault("writer", {})
    writer_section["append_unmapped_columns"] = True
    writer_section["unmapped_prefix"] = "raw_"
    scripts["columns/member_id.py"] = textwrap.dedent(
        """
        def detect_member_id(*, header, **_):
            scores = {}
            if header and "Member" in header:
                scores["member_id"] = 1.0
            return {"scores": scores}

        def transform(*, values, **_):
            return {"values": list(values), "warnings": []}
        """
    ).strip() + "\n"
    scripts["columns/first_name.py"] = textwrap.dedent(
        """
        def detect_first_name(*, header, **_):
            scores = {}
            if header and "First" in header:
                scores["first_name"] = 1.0
            return {"scores": scores}

        def transform(*, values, **_):
            return {"values": list(values), "warnings": []}
        """
    ).strip() + "\n"
    package_bytes = _build_package(manifest, scripts)

    slug = f"unmapped-{uuid4().hex[:8]}"
    headers = _csrf_headers(async_client)
    create_response = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": slug,
            "title": "Unmapped Columns Config",
            "manifest_json": json.dumps(manifest),
        },
        files={
            "package": ("package.zip", package_bytes, "application/zip"),
        },
        headers=headers,
    )
    assert create_response.status_code == 201, create_response.text
    version_id = create_response.json()["versions"][0]["config_version_id"]

    job_response = await async_client.post(
        _jobs_endpoint(workspace_id),
        json={"config_version_id": version_id},
        headers=headers,
    )
    assert job_response.status_code == 202, job_response.text
    job_id = job_response.json()["job_id"]
    job_payload = await _wait_for_job(
        async_client,
        workspace_id=workspace_id,
        job_id=job_id,
        headers=headers,
    )
    if job_payload.get("status") != "succeeded":
        pytest.fail(f"job failed: {job_payload}")
    artifact_response = await async_client.get(
        f"{_jobs_endpoint(workspace_id)}/{job_id}/artifact"
    )
    assert artifact_response.status_code == 200
    artifact = artifact_response.json()

    appended = artifact["output"]["column_plan"]["appended_unmapped"]
    assert appended, "Expected unmapped columns to be recorded"
    assert len(appended) == 1
    appended_entry = appended[0]
    assert appended_entry["source_header"] == "Legacy Code"
    assert appended_entry["output_header"].startswith("raw_")
    assert appended_entry["order"] == 1
    assert appended_entry["column"].endswith(".col.3")

    summary = artifact["summary"]
    assert summary["columns_written"] == 3

    output_response = await async_client.get(
        f"{_jobs_endpoint(workspace_id)}/{job_id}/output"
    )
    assert output_response.status_code == 200
    workbook = load_workbook(io.BytesIO(output_response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][:3] == (
        "Member ID",
        "First Name",
        appended_entry["output_header"],
    )
    workbook.close()


async def test_job_submission_idempotent_on_input_hash(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest(sheet_title="Idempotent")
    package_bytes = _build_package(manifest, scripts)

    slug = f"idem-{uuid4().hex[:8]}"
    headers = _csrf_headers(async_client)
    create_response = await async_client.post(
        _config_endpoint(workspace_id),
        data={
            "slug": slug,
            "title": "Idempotent Config",
            "manifest_json": json.dumps(manifest),
        },
        files={
            "package": ("package.zip", package_bytes, "application/zip"),
        },
        headers=headers,
    )
    assert create_response.status_code == 201
    version_id = create_response.json()["versions"][0]["config_version_id"]

    input_hash = "abc123"
    first = await async_client.post(
        _jobs_endpoint(workspace_id),
        json={"config_version_id": version_id, "input_hash": input_hash},
        headers=headers,
    )
    assert first.status_code == 202
    first_job_id = first.json()["job_id"]
    first_payload = await _wait_for_job(
        async_client,
        workspace_id=workspace_id,
        job_id=first_job_id,
        headers=headers,
    )
    assert first_payload["status"] == "succeeded"

    second = await async_client.post(
        _jobs_endpoint(workspace_id),
        json={"config_version_id": version_id, "input_hash": input_hash},
        headers=headers,
    )
    assert second.status_code == 202
    second_payload = second.json()

    assert second_payload["job_id"] == first_job_id
    assert second_payload["attempt"] == 1
    assert second_payload["status"] == "succeeded"


async def test_validate_endpoint_accepts_columns_without_transform(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    await login(async_client, email=owner["email"], password=owner["password"])

    manifest, scripts = _build_manifest()
    scripts["columns/member_id.py"] = _column_script("member_id", include_transform=False)
    package_bytes = _build_package(manifest, scripts)

    headers = _csrf_headers(async_client)
    response = await async_client.post(
        f"{_config_endpoint(workspace_id)}/validate",
        data={"manifest_json": json.dumps(manifest)},
        files={"package": ("package.zip", package_bytes, "application/zip")},
        headers=headers,
    )
    assert response.status_code == 200
    payload = response.json()
    diagnostics = payload.get("diagnostics", [])
    assert isinstance(diagnostics, list)
    assert all(diag.get("code") != "column.transform.missing" for diag in diagnostics)
