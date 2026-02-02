"""Service layer for document upload and retrieval."""

from __future__ import annotations

import logging
import tempfile
import unicodedata
from collections.abc import Iterator, Mapping, Sequence
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from fastapi import UploadFile
from pydantic import ValidationError
from sqlalchemy import case, func, select
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.sql import Select

from ade_api.common.ids import generate_uuid7
from ade_api.common.list_filters import FilterItem, FilterJoinOperator
from ade_api.common.cursor_listing import (
    ResolvedCursorSort,
    cursor_field,
    paginate_query_cursor,
    parse_int,
    parse_str,
    resolve_cursor_sort,
)
from ade_api.common.logging import log_context
from ade_api.common.workbook_preview import (
    WorkbookSheetPreview,
    build_workbook_preview_from_csv,
    build_workbook_preview_from_xlsx,
)
from ade_storage import AzureBlobStorage, StorageError, StorageLimitError, StoredObject
from ade_db.models import (
    File,
    FileComment,
    FileCommentMention,
    FileKind,
    FileTag,
    FileVersion,
    FileVersionOrigin,
    Run,
    RunField,
    RunMetrics,
    RunStatus,
    RunTableColumn,
    User,
    WorkspaceMembership,
)
from ade_api.settings import Settings
from ade_api.features.runs.schemas import RunColumnResource, RunFieldResource, RunMetricsResource

from .changes import (
    DocumentChangeDelta,
    fetch_document_change_delta,
    get_latest_document_change_id,
)
from .exceptions import (
    DocumentFileMissingError,
    DocumentTooLargeError,
    DocumentNameConflictError,
    DocumentNotFoundError,
    DocumentPreviewParseError,
    DocumentPreviewSheetNotFoundError,
    DocumentPreviewUnsupportedError,
    DocumentVersionNotFoundError,
    DocumentWorksheetParseError,
    InvalidDocumentCommentMentionsError,
    InvalidDocumentTagsError,
)
from .filters import apply_document_filters
from .repository import DocumentsRepository
from .schemas import (
    DocumentCommentOut,
    DocumentCommentPage,
    DocumentConflictMode,
    DocumentFileType,
    DocumentListPage,
    DocumentListRow,
    DocumentOut,
    DocumentRunSummary,
    DocumentSheet,
    DocumentUpdateRequest,
    DocumentUploadRunOptions,
    TagCatalogItem,
    TagCatalogPage,
)
from .tags import (
    MAX_TAGS_PER_DOCUMENT,
    TagValidationError,
    normalize_tag_list,
    normalize_tag_query,
)

logger = logging.getLogger(__name__)

UPLOAD_RUN_OPTIONS_KEY = "__ade_run_options"

_FALLBACK_FILENAME = "upload"
_MAX_FILENAME_LENGTH = 255


def _run_with_timeout(func, *, timeout: float, **kwargs):
    """Run a callable with a timeout to avoid hanging on large workbook operations."""
    if timeout <= 0:
        return func(**kwargs)
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(func, **kwargs)
        return future.result(timeout=timeout)


def _map_document_row(row: Mapping[str, Any]) -> File:
    document = row[File]
    setattr(document, "_last_run_at", row.get("last_run_at"))
    return document


@dataclass(slots=True)
class StagedUpload:
    file_id: UUID
    blob_name: str
    stored: StoredObject
    overwrite_existing: bool = False


class UploadAction(str, Enum):
    NEW = "new"
    NEW_VERSION = "new_version"


@dataclass(slots=True)
class UploadPlan:
    action: UploadAction
    file: File | None
    file_id: UUID
    blob_name: str
    name: str
    name_key: str
    source_file_id: UUID | None = None


class DocumentsService:
    """Manage document metadata and backing file storage."""

    def __init__(
        self,
        *,
        session: Session,
        settings: Settings,
        storage: AzureBlobStorage,
    ) -> None:
        self._session = session
        self._settings = settings
        self._storage = storage
        self._repository = DocumentsRepository(session)

    @staticmethod
    def build_upload_metadata(
        metadata: Mapping[str, Any] | None,
        run_options: DocumentUploadRunOptions | None,
    ) -> dict[str, Any]:
        payload = dict(metadata or {})
        if run_options is not None:
            payload[UPLOAD_RUN_OPTIONS_KEY] = run_options.model_dump(exclude_none=True)
        return payload

    @staticmethod
    def read_upload_run_options(
        metadata: Mapping[str, Any] | None,
    ) -> DocumentUploadRunOptions | None:
        if not metadata:
            return None
        raw = metadata.get(UPLOAD_RUN_OPTIONS_KEY)
        if raw is None:
            return None
        try:
            return DocumentUploadRunOptions.model_validate(raw)
        except ValidationError:
            return None

    def plan_upload(
        self,
        *,
        workspace_id: UUID,
        filename: str | None,
        conflict_mode: DocumentConflictMode | None = None,
    ) -> UploadPlan:
        mode = conflict_mode or DocumentConflictMode.REJECT
        normalized_name = self._normalise_filename(filename)
        name_key = self._build_name_key(normalized_name)
        existing = self._find_by_name_key(workspace_id=workspace_id, name_key=name_key)

        if existing is not None:
            if mode == DocumentConflictMode.REJECT:
                raise DocumentNameConflictError(
                    document_id=existing.id,
                    name=existing.name,
                )
            if mode == DocumentConflictMode.UPLOAD_NEW_VERSION:
                return UploadPlan(
                    action=UploadAction.NEW_VERSION,
                    file=existing,
                    file_id=existing.id,
                    blob_name=existing.blob_name,
                    name=existing.name,
                    name_key=existing.name_key,
                )
            if mode == DocumentConflictMode.KEEP_BOTH:
                disambiguated, disambiguated_key = self._disambiguate_name(
                    workspace_id=workspace_id,
                    base_name=normalized_name,
                )
                file_id = generate_uuid7()
                return UploadPlan(
                    action=UploadAction.NEW,
                    file=None,
                    file_id=file_id,
                    blob_name=self._file_blob_name(workspace_id, file_id),
                    name=disambiguated,
                    name_key=disambiguated_key,
                    source_file_id=existing.id,
                )

        file_id = generate_uuid7()
        return UploadPlan(
            action=UploadAction.NEW,
            file=None,
            file_id=file_id,
            blob_name=self._file_blob_name(workspace_id, file_id),
            name=normalized_name,
            name_key=name_key,
        )

    def plan_upload_for_version(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> UploadPlan:
        document = self._get_document(workspace_id, document_id)
        return UploadPlan(
            action=UploadAction.NEW_VERSION,
            file=document,
            file_id=document.id,
            blob_name=document.blob_name,
            name=document.name,
            name_key=document.name_key,
        )

    def create_document(
        self,
        *,
        workspace_id: UUID,
        upload: UploadFile,
        plan: UploadPlan,
        metadata: Mapping[str, Any] | None = None,
        actor: User | None = None,
        staged: StagedUpload | None = None,
    ) -> DocumentOut:
        """Persist ``upload`` to storage and return the resulting metadata record."""

        actor_id: UUID | None = actor.id if actor is not None else None
        logger.debug(
            "document.create.start",
            extra=log_context(
                workspace_id=workspace_id,
                user_id=actor_id,
                upload_filename=upload.filename,
                content_type=upload.content_type,
            ),
        )

        metadata_payload = dict(metadata or {})
        now = datetime.now(tz=UTC)

        owns_stage = staged is None
        staged_upload = staged or self.stage_upload(upload=upload, plan=plan)
        stored = staged_upload.stored

        upload_name = self._normalise_filename(upload.filename)
        content_type = self._normalise_content_type(upload.content_type)
        storage_version_id = stored.version_id

        try:
            if plan.action == UploadAction.NEW:
                document = File(
                    id=plan.file_id,
                    workspace_id=workspace_id,
                    kind=FileKind.INPUT,
                    name=plan.name,
                    name_key=plan.name_key,
                    blob_name=plan.blob_name,
                    source_file_id=plan.source_file_id,
                    attributes=metadata_payload,
                    uploaded_by_user_id=actor_id,
                )
                self._session.add(document)
                self._session.flush()
                version_no = 1
            else:
                document = plan.file
                if document is None:
                    raise RuntimeError("Upload plan did not include a document.")
                version_no = self._next_version_no(document_id=document.id)
                if metadata_payload:
                    document.attributes = metadata_payload
                document = self._session.merge(document)

            file_version = FileVersion(
                file_id=document.id,
                version_no=version_no,
                origin=FileVersionOrigin.UPLOADED,
                created_by_user_id=actor_id,
                sha256=stored.sha256,
                byte_size=stored.byte_size,
                content_type=content_type,
                filename_at_upload=upload_name,
                storage_version_id=storage_version_id,
            )
            self._session.add(file_version)
            self._session.flush()
            document.current_version_id = file_version.id
            self._session.flush()
        except Exception:
            if owns_stage:
                self.discard_staged_upload(staged=staged_upload)
            raise

        stmt = self._repository.base_query(workspace_id).where(File.id == document.id)
        result = self._session.execute(stmt)
        hydrated = result.scalar_one()

        payload = DocumentOut.model_validate(hydrated)
        self._attach_last_runs(workspace_id, [payload])
        self._apply_derived_fields(payload)
        payload.list_row = self._build_list_row(payload)

        logger.info(
            "document.create.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document.id,
                user_id=actor_id,
                content_type=payload.content_type,
                byte_size=payload.byte_size,
            ),
        )

        return payload

    def stage_upload(
        self,
        *,
        upload: UploadFile,
        plan: UploadPlan,
    ) -> StagedUpload:
        if upload.file is None:  # pragma: no cover - UploadFile always supplies file
            raise RuntimeError("Upload stream is not available")
        try:
            stored = self._storage.write(
                plan.blob_name,
                upload.file,
                max_bytes=self._settings.storage_upload_max_bytes,
            )
        except StorageLimitError as exc:
            raise DocumentTooLargeError(limit=exc.limit, received=exc.received) from exc
        return StagedUpload(
            file_id=plan.file_id,
            blob_name=plan.blob_name,
            stored=stored,
            overwrite_existing=plan.action == UploadAction.NEW_VERSION,
        )

    def discard_staged_upload(self, *, staged: StagedUpload) -> None:
        if staged.overwrite_existing and staged.stored.version_id is None:
            return
        try:
            self._storage.delete(staged.blob_name, version_id=staged.stored.version_id)
        except StorageError:
            logger.warning(
                "document.upload.discard_failed",
                extra={"blob_name": staged.blob_name},
                exc_info=True,
            )

    def list_documents(
        self,
        *,
        workspace_id: UUID,
        limit: int,
        cursor: str | None,
        resolved_sort: ResolvedCursorSort[File],
        filters: list[FilterItem],
        join_operator: FilterJoinOperator,
        q: str | None,
        include_total: bool,
        include_facets: bool,
        include_run_metrics: bool = False,
        include_run_table_columns: bool = False,
        include_run_fields: bool = False,
    ) -> DocumentListPage:
        """Return paginated documents with the shared envelope."""

        logger.debug(
            "document.list.start",
            extra=log_context(
                workspace_id=workspace_id,
                limit=limit,
                cursor=cursor,
                order_by=str(resolved_sort.order_by),
            ),
        )

        stmt = self._repository.base_query(workspace_id).where(File.deleted_at.is_(None))
        stmt = apply_document_filters(
            stmt,
            filters,
            join_operator=join_operator,
            q=q,
        )
        last_run_at_expr = (
            select(
                func.max(func.coalesce(Run.completed_at, Run.started_at, Run.created_at))
            )
            .select_from(Run)
            .join(FileVersion, Run.input_file_version_id == FileVersion.id)
            .where(
                FileVersion.file_id == File.id,
                Run.workspace_id == File.workspace_id,
            )
            .scalar_subquery()
            .label("last_run_at")
        )
        stmt = stmt.add_columns(last_run_at_expr)

        facets = self._build_document_facets(stmt) if include_facets else None
        changes_cursor = get_latest_document_change_id(self._session, workspace_id)
        page_result = paginate_query_cursor(
            self._session,
            stmt,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
            changes_cursor=changes_cursor,
            row_mapper=lambda row: _map_document_row(row),
        )
        raw_items = list(page_result.items)
        items = [DocumentOut.model_validate(item) for item in raw_items]
        self._attach_last_runs(
            workspace_id,
            items,
            include_run_metrics=include_run_metrics,
            include_run_table_columns=include_run_table_columns,
            include_run_fields=include_run_fields,
        )
        for item in items:
            self._apply_derived_fields(item)

        logger.info(
            "document.list.success",
            extra=log_context(
                workspace_id=workspace_id,
                limit=page_result.meta.limit,
                count=len(items),
                has_more=page_result.meta.has_more,
            ),
        )

        rows = [self._build_list_row(item) for item in items]

        return DocumentListPage(items=rows, meta=page_result.meta, facets=facets)

    def get_document_change_delta(
        self,
        *,
        workspace_id: UUID,
        since: int,
        limit: int | None = None,
    ) -> DocumentChangeDelta:
        return fetch_document_change_delta(
            self._session,
            workspace_id=workspace_id,
            since=since,
            limit=limit,
        )

    def list_document_comments(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        limit: int,
        cursor: str | None,
        resolved_sort: ResolvedCursorSort[FileComment],
        include_total: bool,
    ) -> DocumentCommentPage:
        self._get_document(workspace_id, document_id)

        stmt = (
            select(FileComment)
            .where(FileComment.workspace_id == workspace_id)
            .where(FileComment.file_id == document_id)
            .options(
                selectinload(FileComment.author_user),
                selectinload(FileComment.mentions).selectinload(
                    FileCommentMention.mentioned_user
                ),
            )
        )

        page_result = paginate_query_cursor(
            self._session,
            stmt,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
            changes_cursor="0",
        )

        items = [DocumentCommentOut.model_validate(item) for item in page_result.items]
        return DocumentCommentPage(items=items, meta=page_result.meta, facets=page_result.facets)

    def create_document_comment(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        body: str,
        mentions: Sequence[UUID] | None,
        actor: User,
    ) -> DocumentCommentOut:
        document = self._get_document(workspace_id, document_id)

        mention_users = self._resolve_comment_mentions(
            workspace_id=workspace_id,
            mentions=mentions,
        )

        comment = FileComment(
            workspace_id=workspace_id,
            file_id=document_id,
            author_user_id=actor.id,
            body=body,
        )
        comment.author_user = actor
        comment.mentions = [
            FileCommentMention(mentioned_user=user) for user in mention_users
        ]

        document.comment_count = (document.comment_count or 0) + 1

        self._session.add(comment)
        self._session.flush()

        return DocumentCommentOut.model_validate(comment)

    def build_list_row_for_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> DocumentListRow | None:
        stmt = (
            self._repository.base_query(workspace_id)
            .where(File.id == document_id)
            .where(File.deleted_at.is_(None))
        )
        result = self._session.execute(stmt)
        document = result.scalar_one_or_none()
        if document is None:
            return None
        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload])
        self._apply_derived_fields(payload)
        return self._build_list_row(payload)

    def get_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        include_run_metrics: bool = False,
        include_run_table_columns: bool = False,
        include_run_fields: bool = False,
    ) -> DocumentOut:
        """Return document metadata for ``document_id``."""

        logger.debug(
            "document.get.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )
        document = self._get_document(workspace_id, document_id)
        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(
            workspace_id,
            [payload],
            include_run_metrics=include_run_metrics,
            include_run_table_columns=include_run_table_columns,
            include_run_fields=include_run_fields,
        )
        self._apply_derived_fields(payload)

        logger.info(
            "document.get.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                byte_size=document.byte_size,
            ),
        )
        return payload

    def get_document_list_row(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        include_run_metrics: bool = False,
        include_run_table_columns: bool = False,
        include_run_fields: bool = False,
    ) -> DocumentListRow:
        """Return a table-ready row projection for ``document_id``."""

        logger.debug(
            "document.list_row.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )
        document = self._get_document(workspace_id, document_id)
        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(
            workspace_id,
            [payload],
            include_run_metrics=include_run_metrics,
            include_run_table_columns=include_run_table_columns,
            include_run_fields=include_run_fields,
        )
        self._apply_derived_fields(payload)
        row = self._build_list_row(payload)

        logger.info(
            "document.list_row.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                byte_size=document.byte_size,
            ),
        )
        return row

    def update_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        payload: DocumentUpdateRequest,
    ) -> DocumentOut:
        document = self._get_document(workspace_id, document_id)
        changed = False

        if "assignee_user_id" in payload.model_fields_set:
            if document.assignee_user_id != payload.assignee_user_id:
                document.assignee_user_id = payload.assignee_user_id
                changed = True
        if "metadata" in payload.model_fields_set and payload.metadata is not None:
            next_metadata = dict(payload.metadata)
            if document.attributes != next_metadata:
                document.attributes = next_metadata
                changed = True

        if changed:
            self._session.flush()
            self._session.refresh(document, attribute_names=["assignee_user"])

        updated = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [updated])
        self._apply_derived_fields(updated)
        return updated

    def list_document_sheets(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> list[DocumentSheet]:
        """Return worksheet descriptors for ``document_id``."""

        logger.debug(
            "document.sheets.list.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )

        document = self._get_document(workspace_id, document_id)
        current_version = document.current_version
        if current_version is None:
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
            )

        suffix = Path(document.name).suffix.lower()
        try:
            with self._download_blob_to_tempfile(
                blob_name=document.blob_name,
                version_id=current_version.storage_version_id,
                suffix=suffix,
            ) as path:
                if suffix == ".xlsx":
                    sheets = _run_with_timeout(
                        self._inspect_workbook,
                        timeout=self._settings.preview_timeout_seconds,
                        path=path,
                    )
                    logger.info(
                        "document.sheets.list.success",
                        extra=log_context(
                            workspace_id=workspace_id,
                            document_id=document_id,
                            sheet_count=len(sheets),
                            kind="workbook",
                        ),
                    )
                    return sheets
        except FuturesTimeout as exc:
            raise DocumentWorksheetParseError(
                document_id=document_id,
                blob_name=document.blob_name,
                reason="timeout",
            ) from exc
        except FileNotFoundError as exc:
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
            ) from exc
        except Exception as exc:  # pragma: no cover - defensive fallback
            raise DocumentWorksheetParseError(
                document_id=document_id,
                blob_name=document.blob_name,
                reason=type(exc).__name__,
            ) from exc

        name = self._default_sheet_name(document.name)
        sheets = [DocumentSheet(name=name, index=0, kind="file", is_active=True)]
        logger.info(
            "document.sheets.list.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                sheet_count=len(sheets),
                kind="fallback",
            ),
        )
        return sheets

    def get_document_preview(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        max_rows: int,
        max_columns: int,
        trim_empty_columns: bool = False,
        trim_empty_rows: bool = False,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
    ) -> WorkbookSheetPreview:
        """Return a table-ready preview for a document workbook."""

        effective_sheet_index = sheet_index
        if sheet_name is None and sheet_index is None:
            effective_sheet_index = 0

        logger.debug(
            "document.preview.start",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                sheet_name=sheet_name,
                sheet_index=effective_sheet_index,
            ),
        )

        document = self._get_document(workspace_id, document_id)
        current_version = document.current_version
        if current_version is None:
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
            )

        suffix = Path(document.name).suffix.lower()
        try:
            with self._download_blob_to_tempfile(
                blob_name=document.blob_name,
                version_id=current_version.storage_version_id,
                suffix=suffix,
            ) as path:
                if suffix == ".xlsx":
                    preview = _run_with_timeout(
                        build_workbook_preview_from_xlsx,
                        timeout=self._settings.preview_timeout_seconds,
                        path=path,
                        max_rows=max_rows,
                        max_columns=max_columns,
                        trim_empty_columns=trim_empty_columns,
                        trim_empty_rows=trim_empty_rows,
                        sheet_name=sheet_name,
                        sheet_index=effective_sheet_index,
                    )
                elif suffix == ".csv":
                    preview = _run_with_timeout(
                        build_workbook_preview_from_csv,
                        timeout=self._settings.preview_timeout_seconds,
                        path=path,
                        max_rows=max_rows,
                        max_columns=max_columns,
                        trim_empty_columns=trim_empty_columns,
                        trim_empty_rows=trim_empty_rows,
                        sheet_name=sheet_name,
                        sheet_index=effective_sheet_index,
                    )
                else:
                    raise DocumentPreviewUnsupportedError(
                        document_id=document_id,
                        file_type=suffix or "unknown",
                    )
        except (KeyError, IndexError) as exc:
            requested = sheet_name if sheet_name is not None else str(effective_sheet_index)
            raise DocumentPreviewSheetNotFoundError(
                document_id=document_id,
                sheet=requested,
            ) from exc
        except FuturesTimeout as exc:
            raise DocumentPreviewParseError(
                document_id=document_id,
                blob_name=document.blob_name,
                reason="timeout",
            ) from exc
        except FileNotFoundError as exc:
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
            ) from exc
        except DocumentPreviewUnsupportedError:
            raise
        except Exception as exc:  # pragma: no cover - defensive fallback
            raise DocumentPreviewParseError(
                document_id=document_id,
                blob_name=document.blob_name,
                reason=type(exc).__name__,
            ) from exc

        logger.info(
            "document.preview.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                sheet_name=preview.name,
                sheet_index=preview.index,
            ),
        )
        return preview

    def stream_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> tuple[DocumentOut, Iterator[bytes]]:
        """Return a document record and iterator for its bytes."""

        logger.debug(
            "document.stream.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )

        document = self._get_document(workspace_id, document_id)
        current_version = document.current_version
        if current_version is None:
            logger.warning(
                "document.stream.missing_file",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                    blob_name=document.blob_name,
                ),
            )
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
            )

        try:
            stream = self._storage.stream(
                document.blob_name,
                version_id=current_version.storage_version_id,
                chunk_size=self._settings.blob_download_chunk_size_bytes,
            )
        except FileNotFoundError as exc:
            logger.warning(
                "document.stream.file_missing",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                    blob_name=document.blob_name,
                ),
            )
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
            ) from exc

        def _guarded() -> Iterator[bytes]:
            try:
                for chunk in stream:
                    yield chunk
            except FileNotFoundError as exc:
                logger.warning(
                    "document.stream.file_lost_during_stream",
                    extra=log_context(
                        workspace_id=workspace_id,
                        document_id=document_id,
                        blob_name=document.blob_name,
                    ),
                )
                raise DocumentFileMissingError(
                    document_id=document_id,
                    blob_name=document.blob_name,
                ) from exc

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload])
        self._apply_derived_fields(payload)

        logger.info(
            "document.stream.ready",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                byte_size=document.byte_size,
                content_type=document.content_type,
            ),
        )
        return payload, _guarded()

    def stream_document_version(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        version_no: int,
    ) -> tuple[DocumentOut, FileVersion, Iterator[bytes]]:
        """Return a document record, version metadata, and iterator for its bytes."""

        logger.debug(
            "document.stream_version.start",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                version_no=version_no,
            ),
        )

        document = self._get_document(workspace_id, document_id)
        version = self._get_document_version(document=document, version_no=version_no)

        try:
            stream = self._storage.stream(
                document.blob_name,
                version_id=version.storage_version_id,
                chunk_size=self._settings.blob_download_chunk_size_bytes,
            )
        except FileNotFoundError as exc:
            logger.warning(
                "document.stream_version.file_missing",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                    version_no=version_no,
                    stored_uri=document.blob_name,
                ),
            )
            raise DocumentFileMissingError(
                document_id=document_id,
                blob_name=document.blob_name,
                version_id=version.storage_version_id,
            ) from exc

        def _guarded() -> Iterator[bytes]:
            try:
                for chunk in stream:
                    yield chunk
            except FileNotFoundError as exc:
                logger.warning(
                    "document.stream_version.file_lost",
                    extra=log_context(
                        workspace_id=workspace_id,
                        document_id=document_id,
                        version_no=version_no,
                        stored_uri=document.blob_name,
                    ),
                )
                raise DocumentFileMissingError(
                    document_id=document_id,
                    blob_name=document.blob_name,
                    version_id=version.storage_version_id,
                ) from exc

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload])
        self._apply_derived_fields(payload)

        logger.info(
            "document.stream_version.ready",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                version_no=version_no,
                byte_size=version.byte_size,
            ),
        )
        return payload, version, _guarded()

    def delete_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        actor: User | None = None,
    ) -> None:
        """Soft delete ``document_id`` and remove the stored file."""

        actor_id: UUID | None = actor.id if actor is not None else None
        logger.debug(
            "document.delete.start",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                user_id=actor_id,
            ),
        )

        document = self._get_document(workspace_id, document_id)
        now = datetime.now(tz=UTC)
        document.deleted_at = now
        if actor is not None:
            document.deleted_by_user_id = actor_id
        self._session.flush()

        self._storage.delete(document.blob_name)

        logger.info(
            "document.delete.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                user_id=actor_id,
                blob_name=document.blob_name,
            ),
        )

    def delete_documents_batch(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
        actor: User | None = None,
    ) -> list[UUID]:
        """Soft delete multiple documents and remove stored files."""

        ordered_ids = list(dict.fromkeys(document_ids))
        if not ordered_ids:
            return []

        actor_id: UUID | None = actor.id if actor is not None else None
        documents = self._require_documents(
            workspace_id=workspace_id,
            document_ids=ordered_ids,
        )
        document_by_id = {doc.id: doc for doc in documents}

        now = datetime.now(tz=UTC)
        for document in documents:
            document.deleted_at = now
            if actor is not None:
                document.deleted_by_user_id = actor_id

        self._session.flush()

        for document in documents:
            self._storage.delete(document.blob_name)

        return ordered_ids

    def replace_document_tags(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        tags: list[str],
    ) -> DocumentOut:
        """Replace tags on a document in a single transaction."""

        try:
            normalized = normalize_tag_list(tags, max_tags=MAX_TAGS_PER_DOCUMENT)
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        document = self._get_document(workspace_id, document_id)
        document.tags = [FileTag(file_id=document.id, tag=tag) for tag in normalized]
        self._session.flush()

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload])
        self._apply_derived_fields(payload)
        return payload

    def patch_document_tags(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        add: list[str] | None = None,
        remove: list[str] | None = None,
    ) -> DocumentOut:
        """Add or remove tags on a document."""

        if not add and not remove:
            raise InvalidDocumentTagsError("add or remove is required")

        try:
            normalized_add = normalize_tag_list(add or [])
            normalized_remove = normalize_tag_list(remove or [])
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        document = self._get_document(workspace_id, document_id)
        existing = {tag.tag for tag in document.tags}
        next_tags = (existing | set(normalized_add)) - set(normalized_remove)
        if len(next_tags) > MAX_TAGS_PER_DOCUMENT:
            raise InvalidDocumentTagsError(f"Too many tags; max {MAX_TAGS_PER_DOCUMENT}.")

        to_remove = existing - next_tags
        if to_remove:
            document.tags = [tag for tag in document.tags if tag.tag not in to_remove]
        to_add = next_tags - existing
        for tag in to_add:
            document.tags.append(FileTag(file_id=document.id, tag=tag))

        self._session.flush()

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload])
        self._apply_derived_fields(payload)
        return payload

    def patch_document_tags_batch(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
        add: list[str] | None = None,
        remove: list[str] | None = None,
    ) -> list[DocumentOut]:
        """Add or remove tags on multiple documents."""

        if not add and not remove:
            raise InvalidDocumentTagsError("add or remove is required")

        try:
            normalized_add = normalize_tag_list(add or [])
            normalized_remove = normalize_tag_list(remove or [])
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        ordered_ids = list(dict.fromkeys(document_ids))
        documents = self._require_documents(
            workspace_id=workspace_id,
            document_ids=ordered_ids,
        )
        document_by_id = {doc.id: doc for doc in documents}

        for document in documents:
            existing = {tag.tag for tag in document.tags}
            next_tags = (existing | set(normalized_add)) - set(normalized_remove)
            if len(next_tags) > MAX_TAGS_PER_DOCUMENT:
                raise InvalidDocumentTagsError(f"Too many tags; max {MAX_TAGS_PER_DOCUMENT}.")

            to_remove = existing - next_tags
            if to_remove:
                document.tags = [tag for tag in document.tags if tag.tag not in to_remove]
            to_add = next_tags - existing
            for tag in to_add:
                document.tags.append(FileTag(file_id=document.id, tag=tag))

        self._session.flush()

        payloads = [DocumentOut.model_validate(document_by_id[doc_id]) for doc_id in ordered_ids]
        self._attach_last_runs(workspace_id, payloads)
        for payload in payloads:
            self._apply_derived_fields(payload)
        return payloads

    def list_tag_catalog(
        self,
        *,
        workspace_id: UUID,
        limit: int,
        cursor: str | None,
        include_total: bool,
        sort: list[str],
        q: str | None = None,
    ) -> TagCatalogPage:
        """Return tag catalog entries with counts."""

        try:
            normalized_q = normalize_tag_query(q)
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        count_expr = func.count(FileTag.file_id).label("document_count")
        stmt = (
            select(
                FileTag.tag.label("tag"),
                count_expr,
            )
            .join(File, FileTag.file_id == File.id)
            .where(
                File.workspace_id == workspace_id,
                File.deleted_at.is_(None),
                File.kind == FileKind.INPUT,
            )
            .group_by(FileTag.tag)
        )

        if normalized_q:
            pattern = f"%{normalized_q}%"
            stmt = stmt.where(FileTag.tag.like(pattern))

        sort_fields = {
            "name": (FileTag.tag.asc(), FileTag.tag.desc()),
            "count": (count_expr.asc(), count_expr.desc()),
        }

        cursor_fields = {
            "id": cursor_field(lambda item: item.tag, parse_str),
            "name": cursor_field(lambda item: item.tag, parse_str),
            "count": cursor_field(lambda item: item.document_count, parse_int),
        }
        resolved_sort = resolve_cursor_sort(
            sort,
            allowed=sort_fields,
            cursor_fields=cursor_fields,
            default=["name"],
            id_field=(FileTag.tag.asc(), FileTag.tag.desc()),
        )
        page_result = paginate_query_cursor(
            self._session,
            stmt,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
            changes_cursor="0",
            row_mapper=lambda row: TagCatalogItem(
                tag=row["tag"],
                document_count=int(row["document_count"] or 0),
            ),
        )

        return TagCatalogPage(items=page_result.items, meta=page_result.meta, facets=page_result.facets)

    def _get_document(self, workspace_id: UUID, document_id: UUID) -> File:
        document = self._repository.get_document(
            workspace_id=workspace_id,
            document_id=document_id,
        )
        if document is None:
            logger.warning(
                "document.get.not_found",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                ),
            )
            raise DocumentNotFoundError(document_id)
        return document

    def _get_document_version(self, *, document: File, version_no: int) -> FileVersion:
        stmt = select(FileVersion).where(
            FileVersion.file_id == document.id,
            FileVersion.version_no == version_no,
        )
        result = self._session.execute(stmt)
        version = result.scalar_one_or_none()
        if version is None:
            raise DocumentVersionNotFoundError(
                document_id=document.id,
                version_no=version_no,
            )
        return version

    def _resolve_comment_mentions(
        self,
        *,
        workspace_id: UUID,
        mentions: Sequence[UUID] | None,
    ) -> list[User]:
        if not mentions:
            return []
        unique_ids = list(dict.fromkeys(mentions))
        stmt = (
            select(User)
            .join(
                WorkspaceMembership,
                WorkspaceMembership.user_id == User.id,
            )
            .where(WorkspaceMembership.workspace_id == workspace_id)
            .where(User.id.in_(unique_ids))
        )
        users = list(self._session.execute(stmt).scalars())
        found = {user.id for user in users}
        missing = [str(user_id) for user_id in unique_ids if user_id not in found]
        if missing:
            raise InvalidDocumentCommentMentionsError(
                f"Unknown or non-member mentions: {', '.join(missing)}"
            )
        user_by_id = {user.id: user for user in users}
        return [user_by_id[user_id] for user_id in unique_ids]

    def _require_documents(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
    ) -> list[File]:
        if not document_ids:
            return []

        stmt = (
            self._repository.base_query(workspace_id)
            .where(File.deleted_at.is_(None))
            .where(File.id.in_(document_ids))
        )
        result = self._session.execute(stmt)
        documents = list(result.scalars())
        found = {doc.id for doc in documents}
        missing = [document_id for document_id in document_ids if document_id not in found]
        if missing:
            logger.warning(
                "document.require_documents.not_found",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=str(missing[0]),
                    missing_count=len(missing),
                ),
            )
            raise DocumentNotFoundError(missing[0])

        return documents

    def _attach_last_runs(
        self,
        workspace_id: UUID,
        documents: Sequence[DocumentOut],
        *,
        include_run_metrics: bool = False,
        include_run_table_columns: bool = False,
        include_run_fields: bool = False,
    ) -> None:
        """Populate last run summaries and metrics on each document."""

        if not documents:
            return

        doc_ids = [doc.id for doc in documents]
        last_runs = self._last_runs_by_document(
            workspace_id=workspace_id,
            document_ids=doc_ids,
        )
        run_ids = [run.id for run in last_runs.values()]
        metrics_by_run_id = (
            self._last_run_metrics(run_ids=run_ids) if include_run_metrics else {}
        )
        table_columns_by_run_id = (
            self._last_run_table_columns(run_ids=run_ids) if include_run_table_columns else {}
        )
        fields_by_run_id = (
            self._last_run_fields(run_ids=run_ids) if include_run_fields else {}
        )
        for document in documents:
            run = last_runs.get(document.id)
            document.last_run = run
            document.last_run_metrics = (
                metrics_by_run_id.get(run.id) if run and include_run_metrics else None
            )
            document.last_run_table_columns = (
                table_columns_by_run_id.get(run.id, [])
                if run and include_run_table_columns
                else None
            )
            document.last_run_fields = (
                fields_by_run_id.get(run.id, []) if run and include_run_fields else None
            )

    def _apply_derived_fields(self, document: DocumentOut) -> None:
        updated_at = document.updated_at
        latest_at = self._last_run_at(document.last_run)
        document.activity_at = latest_at if latest_at and latest_at > updated_at else updated_at

    def _build_list_row(self, document: DocumentOut) -> DocumentListRow:
        activity_at = document.activity_at or document.updated_at
        return DocumentListRow(
            id=document.id,
            workspace_id=document.workspace_id,
            name=document.name,
            file_type=self._derive_file_type(document.name),
            uploader=document.uploader,
            assignee=document.assignee,
            tags=document.tags,
            byte_size=document.byte_size,
            current_version_no=document.current_version_no,
            comment_count=document.comment_count,
            created_at=document.created_at,
            updated_at=document.updated_at,
            activity_at=activity_at,
            last_run=document.last_run,
            last_run_metrics=document.last_run_metrics,
            last_run_table_columns=document.last_run_table_columns,
            last_run_fields=document.last_run_fields,
        )

    def _build_document_facets(self, stmt: Select) -> dict[str, Any]:
        filtered_ids = stmt.order_by(None).with_only_columns(File.id).distinct().subquery()

        def coerce_value(value: Any) -> Any:
            if isinstance(value, Enum):
                return value.value
            if isinstance(value, UUID):
                return str(value)
            return value

        def build_buckets(expr) -> list[dict[str, Any]]:
            rows = self._session.execute(
                select(
                    expr.label("value"),
                    func.count().label("count"),
                )
                .select_from(File)
                .join(filtered_ids, filtered_ids.c.id == File.id)
                .group_by(expr)
            ).all()
            buckets = [
                {"value": coerce_value(value), "count": int(count or 0)}
                for value, count in rows
            ]
            buckets.sort(key=lambda bucket: str(bucket["value"]))
            return buckets

        lower_name = func.lower(File.name)
        file_type_expr = case(
            (lower_name.like("%.xlsx"), DocumentFileType.XLSX.value),
            (lower_name.like("%.xls"), DocumentFileType.XLS.value),
            (lower_name.like("%.csv"), DocumentFileType.CSV.value),
            (lower_name.like("%.pdf"), DocumentFileType.PDF.value),
            else_=DocumentFileType.UNKNOWN.value,
        )
        file_type_subquery = (
            select(
                File.id.label("document_id"),
                file_type_expr.label("file_type"),
            )
            .select_from(File)
            .subquery()
        )
        file_type_rows = self._session.execute(
            select(
                file_type_subquery.c.file_type.label("value"),
                func.count().label("count"),
            )
            .select_from(file_type_subquery)
            .join(filtered_ids, filtered_ids.c.id == file_type_subquery.c.document_id)
            .group_by(file_type_subquery.c.file_type)
        ).all()
        file_type_buckets = [
            {"value": coerce_value(value), "count": int(count or 0)}
            for value, count in file_type_rows
        ]
        file_type_buckets.sort(key=lambda bucket: str(bucket["value"]))

        timestamp = func.coalesce(Run.completed_at, Run.started_at, Run.created_at)
        ranked_runs = (
            select(
                FileVersion.file_id.label("document_id"),
                Run.status.label("status"),
                func.row_number()
                .over(
                    partition_by=FileVersion.file_id,
                    order_by=timestamp.desc(),
                )
                .label("rank"),
            )
            .select_from(Run)
            .join(FileVersion, Run.input_file_version_id == FileVersion.id)
            .where(FileVersion.file_id.in_(select(filtered_ids.c.id)))
            .subquery()
        )
        status_rows = self._session.execute(
            select(
                ranked_runs.c.status.label("value"),
                func.count().label("count"),
            )
            .select_from(File)
            .join(filtered_ids, filtered_ids.c.id == File.id)
            .outerjoin(
                ranked_runs,
                (ranked_runs.c.document_id == File.id)
                & (ranked_runs.c.rank == 1),
            )
            .group_by(ranked_runs.c.status)
        ).all()
        status_buckets = [
            {"value": coerce_value(value), "count": int(count or 0)}
            for value, count in status_rows
        ]
        status_buckets.sort(key=lambda bucket: str(bucket["value"]))

        return {
            "lastRunPhase": {"buckets": status_buckets},
            "fileType": {"buckets": file_type_buckets},
        }

    @staticmethod
    def _derive_file_type(name: str) -> DocumentFileType:
        suffix = Path(name).suffix.lower().lstrip(".")
        if suffix == "xlsx":
            return DocumentFileType.XLSX
        if suffix == "xls":
            return DocumentFileType.XLS
        if suffix == "csv":
            return DocumentFileType.CSV
        if suffix == "pdf":
            return DocumentFileType.PDF
        return DocumentFileType.UNKNOWN

    @staticmethod
    def _last_run_at(run: DocumentRunSummary | None) -> datetime | None:
        if run is None:
            return None
        return run.completed_at or run.started_at or run.created_at

    def _last_runs_by_document(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
    ) -> dict[UUID, DocumentRunSummary]:
        if not document_ids:
            return {}

        timestamp = func.coalesce(Run.completed_at, Run.started_at, Run.created_at)
        ranked_runs = (
            select(
                FileVersion.file_id.label("document_id"),
                Run.id.label("run_id"),
                Run.status.label("status"),
                Run.error_message.label("error_message"),
                Run.completed_at.label("completed_at"),
                Run.started_at.label("started_at"),
                Run.created_at.label("created_at"),
                func.row_number()
                .over(
                    partition_by=FileVersion.file_id,
                    order_by=timestamp.desc(),
                )
                .label("rank"),
            )
            .where(
                Run.workspace_id == workspace_id,
                FileVersion.file_id.in_(list(document_ids)),
            )
            .select_from(Run)
            .join(FileVersion, Run.input_file_version_id == FileVersion.id)
            .subquery()
        )

        stmt = select(ranked_runs).where(ranked_runs.c.rank == 1)
        result = self._session.execute(stmt)

        runs_by_doc: dict[UUID, DocumentRunSummary] = {}
        for row in result.mappings():
            document_id = row["document_id"]
            if not isinstance(document_id, UUID):
                document_id = UUID(str(document_id))
            status = row["status"]
            if not isinstance(status, RunStatus):
                status = RunStatus(str(status))
            runs_by_doc[document_id] = DocumentRunSummary(
                id=row["run_id"],
                status=status,
                created_at=self._ensure_utc(row.get("created_at")),
                started_at=self._ensure_utc(row.get("started_at")),
                completed_at=self._ensure_utc(row.get("completed_at")),
                error_message=row.get("error_message"),
            )

        return runs_by_doc

    def _last_run_metrics(self, *, run_ids: Sequence[UUID]) -> dict[UUID, RunMetricsResource]:
        unique_ids = list({run_id for run_id in run_ids if run_id is not None})
        if not unique_ids:
            return {}
        stmt = select(RunMetrics).where(RunMetrics.run_id.in_(unique_ids))
        result = self._session.execute(stmt)
        metrics_by_run: dict[UUID, RunMetricsResource] = {}
        for metrics in result.scalars():
            metrics_by_run[metrics.run_id] = RunMetricsResource.model_validate(metrics)
        return metrics_by_run

    def _last_run_fields(self, *, run_ids: Sequence[UUID]) -> dict[UUID, list[RunFieldResource]]:
        unique_ids = list({run_id for run_id in run_ids if run_id is not None})
        if not unique_ids:
            return {}
        stmt = (
            select(RunField)
            .where(RunField.run_id.in_(unique_ids))
            .order_by(RunField.run_id.asc(), RunField.field.asc())
        )
        result = self._session.execute(stmt)
        fields_by_run: dict[UUID, list[RunFieldResource]] = {}
        for field in result.scalars():
            fields_by_run.setdefault(field.run_id, []).append(
                RunFieldResource.model_validate(field)
            )
        return fields_by_run

    def _last_run_table_columns(
        self, *, run_ids: Sequence[UUID]
    ) -> dict[UUID, list[RunColumnResource]]:
        unique_ids = list({run_id for run_id in run_ids if run_id is not None})
        if not unique_ids:
            return {}
        stmt = (
            select(RunTableColumn)
            .where(RunTableColumn.run_id.in_(unique_ids))
            .order_by(
                RunTableColumn.run_id.asc(),
                RunTableColumn.workbook_index.asc(),
                RunTableColumn.sheet_index.asc(),
                RunTableColumn.table_index.asc(),
                RunTableColumn.column_index.asc(),
            )
        )
        result = self._session.execute(stmt)
        columns_by_run: dict[UUID, list[RunColumnResource]] = {}
        for column in result.scalars():
            columns_by_run.setdefault(column.run_id, []).append(
                RunColumnResource.model_validate(column)
            )
        return columns_by_run

    @staticmethod
    def _ensure_utc(dt: datetime | None) -> datetime | None:
        if dt is None:
            return None
        if dt.tzinfo is None:
            return dt.replace(tzinfo=UTC)
        return dt

    def _normalise_filename(self, name: str | None) -> str:
        """Return a safe, display-friendly filename for stored documents."""

        if name is None:
            return _FALLBACK_FILENAME

        candidate = name.strip()
        if not candidate:
            return _FALLBACK_FILENAME

        # Strip control characters (including newlines) to avoid header injection and
        # other control sequence issues when the filename is rendered in responses.
        filtered = "".join(ch for ch in candidate if unicodedata.category(ch)[0] != "C").strip()

        if not filtered:
            return _FALLBACK_FILENAME

        if len(filtered) > _MAX_FILENAME_LENGTH:
            filtered = filtered[:_MAX_FILENAME_LENGTH].rstrip()

        return filtered or _FALLBACK_FILENAME

    def _normalise_content_type(self, content_type: str | None) -> str | None:
        if content_type is None:
            return None
        candidate = content_type.strip()
        return candidate or None

    def _file_blob_name(self, workspace_id: UUID, file_id: UUID) -> str:
        return f"{workspace_id}/files/{file_id}"

    def _build_name_key(self, name: str) -> str:
        normalized = unicodedata.normalize("NFKC", name)
        collapsed = " ".join(normalized.split())
        return collapsed.casefold()

    def _find_by_name_key(self, *, workspace_id: UUID, name_key: str) -> File | None:
        stmt = (
            self._repository.base_query(workspace_id)
            .where(File.deleted_at.is_(None))
            .where(File.name_key == name_key)
        )
        result = self._session.execute(stmt)
        return result.scalar_one_or_none()

    def _disambiguate_name(self, *, workspace_id: UUID, base_name: str) -> tuple[str, str]:
        stem = Path(base_name).stem.strip() or "Document"
        suffix = Path(base_name).suffix
        for index in range(2, 1000):
            label = f" ({index})"
            max_stem = max(1, _MAX_FILENAME_LENGTH - len(label) - len(suffix))
            trimmed = stem[:max_stem].rstrip() or "Document"
            candidate = f"{trimmed}{label}{suffix}"
            name_key = self._build_name_key(candidate)
            if self._find_by_name_key(workspace_id=workspace_id, name_key=name_key) is None:
                return candidate, name_key
        raise RuntimeError("Unable to disambiguate document name.")

    def _next_version_no(self, *, document_id: UUID) -> int:
        stmt = select(func.max(FileVersion.version_no)).where(FileVersion.file_id == document_id)
        current = self._session.execute(stmt).scalar_one()
        return int(current or 0) + 1

    @contextmanager
    def _download_blob_to_tempfile(
        self,
        *,
        blob_name: str,
        version_id: str | None,
        suffix: str | None = None,
    ) -> Iterator[Path]:
        suffix = suffix or ""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / f"document{suffix}"
            with path.open("wb") as handle:
                for chunk in self._storage.stream(
                    blob_name,
                    version_id=version_id,
                    chunk_size=self._settings.blob_download_chunk_size_bytes,
                ):
                    handle.write(chunk)
            yield path

    @staticmethod
    def _inspect_workbook(path: Path) -> list[DocumentSheet]:
        with path.open("rb") as raw:
            workbook = openpyxl.load_workbook(
                raw,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            try:
                sheetnames = workbook.sheetnames
                active = workbook.active.title if sheetnames else None
                return [
                    DocumentSheet(
                        name=title,
                        index=index,
                        kind="worksheet",
                        is_active=title == active,
                    )
                    for index, title in enumerate(sheetnames)
                ]
            finally:
                workbook.close()

    @staticmethod
    def _default_sheet_name(name: str | None) -> str:
        stem = Path(name or "Sheet").stem.strip() or "Sheet"
        return stem[:64]


__all__ = ["DocumentsService"]
