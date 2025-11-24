"""Output composition helpers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from ade_engine.core.manifest import ManifestContext
from ade_engine.core.models import JobContext
from .models import FileExtraction
from .util import unique_sheet_name


def _combined_headers(manifest: ManifestContext, extractions: list[FileExtraction]) -> list[str]:
    """Build a shared header row for combined outputs."""

    order = manifest.column_order
    meta = manifest.column_meta
    headers = [meta.get(field, {}).get("label", field) for field in order]
    extras: list[str] = []
    seen: set[str] = set()
    for extraction in extractions:
        for extra in extraction.extra_columns:
            if extra.output_header not in seen:
                seen.add(extra.output_header)
                extras.append(extra.output_header)
    headers.extend(extras)
    return headers


def _append_rows(sheet, headers: list[str], order_len: int, extraction: FileExtraction) -> None:
    """Append rows to a sheet using a unified header layout."""

    header_to_index = {name: idx for idx, name in enumerate(headers)}
    for row in extraction.rows:
        base = row[:order_len]
        extras_values = row[order_len:]
        extra_map = {
            extra.output_header: extras_values[idx]
            for idx, extra in enumerate(extraction.extra_columns)
            if idx < len(extras_values)
        }
        padded = list(base)
        extra_start = len(base)
        for name in headers[extra_start:]:
            padded.append(extra_map.get(name, ""))
        sheet.append(padded)


def write_outputs(
    job: JobContext,
    manifest: ManifestContext,
    extractions: list[FileExtraction],
) -> Path:
    """Persist normalized rows into an Excel workbook honoring writer settings."""

    writer_cfg = manifest.writer
    raw_writer = (
        manifest.raw.get("engine", {}).get("writer", {})
        if isinstance(manifest.raw, dict)
        else {}
    )
    output_sheet = writer_cfg.output_sheet if isinstance(writer_cfg.output_sheet, str) else None
    output_sheet_configured = isinstance(raw_writer, dict) and "output_sheet" in raw_writer
    in_memory = writer_cfg.mode == "in_memory"

    output_path = job.paths.output_dir / "normalized.xlsx"
    used_sheet_names: set[str] = set()

    tables_payload = []
    if in_memory:
        for extraction in extractions:
            tables_payload.append(
                {
                    "sheet": extraction.sheet_name,
                    "rows": [list(row) for row in extraction.rows],
                    "header": output_headers(manifest, extraction),
                }
            )
        job.metadata["output_tables"] = tables_payload

    workbook = Workbook(write_only=not in_memory)

    try:
        if output_sheet and output_sheet_configured and len(extractions) > 1:
            headers = _combined_headers(manifest, extractions)
            sheet = workbook.create_sheet(title=output_sheet[:31])
            sheet.append(headers)
            for extraction in extractions:
                _append_rows(sheet, headers, len(manifest.column_order), extraction)
        else:
            for extraction in extractions:
                sheet_title = output_sheet if output_sheet_configured and output_sheet else extraction.sheet_name
                sheet_title = unique_sheet_name(sheet_title, used_sheet_names)
                sheet = workbook.create_sheet(title=sheet_title)
                header_cells = output_headers(manifest, extraction)
                sheet.append(header_cells)
                for row in extraction.rows:
                    sheet.append(row)

        tmp_path = output_path.with_suffix(".xlsx.tmp")
        workbook.save(tmp_path)
        tmp_path.replace(output_path)
        return output_path
    finally:
        workbook.close()


def output_headers(manifest: ManifestContext, extraction: FileExtraction) -> list[str]:
    """Build output headers combining manifest labels and unmapped columns."""

    order = manifest.column_order
    meta = manifest.column_meta
    headers = [meta.get(field, {}).get("label", field) for field in order]
    headers.extend(extra.output_header for extra in extraction.extra_columns)
    return headers


__all__ = ["output_headers", "write_outputs"]
