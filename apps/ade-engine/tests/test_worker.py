from __future__ import annotations

import json
import sys
import textwrap
from pathlib import Path

import openpyxl
import pytest

from ade_engine.schemas import ADE_TELEMETRY_EVENT_SCHEMA, TelemetryEnvelope

from ade_engine.worker import run_job


def _setup_config_package(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    *,
    include_hooks: bool = False,
) -> Path:
    pkg_root = tmp_path / ("config_pkg_hooks" if include_hooks else "config_pkg")
    config_pkg = pkg_root / "ade_config"
    detectors_dir = config_pkg / "column_detectors"
    detectors_dir.mkdir(parents=True)
    (config_pkg / "__init__.py").write_text("", encoding="utf-8")
    (detectors_dir / "__init__.py").write_text("", encoding="utf-8")

    member_module = textwrap.dedent(
        """
        def detect_from_header(**kwargs):
            header = kwargs.get("header") or ""
            if header == "id":
                return {"scores": {kwargs["field_name"]: 1.0}}
            return {"scores": {}}

        def transform(*, value, row, field_name, **_):
            if value is None:
                return None
            normalized = str(value).strip()
            row[field_name] = normalized
            return {field_name: normalized}
        """
    )
    (detectors_dir / "member_id.py").write_text(member_module, encoding="utf-8")

    email_module = textwrap.dedent(
        """
        def detect_from_header(**kwargs):
            header = kwargs.get("header") or ""
            if header == "email":
                return {"scores": {kwargs["field_name"]: 1.0}}
            sample = kwargs.get("column_values_sample", [])
            if any("@" in str(value) for value in sample if value):
                return {"scores": {kwargs["field_name"]: 0.5}}
            return {"scores": {}}

        def transform(*, value, row, field_name, **_):
            if value is None:
                return None
            normalized = str(value).strip().lower()
            row[field_name] = normalized
            return {field_name: normalized}

        def validate(*, value, field_meta=None, row_index, field_name, **_):
            issues = []
            if field_meta and field_meta.get("required") and not value:
                issues.append({"code": "required_missing"})
            if value and "@" not in value:
                issues.append({"code": "invalid_email"})
            for issue in issues:
                issue.setdefault("field", field_name)
                issue.setdefault("row_index", row_index)
            return issues
        """
    )
    (detectors_dir / "email.py").write_text(email_module, encoding="utf-8")

    manifest: dict[str, object] = {
        "schema_version": "ade.manifest@1",
        "script_api": 1,
        "engine": {
            "defaults": {
                "mapping_score_threshold": 0.25,
                "detector_sample_size": 8,
            },
            "writer": {
                "append_unmapped_columns": True,
                "unmapped_prefix": "raw_",
            },
        },
        "hooks": {},
        "columns": {
            "order": ["member_id", "email"],
            "meta": {
                "member_id": {
                    "label": "Member ID",
                    "synonyms": ["ID"],
                    "script": "column_detectors/member_id.py",
                },
                "email": {
                    "label": "Email",
                    "required": True,
                    "script": "column_detectors/email.py",
                },
            },
        },
    }

    if include_hooks:
        hooks_dir = config_pkg / "hooks"
        hooks_dir.mkdir(parents=True)
        (hooks_dir / "__init__.py").write_text("", encoding="utf-8")
        (hooks_dir / "on_job_start.py").write_text(
            "def run(*, artifact, **_):\n    artifact.note('start hook')\n",
            encoding="utf-8",
        )
        (hooks_dir / "on_job_end.py").write_text(
            "def run(*, artifact, result, **_):\n    artifact.note('end hook', status=result.status)\n",
            encoding="utf-8",
        )
        manifest["hooks"] = {
            "on_job_start": [{"script": "hooks/on_job_start.py"}],
            "on_job_end": [{"script": "hooks/on_job_end.py"}],
        }

    manifest_path = config_pkg / "manifest.json"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    monkeypatch.syspath_prepend(str(pkg_root))
    for name in list(sys.modules):
        if name == "ade_config" or name.startswith("ade_config."):
            monkeypatch.delitem(sys.modules, name, raising=False)

    return manifest_path


def test_run_job_normalizes_csv(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    manifest_path = _setup_config_package(tmp_path, monkeypatch)

    jobs_dir = tmp_path / "jobs"
    job_dir = jobs_dir / "job-1"
    input_dir = job_dir / "input"
    logs_dir = job_dir / "logs"
    input_dir.mkdir(parents=True)
    logs_dir.mkdir(parents=True)

    csv_path = input_dir / "employees.csv"
    csv_path.write_text(
        "ID,Email,Name\n123,USER@EXAMPLE.COM,Alice\n456,invalid,Bob\n",
        encoding="utf-8",
    )

    result = run_job(
        "job-1",
        jobs_dir=jobs_dir,
        manifest_path=manifest_path,
        config_package="ade_config",
    )

    assert result.status == "succeeded"
    assert result.processed_files == ("employees.csv",)
    workbook = openpyxl.load_workbook(result.output_paths[0], read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Member ID", "Email", "raw_Name")
    assert rows[1][:2] == ("123", "user@example.com")
    assert rows[2][1] == "invalid"
    workbook.close()

    artifact = json.loads(result.artifact_path.read_text(encoding="utf-8"))
    assert artifact["job"]["status"] == "succeeded"
    table = artifact["tables"][0]
    assert table["mapping"][0]["field"] == "member_id"
    assert any(entry["code"] == "invalid_email" for entry in table["validation"])

    envelopes = [
        TelemetryEnvelope.model_validate_json(line)
        for line in result.events_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    assert all(envelope.schema == ADE_TELEMETRY_EVENT_SCHEMA for envelope in envelopes)
    assert any(env.event.name == "job_started" for env in envelopes)
    assert any(env.event.name == "job_completed" for env in envelopes)
    assert any(env.event.name == "validation_issue" for env in envelopes)


def test_hooks_are_executed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    manifest_path = _setup_config_package(tmp_path, monkeypatch, include_hooks=True)

    jobs_dir = tmp_path / "jobs"
    job_dir = jobs_dir / "job-2"
    input_dir = job_dir / "input"
    logs_dir = job_dir / "logs"
    input_dir.mkdir(parents=True)
    logs_dir.mkdir(parents=True)
    (input_dir / "data.csv").write_text("ID\n42\n", encoding="utf-8")

    result = run_job(
        "job-2",
        jobs_dir=jobs_dir,
        manifest_path=manifest_path,
        config_package="ade_config",
    )

    assert result.status == "succeeded"
    artifact = json.loads(result.artifact_path.read_text(encoding="utf-8"))
    notes = [entry["message"] for entry in artifact["notes"]]
    assert "start hook" in notes
    assert any(
        entry["message"] == "end hook" and entry.get("details", {}).get("status") == "succeeded"
        for entry in artifact["notes"]
    )
