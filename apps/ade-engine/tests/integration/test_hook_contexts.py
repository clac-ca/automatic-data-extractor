from __future__ import annotations

from pathlib import Path

import openpyxl

from ade_engine.application.engine import Engine
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.run import RunRequest, RunStatus


def _write_config_package(root: Path) -> None:
    pkg = root / "ade_config"
    pkg.mkdir(parents=True, exist_ok=True)
    (pkg / "__init__.py").write_text("", encoding="utf-8")

    row_detectors = pkg / "row_detectors"
    row_detectors.mkdir(parents=True, exist_ok=True)
    (row_detectors / "__init__.py").write_text("", encoding="utf-8")
    (row_detectors / "two_tables.py").write_text(
        """
from ade_engine.models import RowKind

def detect_two_tables(*, row_index, **_):
    if row_index in (1, 5):
        return {RowKind.HEADER.value: 1.0}
    if row_index in (2, 3, 6):
        return {RowKind.DATA.value: 1.0}
    return {}

def register(registry):
    registry.register_row_detector(detect_two_tables, row_kind=RowKind.UNKNOWN.value, priority=0)
""",
        encoding="utf-8",
    )

    hooks = pkg / "hooks"
    hooks.mkdir(parents=True, exist_ok=True)
    (hooks / "__init__.py").write_text("", encoding="utf-8")
    (hooks / "capture_sheet_tables.py").write_text(
        """
def on_sheet_end(*, output_sheet, tables, **_):
    indices = [t.table_index for t in tables]
    output_sheet.append([\"table_indices\", *indices])

def register(registry):
    registry.register_hook(on_sheet_end, hook=\"on_sheet_end\", priority=0)
""",
        encoding="utf-8",
    )


def _write_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.append([3, 4])
    ws.append([None, None])
    ws.append(["C", "D"])
    ws.append([5, 6])
    wb.save(path)


def test_on_sheet_end_receives_tables_in_order(tmp_path: Path) -> None:
    config_root = tmp_path / "cfg"
    config_root.mkdir()
    _write_config_package(config_root)

    source = tmp_path / "input.xlsx"
    _write_workbook(source)

    engine = Engine(settings=Settings())
    output_dir = tmp_path / "out"
    logs_dir = tmp_path / "logs"
    req = RunRequest(config_package=config_root, input_file=source, output_dir=output_dir, logs_dir=logs_dir)

    result = engine.run(req)
    assert result.status == RunStatus.SUCCEEDED

    output_file = output_dir / f"{source.stem}_normalized.xlsx"
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
    assert last_row[:3] == ("table_indices", 0, 1)

    wb.close()
