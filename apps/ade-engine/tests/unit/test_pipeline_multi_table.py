from __future__ import annotations

from openpyxl import Workbook

from ade_engine.application.pipeline.pipeline import Pipeline
from ade_engine.extensions.registry import Registry
from ade_engine.infrastructure.observability.logger import NullLogger
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.extension_contexts import FieldDef, RowKind


def test_process_sheet_renders_multiple_tables_with_blank_row():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index in (0, 3):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (1, 2, 4):
            return {RowKind.DATA.value: 1.0}
        return {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(registry=registry, settings=Settings(), logger=logger)

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["A", "B"])
    source_ws.append([1, 2])
    source_ws.append([3, 4])
    source_ws.append(["C", "D"])
    source_ws.append([5, 6])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    tables = pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert [t.table_index for t in tables] == [0, 1]

    emitted = list(output_ws.iter_rows(min_row=1, max_row=6, max_col=2, values_only=True))
    assert emitted == [
        ("A", "B"),
        (1, 2),
        (3, 4),
        (None, None),  # blank separator row
        ("C", "D"),
        (5, 6),
    ]


def test_process_sheet_handles_mixed_numeric_types():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index == 0:
            return {RowKind.HEADER.value: 1.0}
        return {RowKind.DATA.value: 1.0}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(registry=registry, settings=Settings(), logger=logger)

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Amount"])
    source_ws.append([170])
    source_ws.append([169.75])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    emitted = list(output_ws.iter_rows(min_row=1, max_row=3, max_col=1, values_only=True))
    assert emitted == [("Amount",), (170.0,), (169.75,)]


def test_on_table_written_receives_written_table_after_output_policies():
    """The on_table_written hook should see exactly what was written to the output sheet."""

    registry = Registry()
    logger = NullLogger()

    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))

    def row_detector(*, row_index, **_):
        if row_index == 0:
            return {RowKind.HEADER.value: 1.0}
        return {RowKind.DATA.value: 1.0}

    def detect_email(*, header_text, **_):
        return {"email": 1.0} if header_text.strip().lower() == "email" else {}

    def detect_name(*, header_text, **_):
        return {"name": 1.0} if header_text.strip().lower() == "name" else {}

    def capture_written(*, table, sheet, state, **_):
        state["written_columns"] = list(table.columns)
        state["written_sheet_title"] = getattr(sheet, "title", getattr(sheet, "name", ""))

    registry.register_row_detector(row_detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.register_column_detector(detect_email, field="email", priority=0)
    registry.register_column_detector(detect_name, field="name", priority=0)
    registry.register_hook(capture_written, hook="on_table_written", priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(
            remove_unmapped_columns=True,
            write_diagnostics_columns=False,
        ),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Email", "Name", "Notes"])
    source_ws.append(["a@example.com", "Alice", "keep?"])
    source_ws.append(["b@example.com", "Bob", "keep?"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    state: dict = {}
    pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state=state,
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert state["written_sheet_title"] == "Sheet1"
    assert state["written_columns"] == ["email", "name"]

    headers = list(output_ws.iter_rows(min_row=1, max_row=1, max_col=2, values_only=True))[0]
    assert headers == ("email", "name")
