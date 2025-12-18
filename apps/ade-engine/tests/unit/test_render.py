from __future__ import annotations

from openpyxl import Workbook
import polars as pl

from ade_engine.application.pipeline.render import SheetWriter, render_table
from ade_engine.extensions.registry import Registry
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.extension_contexts import FieldDef
from ade_engine.models.table import TableRegion, TableResult


class DummyLogger:
    def event(self, *args, **kwargs):
        pass


def _registry_with_fields(*fields: str) -> Registry:
    reg = Registry()
    for name in fields:
        reg.register_field(FieldDef(name=name))
    reg.finalize()
    return reg


def test_render_writes_non_reserved_columns_by_default():
    wb = Workbook()
    ws = wb.active

    table = pl.DataFrame(
        {
            "email": ["a@example.com"],
            "name": ["Alice"],
            "Notes": ["note-1"],
            "__ade_issue__email": ["bad email"],
        }
    )

    table_result = TableResult(
        sheet_name="Sheet1",
        table=table,
        source_region=TableRegion(
            min_row=1,
            min_col=1,
            max_row=1 + table.height,
            max_col=len(table.columns),
        ),
        source_columns=[],
        row_count=table.height,
    )
    reg = _registry_with_fields("email", "name")

    write_table = render_table(
        table_result=table_result,
        writer=SheetWriter(ws),
        registry=reg,
        settings=Settings(),
        logger=DummyLogger(),
    )

    assert write_table.columns == ["email", "name", "Notes"]
    assert [cell.value for cell in ws[1]] == ["email", "name", "Notes"]
    assert [cell.value for cell in ws[2]] == ["a@example.com", "Alice", "note-1"]
    assert table_result.output_region is not None
    assert table_result.output_region.a1 == "A1:C2"


def test_render_remove_unmapped_columns_drops_non_canonical():
    wb = Workbook()
    ws = wb.active

    table = pl.DataFrame({"email": ["a@example.com"], "name": ["Alice"], "Notes": ["note-1"]})
    table_result = TableResult(
        sheet_name="Sheet1",
        table=table,
        source_region=TableRegion(
            min_row=1,
            min_col=1,
            max_row=1 + table.height,
            max_col=len(table.columns),
        ),
        source_columns=[],
        row_count=table.height,
    )
    reg = _registry_with_fields("email", "name")

    write_table = render_table(
        table_result=table_result,
        writer=SheetWriter(ws),
        registry=reg,
        settings=Settings(remove_unmapped_columns=True),
        logger=DummyLogger(),
    )

    assert write_table.columns == ["email", "name"]
    assert [cell.value for cell in ws[1]] == ["email", "name"]


def test_render_write_diagnostics_columns_keeps_reserved_columns():
    wb = Workbook()
    ws = wb.active

    table = pl.DataFrame({"email": ["a@example.com"], "__ade_issue__email": ["bad email"]})
    table_result = TableResult(
        sheet_name="Sheet1",
        table=table,
        source_region=TableRegion(
            min_row=1,
            min_col=1,
            max_row=1 + table.height,
            max_col=len(table.columns),
        ),
        source_columns=[],
        row_count=table.height,
    )
    reg = _registry_with_fields("email")

    write_table = render_table(
        table_result=table_result,
        writer=SheetWriter(ws),
        registry=reg,
        settings=Settings(write_diagnostics_columns=True),
        logger=DummyLogger(),
    )

    assert write_table.columns == ["email", "__ade_issue__email"]
    assert [cell.value for cell in ws[1]] == ["email", "__ade_issue__email"]
