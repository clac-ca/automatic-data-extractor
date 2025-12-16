from __future__ import annotations

from openpyxl import Workbook

from ade_engine.application.pipeline.pipeline import Pipeline
from ade_engine.extensions.registry import Registry
from ade_engine.infrastructure.observability.logger import NullLogger
from ade_engine.infrastructure.settings import Settings


def _make_pipeline(**settings_kwargs):
    settings = Settings(**settings_kwargs)
    return Pipeline(registry=Registry(), settings=settings, logger=NullLogger())


def test_materialize_rows_trims_to_populated_extent():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=5, value="x")

    pipe = _make_pipeline(max_empty_rows_run=10, max_empty_cols_run=10)
    rows = pipe._materialize_rows(ws)

    # Empty leading rows are kept; trailing empties per row are trimmed.
    assert len(rows) == 3
    assert rows[2][4] == "x"


def test_materialize_rows_empty_sheet_returns_empty_list():
    wb = Workbook()
    ws = wb.active

    pipe = _make_pipeline(max_empty_rows_run=5)
    rows = pipe._materialize_rows(ws)

    assert rows == []


def test_materialize_rows_stops_after_long_empty_row_run():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="header")
    ws.cell(row=3, column=1, value="value1")
    # add two explicit empty rows to trigger stop when limit is 2
    ws.append([])
    ws.append([])

    pipe = _make_pipeline(max_empty_rows_run=2)
    rows = pipe._materialize_rows(ws)

    # Sequence: empty row1, header, value1, empty, empty (triggers stop before append)
    assert len(rows) == 4 - 1  # last empty that triggers stop is not kept


def test_materialize_rows_breaks_after_long_empty_col_run():
    wb = Workbook()
    ws = wb.active
    ws.append(["v1", None, None, None, None, None, "late"])

    pipe = _make_pipeline(max_empty_cols_run=3)
    rows = pipe._materialize_rows(ws)

    # Row is trimmed to last value before the empty-column cutoff; far "late" is ignored.
    assert rows[0] == ["v1"]
