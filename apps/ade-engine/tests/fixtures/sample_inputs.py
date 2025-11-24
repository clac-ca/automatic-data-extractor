from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook


def sample_csv(tmp_path: Path) -> Path:
    path = tmp_path / "input.csv"
    path.write_text("member_id,value,note\n1001,42,alpha\n1002,24,beta\n")
    return path


def sample_large_csv(tmp_path: Path, *, rows: int = 5000) -> Path:
    path = tmp_path / "large.csv"
    header = "member_id,value\n"
    body = "\n".join(f"{1000 + idx},{idx}" for idx in range(rows))
    path.write_text(header + body + "\n")
    return path


def sample_xlsx_single_sheet(tmp_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Members"
    sheet.append(["member_id", "value", "note"])
    sheet.append(["1001", 42, "alpha"])
    sheet.append(["1002", 24, "beta"])

    path = tmp_path / "input.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def sample_xlsx_multi_sheet(tmp_path: Path, *, extra_rows: Iterable[list] | None = None) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Primary"
    sheet.append(["member_id", "value", "note", "surplus"])
    sheet.append(["1001", 10, "a", "keep-me"])
    sheet.append(["1002", 20, "b", "keep-me-too"])
    for row in extra_rows or []:
        sheet.append(row)

    secondary = workbook.create_sheet(title="Empty")
    secondary.append(["member_id", "value"])

    path = tmp_path / "multi.xlsx"
    workbook.save(path)
    workbook.close()
    return path
