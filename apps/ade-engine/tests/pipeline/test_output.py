from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from ade_engine.model import JobContext, JobPaths
from ade_engine.pipeline.models import ColumnMapping, ExtraColumn, FileExtraction
from ade_engine.pipeline.output import output_headers, write_outputs
from ade_engine.schemas.models import ManifestContext


def _job(tmp_path: Path) -> JobContext:
    output_dir = tmp_path / "output"
    output_dir.mkdir(parents=True)
    paths = JobPaths(
        jobs_root=tmp_path,
        job_dir=tmp_path,
        input_dir=tmp_path / "input",
        output_dir=output_dir,
        logs_dir=tmp_path / "logs",
        artifact_path=tmp_path / "logs" / "artifact.json",
        events_path=tmp_path / "logs" / "events.ndjson",
    )
    return JobContext(job_id="job", manifest={}, paths=paths, started_at=datetime.now(timezone.utc))


def test_output_headers_combines_manifest_and_extras() -> None:
    extraction = FileExtraction(
        source_name="file.csv",
        sheet_name="Sheet",
        mapped_columns=[
            ColumnMapping(field="member_id", header="ID", index=0, score=1.0, contributions=tuple())
        ],
        extra_columns=[ExtraColumn(header="Original", index=1, output_header="raw_original")],
        rows=[["123", "foo"]],
        header_row=["ID", "Original"],
        validation_issues=[],
    )
    manifest = ManifestContext(
        raw={"columns": {"order": ["member_id"], "meta": {"member_id": {"label": "Member"}}}},
        version=None,
        model=None,
    )
    headers = output_headers(manifest, extraction)
    assert headers == ["Member", "raw_original"]


def test_write_outputs_creates_workbook(tmp_path: Path) -> None:
    job = _job(tmp_path)
    extraction = FileExtraction(
        source_name="file.csv",
        sheet_name="Employees",
        mapped_columns=[
            ColumnMapping(field="member_id", header="ID", index=0, score=1.0, contributions=tuple())
        ],
        extra_columns=[],
        rows=[["123"]],
        header_row=["ID"],
        validation_issues=[],
    )
    manifest = ManifestContext(
        raw={"columns": {"order": ["member_id"], "meta": {"member_id": {"label": "Member"}}}},
        version=None,
        model=None,
    )

    output_path = write_outputs(job, manifest, [extraction])
    workbook = openpyxl.load_workbook(output_path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    first_row = next(sheet.iter_rows(values_only=True))
    assert first_row == ("Member",)
    workbook.close()


def test_write_outputs_dedupes_sheet_names(tmp_path: Path) -> None:
    job = _job(tmp_path)
    manifest = ManifestContext(
        raw={"columns": {"order": ["member_id"], "meta": {"member_id": {"label": "Member"}}}},
        version=None,
        model=None,
    )
    base_extraction = FileExtraction(
        source_name="file.csv",
        sheet_name="Employees",
        mapped_columns=[
            ColumnMapping(field="member_id", header="ID", index=0, score=1.0, contributions=tuple())
        ],
        extra_columns=[],
        rows=[["123"]],
        header_row=["ID"],
        validation_issues=[],
    )
    second = FileExtraction(
        source_name="file2.csv",
        sheet_name="Employees",
        mapped_columns=list(base_extraction.mapped_columns),
        extra_columns=list(base_extraction.extra_columns),
        rows=[["456"]],
        header_row=["ID"],
        validation_issues=[],
    )

    output_path = write_outputs(job, manifest, [base_extraction, second])
    workbook = openpyxl.load_workbook(output_path, read_only=True)
    assert workbook.sheetnames == ["Employees", "Employees-2"]
    workbook.close()
