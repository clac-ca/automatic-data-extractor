import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from ade_engine.config.loader import load_config_runtime
from ade_engine.core.pipeline import map_extracted_tables, normalize_table, write_workbook
from ade_engine.core.types import ExtractedTable, RunContext, RunPaths, RunRequest
from ade_engine.infra.event_emitter import ConfigEventEmitter, EngineEventEmitter
from ade_engine.infra.telemetry import FileEventSink


def _clear_import_cache(prefix: str = "ade_config") -> None:
    for name in list(sys.modules):
        if name == prefix or name.startswith(f"{prefix}."):
            sys.modules.pop(name)


def _bootstrap_package(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    pkg_dir = tmp_path / "ade_config"
    pkg_dir.mkdir()
    (pkg_dir / "__init__.py").write_text("")
    monkeypatch.syspath_prepend(str(tmp_path))
    _clear_import_cache()
    return pkg_dir


def _write_manifest(pkg_dir: Path, *, order: list[str], writer: dict | None = None, hooks: dict | None = None) -> Path:
    manifest = {
        "schema": "ade.manifest/v1",
        "version": "1.0.0",
        "script_api_version": 3,
        "columns": {
            "order": order,
            "fields": {field: {"label": field, "module": f"column_detectors.{field}", "required": False} for field in order},
        },
        "hooks": hooks
        or {
            "on_run_start": [],
            "on_after_extract": [],
            "on_after_mapping": [],
            "on_before_save": [],
            "on_run_end": [],
        },
        "writer": writer or {"append_unmapped_columns": True, "unmapped_prefix": "raw_"},
    }
    manifest_path = pkg_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    return manifest_path


def _write_column_detector(pkg_dir: Path, field: str, body: str) -> None:
    detector_dir = pkg_dir / "column_detectors"
    detector_dir.mkdir(exist_ok=True)
    (detector_dir / "__init__.py").write_text("")
    (detector_dir / f"{field}.py").write_text(body)


def _write_hook(pkg_dir: Path, name: str, body: str) -> None:
    hook_dir = pkg_dir / "hooks"
    hook_dir.mkdir(exist_ok=True)
    (hook_dir / "__init__.py").write_text("")
    (hook_dir / f"{name}.py").write_text(body)


def _run_context(tmp_path: Path, manifest: object, request: RunRequest) -> RunContext:
    paths = RunPaths(
        input_file=request.input_file or tmp_path / "input.csv",
        output_dir=tmp_path / "out",
        logs_dir=tmp_path / "logs",
    )
    return RunContext(
        run_id=uuid4(),
        metadata={},
        manifest=manifest,
        paths=paths,
        started_at=datetime.utcnow(),
    )


def _event_emitters(run: RunContext) -> tuple[EngineEventEmitter, ConfigEventEmitter]:
    sink = FileEventSink(path=run.paths.logs_dir / "events.ndjson")
    engine_emitter = EngineEventEmitter(run=run, event_sink=sink)
    return engine_emitter, engine_emitter.config_emitter()


def test_writes_sheet_with_unmapped_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    pkg_dir = _bootstrap_package(tmp_path, monkeypatch)
    manifest_path = _write_manifest(
        pkg_dir,
        order=["alpha", "beta"],
        writer={"append_unmapped_columns": True, "unmapped_prefix": "raw_"},
    )
    _write_column_detector(
        pkg_dir,
        "alpha",
        """
def detect_header(*, header, logger, event_emitter, **_):
    return 1.0 if header.lower() == "alpha" else 0.0
""",
    )
    _write_column_detector(
        pkg_dir,
        "beta",
        """
def detect_header(*, header, logger, event_emitter, **_):
    return 1.0 if header.lower() == "beta" else 0.0
""",
    )

    runtime = load_config_runtime(manifest_path=manifest_path)
    request = RunRequest(input_file=tmp_path / "input.csv")
    run = _run_context(tmp_path, runtime.manifest, request)
    engine_emitter, config_emitter = _event_emitters(run)
    logger = logging.getLogger("test_write")

    raw = ExtractedTable(
        source_file=tmp_path / "input.csv",
        source_sheet=None,
        table_index=0,
        header_row=["Alpha", "Beta", "Gamma"],
        data_rows=[["a1", "b1", "c1"]],
        header_row_index=1,
        first_data_row_index=2,
        last_data_row_index=2,
    )

    mapped = map_extracted_tables(
        tables=[raw],
        runtime=runtime,
        run=run,
        event_emitter=engine_emitter,
        config_event_emitter=config_emitter,
    )[0]
    normalized = normalize_table(
        ctx=run,
        cfg=runtime,
        mapped=mapped,
        event_emitter=engine_emitter,
        config_event_emitter=config_emitter,
    )

    output_path = write_workbook(
        ctx=run,
        cfg=runtime,
        tables=[normalized],
        input_file_name="sample.csv",
        event_emitter=config_emitter,
        logger=logger,
    )

    workbook = load_workbook(output_path)
    sheet = workbook["input"]

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == ["alpha", "beta", "raw_3"]

    row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert row == ["a1", "b1", "c1"]
    assert normalized.output_sheet_name == "input"


def test_omits_unmapped_columns_when_disabled(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    pkg_dir = _bootstrap_package(tmp_path, monkeypatch)
    manifest_path = _write_manifest(
        pkg_dir,
        order=["alpha"],
        writer={"append_unmapped_columns": False, "unmapped_prefix": "raw_"},
    )
    _write_column_detector(
        pkg_dir,
        "alpha",
        """
def detect_header(*, header, logger, event_emitter, **_):
    return 1.0 if header.lower() == "alpha" else 0.0
""",
    )

    runtime = load_config_runtime(manifest_path=manifest_path)
    request = RunRequest(input_file=tmp_path / "input.csv")
    run = _run_context(tmp_path, runtime.manifest, request)
    engine_emitter, config_emitter = _event_emitters(run)

    raw = ExtractedTable(
        source_file=tmp_path / "input.csv",
        source_sheet=None,
        table_index=0,
        header_row=["Alpha", "Extra"],
        data_rows=[["a1", "b1"]],
        header_row_index=1,
        first_data_row_index=2,
        last_data_row_index=2,
    )

    mapped = map_extracted_tables(
        tables=[raw],
        runtime=runtime,
        run=run,
        event_emitter=engine_emitter,
        config_event_emitter=config_emitter,
    )[0]
    normalized = normalize_table(
        ctx=run,
        cfg=runtime,
        mapped=mapped,
        event_emitter=engine_emitter,
        config_event_emitter=config_emitter,
    )

    output_path = write_workbook(
        ctx=run,
        cfg=runtime,
        tables=[normalized],
        input_file_name="input.csv",
        event_emitter=config_emitter,
        logger=logging.getLogger("test_write"),
    )

    workbook = load_workbook(output_path)
    try:
        sheet = workbook["input"]
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert header == ["alpha"]
        assert row == ["a1"]
    finally:
        workbook.close()


def test_creates_unique_sheets_and_runs_hooks(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    pkg_dir = _bootstrap_package(tmp_path, monkeypatch)
    _write_hook(
        pkg_dir,
        "on_before_save",
        """
def run(*, workbook, tables, logger, event_emitter, **_):
    sheet = workbook.create_sheet(title="Hooked")
    sheet["A1"] = tables[0].output_sheet_name
""",
    )
    manifest_path = _write_manifest(
        pkg_dir,
        order=["alpha"],
        writer={"append_unmapped_columns": True, "unmapped_prefix": "raw_"},
        hooks={
            "on_run_start": [],
            "on_after_extract": [],
            "on_after_mapping": [],
            "on_before_save": ["hooks.on_before_save"],
            "on_run_end": [],
        },
    )
    _write_column_detector(
        pkg_dir,
        "alpha",
        """
def detect_header(*, header, logger, event_emitter, **_):
    return 1.0 if header.lower().startswith("alpha") else 0.0
""",
    )

    runtime = load_config_runtime(manifest_path=manifest_path)
    request = RunRequest(input_file=tmp_path / "input.csv")
    run = _run_context(tmp_path, runtime.manifest, request)
    engine_emitter, config_emitter = _event_emitters(run)
    logger = logging.getLogger("test_write")

    raw_one = ExtractedTable(
        source_file=tmp_path / "first.csv",
        source_sheet=None,
        table_index=0,
        header_row=["Alpha"],
        data_rows=[["a1"]],
        header_row_index=1,
        first_data_row_index=2,
        last_data_row_index=2,
    )
    raw_two = ExtractedTable(
        source_file=tmp_path / "second.csv",
        source_sheet="Sheet1",
        table_index=1,
        header_row=["Alpha"],
        data_rows=[["b1"]],
        header_row_index=1,
        first_data_row_index=2,
        last_data_row_index=2,
    )

    mapped_tables = map_extracted_tables(
        tables=[raw_one, raw_two],
        runtime=runtime,
        run=run,
        event_emitter=engine_emitter,
        config_event_emitter=config_emitter,
    )
    normalized_tables = [
        normalize_table(
            ctx=run,
            cfg=runtime,
            mapped=mapped,
            event_emitter=engine_emitter,
            config_event_emitter=config_emitter,
        )
        for mapped in mapped_tables
    ]

    output_path = write_workbook(
        ctx=run,
        cfg=runtime,
        tables=normalized_tables,
        input_file_name=normalized_tables[0].mapped.extracted.source_file.name,
        event_emitter=config_emitter,
        logger=logger,
    )

    workbook = load_workbook(output_path)
    sheet_names = workbook.sheetnames
    assert "Hooked" in sheet_names
    assert normalized_tables[0].output_sheet_name == "first"
    assert normalized_tables[1].output_sheet_name == "Sheet1"

    first_sheet = workbook[normalized_tables[0].output_sheet_name]
    first_header = [cell.value for cell in next(first_sheet.iter_rows(min_row=1, max_row=1))]
    assert first_header == ["alpha"]
    first_row = [cell.value for cell in next(first_sheet.iter_rows(min_row=2, max_row=2))]
    assert first_row == ["a1"]

    hook_sheet = workbook["Hooked"]
    assert hook_sheet["A1"].value == normalized_tables[0].output_sheet_name
