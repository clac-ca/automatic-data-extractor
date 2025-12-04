from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from ade_engine import Engine, RunRequest, RunStatus, run
from ade_engine.core.types import RunPhase
from ade_engine.schemas.telemetry import AdeEvent
from fixtures.config_factories import clear_config_import, make_minimal_config
from fixtures.sample_inputs import (
    sample_csv,
    sample_large_csv,
    sample_xlsx_multi_sheet,
    sample_xlsx_single_sheet,
)


@pytest.fixture(autouse=True)
def _cleanup_imports() -> None:
    yield
    clear_config_import()


def _parse_events(path: Path) -> list[AdeEvent]:
    return [AdeEvent.model_validate_json(line) for line in path.read_text().splitlines() if line]


def test_engine_run_end_to_end(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config = make_minimal_config(tmp_path, monkeypatch, include_transform=True, include_validator=True)
    source = sample_xlsx_single_sheet(tmp_path)

    result = run(
        manifest_path=config.manifest_path,
        input_file=source,
        output_dir=tmp_path / "out",
        logs_dir=tmp_path / "logs",
        metadata={"test_case": "end_to_end"},
    )

    assert result.status is RunStatus.SUCCEEDED
    workbook_path = Path(result.output_paths[0])
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path)
    try:
        sheet = workbook["Normalized"]
        assert [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))] == ["member_id", "value", "raw_3"]
    finally:
        workbook.close()

    events_path = Path(result.logs_dir) / "events.ndjson"
    events = _parse_events(events_path)

    table_event = next(evt for evt in events if evt.type == "engine.table.summary")
    table = table_event.payload_dict()
    assert table["counts"]["rows"]["total"] == 2
    summary_event = next(evt for evt in events if evt.type == "engine.run.summary")
    assert summary_event.payload_dict().get("counts", {}).get("tables", {}).get("total") == 1


def test_engine_run_hook_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config = make_minimal_config(tmp_path, monkeypatch, include_hooks=True, failing_hook=True)
    source = sample_csv(tmp_path)

    request = RunRequest(
        manifest_path=config.manifest_path,
        input_file=source,
        output_dir=tmp_path / "out",
        logs_dir=tmp_path / "logs",
    )
    result = Engine().run(request)

    assert result.status is RunStatus.FAILED
    assert result.error is not None
    events_path = Path(result.logs_dir) / "events.ndjson"
    events = _parse_events(events_path)
    completed = next(evt for evt in events if evt.type == "engine.complete")
    assert completed.payload_dict().get("status") == "failed"


def test_engine_mapping_snapshot(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config = make_minimal_config(tmp_path, monkeypatch, include_transform=False, include_validator=False)
    source = sample_xlsx_multi_sheet(tmp_path)

    result = run(
        manifest_path=config.manifest_path,
        input_file=source,
        output_dir=tmp_path / "out",
        logs_dir=tmp_path / "logs",
    )

    events_path = Path(result.logs_dir) / "events.ndjson"
    events = _parse_events(events_path)
    table_event = next(evt for evt in events if evt.type == "engine.table.summary")
    table = table_event.payload_dict()

    mapped_fields = [
        {key: field.get(key) for key in ("field", "header", "score", "source_column_index")}
        for field in table.get("fields", [])
        if field.get("mapped")
    ]
    unmapped_columns = [
        {
            "header": column.get("header"),
            "source_column_index": column.get("source_column_index"),
            "output_header": column.get("output_header"),
        }
        for column in table.get("columns", [])
        if not column.get("mapped")
    ]

    assert mapped_fields == [
        {"field": "member_id", "header": "member_id", "score": 1.0, "source_column_index": 0},
        {"field": "value", "header": "value", "score": 1.0, "source_column_index": 1},
    ]
    assert unmapped_columns == [
        {"header": "note", "source_column_index": 2, "output_header": "raw_3"},
        {"header": "surplus", "source_column_index": 3, "output_header": "raw_4"},
    ]


def test_engine_large_input_smoke(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config = make_minimal_config(tmp_path, monkeypatch)
    source = sample_large_csv(tmp_path, rows=2000)

    result = run(
        manifest_path=config.manifest_path,
        input_file=source,
        output_dir=tmp_path / "out",
        logs_dir=tmp_path / "logs",
    )

    assert result.status is RunStatus.SUCCEEDED
    assert Path(result.output_paths[0]).exists()


def test_engine_reports_pipeline_stage_on_mapping_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config = make_minimal_config(tmp_path, monkeypatch)
    failing_detector = config.package_dir / "column_detectors" / "member_id.py"
    failing_detector.write_text(
        """
def detect_header(*, header, **_):
    raise RuntimeError("boom")
"""
    )

    source = sample_csv(tmp_path)
    request = RunRequest(
        manifest_path=config.manifest_path,
        input_file=source,
        output_dir=tmp_path / "out",
        logs_dir=tmp_path / "logs",
    )

    result = Engine().run(request)

    assert result.status is RunStatus.FAILED
    assert result.error is not None
    assert result.error.stage is RunPhase.MAPPING
