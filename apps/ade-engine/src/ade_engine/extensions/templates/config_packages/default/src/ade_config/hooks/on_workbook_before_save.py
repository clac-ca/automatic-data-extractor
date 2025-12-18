from __future__ import annotations


# -----------------------------------------------------------------------------
# Hook: on_workbook_before_save
#
# When it runs:
# - Called once per input workbook, after all sheets/tables have been processed
#   and written to the OUTPUT workbook, and immediately before the workbook is
#   saved to disk (see `metadata["output_file"]`).
#
# What it's good for:
# - Workbook-wide formatting (freeze panes, filters, print setup, styles)
# - Adding summary/notes sheets using information collected in `state`
# - Setting workbook properties (title/author) or hiding helper sheets
#
# Notes:
# - `workbook` is the output openpyxl `Workbook`.
# - `sheet` and `table` are always `None` for this hook.
# -----------------------------------------------------------------------------


def register(registry):
    registry.register_hook(on_workbook_before_save, hook="on_workbook_before_save", priority=0)

    # Examples (uncomment to enable)
    # registry.register_hook(on_workbook_before_save_example_1_set_properties, hook="on_workbook_before_save", priority=0)
    # registry.register_hook(on_workbook_before_save_example_2_apply_worksheet_ux, hook="on_workbook_before_save", priority=0)
    # registry.register_hook(on_workbook_before_save_example_3_format_headers_and_widths, hook="on_workbook_before_save", priority=0)
    # registry.register_hook(on_workbook_before_save_example_4_write_notes_sheet, hook="on_workbook_before_save", priority=0)


def on_workbook_before_save(
    *,
    hook_name,  # HookName enum value for this stage
    settings,  # Engine Settings
    metadata: dict,  # Run metadata (filenames, etc.)
    state: dict,  # Mutable dict shared across the run
    workbook,  # Output workbook (openpyxl Workbook)
    sheet,  # Always None for this hook
    table,  # Always None for this hook
    write_table,  # Always None for this hook
    input_file_name: str | None,  # Input filename (basename) if known
    logger,  # RunLogger (structured events + text logs)
) -> None:
    """Finalize the output workbook before it is saved."""

    if logger:
        logger.info("Config hook: workbook before save (%s)", metadata.get("output_file", ""))



def on_workbook_before_save_example_1_set_properties(
    *,
    hook_name,
    settings,
    metadata: dict,
    state: dict,
    workbook,
    sheet,
    table,
    write_table,
    input_file_name: str | None,
    logger,
) -> None:
    """Example: set workbook properties and choose which sheet opens by default."""

    workbook.properties.creator = "Automatic Data Extractor (ADE)"
    workbook.properties.title = f"Normalized - {metadata.get('input_file_name', '')}"
    workbook.active = 0


def on_workbook_before_save_example_2_apply_worksheet_ux(
    *,
    hook_name,
    settings,
    metadata: dict,
    state: dict,
    workbook,
    sheet,
    table,
    write_table,
    input_file_name: str | None,
    logger,
) -> None:
    """Example: apply consistent worksheet UX settings."""

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"  # keep header row visible while scrolling
        ws.sheet_view.showGridLines = False
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions


def on_workbook_before_save_example_3_format_headers_and_widths(
    *,
    hook_name,
    settings,
    metadata: dict,
    state: dict,
    workbook,
    sheet,
    table,
    write_table,
    input_file_name: str | None,
    logger,
) -> None:
    """Example: make header rows bold and set simple column widths."""

    from openpyxl.styles import Alignment, Font

    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical="top")
    for ws in workbook.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            if cell.value:
                width = min(max(len(str(cell.value)) + 2, 12), 40)
                ws.column_dimensions[cell.column_letter].width = width


def on_workbook_before_save_example_4_write_notes_sheet(
    *,
    hook_name,
    settings,
    metadata: dict,
    state: dict,
    workbook,
    sheet,
    table,
    write_table,
    input_file_name: str | None,
    logger,
) -> None:
    """Example: write a simple Notes sheet from items collected in `state`."""

    notes = state.get("notes") or []
    if not notes:
        return

    notes_ws = workbook.create_sheet(title="Notes")
    notes_ws.append(["note"])
    for note in notes:
        notes_ws.append([note])
