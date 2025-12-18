"""ADE config hook: `on_table_written`

When this runs
--------------
- Once per detected table region, after ADE writes the normalized table into the
  OUTPUT workbook (openpyxl) and before the workbook is saved.
- The table's cells already exist in `sheet` when this hook runs.

What it's useful for
--------------------
- Excel-only formatting (freeze panes, filters, column widths, styles)
- Adding Excel "structured tables" (Insert → Table) for better UX
- Adding comments/notes for humans reading the output workbook
- Hiding helper/diagnostic columns that you still want to keep written
- Collecting per-table facts into `state` for later hooks (e.g., summary sheets)

Best practice
-------------
- Prefer not to change *table values* here. Do value changes earlier in the pipeline
  (`on_table_mapped` / `on_table_transformed` / `on_table_validated`).
- This hook is best for Excel UX and presentation.

Notes
-----
- ADE passes `table_region` (a `TableRegion`) to this hook. It is the authoritative,
  openpyxl-friendly range for the table that was just written, so examples avoid
  brittle `sheet.max_row` inference.

Reference docs
--------------
- Worksheet Tables: https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html
- Conditional formatting: https://openpyxl.readthedocs.io/en/3.1/formatting.html
"""

from __future__ import annotations

from typing import Mapping, MutableMapping

import openpyxl
import openpyxl.worksheet.worksheet
import polars as pl
from ade_engine.models import TableRegion

# `TableRegion` (engine-owned, openpyxl-friendly coordinates):
# - min_row, min_col, max_row, max_col (1-based, inclusive)
# - convenience properties: a1, cell_range, width, height
# - header/data helpers: header_row, data_first_row, data_min_row, has_data_rows, data_row_count


# -----------------------------------------------------------------------------
# Shared state namespacing
# -----------------------------------------------------------------------------
# `state` is a mutable dict shared across the run.
# Best practice: store everything your config package needs under ONE top-level key.
#
# IMPORTANT: Keep this constant the same across *all* your hooks so they can share state.
STATE_NAMESPACE = "ade.config_package_template"
STATE_SCHEMA_VERSION = 1


# --- ADE diagnostic column conventions ---------------------------------------
#
# When validators are present, ADE may append helper columns such as:
#
# - `__ade_has_issues` (bool): True if any field in the row has an issue
# - `__ade_issue_count` (int): count of fields with issues in the row
# - `__ade_issue__<field>` (str): per-field issue message (blank/None when ok)
#
# Output policies might drop these columns; every example below checks for
# presence and becomes a no-op if they're not written.

ADE_DIAGNOSTIC_PREFIX = "__ade_"
ADE_HAS_ISSUES_COL = "__ade_has_issues"
ADE_ISSUE_COUNT_COL = "__ade_issue_count"
ADE_ISSUE_PREFIX = "__ade_issue__"


# --- Public registration -----------------------------------------------------

def register(registry) -> None:
    """
    Register hook(s) with the ADE registry.

    Keep the default hook minimal and safe. Enable examples by uncommenting
    registrations below (you can also adjust priorities to control ordering).
    """
    registry.register_hook(on_table_written, hook="on_table_written", priority=0)

    # --- Examples (uncomment to enable) -------------------------------------
    # registry.register_hook(on_table_written_example_1_log_output_range, hook="on_table_written", priority=10)
    # registry.register_hook(on_table_written_example_2_freeze_header_and_add_filters, hook="on_table_written", priority=20)
    # registry.register_hook(on_table_written_example_3_style_header_row, hook="on_table_written", priority=30)
    # registry.register_hook(on_table_written_example_4_hide_diagnostic_columns, hook="on_table_written", priority=40)
    # registry.register_hook(on_table_written_example_5_collect_table_facts, hook="on_table_written", priority=50)

    # High-value openpyxl examples for Excel UX:
    # registry.register_hook(on_table_written_example_6_add_excel_structured_table, hook="on_table_written", priority=60)
    # registry.register_hook(on_table_written_example_7_add_header_comments, hook="on_table_written", priority=70)
    # registry.register_hook(on_table_written_example_8_highlight_and_comment_validation_issues, hook="on_table_written", priority=80)
    # registry.register_hook(on_table_written_example_9_autofit_column_widths, hook="on_table_written", priority=90)


# --- Core hook ---------------------------------------------------------------

def on_table_written(
    *,
    table: pl.DataFrame,  # Exact DF that was written (after output policies)
    sheet: openpyxl.worksheet.worksheet.Worksheet,  # Output worksheet (openpyxl Worksheet)
    workbook: openpyxl.Workbook,  # Output workbook (openpyxl Workbook)
    table_region: TableRegion,  # Output header+data bounds (1-based, inclusive)
    table_index: int,  # 0-based table index within the sheet
    input_file_name: str,  # Input filename (basename)
    settings,  # Engine Settings
    metadata: dict,  # Run/sheet metadata (filenames, sheet_index, etc.)
    state: dict,  # Mutable dict shared across the run
    logger,  # RunLogger (structured events + text logs)
) -> None:
    """
    Called after ADE writes a single normalized table to the output worksheet.

    Keep this function *small*. Use additional hook registrations for examples
    and project-specific behavior.

    Tip: `table.columns` and `table.height` reflect what you see in `sheet`.
    """
    _ = (settings, metadata, workbook, input_file_name)  # unused by default

    cfg = state.get(STATE_NAMESPACE)
    if not isinstance(cfg, MutableMapping):
        cfg = {}
        state[STATE_NAMESPACE] = cfg
    cfg.setdefault("schema_version", STATE_SCHEMA_VERSION)

    counters = cfg.get("counters")
    if not isinstance(counters, MutableMapping):
        counters = {}
        cfg["counters"] = counters
    counters["tables_written_seen"] = int(counters.get("tables_written_seen", 0) or 0) + 1

    sheet_name = getattr(sheet, "title", getattr(sheet, "name", ""))
    range_ref = table_region.a1
    if logger:
        logger.info(
            "Config hook: table written (sheet=%s, table_index=%s, range=%s, rows=%d, columns=%d)",
            sheet_name,
            table_index,
            range_ref,
            int(table.height),
            len(table.columns),
        )

    return None


# --- Example hooks -----------------------------------------------------------

def on_table_written_example_1_log_output_range(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """Example 1: log the output range for THIS table."""
    _ = (settings, metadata, state, workbook, sheet, table, table_index, input_file_name)

    if logger:
        logger.info("Output range for last table: %s", table_region.a1)


def on_table_written_example_2_freeze_header_and_add_filters(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 2: freeze the header row and enable filters for this table range.

    Note:
    - If you also add an Excel *structured table* (example 6), Excel will add
      filters automatically for that table.
    """
    _ = (settings, metadata, state, workbook, table, table_index, input_file_name, logger)

    # Freeze panes at the first cell *below* the header row.
    sheet.freeze_panes = sheet.cell(row=table_region.data_first_row, column=table_region.min_col).coordinate

    # Enable worksheet auto-filter for the table region.
    sheet.auto_filter.ref = table_region.a1


def on_table_written_example_3_style_header_row(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 3: make the header row look nice (fill, font, wrap, alignment).

    This is intentionally "manual" styling so developers can learn openpyxl
    basics. If you use an Excel structured table (example 6), the table style
    will also style the header row.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    _ = (settings, metadata, state, workbook, table, table_index, input_file_name, logger)

    # Style constants (easy to tweak).
    header_fill = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(table_region.min_col, table_region.max_col + 1):
        cell = sheet.cell(row=table_region.header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Optional: slightly taller header row for wrapped text.
    sheet.row_dimensions[table_region.header_row].height = 24


def on_table_written_example_4_hide_diagnostic_columns(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 4: hide ADE diagnostic columns (if you choose to keep them written).
    """
    from openpyxl.utils import get_column_letter
    _ = (settings, metadata, state, workbook, sheet, table_index, input_file_name, logger)

    for offset, col_name in enumerate(table.columns):
        if str(col_name).startswith(ADE_DIAGNOSTIC_PREFIX):
            col_idx = table_region.min_col + offset
            sheet.column_dimensions[get_column_letter(col_idx)].hidden = True


def on_table_written_example_5_collect_table_facts(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 5: collect per-table facts into shared `state` for later hooks.
    """
    cfg = state.get(STATE_NAMESPACE)
    if not isinstance(cfg, MutableMapping):
        cfg = {}
        state[STATE_NAMESPACE] = cfg

    _ = (settings, metadata, workbook, sheet, table_index, logger)
    table_region_ref = table_region.a1

    tables = cfg.get("tables_written")
    if not isinstance(tables, list):
        tables = []
        cfg["tables_written"] = tables

    tables.append(
        {
            "input_file": input_file_name,
            "sheet": getattr(sheet, "title", ""),
            "table_index": table_index,
            "rows": int(table.height),
            "columns": list(table.columns),
            "range": table_region_ref,
        }
    )


def on_table_written_example_6_add_excel_structured_table(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 6: convert the written range into an Excel "structured table".

    Why this is great:
    - Users get built-in filter dropdowns
    - Banded rows / table styles
    - Structured references in formulas
    - Better UX when people copy/paste or build pivot tables
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import re

    cfg = state.get(STATE_NAMESPACE)
    if not isinstance(cfg, MutableMapping):
        cfg = {}
        state[STATE_NAMESPACE] = cfg

    _ = (settings, metadata, input_file_name)
    ref = table_region.a1

    # Excel tables are most useful when there is at least one data row.
    if int(table.height) <= 0 or not table_region.has_data_rows:
        if logger:
            logger.info("Skipping Excel table creation (no data rows).")
        return

    sheet_name = getattr(sheet, "title", "Sheet")
    base = f"{sheet_name}_Table_{table_index}"

    # Make a valid Excel table name (simple + conservative).
    # Rules (high level): starts with letter/_/\, contains letters/numbers/._, no spaces,
    # can't look like A1 or R1C1, max 255 chars.
    name = re.sub(r"\s+", "_", (base or "").strip()) or "Table"
    name = re.sub(r"[^A-Za-z0-9_.]", "_", name)
    if not re.match(r"^[A-Za-z_]", name):
        name = f"T_{name}"
    if re.match(r"^[A-Za-z]{1,3}\d+$", name) or re.match(r"^[Rr]\d+[Cc]\d+$", name):
        name = f"T_{name}"
    name = name[:255]

    existing: set[str] = set()
    for ws in getattr(workbook, "worksheets", []) or []:
        tables = getattr(ws, "tables", None)
        if not tables:
            continue
        if hasattr(tables, "keys"):
            try:
                existing.update(str(k) for k in tables.keys())
            except Exception:
                pass

    table_name = name
    if table_name in existing:
        for i in range(2, 10_000):
            suffix = f"_{i}"
            candidate = f"{name[: 255 - len(suffix)]}{suffix}"
            if candidate not in existing:
                table_name = candidate
                break

    tab = Table(displayName=table_name, ref=ref)

    # Built-in Excel table styles (pick any "TableStyle..." string)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style

    # Important: use ws.add_table()
    sheet.add_table(tab)

    excel_tables = cfg.get("excel_tables")
    if not isinstance(excel_tables, list):
        excel_tables = []
        cfg["excel_tables"] = excel_tables
    excel_tables.append({"sheet": sheet_name, "name": table_name, "ref": ref})

    if logger:
        logger.info("Added Excel structured table: %s (%s)", table_name, ref)


def on_table_written_example_7_add_header_comments(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 7: add helpful comments to header cells.

    Header comments are a great "self-documenting output" technique—especially
    when workbook consumers may not know ADE or your schema.
    """
    from openpyxl.comments import Comment

    cfg = state.get(STATE_NAMESPACE)
    if not isinstance(cfg, MutableMapping):
        cfg = {}
        state[STATE_NAMESPACE] = cfg

    _ = (settings, metadata, workbook, table_index, input_file_name)
    if not table.columns:
        return

    # Built-in comments for ADE diagnostic columns.
    default_comments: dict[str, str] = {
        ADE_HAS_ISSUES_COL: "TRUE when any field in this row failed validation.",
        ADE_ISSUE_COUNT_COL: "Number of fields with validation issues in this row.",
    }

    # Merge in any project-specific comments from shared state.
    # Example: state[STATE_NAMESPACE]["header_comments"] = {"amount": "USD, must be >= 0"}
    extra: Mapping[str, str] = cfg.get("header_comments", {}) or {}
    comments: dict[str, str] = {**default_comments, **dict(extra)}

    col_to_idx = {name: table_region.min_col + i for i, name in enumerate(table.columns)}

    added = 0
    for col_name, comment_text in comments.items():
        col_idx = col_to_idx.get(col_name)
        if not col_idx:
            continue
        cell = sheet.cell(row=table_region.header_row, column=col_idx)
        comment = Comment(text=comment_text, author="ADE")
        comment.width = 320
        comment.height = 120
        cell.comment = comment
        added += 1

    if logger and added:
        logger.info("Added %d header comments.", added)


def on_table_written_example_8_highlight_and_comment_validation_issues(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 8: use ADE diagnostics to improve UX for validation failures.

    This example does two things:
    1) Adds row-level conditional formatting when __ade_has_issues is TRUE.
    2) Adds cell-level comments on specific cells that failed validation, using
       the message from __ade_issue__<field>.

    NOTE: Adding thousands of comments can produce very large Excel files.
    This example includes a safety limit.
    """
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    cfg = state.get(STATE_NAMESPACE)
    if not isinstance(cfg, MutableMapping):
        cfg = {}
        state[STATE_NAMESPACE] = cfg

    _ = (settings, metadata, workbook, table_index, input_file_name)
    if not table_region.has_data_rows or not table.columns:
        return

    from openpyxl.worksheet.cell_range import CellRange

    data_ref = CellRange(
        min_row=table_region.data_first_row,
        max_row=table_region.max_row,
        min_col=table_region.min_col,
        max_col=table_region.max_col,
    ).coord
    data_first_row = table_region.data_first_row
    last_row = table_region.max_row

    col_to_idx = {name: table_region.min_col + i for i, name in enumerate(table.columns)}
    has_issues_idx = col_to_idx.get(ADE_HAS_ISSUES_COL)
    if not has_issues_idx:
        return  # diagnostics not present / dropped

    issue_cols = [c for c in table.columns if str(c).startswith(ADE_ISSUE_PREFIX)]
    if not issue_cols:
        return

    # 1) Row-level conditional formatting
    #
    # Use an absolute column reference ($X) and a relative row number so the rule
    # applies per-row over the whole range.
    issue_col_letter = get_column_letter(has_issues_idx)
    warn_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")  # light yellow
    rule = FormulaRule(
        formula=[f'=${issue_col_letter}{data_first_row}=TRUE'],
        fill=warn_fill,
    )
    sheet.conditional_formatting.add(data_ref, rule)

    # 2) Per-cell comments for issue messages
    max_comments = int(cfg.get("max_issue_comments", 500) or 500)
    comments_added = 0

    # Iterate only the __ade_has_issues column to find problematic rows efficiently.
    for offset, (flag,) in enumerate(
        sheet.iter_rows(
            min_row=data_first_row,
            max_row=last_row,
            min_col=has_issues_idx,
            max_col=has_issues_idx,
            values_only=True,
        ),
        start=0,
    ):
        if not flag:
            continue

        excel_row = data_first_row + offset

        for issue_col in issue_cols:
            msg_idx = col_to_idx.get(issue_col)
            if not msg_idx:
                continue

            msg = sheet.cell(row=excel_row, column=msg_idx).value
            if msg in (None, ""):
                continue

            # Map "__ade_issue__amount" -> "amount"
            field = str(issue_col)[len(ADE_ISSUE_PREFIX) :]
            target_idx = col_to_idx.get(field)
            if not target_idx:
                continue  # field column renamed/dropped

            target_cell = sheet.cell(row=excel_row, column=target_idx)
            comment = Comment(text=str(msg), author="ADE Validation")
            comment.width = 360
            comment.height = 140
            target_cell.comment = comment
            comments_added += 1

            if comments_added >= max_comments:
                break

        if comments_added >= max_comments:
            break

    if logger and comments_added:
        logger.info("Added %d validation comments (limit=%d).", comments_added, max_comments)
        if comments_added >= max_comments:
            logger.warning(
                "Stopped adding validation comments at limit=%d. "
                "Consider relying on conditional formatting instead for huge tables.",
                max_comments,
            )


def on_table_written_example_9_autofit_column_widths(
    *,
    table: pl.DataFrame,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    workbook: openpyxl.Workbook,
    table_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings,
    metadata: dict,
    state: dict,
    logger,
) -> None:
    """
    Example 9: approximate Excel "auto-fit" for column widths.

    Excel doesn't auto-size columns when writing files programmatically. This
    heuristic measures string lengths in the header + a sample of data rows.

    The result is an approximation (often good enough for reports).
    """
    from openpyxl.utils import get_column_letter

    _ = (settings, metadata, workbook, table_index, input_file_name)

    cfg = state.get(STATE_NAMESPACE)
    if not isinstance(cfg, MutableMapping):
        cfg = {}
        state[STATE_NAMESPACE] = cfg

    sample_rows = int(cfg.get("autofit_sample_rows", 200) or 200)
    max_width = float(cfg.get("autofit_max_width", 60) or 60)
    padding = float(cfg.get("autofit_padding", 2) or 2)
    skip_diagnostics = bool(cfg.get("autofit_skip_diagnostics", True))

    scan_last_row = min(table_region.max_row, table_region.data_first_row + max(sample_rows, 0) - 1)

    for col in range(table_region.min_col, table_region.max_col + 1):
        col_name = table.columns[col - table_region.min_col]
        if str(col_name).startswith(ADE_DIAGNOSTIC_PREFIX) and skip_diagnostics:
            continue

        best = 0

        header_val = sheet.cell(row=table_region.header_row, column=col).value
        best = max(best, len(str(header_val)) if header_val is not None else 0)

        if table_region.has_data_rows and scan_last_row >= table_region.data_first_row:
            for r in range(table_region.data_first_row, scan_last_row + 1):
                v = sheet.cell(row=r, column=col).value
                if v is None:
                    continue
                best = max(best, len(str(v)))

        width = min(max_width, best + padding)
        sheet.column_dimensions[get_column_letter(col)].width = width

    if logger:
        logger.info("Auto-fit applied (sample_rows=%d, max_width=%s).", sample_rows, max_width)
