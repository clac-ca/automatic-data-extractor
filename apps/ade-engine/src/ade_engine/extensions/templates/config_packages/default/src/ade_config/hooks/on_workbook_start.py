"""
ADE hook: ``on_workbook_start``.

This hook runs once per input file after ADE opens the *source* workbook and before any
sheets or tables are processed.

Use this hook for run-scoped initialization and safe workbook inspection: seeding shared
``state``, computing flags used by downstream hooks/detectors, initializing reusable
clients (without performing work yet), and failing fast when required sheets or workbook
structure is missing.

Contract
--------
- Called once per source workbook.
- Must return ``None`` (any other return value is an error).

Guidance
--------
- Keep work inexpensive and deterministic (avoid heavy scanning).
- Treat the source workbook as read-only.
- Prefer caching derived facts in ``state`` for reuse by later hooks.
"""

from __future__ import annotations

import uuid
from collections.abc import Mapping, MutableMapping
from datetime import UTC, datetime
from typing import Any, TYPE_CHECKING

import polars as pl

if TYPE_CHECKING:
    import openpyxl

    from ade_engine.extensions.registry import Registry
    from ade_engine.infrastructure.observability.logger import RunLogger
    from ade_engine.infrastructure.settings import Settings


def register(registry: Registry) -> None:
    """Register hook(s) with the ADE registry."""
    registry.register_hook(on_workbook_start, hook="on_workbook_start", priority=0)

    # ---------------------------------------------------------------------
    # Examples (uncomment to enable, then customize as needed)
    # ---------------------------------------------------------------------
    # registry.register_hook(on_workbook_start_example_0_seed_state_and_log_workbook_basics, hook="on_workbook_start", priority=10)
    # registry.register_hook(on_workbook_start_example_1_fail_fast_missing_sheets, hook="on_workbook_start", priority=20)
    # registry.register_hook(on_workbook_start_example_2_detect_workbook_flavor_and_flags, hook="on_workbook_start", priority=30)
    # registry.register_hook(on_workbook_start_example_3_emit_structured_event, hook="on_workbook_start", priority=40)
    # registry.register_hook(on_workbook_start_example_4_init_openai_client, hook="on_workbook_start", priority=50)
    # registry.register_hook(on_workbook_start_example_5_load_reference_sheet_into_polars, hook="on_workbook_start", priority=60)


def on_workbook_start(
    *,
    workbook: openpyxl.Workbook,  # Source workbook (openpyxl Workbook)
    input_file_name: str,  # Input filename (basename)
    settings: Settings,  # Engine settings
    metadata: Mapping[str, Any],  # Run metadata (filenames, IDs, etc.)
    state: MutableMapping[str, Any],  # Mutable dict shared across the run
    logger: RunLogger,  # RunLogger (structured events + text logs)
) -> None:  # noqa: ARG001
    """Default implementation is a placeholder (no-op). Add run-scoped init here."""

    return None


# ----------------------------
# Examples (uncomment in register() to enable)
# ----------------------------


def on_workbook_start_example_0_seed_state_and_log_workbook_basics(
    *,
    workbook: openpyxl.Workbook,  # Source workbook (openpyxl Workbook)
    input_file_name: str,  # Input filename (basename)
    settings: Settings,  # Engine settings
    metadata: Mapping[str, Any],  # Run metadata (filenames, IDs, etc.)
    state: MutableMapping[str, Any],  # Mutable dict shared across the run
    logger: RunLogger,  # RunLogger (structured events + text logs)
) -> None:  # noqa: ARG001
    """Example 0 (recommended): seed run state and log workbook basics."""

    # `state` is already the shared mutable dict for the run.
    cfg = state

    # ------------------------------------------------------------------
    # 1) Establish run identity & timestamps (helpful for logs + debugging)
    # ------------------------------------------------------------------
    run_started_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    cfg.setdefault("run_started_at", run_started_at)
    cfg.setdefault("run_id", metadata.get("run_id") or uuid.uuid4().hex[:12])

    # ------------------------------------------------------------------
    # 2) Record the “best available” input label for logs/events
    # ------------------------------------------------------------------
    input_label = ""
    for key in ("input_file", "source_file", "filename", "path"):
        val = metadata.get(key)
        if isinstance(val, str) and val.strip():
            input_label = val.strip()
            break
    if not input_label:
        input_label = input_file_name

    cfg.setdefault("input", {})
    if isinstance(cfg.get("input"), MutableMapping):
        cfg["input"].update({"file": input_label, "basename": input_file_name})  # type: ignore[union-attr]
    else:
        cfg["input"] = {"file": input_label, "basename": input_file_name}

    # ------------------------------------------------------------------
    # 3) Inspect workbook shape (safe + fast)
    # ------------------------------------------------------------------
    sheet_names = list(getattr(workbook, "sheetnames", []) or [])

    # Safe, compact document properties (avoid logging full objects).
    props_out: dict[str, Any] = {}
    props = getattr(workbook, "properties", None)
    if props is not None:
        for key in (
            "title",
            "subject",
            "creator",
            "description",
            "keywords",
            "category",
            "created",
            "modified",
            "lastModifiedBy",
            "revision",
        ):
            val = getattr(props, key, None)
            if val not in (None, ""):
                props_out[key] = val

    cfg["workbook"] = {
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "properties": props_out,
    }

    # (Optional) Collect Excel "Table" objects per sheet (Insert → Table).
    # These are NOT the same thing as ADE tables, but can be useful for validation/routing.
    excel_tables_by_sheet: dict[str, list[str]] = {}
    for ws in getattr(workbook, "worksheets", []) or []:
        ws_name = getattr(ws, "title", None) or "<unknown>"
        tables = getattr(ws, "tables", None)

        names: list[str] = []
        if tables is None:
            names = []
        elif hasattr(tables, "keys"):
            try:
                names = sorted([str(k) for k in tables.keys()])
            except Exception:
                names = []
        else:
            try:
                names = sorted([str(k) for k in tables])
            except Exception:
                names = []

        excel_tables_by_sheet[str(ws_name)] = names

    cfg["excel_tables_by_sheet"] = excel_tables_by_sheet

    # ------------------------------------------------------------------
    # 4) Seed shared state structures for downstream hooks
    # ------------------------------------------------------------------
    cfg.setdefault("notes", [])  # downstream hooks can append strings here

    counters = cfg.get("counters")
    if not isinstance(counters, MutableMapping):
        counters = {}
        cfg["counters"] = counters
    counters.setdefault("sheets_seen", 0)
    counters.setdefault("tables_seen", 0)

    # A good spot to keep reusable clients/resources.
    # (Avoid storing secrets here; store secrets in env vars or secure settings.)
    cfg.setdefault("clients", {})  # e.g., {"openai": <OpenAI>, "http": <httpx.Client>}

    # ------------------------------------------------------------------
    # 5) Emit a simple log line
    # ------------------------------------------------------------------
    if logger:
        logger.info(
            "Config hook: workbook start (input=%s, sheets=%d)",
            input_label,
            len(sheet_names),
        )

    return None


def on_workbook_start_example_1_fail_fast_missing_sheets(
    *,
    workbook: openpyxl.Workbook,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example 1: Fail fast if required worksheets are missing."""

    required = {"Orders", "Customers"}  # customize
    sheet_names = set(getattr(workbook, "sheetnames", []) or [])
    missing = sorted(required - sheet_names)
    if missing:
        raise ValueError("Input workbook is missing required worksheet(s): " + ", ".join(missing))

    return None


def on_workbook_start_example_2_detect_workbook_flavor_and_flags(
    *,
    workbook: openpyxl.Workbook,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example 2: Set workbook-level flags for later hooks/detectors."""

    cfg = state

    sheet_names = [str(s).lower() for s in (getattr(workbook, "sheetnames", []) or [])]

    # Toy heuristics (replace with your real template/vendor detection).
    if any("acme" in s for s in sheet_names):
        vendor = "acme"
    elif any("globex" in s for s in sheet_names):
        vendor = "globex"
    else:
        vendor = "unknown"

    flags = cfg.get("flags")
    if not isinstance(flags, MutableMapping):
        flags = {}
        cfg["flags"] = flags
    flags.update(
        {
            "vendor": vendor,
            "treat_blank_as_null": True,
        }
    )

    if logger:
        logger.info("Detected workbook vendor=%s", vendor)

    return None


def on_workbook_start_example_3_emit_structured_event(
    *,
    workbook: openpyxl.Workbook,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example 3: Emit a config-scoped structured event (no strict schema required)."""

    if not logger or not hasattr(logger, "event"):
        return None

    input_label = ""
    for key in ("input_file", "source_file", "filename", "path"):
        val = metadata.get(key)
        if isinstance(val, str) and val.strip():
            input_label = val.strip()
            break
    if not input_label:
        input_label = input_file_name

    logger.event(
        "engine.config.workbook.start",
        data={
            "input_file": input_label,
            "sheet_count": len(getattr(workbook, "sheetnames", []) or []),
            "sheet_names": list(getattr(workbook, "sheetnames", []) or []),
        },
    )

    return None


def on_workbook_start_example_4_init_openai_client(
    *,
    workbook: openpyxl.Workbook,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example 4: Initialize an OpenAI client once and store it in `state`.

    Notes:
    - This example does NOT make a network call. It just prepares the client.
    - The OpenAI SDK reads `OPENAI_API_KEY` from the environment automatically.
    """
    import os

    cfg = state

    clients = cfg.get("clients")
    if not isinstance(clients, MutableMapping):
        clients = {}
        cfg["clients"] = clients

    if "openai" in clients:
        return None  # already initialized

    if not os.getenv("OPENAI_API_KEY"):
        if logger and hasattr(logger, "warning"):
            logger.warning("OPENAI_API_KEY is not set; skipping OpenAI client init.")
        return None

    try:
        # Official OpenAI Python SDK:
        #   pip install openai
        #   from openai import OpenAI
        from openai import OpenAI  # type: ignore
    except Exception as exc:
        if logger and hasattr(logger, "warning"):
            logger.warning("OpenAI SDK not installed; skipping OpenAI init (%s).", exc)
        return None

    clients["openai"] = OpenAI()

    # Store per-run defaults for downstream usage (detectors can read these).
    llm = cfg.get("llm")
    if not isinstance(llm, MutableMapping):
        llm = {}
        cfg["llm"] = llm
    llm.setdefault("model", getattr(settings, "openai_model", "gpt-4o-mini"))
    llm.setdefault(
        "system_prompt",
        "You are a helpful assistant that summarizes Excel workbooks for data processing.",
    )

    if logger:
        logger.info("Initialized OpenAI client for downstream use (model=%s)", llm["model"])

    return None


def on_workbook_start_example_5_load_reference_sheet_into_polars(
    *,
    workbook: openpyxl.Workbook,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> None:  # noqa: ARG001
    """Example 5: Load a small reference sheet into Polars and stash it in `state`.

    Pattern:
    - Read the lookup sheet with openpyxl (`iter_rows(values_only=True)`)
    - Build a small Polars DataFrame
    - Cache it in `state` so downstream hooks can join/enrich quickly

    Best practice:
    - Keep this small. For large sheets, defer reading until you need it.
    """
    cfg = state

    # Only do this once per run.
    if "reference" in cfg:
        return None

    try:
        import polars as pl  # type: ignore
    except Exception as exc:
        if logger and hasattr(logger, "warning"):
            logger.warning("Polars is not installed; skipping reference load (%s).", exc)
        cfg["reference"] = {"df": None, "sheet": None}
        return None

    sheet_name = getattr(settings, "reference_sheet_name", "Reference")
    if sheet_name not in (getattr(workbook, "sheetnames", []) or []):
        if logger:
            logger.info("No %r sheet found; skipping reference load.", sheet_name)
        cfg["reference"] = {"df": None, "sheet": sheet_name}
        return None

    ws = workbook[sheet_name]

    # Beginner-friendly: read rows as values (no cell objects).
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        cfg["reference"] = {"df": pl.DataFrame(), "sheet": sheet_name}
        return None

    # Normalize column names (avoid None and strip whitespace).
    columns = [str(c).strip() if c is not None else f"col_{i + 1}" for i, c in enumerate(header)]

    data_rows: list[list[Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        # Skip completely empty rows.
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        data_rows.append(list(row))

    df: pl.DataFrame = pl.DataFrame(data_rows, schema=columns, orient="row")
    cfg["reference"] = {"df": df, "sheet": sheet_name}

    if logger:
        logger.info("Loaded reference sheet %r into Polars (%d rows).", sheet_name, df.height)

    return None
