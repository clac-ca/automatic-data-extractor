"""
ADE hook: ``on_table_transformed``.

This hook runs once per detected table after column mapping and per-column transforms,
and before validation. The table is fully materialized with canonical columns (when
available), making this the best stage for table-wide, cross-column finishing work.

Use this hook for actions that are awkward or impossible in per-column transforms, such
as deriving/splitting fields across multiple columns, final type normalization, adding
missing canonical columns for validator consistency, dropping obvious non-data rows, and
optionally enriching via workbook context (read-only) or cached lookups.

Contract
--------
- Called once per transformed table.
- Return a replacement ``pl.DataFrame`` to update the table, or ``None`` to leave it
  unchanged. Any other return value is an error.

Guidance
--------
- Keep logic deterministic and idempotent.
- Prefer vectorized Polars expressions; avoid Python row loops.
- Treat the workbook as read-only and use ``source_region`` only as source context.
- When constructing literals in expressions, use ``pl.lit(...)``.
"""


from __future__ import annotations

from collections.abc import Mapping, MutableMapping
from datetime import date, datetime
from typing import Any, TYPE_CHECKING

import polars as pl
from ade_engine.models import TableRegion

if TYPE_CHECKING:
    import openpyxl
    import openpyxl.worksheet.worksheet

    from ade_engine.extensions.registry import Registry
    from ade_engine.infrastructure.observability.logger import RunLogger
    from ade_engine.infrastructure.settings import Settings

# -----------------------------------------------------------------------------
# Polars convenience: string dtype selection across versions
# -----------------------------------------------------------------------------
_TEXT_DTYPES: list[pl.DataType] = []
for _attr in ("String", "Utf8"):
    _dt = getattr(pl, _attr, None)
    if _dt is not None:
        _TEXT_DTYPES.append(_dt)
TEXT_DTYPE: pl.DataType = _TEXT_DTYPES[0] if _TEXT_DTYPES else pl.Utf8  # type: ignore[attr-defined]


def register(registry: Registry) -> None:
    """Register this config package's `on_table_transformed` hook(s)."""
    registry.register_hook(on_table_transformed, hook="on_table_transformed", priority=0)

    # ---------------------------------------------------------------------
    # Examples (uncomment to enable, then customize as needed)
    # ---------------------------------------------------------------------
    # registry.register_hook(on_table_transformed_example_1_normalize_all_text, hook="on_table_transformed", priority=10)
    # registry.register_hook(on_table_transformed_example_2_add_provenance_columns, hook="on_table_transformed", priority=20)
    # registry.register_hook(on_table_transformed_example_3_derive_full_name, hook="on_table_transformed", priority=30)
    # registry.register_hook(on_table_transformed_example_4_parse_date_multi_format, hook="on_table_transformed", priority=40)
    # registry.register_hook(on_table_transformed_example_5_parse_currency_amount, hook="on_table_transformed", priority=50)
    # registry.register_hook(on_table_transformed_example_6_compute_line_total, hook="on_table_transformed", priority=60)
    # registry.register_hook(on_table_transformed_example_7_drop_non_data_rows, hook="on_table_transformed", priority=70)
    # registry.register_hook(on_table_transformed_example_8_add_report_date_from_sheet_header, hook="on_table_transformed", priority=80)
    # registry.register_hook(on_table_transformed_example_9_join_lookup_from_reference_sheet, hook="on_table_transformed", priority=90)

    # Advanced (network calls; non-deterministic; read docstring before enabling)
    # registry.register_hook(on_table_transformed_example_10_geocode_address_google, hook="on_table_transformed", priority=100)


def on_table_transformed(
    *,
    table: pl.DataFrame,  # Current table DF (post-transforms; pre-validation)
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,  # Source worksheet (openpyxl Worksheet)
    source_workbook: openpyxl.Workbook,  # Input workbook (openpyxl Workbook)
    source_region: TableRegion,  # Excel coords via .min_row/.max_row/.min_col/.max_col; helpers .a1/.header_row/.data_first_row
    table_index: int,  # 0-based table index within the source_sheet
    input_file_name: str,  # Input filename (basename)
    settings: Settings,  # Engine Settings
    metadata: Mapping[str, Any],  # Run/sheet metadata (filenames, sheet_index, etc.)
    state: MutableMapping[str, Any],  # Mutable dict shared across the run
    logger: RunLogger,  # RunLogger (structured events + text logs)
) -> pl.DataFrame | None:  # noqa: ARG001
    """Default implementation is a placeholder (no-op)."""
    return None


# ----------------------------
# Examples (uncomment in register() to enable)
# ----------------------------


def on_table_transformed_example_1_normalize_all_text(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame:
    """
    Example 1: Normalize all text columns.

    Teaches:
    - Dtype-based column targeting (string columns)
    - Safe string cleanup pipeline:
      - strip leading/trailing whitespace
      - collapse internal whitespace
      - convert common "null-like" tokens to null
    - Using `.name.keep()` when broadcasting `when/then/otherwise` over many columns
    """

    # Select string columns in a version-tolerant way.
    sel = pl.col(_TEXT_DTYPES)

    cleaned = sel.cast(TEXT_DTYPE, strict=False).str.strip_chars().str.replace_all(r"\s+", " ")
    lowered = cleaned.str.to_lowercase()

    null_like = lowered.is_in(["", "na", "n/a", "null", "none", "-", "--"])

    out = table.with_columns(pl.when(null_like).then(pl.lit(None)).otherwise(cleaned).name.keep())

    if logger:
        logger.info("Normalized all text columns (strip/collapse/null-like tokens).")
    return out


def on_table_transformed_example_2_add_provenance_columns(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame:
    """
    Example 2: Add provenance columns for traceability.

    Common in ingestion pipelines: you’ll thank yourself later when debugging.

    Teaches:
    - Adding constant columns with `pl.lit(...)`
    - Adding a stable row index with `pl.int_range(pl.len())` or `with_row_index`
    """

    sheet_name = (getattr(source_sheet, "title", None) or getattr(source_sheet, "name", None) or "").strip()

    # 0-based index within this materialized table
    row_index_expr = pl.int_range(pl.len(), dtype=pl.UInt32).alias("__row_index")

    return table.with_columns(
        row_index_expr,
        pl.lit(input_file_name).cast(TEXT_DTYPE, strict=False).alias("__source_file"),
        pl.lit(sheet_name).cast(TEXT_DTYPE, strict=False).alias("__source_sheet"),
        pl.lit(metadata.get("sheet_index")).cast(pl.Int64).alias("__source_sheet_index"),
    )


def on_table_transformed_example_3_derive_full_name(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """
    Example 3: Derive/backfill `full_name`.

    Teaches:
    - Building reusable expressions (first/last/full) without creating temp columns
    - `pl.concat_str(..., ignore_nulls=True)` for robust name building
    - Conditional update with `when/then/otherwise`
    """

    required = {"full_name", "first_name", "last_name"}
    if not required.issubset(set(table.columns)):
        return None

    first = pl.col("first_name").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    first = pl.when(first == "").then(pl.lit(None)).otherwise(first)

    last = pl.col("last_name").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    last = pl.when(last == "").then(pl.lit(None)).otherwise(last)

    full = pl.col("full_name").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    full = pl.when(full == "").then(pl.lit(None)).otherwise(full)

    derived = pl.concat_str([first, last], separator=" ", ignore_nulls=True).str.strip_chars()
    derived = pl.when(derived == "").then(pl.lit(None)).otherwise(derived)

    out = table.with_columns(
        pl.when(full.is_null()).then(derived).otherwise(full).alias("full_name")
    )

    if logger:
        logger.info("Backfilled full_name from first_name/last_name where missing.")
    return out


def on_table_transformed_example_4_parse_date_multi_format(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """
    Example 4: Parse a date column using multiple common formats.

    Teaches:
    - `str.strptime(pl.Date, ...)` with `strict=False`
    - Fallback parsing via `pl.coalesce(...)` (first successful parse wins)
    """

    col_name = "transaction_date"
    if col_name not in table.columns:
        return None

    raw = pl.col(col_name).cast(TEXT_DTYPE, strict=False).str.strip_chars()

    parsed = pl.coalesce(
        [
            raw.str.strptime(pl.Date, format="%Y-%m-%d", strict=False),
            raw.str.strptime(pl.Date, format="%m/%d/%Y", strict=False),
            raw.str.strptime(pl.Date, format="%d/%m/%Y", strict=False),
            raw.str.strptime(pl.Date, format="%Y%m%d", strict=False),
        ]
    )

    out = table.with_columns(parsed.alias(col_name))

    if logger:
        logger.info("Parsed %s to Date (multi-format, strict=False).", col_name)
    return out


def on_table_transformed_example_5_parse_currency_amount(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """
    Example 5: Parse a currency-like string column into Float64.

    Teaches:
    - Robust currency cleanup (commas, symbols, whitespace, parentheses negatives)
    - `cast(..., strict=False)` to avoid raising on bad cells
    - Shows how to handle EU-style decimals "1.234,56" as well

    Customize:
    - column name(s)
    - allowed symbols
    - whether you want Decimal instead of Float64
    """

    col_name = "amount"
    if col_name not in table.columns:
        return None

    raw = pl.col(col_name).cast(TEXT_DTYPE, strict=False).str.strip_chars()

    # Normalize whitespace early
    s = raw.str.replace_all(r"\s+", "")

    # Detect negatives expressed as parentheses, e.g. "(1,234.56)"
    is_paren_neg = s.str.starts_with("(") & s.str.ends_with(")")

    # Heuristic: if both '.' and ',' appear and it ends with comma-decimals, treat as EU style
    # Example: "€1.234,56" -> "1234.56"
    looks_eu = s.str.contains(r"\.") & s.str.contains(r",\d{1,4}\)?$")

    s_std = (
        pl.when(looks_eu)
        .then(
            s.str.replace_all(r"\.", "").str.replace_all(  # remove thousands dots
                ",", "."
            )  # decimal comma -> decimal dot
        )
        .otherwise(s)
    )

    # Remove parentheses and common currency symbols / thousands separators
    cleaned = (
        s_std.str.replace_all(r"[()]", "")
        .str.replace_all(r"[,$€£]", "")
        .str.replace_all(r"[^\d\.\-]", "")  # keep digits/dot/minus only
    )

    num = cleaned.cast(pl.Float64, strict=False)
    num = pl.when(is_paren_neg & num.is_not_null()).then(-num).otherwise(num)

    out = table.with_columns(num.alias(col_name))

    if logger:
        logger.info("Parsed %s to Float64 (currency cleanup, strict=False).", col_name)
    return out


def on_table_transformed_example_6_compute_line_total(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """
    Example 6: Compute a derived numeric column from multiple columns.

    Use case:
    - Invoices / purchases: line_total = quantity * unit_price

    Teaches:
    - Cross-column math
    - `coalesce` to keep existing line_total if present
    """

    required = {"quantity", "unit_price"}
    if not required.issubset(table.columns):
        return None

    qty = pl.col("quantity").cast(pl.Float64, strict=False)
    unit = pl.col("unit_price").cast(pl.Float64, strict=False)
    derived = (qty * unit).round(2)

    if "line_total" in table.columns:
        out = table.with_columns(
            pl.coalesce([pl.col("line_total").cast(pl.Float64, strict=False), derived]).alias(
                "line_total"
            )
        )
    else:
        out = table.with_columns(derived.alias("line_total"))

    if logger:
        logger.info("Computed line_total = quantity * unit_price (rounded to 2 decimals).")
    return out


def on_table_transformed_example_7_drop_non_data_rows(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame:
    """
    Example 7: Drop obvious non-data rows.

    Teaches:
    - Drop rows that are completely null (any_horizontal/all().is_not_null())
    - Drop repeated header rows embedded inside data (common in exported PDFs / print views)
    - Optional drop of "TOTAL"/"SUBTOTAL" rows (conservative heuristic)

    IMPORTANT:
    - Be conservative. Dropping rows is irreversible and can hide issues.
      Prefer logging counts, and tune heuristics per dataset.
    """

    before = table.height
    out = table

    # 1) Drop all-null rows (keep rows where ANY column is non-null)
    if out.width > 0:
        out = out.filter(pl.any_horizontal([pl.col(c).is_not_null() for c in out.columns]))

    # 2) Drop repeated header rows:
    #    A row is "header-like" if many cells equal their column names (case-insensitive).
    cols = out.columns
    if cols:
        comparisons: list[pl.Expr] = []
        for c in cols:
            comparisons.append(
                pl.col(c).cast(TEXT_DTYPE, strict=False).str.strip_chars().str.to_lowercase()
                == pl.lit(c.lower())
            )

        # Count how many columns match their own name for each row
        match_count = pl.sum_horizontal([e.cast(pl.Int32) for e in comparisons])
        # threshold: 80% of columns match -> treat as repeated header
        threshold = max(1, int(len(cols) * 0.8))
        is_repeated_header = match_count >= threshold

        out = out.filter(~is_repeated_header)

    # 3) OPTIONAL: drop totals/subtotals based on first string column (very conservative)
    #    Adjust column choice / patterns for your data.
    text_cols = out.select(pl.col(_TEXT_DTYPES)).columns if out.width > 0 else []
    if text_cols:
        c0 = text_cols[0]
        marker = pl.col(c0).cast(TEXT_DTYPE, strict=False).str.strip_chars().str.to_lowercase()
        is_total = marker.is_in(["total", "subtotal", "grand total"]) | marker.str.starts_with(
            "total "
        )
        out = out.filter(~is_total)

    if logger:
        logger.info(
            "Dropped non-data rows: %d -> %d (removed %d).",
            int(before),
            int(out.height),
            int(before - out.height),
        )
    return out


def on_table_transformed_example_8_add_report_date_from_sheet_header(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """
    Example 8: Use openpyxl to read a sheet-level constant (e.g., report_date) and add it.

    This is a *very* common real-world case: the report date / client / currency / batch id
    lives in header cells above the table.

    Pattern:
    - find a label like "Report Date" in the top-left header region
    - read the adjacent cell value
    - normalize and add as a constant column for all rows

    Works well *before validation*, because validators can now assume report_date exists.
    """

    # Find a label like "Report Date" in a small top-left region, then read the adjacent cell.
    target = "report date"
    found_label = False
    report_date_value: Any | None = None

    max_scan_row = min(40, int(source_region.header_row))

    try:
        rows = source_sheet.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=20)
    except Exception:
        rows = None

    if rows:
        for row in rows:
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if str(v).strip().lower() == target:
                    row_idx = cell.row
                    col_idx = cell.column
                    if row_idx is None or col_idx is None:
                        continue
                    found_label = True
                    try:
                        report_date_value = source_sheet.cell(row=row_idx, column=col_idx + 1).value
                    except Exception:
                        report_date_value = None
                    break
            if found_label:
                break

    if not found_label or report_date_value is None:
        return None

    # openpyxl often returns datetime/date for date cells; normalize to date.
    if isinstance(report_date_value, datetime):
        report_date_value = report_date_value.date()
    if isinstance(report_date_value, date):
        out = table.with_columns(pl.lit(report_date_value, dtype=pl.Date).alias("report_date"))
        if logger:
            logger.info("Added report_date from source_sheet header (date object).")
        return out

    # Otherwise treat as a string and let Polars parse it.
    s = str(report_date_value).strip()
    if not s:
        return None

    out = table.with_columns(pl.lit(s).cast(TEXT_DTYPE, strict=False).alias("__report_date_raw"))
    out = out.with_columns(
        pl.col("__report_date_raw").str.to_date(strict=False).alias("report_date")
    )
    out = out.drop("__report_date_raw")

    if logger:
        logger.info("Added report_date from source_sheet header (parsed from string).")
    return out


def on_table_transformed_example_9_join_lookup_from_reference_sheet(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """
    Example 9: Read a lookup table from another worksheet (openpyxl), cache it, join via Polars.

    This is an excellent “template-quality” example because it demonstrates:
    - openpyxl: reading structured rows with iter_rows(values_only=True)
    - caching in `state` so we don't reread the lookup for every table
    - polars: join enrichment (and join cardinality validation when available)

    Scenario:
    - Your data table has a `state_code` (e.g., "CA")
    - The workbook has a sheet "Lookups" with columns A=state_code, B=state_name

    Customize:
    - sheet name, columns, header row, and the join keys
    """

    join_key = "state_code"
    if join_key not in table.columns:
        return None

    lookup_sheet_name = "Lookups"
    sheetnames = getattr(source_workbook, "sheetnames", []) or []
    if lookup_sheet_name not in sheetnames:
        return None

    cfg = state

    cache = cfg.get("cache")
    if not isinstance(cache, MutableMapping):
        cache = {}
        cfg["cache"] = cache

    # Cache per input file + lookup source_sheet name (adjust keying for your environment).
    cache_key = ("lookup_df", input_file_name, lookup_sheet_name)
    lookup_df = cache.get(cache_key)

    if lookup_df is None and cache_key not in cache:
        ws_lookup = source_workbook[lookup_sheet_name]

        rows: list[dict[str, Any]] = []
        # Assume header is row 1, data starts at row 2, columns A..B are the mapping.
        for code, name in ws_lookup.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
            if code is None:
                continue
            code_s = str(code).strip()
            if not code_s:
                continue
            name_s = None if name is None else str(name).strip()
            rows.append({"state_code": code_s, "state_name": name_s})

        if not rows:
            if logger:
                logger.warning(
                    "Lookup source_sheet %r found but contained no usable rows.", lookup_sheet_name
                )
            cache[cache_key] = None
            return None

        lookup_df = pl.DataFrame(rows).unique(subset=["state_code"], keep="first")
        cache[cache_key] = lookup_df

        if logger:
            logger.info(
                "Loaded lookup table from source_sheet %r (%d rows).",
                lookup_sheet_name,
                int(lookup_df.height),
            )

    if lookup_df is None:
        return None

    if lookup_df is None or not isinstance(lookup_df, pl.DataFrame):
        return None

    # Normalize join key in the main table (trim + uppercase, for example).
    main = table.with_columns(
        pl.col(join_key)
        .cast(TEXT_DTYPE, strict=False)
        .str.strip_chars()
        .str.to_uppercase()
        .alias(join_key)
    )
    other = lookup_df.with_columns(
        pl.col(join_key)
        .cast(TEXT_DTYPE, strict=False)
        .str.strip_chars()
        .str.to_uppercase()
        .alias(join_key)
    )

    try:
        out = main.join(other, on=join_key, how="left", validate="m:1")
    except TypeError:
        out = main.join(other, on=join_key, how="left")

    if logger:
        logger.info("Enriched table via lookup join on %r -> added state_name.", join_key)
    return out


# ---------------------------------------------------------
# Advanced example (external enrichment; non-deterministic)
# ---------------------------------------------------------


def on_table_transformed_example_10_geocode_address_google(
    *,
    table: pl.DataFrame,
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_workbook: openpyxl.Workbook,
    source_region: TableRegion,  # See `TableRegion` notes above
    table_index: int,
    input_file_name: str,
    settings: Settings,
    metadata: Mapping[str, Any],
    state: MutableMapping[str, Any],
    logger: RunLogger,
) -> pl.DataFrame | None:
    """Example 10 (advanced): use Google Geocoding to canonicalize `address_full`.

    What this teaches
    - Detect whether a canonical column exists (`address_full`)
    - Dedupe values and cache results in `state` (efficient + deterministic-ish)
    - Join results back and update/add columns with Polars

    Important
    - This makes network calls: it is non-deterministic and can be rate-limited.
    - Prefer offline enrichment where possible, or cache aggressively.
    - Requires a Google API key (billing/quota applies).

    Tip
    - Keep this disabled in templates unless you explicitly want to teach “how to do enrichment.”
    """

    import json
    import os
    from urllib.parse import urlencode
    from urllib.request import Request, urlopen

    if "address_full" not in table.columns:
        return None

    api_key = getattr(settings, "google_geocoding_api_key", None) or os.getenv(
        "GOOGLE_GEOCODING_API_KEY"
    )
    if not api_key:
        if logger:
            logger.warning(
                "Geocoding skipped: set settings.google_geocoding_api_key or env GOOGLE_GEOCODING_API_KEY"
            )
        return None

    cfg = state

    cache = cfg.get("google_geocoding_cache")
    if not isinstance(cache, dict):
        cache = {}
        cfg["google_geocoding_cache"] = cache

    address_key = (
        pl.col("address_full")
        .cast(TEXT_DTYPE, strict=False)
        .str.strip_chars()
        .str.replace_all(r"\s+", " ")
    )
    address_key = (
        pl.when(address_key.is_null() | (address_key == ""))
        .then(pl.lit(None))
        .otherwise(address_key)
    )

    unique_addresses: list[str] = (
        table.select(address_key.alias("__address_key"))
        .filter(pl.col("__address_key").is_not_null())
        .unique()
        .get_column("__address_key")
        .to_list()
    )
    if not unique_addresses:
        return None

    unique_addresses = sorted(unique_addresses)
    max_requests = int(getattr(settings, "google_geocoding_max_requests", 200) or 200)
    if max_requests > 0 and len(unique_addresses) > max_requests:
        if logger:
            logger.warning(
                "Geocoding: limiting unique addresses %d -> %d (set settings.google_geocoding_max_requests to change).",
                len(unique_addresses),
                max_requests,
            )
        unique_addresses = unique_addresses[:max_requests]

    for addr in unique_addresses:
        if addr in cache:
            continue

        url = "https://maps.googleapis.com/maps/api/geocode/json?" + urlencode(
            {"address": addr, "key": api_key}
        )
        try:
            req = Request(url, headers={"User-Agent": "ade-config-template"})
            with urlopen(req, timeout=10) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
        except Exception as exc:
            if logger:
                logger.warning("Geocoding request failed for %r: %s", addr, exc)
            cache[addr] = {}
            continue

        if payload.get("status") != "OK" or not payload.get("results"):
            cache[addr] = {}
            continue

        result = payload["results"][0]
        components = result.get("address_components") or []

        by_type: dict[str, dict] = {}
        for comp in components:
            for t in comp.get("types", []):
                by_type.setdefault(t, comp)

        city_comp = by_type.get("locality") or by_type.get("postal_town")
        region_comp = by_type.get("administrative_area_level_1")  # state/province
        postal_comp = by_type.get("postal_code")
        country_comp = by_type.get("country")

        cache[addr] = {
            "formatted_address": result.get("formatted_address"),
            "city": city_comp.get("long_name") if city_comp else None,
            "province_state": (region_comp.get("short_name") or region_comp.get("long_name"))
            if region_comp
            else None,
            "postal_code": postal_comp.get("long_name") if postal_comp else None,
            "country": country_comp.get("long_name") if country_comp else None,
        }

    mapping_df = pl.DataFrame(
        [
            {
                "__address_key": addr,
                "__geo_formatted_address": cache.get(addr, {}).get("formatted_address"),
                "__geo_city": cache.get(addr, {}).get("city"),
                "__geo_province_state": cache.get(addr, {}).get("province_state"),
                "__geo_postal_code": cache.get(addr, {}).get("postal_code"),
                "__geo_country": cache.get(addr, {}).get("country"),
            }
            for addr in unique_addresses
        ]
    )

    out = table.with_columns(address_key.alias("__address_key"))
    try:
        out = out.join(mapping_df, on="__address_key", how="left", validate="m:1")
    except TypeError:
        out = out.join(mapping_df, on="__address_key", how="left")

    formatted = pl.col("__geo_formatted_address").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    formatted = (
        pl.when(formatted.is_null() | (formatted == "")).then(pl.lit(None)).otherwise(formatted)
    )

    original = pl.col("address_full").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    original = pl.when(original.is_null() | (original == "")).then(pl.lit(None)).otherwise(original)

    out = out.with_columns(pl.coalesce([formatted, original]).alias("address_full"))

    geo_city = pl.col("__geo_city").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    geo_region = pl.col("__geo_province_state").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    geo_postal = pl.col("__geo_postal_code").cast(TEXT_DTYPE, strict=False).str.strip_chars()
    geo_country = pl.col("__geo_country").cast(TEXT_DTYPE, strict=False).str.strip_chars()

    out = out.with_columns(
        (
            pl.coalesce([geo_city, pl.col("city").cast(TEXT_DTYPE, strict=False).str.strip_chars()])
            if "city" in out.columns
            else geo_city
        ).alias("city"),
        (
            pl.coalesce(
                [
                    geo_region,
                    pl.col("province_state").cast(TEXT_DTYPE, strict=False).str.strip_chars(),
                ]
            )
            if "province_state" in out.columns
            else geo_region
        ).alias("province_state"),
        (
            pl.coalesce(
                [geo_postal, pl.col("postal_code").cast(TEXT_DTYPE, strict=False).str.strip_chars()]
            )
            if "postal_code" in out.columns
            else geo_postal
        ).alias("postal_code"),
        (
            pl.coalesce(
                [geo_country, pl.col("country").cast(TEXT_DTYPE, strict=False).str.strip_chars()]
            )
            if "country" in out.columns
            else geo_country
        ).alias("country"),
    )

    return out.drop(
        [
            "__address_key",
            "__geo_formatted_address",
            "__geo_city",
            "__geo_province_state",
            "__geo_postal_code",
            "__geo_country",
        ]
    )
