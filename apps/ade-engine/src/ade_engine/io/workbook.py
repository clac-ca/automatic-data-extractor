"""Workbook IO helpers for :class:`~ade_engine.engine.ADEngine`."""

from __future__ import annotations

import csv
from contextlib import contextmanager, suppress
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from ade_engine.exceptions import InputError


def load_source_workbook(path: Path) -> Workbook:
    """Load source data from CSV/XLSX into a workbook."""

    if path.suffix.lower() == ".csv":
        wb = Workbook()
        ws = wb.active
        ws.title = path.stem
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                ws.append(row)
        return wb

    return openpyxl.load_workbook(filename=path, read_only=True, data_only=True)


@contextmanager
def open_source_workbook(path: Path):
    """Context manager for safely opening source workbooks."""

    workbook = load_source_workbook(path)
    try:
        yield workbook
    finally:
        with suppress(Exception):
            workbook.close()


def create_output_workbook() -> Workbook:
    """Create a clean output workbook with no default sheet."""

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])
    return workbook


def resolve_sheet_names(workbook: Workbook, requested: list[str] | None) -> list[str]:
    """Determine which sheets to process, preserving source order."""

    visible = [ws.title for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    if not requested:
        return visible

    cleaned = [name.strip() for name in requested if isinstance(name, str) and name.strip()]
    unique_requested = list(dict.fromkeys(cleaned))  # preserve order, drop duplicates

    missing = [name for name in unique_requested if name not in visible]
    if missing:
        raise InputError(f"Worksheet(s) not found: {', '.join(missing)}")

    order_index = {name: idx for idx, name in enumerate(visible)}
    return sorted(unique_requested, key=lambda n: order_index[n])


class WorkbookIO:
    """Adapter for workbook operations (open/save)."""

    def open_source(self, path: Path):
        return open_source_workbook(path)

    def create_output(self) -> Workbook:
        return create_output_workbook()

    def save_output(self, workbook: Workbook, path: Path) -> None:
        workbook.save(path)
        with suppress(Exception):
            workbook.close()

    def resolve_sheet_names(self, workbook: Workbook, requested: list[str] | None) -> list[str]:
        return resolve_sheet_names(workbook, requested)


__all__ = [
    "WorkbookIO",
    "create_output_workbook",
    "load_source_workbook",
    "open_source_workbook",
    "resolve_sheet_names",
]
