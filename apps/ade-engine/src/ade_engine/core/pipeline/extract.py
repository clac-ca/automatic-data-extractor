"""Row detector integration and ExtractedTable construction.

This module bridges streamed sheet rows with config-provided row detectors to
produce :class:`ade_engine.core.types.ExtractedTable` instances. The implementation is
intentionally lightweight but deterministic, mirroring the behavior outlined in
``docs/03-io-and-table-detection.md``.
"""

from __future__ import annotations

import importlib
import inspect
import logging
from dataclasses import dataclass
from importlib import resources
from pathlib import Path
from types import ModuleType
from typing import Any, Callable, Iterable, Iterator, Mapping

from openpyxl import load_workbook

from ade_engine.config.loader import ConfigRuntime
from ade_engine.core.errors import ConfigError, InputError
from ade_engine.core.types import ExtractedTable, RunContext, RunRequest
from ade_engine.infra.io import iter_csv_rows, iter_sheet_rows, list_input_files
from ade_engine.infra.telemetry import EventEmitter

RowDetectorFn = Callable[..., Mapping[str, Any]]

HEADER_SCORE_THRESHOLD = 0.6
DATA_SCORE_THRESHOLD = 0.5


def _validate_keyword_only(func: Callable[..., object], *, label: str) -> None:
    signature = inspect.signature(func)
    invalid_params = [
        p
        for p in signature.parameters.values()
        if p.kind in (inspect.Parameter.POSITIONAL_OR_KEYWORD, inspect.Parameter.POSITIONAL_ONLY)
    ]
    if invalid_params:
        names = ", ".join(p.name for p in invalid_params)
        raise ConfigError(f"{label} must declare keyword-only parameters (invalid: {names})")

    if not any(p.kind is inspect.Parameter.VAR_KEYWORD for p in signature.parameters.values()):
        raise ConfigError(f"{label} must accept **_ for forwards compatibility")


def _load_row_detectors(package: ModuleType) -> list[RowDetectorFn]:
    try:
        detector_pkg = importlib.import_module(f"{package.__name__}.row_detectors")
    except ModuleNotFoundError:
        return []

    detectors: list[RowDetectorFn] = []
    for entry in resources.files(detector_pkg).iterdir():
        if entry.name.startswith("_") or entry.suffix != ".py":
            continue

        module = importlib.import_module(f"{detector_pkg.__name__}.{entry.stem}")
        for name, attr in inspect.getmembers(module, predicate=callable):
            if not name.startswith("detect_"):
                continue
            _validate_keyword_only(attr, label=f"Row detector '{module.__name__}.{name}'")
            detectors.append(attr)

    return detectors


@dataclass(frozen=True)
class _RowScore:
    row_index: int
    values: list[Any]
    header_score: float
    data_score: float
    contributions: dict[str, dict[str, float]]


def _resolve_sheet_names(path: Path, requested: Iterable[str] | None) -> list[str]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        available = [name for name in workbook.sheetnames if workbook[name].sheet_state == "visible"]
        if requested is None:
            return available

        missing = [name for name in requested if name not in available]
        if missing:
            raise InputError(f"Worksheet(s) {', '.join(missing)} not found in `{path}`")
        return list(requested)
    finally:
        workbook.close()


def _score_rows(
    *,
    rows: Iterator[tuple[int, list[Any]]],
    detectors: list[RowDetectorFn],
    run: RunContext,
    manifest: Any,
    logger: logging.Logger,
    state: dict[str, Any],
    event_emitter: EventEmitter,
) -> list[_RowScore]:
    scored: list[_RowScore] = []
    for row_index, values in rows:
        label_totals: dict[str, float] = {}
        contributions: dict[str, dict[str, float]] = {}
        for detector in detectors:
            result = detector(
                run=run,
                state=state,
                row_index=row_index,
                row_values=values,
                manifest=manifest,
                logger=logger,
                event_emitter=event_emitter,
            )
            scores = result.get("scores", {}) if isinstance(result, Mapping) else {}
            for label, delta in scores.items():
                label_totals[label] = label_totals.get(label, 0.0) + float(delta)
                detector_name = f"{detector.__module__}.{detector.__name__}"
                bucket = contributions.setdefault(detector_name, {})
                bucket[label] = float(delta)

        scored.append(
            _RowScore(
                row_index=row_index,
                values=list(values),
                header_score=label_totals.get("header", 0.0),
                data_score=label_totals.get("data", 0.0),
                contributions=contributions,
            )
        )

    return scored


def _detect_tables_for_sheet(
    *,
    scored_rows: list[_RowScore],
    source_file: Path,
    source_sheet: str | None,
    header_threshold: float = HEADER_SCORE_THRESHOLD,
    data_threshold: float = DATA_SCORE_THRESHOLD,
    event_emitter: EventEmitter,
) -> list[ExtractedTable]:
    tables: list[ExtractedTable] = []
    table_index = 0
    position = 0
    while position < len(scored_rows):
        row = scored_rows[position]
        if row.header_score < header_threshold:
            position += 1
            continue

        data_rows: list[_RowScore] = []
        lookahead = position + 1
        while lookahead < len(scored_rows) and scored_rows[lookahead].data_score >= data_threshold:
            data_rows.append(scored_rows[lookahead])
            lookahead += 1

        if data_rows:
            table = ExtractedTable(
                source_file=source_file,
                source_sheet=source_sheet,
                table_index=table_index,
                header_row=["" if cell is None else str(cell) for cell in row.values],
                data_rows=[dr.values for dr in data_rows],
                header_row_index=row.row_index,
                first_data_row_index=data_rows[0].row_index,
                last_data_row_index=data_rows[-1].row_index,
            )
            tables.append(table)
            _emit_row_detector_score(
                event_emitter=event_emitter,
                table=table,
                header_threshold=header_threshold,
                data_threshold=data_threshold,
                header_row=row,
                data_rows=data_rows,
            )
            table_index += 1
        position = lookahead if lookahead > position else position + 1

    return tables


def _serialize_row_contributions(contributions: dict[str, dict[str, float]]) -> list[dict[str, Any]]:
    return [{"detector": detector, "scores": scores} for detector, scores in contributions.items()]


def _emit_row_detector_score(
    *,
    event_emitter: EventEmitter,
    table: ExtractedTable,
    header_threshold: float,
    data_threshold: float,
    header_row: _RowScore,
    data_rows: list[_RowScore],
) -> None:
    event_emitter.custom(
        "row_detector.score",
        source_file=str(table.source_file),
        source_sheet=table.source_sheet,
        table_index=table.table_index,
        thresholds={"header": header_threshold, "data": data_threshold},
        header_row_index=header_row.row_index,
        data_row_start_index=data_rows[0].row_index if data_rows else None,
        data_row_end_index=data_rows[-1].row_index if data_rows else None,
        trigger={
            "row_index": header_row.row_index,
            "header_score": header_row.header_score,
            "data_score": header_row.data_score,
            "contributions": _serialize_row_contributions(header_row.contributions),
            "sample": header_row.values[:5],
        },
    )


def extract_raw_tables(
    *,
    request: RunRequest,
    run: RunContext,
    runtime: ConfigRuntime,
    logger: logging.Logger | None = None,
    event_emitter: EventEmitter,
) -> list[ExtractedTable]:
    """Detect tables across CSV/XLSX inputs using config row detectors."""

    logger = logger or logging.getLogger(__name__)
    detectors = _load_row_detectors(runtime.package)
    state: dict[str, Any] = {}

    if request.input_files:
        source_files = [Path(path).resolve() for path in request.input_files]
    elif request.input_dir:
        source_files = list_input_files(request.input_dir)
    else:
        raise InputError("RunRequest must include either input_files or input_dir")

    detected: list[ExtractedTable] = []
    for source_file in source_files:
        suffix = source_file.suffix.lower()
        if suffix == ".csv":
            scored_rows = _score_rows(
                rows=iter_csv_rows(source_file),
                detectors=detectors,
                run=run,
                manifest=runtime.manifest,
                logger=logger,
                state=state,
                event_emitter=event_emitter,
            )
            detected.extend(
                _detect_tables_for_sheet(
                    scored_rows=scored_rows,
                    source_file=source_file,
                    source_sheet=None,
                    event_emitter=event_emitter,
                )
            )
            continue

        sheet_names = _resolve_sheet_names(source_file, request.input_sheets)
        for sheet_name in sheet_names:
            scored_rows = _score_rows(
                rows=iter_sheet_rows(source_file, sheet_name),
                detectors=detectors,
                run=run,
                manifest=runtime.manifest,
                logger=logger,
                state=state,
                event_emitter=event_emitter,
            )
            detected.extend(
                _detect_tables_for_sheet(
                    scored_rows=scored_rows,
                    source_file=source_file,
                    source_sheet=sheet_name,
                    event_emitter=event_emitter,
                )
            )

    return detected


__all__ = ["extract_raw_tables"]
