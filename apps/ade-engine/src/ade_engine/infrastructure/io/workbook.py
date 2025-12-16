"""Workbook IO helpers for :class:`~ade_engine.application.engine.Engine`."""

from __future__ import annotations

import csv
from contextlib import contextmanager, suppress
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from ade_engine.models.errors import InputError


def load_source_workbook(path: Path) -> Workbook:
    """Load source data from CSV/XLSX into a workbook."""

    if path.suffix.lower() == ".csv":
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise InputError("Failed to initialize worksheet for CSV input")
        ws.title = path.stem
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                ws.append(row)
        return wb

    return openpyxl.load_workbook(filename=path, read_only=True, data_only=True)


@contextmanager
def open_source_workbook(path: Path):
    """Context manager for safely opening source workbooks."""

    workbook = load_source_workbook(path)
    try:
        yield workbook
    finally:
        with suppress(Exception):
            workbook.close()


def create_output_workbook() -> Workbook:
    """Create a clean output workbook with no default sheet."""

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])
    return workbook


def resolve_sheet_names(workbook: Workbook, requested: list[str] | None) -> list[str]:
    """Determine which sheets to process, preserving source order."""

    visible = [ws.title for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    if not requested:
        return visible

    cleaned = [name.strip() for name in requested if isinstance(name, str) and name.strip()]
    unique_requested = list(dict.fromkeys(cleaned))  # preserve order, drop duplicates

    missing = [name for name in unique_requested if name not in visible]
    if missing:
        raise InputError(f"Worksheet(s) not found: {', '.join(missing)}")

    order_index = {name: idx for idx, name in enumerate(visible)}
    return sorted(unique_requested, key=lambda n: order_index[n])


__all__ = [
    "create_output_workbook",
    "load_source_workbook",
    "open_source_workbook",
    "resolve_sheet_names",
]
