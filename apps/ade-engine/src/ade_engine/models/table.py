from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

import polars as pl


@dataclass(frozen=True, slots=True)
class TableRegion:
    """Rectangular table bounds in Excel coordinates (1-based, inclusive)."""

    min_row: int
    min_col: int
    max_row: int
    max_col: int
    header_row_count: int = 1

    def __post_init__(self) -> None:
        if min(self.min_row, self.min_col, self.max_row, self.max_col) < 1:
            raise ValueError("TableRegion coordinates must be >= 1 (Excel-style).")
        if self.min_row > self.max_row:
            raise ValueError("min_row must be <= max_row.")
        if self.min_col > self.max_col:
            raise ValueError("min_col must be <= max_col.")
        if self.header_row_count < 1:
            raise ValueError("header_row_count must be >= 1.")
        if self.header_row_count > self.height:
            raise ValueError("header_row_count must be <= table height.")

    @property
    def cell_range(self) -> CellRange:
        return CellRange(
            min_col=self.min_col,
            min_row=self.min_row,
            max_col=self.max_col,
            max_row=self.max_row,
        )

    @property
    def a1(self) -> str:
        return self.cell_range.coord

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def header_row(self) -> int:
        return self.min_row

    @property
    def data_first_row(self) -> int:
        return self.min_row + self.header_row_count

    @property
    def data_min_row(self) -> int:
        return min(self.min_row + self.header_row_count, self.max_row)

    @property
    def has_data_rows(self) -> bool:
        return self.max_row >= self.data_first_row

    @property
    def data_row_count(self) -> int:
        return max(0, self.max_row - self.min_row + 1 - self.header_row_count)

    def iter_values(self, ws: Worksheet, *, values_only: bool = True):
        return ws.iter_rows(
            min_row=self.min_row,
            max_row=self.max_row,
            min_col=self.min_col,
            max_col=self.max_col,
            values_only=values_only,
        )


@dataclass
class SourceColumn:
    index: int
    header: Any
    values: list[Any]


@dataclass
class MappedColumn:
    field_name: str
    source_index: int
    header: Any
    values: list[Any]
    score: float | None = None


@dataclass
class TableResult:
    """Processed table facts for reporting/debugging.

    Table values are stored only in ``table`` (a single Polars DataFrame).
    """

    sheet_name: str
    table: pl.DataFrame
    source_region: TableRegion
    source_columns: list[SourceColumn]
    table_index: int = 0
    sheet_index: int = 0
    mapped_columns: list[MappedColumn] = field(default_factory=list)
    unmapped_columns: list[SourceColumn] = field(default_factory=list)
    column_scores: dict[int, dict[str, float]] = field(default_factory=dict)
    duplicate_unmapped_indices: set[int] = field(default_factory=set)
    row_count: int = 0
    output_region: TableRegion | None = None
    output_sheet_name: str | None = None

    def mapping_lookup(self) -> dict[str, int | None]:
        return {col.field_name: col.source_index for col in self.mapped_columns}


__all__ = ["SourceColumn", "MappedColumn", "TableRegion", "TableResult"]
