from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Sequence

import polars as pl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ade_engine.extensions.registry import Registry
from ade_engine.infrastructure.observability.logger import RunLogger
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.table import TableResult

_RESERVED_PREFIX = "__ade_"


@dataclass
class SheetWriter:
    """Simple worksheet writer that tracks an explicit row cursor.

    openpyxl's ``Worksheet.append`` advances an internal cursor even when appending
    an empty row; ``max_row`` does not. Tracking our own cursor keeps output ranges
    accurate without relying on openpyxl internals.
    """

    worksheet: Worksheet
    row: int = 0  # last written row index (1-based); 0 means "nothing written yet"

    def write_row(self, values: Sequence[Any]) -> int:
        self.worksheet.append(list(values))
        self.row += 1
        return self.row

    def blank_row(self) -> int:
        return self.write_row([])


def derive_write_table(*, table: pl.DataFrame, registry: Registry, settings: Settings) -> pl.DataFrame:
    write_table = table

    if settings.remove_unmapped_columns:
        canonical_fields = set(registry.fields.keys())
        drop_cols = [
            c for c in write_table.columns if not c.startswith(_RESERVED_PREFIX) and c not in canonical_fields
        ]
        if drop_cols:
            write_table = write_table.drop(drop_cols)

    if not settings.write_diagnostics_columns:
        reserved_cols = [c for c in write_table.columns if c.startswith(_RESERVED_PREFIX)]
        if reserved_cols:
            write_table = write_table.drop(reserved_cols)

    return write_table


def render_table(
    *,
    table_result: TableResult,
    writer: SheetWriter,
    registry: Registry,
    settings: Settings,
    logger: RunLogger,
) -> pl.DataFrame:
    """Write the table to ``writer.worksheet`` and update ``table_result`` facts."""

    write_table = derive_write_table(table=table_result.table, registry=registry, settings=settings)

    start_row = writer.row + 1

    headers = list(write_table.columns)
    writer.write_row(headers)

    for row in write_table.iter_rows(named=False):
        writer.write_row(row)

    rows_written = 1 + write_table.height
    col_count = len(headers)
    if col_count > 0 and rows_written > 0:
        end_row = start_row + rows_written - 1
        end_col = get_column_letter(col_count)
        output_range = f"A{start_row}:{end_col}{end_row}"
    else:
        output_range = ""

    table_result.output_range = output_range or None
    table_result.output_sheet_name = writer.worksheet.title

    logger.event(
        "table.written",
        message=f"Rendered table with {len(headers)} columns",
        data={
            "sheet_name": writer.worksheet.title,
            "table_index": table_result.table_index,
            "output_range": output_range,
        },
    )

    return write_table


__all__ = ["SheetWriter", "derive_write_table", "render_table"]
