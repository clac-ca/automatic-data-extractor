from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook

from ade_engine.core.errors import InputError

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm"}


def _ensure_supported(path: Path) -> None:
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise InputError(f"File `{path}` has unsupported extension `{path.suffix}`")


def list_input_files(input_dir: Path) -> list[Path]:
    """Return a deterministic list of supported source files under ``input_dir``."""

    if not input_dir.exists():
        raise InputError(f"Input directory does not exist: {input_dir}")
    if not input_dir.is_dir():
        raise InputError(f"Input path is not a directory: {input_dir}")

    discovered: list[Path] = []
    for child in input_dir.iterdir():
        if child.name.startswith("."):
            continue
        if not child.is_file():
            continue
        _ensure_supported(child)
        discovered.append(child.resolve())

    return sorted(discovered)


def iter_csv_rows(path: Path) -> Iterator[tuple[int, list]]:
    """Stream 1-based row index and values from a CSV file."""

    if not path.exists():
        raise InputError(f"Source file not found: {path}")
    if not path.is_file():
        raise InputError(f"Source path is not a file: {path}")
    _ensure_supported(path)

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row_index, row_values in enumerate(reader, start=1):
            yield row_index, list(row_values)


def iter_sheet_rows(path: Path, sheet_name: str) -> Iterable[tuple[int, list]]:
    """Stream 1-based row index and values from a worksheet within an XLSX file."""

    if not path.exists():
        raise InputError(f"Source file not found: {path}")
    if not path.is_file():
        raise InputError(f"Source path is not a file: {path}")
    _ensure_supported(path)

    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - exercised via error mapping tests later
        raise InputError(f"Could not read workbook `{path}`: {exc}") from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise InputError(f"Worksheet `{sheet_name}` not found in `{path}`")

        worksheet = workbook[sheet_name]
        for row_index, row_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            yield row_index, [cell for cell in row_values]
    finally:
        workbook.close()
