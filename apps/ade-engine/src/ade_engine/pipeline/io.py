"""Input discovery and ingestion helpers."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Iterable, Iterator

import openpyxl


def list_input_files(input_dir: Path) -> list[Path]:
    """Return sorted job input files limited to CSV/XLSX types."""

    if not input_dir.exists():
        return []
    candidates = [
        path
        for path in sorted(input_dir.iterdir())
        if path.suffix.lower() in {".csv", ".xlsx"}
    ]
    return [path for path in candidates if path.is_file()]


def read_table(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[list[str], list[list[Any]]]:
    """Read a CSV or XLSX file returning the header row and data rows."""

    tables = list(iter_tables(path, sheet_names=[sheet_name] if sheet_name else None))
    if not tables:
        raise RuntimeError(f"Input file '{path.name}' is empty")
    _, header, rows = tables[0]
    return header, rows


def iter_tables(
    path: Path,
    *,
    sheet_names: Iterable[str] | None = None,
) -> Iterator[tuple[str | None, list[str], list[list[Any]]]]:
    """Yield (sheet name, header, rows) tuples for each worksheet in the input."""

    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            try:
                header = next(reader)
            except StopIteration:
                return
            data_rows = [list(row) for row in reader]
        yield None, [str(value) if value is not None else "" for value in header], data_rows
        return

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        targets = list(sheet_names) if sheet_names else list(workbook.sheetnames)
        missing = [name for name in targets if name not in workbook.sheetnames]
        if missing:
            raise RuntimeError(
                f"Worksheet '{missing[0]}' not found in {path.name}"
            )

        for name in targets:
            sheet = workbook[name]
            iterator = sheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration:
                raise RuntimeError(f"Input sheet '{name}' in '{path.name}' is empty")
            header_row = [str(value) if value is not None else "" for value in header]
            data_rows = [
                [cell if cell is not None else "" for cell in row] for row in iterator
            ]
            yield name, header_row, data_rows
    finally:
        workbook.close()


def sheet_name(stem: str) -> str:
    """Normalize worksheet names to Excel-safe identifiers."""

    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", stem).strip()
    cleaned = cleaned or "Sheet"
    return cleaned[:31]


__all__ = ["iter_tables", "list_input_files", "read_table", "sheet_name"]
