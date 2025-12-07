"""Table detection utilities."""

from __future__ import annotations

import importlib
from dataclasses import dataclass
from importlib import resources
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, Mapping

from openpyxl.worksheet.worksheet import Worksheet

from ade_engine.config.loader import ConfigRuntime
from ade_engine.config.validators import require_keyword_only
from ade_engine.core.errors import ConfigError
from ade_engine.types.contexts import RunContext
from ade_engine.types.origin import TableRegion

RowDetectorFn = Callable[..., Any]

HEADER_SCORE_THRESHOLD = 0.6
DATA_SCORE_THRESHOLD = 0.5


@dataclass(frozen=True)
class RowDetector:
    func: RowDetectorFn
    default_label: str | None
    qualified_name: str


@dataclass(frozen=True)
class RowScore:
    row_index: int
    values: list[Any]
    header_score: float
    data_score: float


def _infer_default_label(detector: Callable[..., object]) -> str | None:
    explicit = (
        getattr(detector, "__row_label__", None)
        or getattr(detector, "row_label", None)
        or getattr(detector, "default_label", None)
    )
    if explicit:
        return str(explicit)

    module = getattr(detector, "__module__", "")
    if module.endswith(".header"):
        return "header"
    if module.endswith(".data"):
        return "data"
    return None


def _discover_row_detectors(package) -> list[RowDetector]:
    """Load and validate row detectors under <package>.row_detectors."""

    try:
        detector_pkg = importlib.import_module(f"{package.__name__}.row_detectors")
    except ModuleNotFoundError:
        return []

    detectors: list[RowDetector] = []
    for entry in sorted(resources.files(detector_pkg).iterdir(), key=lambda e: e.name):
        if entry.name.startswith("_") or entry.suffix != ".py":
            continue

        module = importlib.import_module(f"{detector_pkg.__name__}.{entry.stem}")
        for name in sorted(dir(module)):
            if not name.startswith("detect_"):
                continue
            attr = getattr(module, name)
            if not callable(attr):
                continue
            require_keyword_only(attr, label=f"Row detector '{module.__name__}.{name}'")
            detectors.append(
                RowDetector(
                    func=attr,
                    default_label=_infer_default_label(attr),
                    qualified_name=f"{module.__name__}.{name}",
                )
            )
    return detectors


def _normalize_row_detector_scores(result: Any, *, detector: RowDetector) -> Mapping[str, float]:
    if isinstance(result, Mapping):
        if "scores" in result:
            raise ConfigError(f"{detector.qualified_name} must return a float or dict of label deltas (no 'scores' wrapper')")
        return {str(k): float(v) for k, v in result.items()}

    if result is None:
        return {}
    try:
        delta = float(result)
    except (TypeError, ValueError) as exc:
        raise ConfigError(f"{detector.qualified_name} returned a non-numeric score: {result!r}") from exc

    if detector.default_label is None:
        raise ConfigError(
            f"{detector.qualified_name} returned a bare score but no default label could be inferred; "
            "return a dict like {'header': score} or set a default label on the detector."
        )
    return {detector.default_label: delta}


def _iter_rows(ws: Worksheet) -> Iterator[tuple[int, list[Any]]]:
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        yield idx, [cell for cell in row]


def _score_rows(
    *,
    rows: Iterable[tuple[int, list[Any]]],
    detectors: list[RowDetector],
    run_ctx: RunContext,
    input_file_name: str | None,
    manifest: Any,
    state: dict[str, Any],
) -> list[RowScore]:
    scored: list[RowScore] = []
    for row_index, values in rows:
        label_totals: dict[str, float] = {}
        for detector in detectors:
            result = detector.func(
                run=run_ctx,
                state=state,
                row_index=row_index,
                row_values=values,
                input_file_name=input_file_name,
                file_name=input_file_name,
                manifest=manifest,
            )
            scores = _normalize_row_detector_scores(result, detector=detector)
            for label, delta in scores.items():
                label_totals[label] = label_totals.get(label, 0.0) + delta

        scored.append(
            RowScore(
                row_index=row_index,
                values=list(values),
                header_score=label_totals.get("header", 0.0),
                data_score=label_totals.get("data", 0.0),
            )
        )
    return scored


def _region_width(header: list[Any], data_rows: list[RowScore]) -> int:
    max_len = len(header)
    for row in data_rows:
        max_len = max(max_len, len(row.values))
    return max_len or 1


def _detect_regions(
    *,
    scored_rows: list[RowScore],
    header_threshold: float,
    data_threshold: float,
) -> list[tuple[int, int, int]]:
    """Return list of (min_row, max_row, max_col) tuples."""

    regions: list[tuple[int, int, int]] = []
    position = 0
    while position < len(scored_rows):
        row = scored_rows[position]
        if row.header_score < header_threshold:
            position += 1
            continue

        data_rows: list[RowScore] = []
        lookahead = position + 1
        while lookahead < len(scored_rows) and scored_rows[lookahead].data_score >= data_threshold:
            data_rows.append(scored_rows[lookahead])
            lookahead += 1

        if data_rows:
            max_col = _region_width(row.values, data_rows)
            regions.append((row.row_index, data_rows[-1].row_index, max_col))
        position = lookahead if lookahead > position else position + 1
    return regions


class TableDetector:
    """Detect table regions on a worksheet using config-provided row detectors."""

    def __init__(
        self,
        *,
        header_threshold: float = HEADER_SCORE_THRESHOLD,
        data_threshold: float = DATA_SCORE_THRESHOLD,
        loader: Callable[[Any], list[RowDetector]] = _discover_row_detectors,
    ):
        self.header_threshold = header_threshold
        self.data_threshold = data_threshold
        self._loader = loader
        self._cache: dict[str, list[RowDetector]] = {}

    def _get_detectors(self, runtime: ConfigRuntime) -> list[RowDetector]:
        package_name = runtime.package.__name__
        if package_name in self._cache:
            return self._cache[package_name]
        detectors = self._loader(runtime.package)
        self._cache[package_name] = detectors
        return detectors

    def detect(
        self,
        *,
        source_path: Path,
        worksheet: Worksheet,
        runtime: ConfigRuntime,
        run_ctx: RunContext,
        logger=None,
    ) -> list[TableRegion]:
        """Return ordered TableRegion entries for the given worksheet."""

        detectors = self._get_detectors(runtime)
        state: dict[str, Any] = run_ctx.state

        rows = _iter_rows(worksheet)
        scored_rows = _score_rows(
            rows=rows,
            detectors=detectors,
            run_ctx=run_ctx,
            input_file_name=source_path.name,
            manifest=runtime.manifest,
            state=state,
        )
        regions: list[TableRegion] = []
        for min_row, max_row, max_col in _detect_regions(
            scored_rows=scored_rows,
            header_threshold=self.header_threshold,
            data_threshold=self.data_threshold,
        ):
            regions.append(TableRegion(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col))
        return regions


__all__ = ["TableDetector", "HEADER_SCORE_THRESHOLD", "DATA_SCORE_THRESHOLD", "RowDetector", "RowScore"]
