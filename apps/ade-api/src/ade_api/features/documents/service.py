"""Service layer for document upload and retrieval."""

from __future__ import annotations

import logging
import unicodedata
from collections.abc import Iterator, Mapping, Sequence
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from fastapi import UploadFile
from pydantic import ValidationError
from sqlalchemy import case, func, literal, select
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.sql import Select

from ade_api.common.etag import build_etag_token, format_weak_etag
from ade_api.common.ids import generate_uuid7
from ade_api.common.list_filters import FilterItem, FilterJoinOperator
from ade_api.common.cursor_listing import (
    ResolvedCursorSort,
    cursor_field,
    paginate_query_cursor,
    parse_int,
    parse_str,
    resolve_cursor_sort,
)
from ade_api.common.logging import log_context
from ade_api.common.workbook_preview import (
    WorkbookSheetPreview,
    build_workbook_preview_from_csv,
    build_workbook_preview_from_xlsx,
)
from ade_api.infra.storage import workspace_documents_root
from ade_api.models import (
    Document,
    DocumentComment,
    DocumentCommentMention,
    DocumentEvent,
    DocumentEventType,
    DocumentSource,
    DocumentTag,
    Environment,
    Run,
    RunStatus,
    User,
    WorkspaceMembership,
)
from ade_api.settings import Settings

from .change_feed import DocumentEventsService
from .exceptions import (
    DocumentFileMissingError,
    DocumentNotFoundError,
    DocumentPreviewParseError,
    DocumentPreviewSheetNotFoundError,
    DocumentPreviewUnsupportedError,
    DocumentWorksheetParseError,
    InvalidDocumentExpirationError,
    InvalidDocumentCommentMentionsError,
    InvalidDocumentTagsError,
)
from .filters import apply_document_filters
from .repository import DocumentsRepository
from .schemas import (
    DocumentChangeEntry,
    DocumentChangesPage,
    DocumentCommentOut,
    DocumentCommentPage,
    DocumentFileType,
    DocumentListPage,
    DocumentListRow,
    DocumentOut,
    DocumentResultSummary,
    DocumentRunPhase,
    DocumentRunPhaseReason,
    DocumentRunSummary,
    DocumentSheet,
    DocumentUpdateRequest,
    DocumentUploadRunOptions,
    TagCatalogItem,
    TagCatalogPage,
)
from .storage import DocumentStorage, StoredDocument
from .tags import (
    MAX_TAGS_PER_DOCUMENT,
    TagValidationError,
    normalize_tag_list,
    normalize_tag_query,
)

logger = logging.getLogger(__name__)

UPLOAD_RUN_OPTIONS_KEY = "__ade_run_options"

_FALLBACK_FILENAME = "upload"
_MAX_FILENAME_LENGTH = 255


def _run_with_timeout(func, *, timeout: float, **kwargs):
    """Run a callable with a timeout to avoid hanging on large workbook operations."""
    if timeout <= 0:
        return func(**kwargs)
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(func, **kwargs)
        return future.result(timeout=timeout)


def _map_document_row(row: Mapping[str, Any]) -> Document:
    document = row[Document]
    setattr(document, "_last_run_at", row.get("last_run_at"))
    return document


@dataclass(slots=True)
class StagedUpload:
    document_id: UUID
    stored_uri: str
    stored: StoredDocument


class DocumentsService:
    """Manage document metadata and backing file storage."""

    def __init__(self, *, session: Session, settings: Settings) -> None:
        self._session = session
        self._settings = settings

        documents_dir = settings.documents_dir
        if documents_dir is None:
            raise RuntimeError("Document storage directory is not configured")

        self._repository = DocumentsRepository(session)
        self._events = DocumentEventsService(session=session, settings=settings)

    @staticmethod
    def build_upload_metadata(
        metadata: Mapping[str, Any] | None,
        run_options: DocumentUploadRunOptions | None,
    ) -> dict[str, Any]:
        payload = dict(metadata or {})
        if run_options is not None:
            payload[UPLOAD_RUN_OPTIONS_KEY] = run_options.model_dump(exclude_none=True)
        return payload

    @staticmethod
    def read_upload_run_options(
        metadata: Mapping[str, Any] | None,
    ) -> DocumentUploadRunOptions | None:
        if not metadata:
            return None
        raw = metadata.get(UPLOAD_RUN_OPTIONS_KEY)
        if raw is None:
            return None
        try:
            return DocumentUploadRunOptions.model_validate(raw)
        except ValidationError:
            return None

    def create_document(
        self,
        *,
        workspace_id: UUID,
        upload: UploadFile,
        metadata: Mapping[str, Any] | None = None,
        expires_at: str | None = None,
        actor: User | None = None,
        staged: StagedUpload | None = None,
    ) -> DocumentOut:
        """Persist ``upload`` to storage and return the resulting metadata record."""

        actor_id: UUID | None = actor.id if actor is not None else None
        logger.debug(
            "document.create.start",
            extra=log_context(
                workspace_id=workspace_id,
                user_id=actor_id,
                upload_filename=upload.filename,
                content_type=upload.content_type,
                expires_at=expires_at,
            ),
        )

        metadata_payload = dict(metadata or {})
        now = datetime.now(tz=UTC)
        expiration = self._resolve_expiration(expires_at, now)

        owns_stage = staged is None
        staged_upload = staged or self.stage_upload(workspace_id=workspace_id, upload=upload)
        document_id = staged_upload.document_id
        stored_uri = staged_upload.stored_uri
        stored = staged_upload.stored

        try:
            document = Document(
                id=document_id,
                workspace_id=workspace_id,
                original_filename=self._normalise_filename(upload.filename),
                content_type=self._normalise_content_type(upload.content_type),
                byte_size=stored.byte_size,
                sha256=stored.sha256,
                stored_uri=stored_uri,
                attributes=metadata_payload,
                uploaded_by_user_id=actor_id,
                source=DocumentSource.MANUAL_UPLOAD,
                expires_at=expiration,
            )
            self._session.add(document)
            self._session.flush()
        except Exception:
            if owns_stage:
                self.discard_staged_upload(workspace_id=workspace_id, staged=staged_upload)
            raise

        stmt = self._repository.base_query(workspace_id).where(Document.id == document_id)
        result = self._session.execute(stmt)
        hydrated = result.scalar_one()

        payload = DocumentOut.model_validate(hydrated)
        self._attach_last_runs(workspace_id, [payload], last_run_ids={document_id: hydrated.last_run_id})
        self._apply_derived_fields(payload)
        payload.list_row = self._build_list_row(payload)

        logger.info(
            "document.create.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                user_id=actor_id,
                content_type=document.content_type,
                byte_size=document.byte_size,
                expires_at=document.expires_at.isoformat() if document.expires_at else None,
            ),
        )

        return payload

    def stage_upload(
        self,
        *,
        workspace_id: UUID,
        upload: UploadFile,
        document_id: UUID | None = None,
    ) -> StagedUpload:
        document_id = document_id or generate_uuid7()
        storage = self._storage_for(workspace_id)
        stored_uri = storage.make_stored_uri(str(document_id))
        if upload.file is None:  # pragma: no cover - UploadFile always supplies file
            raise RuntimeError("Upload stream is not available")
        stored = storage.write(
            stored_uri,
            upload.file,
            max_bytes=self._settings.storage_upload_max_bytes,
        )
        return StagedUpload(document_id=document_id, stored_uri=stored_uri, stored=stored)

    def discard_staged_upload(self, *, workspace_id: UUID, staged: StagedUpload) -> None:
        storage = self._storage_for(workspace_id)
        storage.delete(staged.stored_uri)

    def list_documents(
        self,
        *,
        workspace_id: UUID,
        limit: int,
        cursor: str | None,
        resolved_sort: ResolvedCursorSort[Document],
        filters: list[FilterItem],
        join_operator: FilterJoinOperator,
        q: str | None,
        include_total: bool,
        include_facets: bool,
    ) -> DocumentListPage:
        """Return paginated documents with the shared envelope."""

        logger.debug(
            "document.list.start",
            extra=log_context(
                workspace_id=workspace_id,
                limit=limit,
                cursor=cursor,
                order_by=str(resolved_sort.order_by),
            ),
        )

        stmt = self._repository.base_query(workspace_id).where(Document.deleted_at.is_(None))
        stmt = apply_document_filters(
            stmt,
            filters,
            join_operator=join_operator,
            q=q,
        )
        last_run_at_expr = (
            select(func.coalesce(Run.completed_at, Run.started_at, Run.created_at))
            .where(Run.id == Document.last_run_id)
            .scalar_subquery()
            .label("last_run_at")
        )
        stmt = stmt.add_columns(last_run_at_expr)

        # Capture the cursor before listing to avoid skipping changes committed during the query.
        changes_cursor = self._events.current_cursor(workspace_id=workspace_id)
        facets = self._build_document_facets(stmt) if include_facets else None
        page_result = paginate_query_cursor(
            self._session,
            stmt,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
            changes_cursor=str(changes_cursor),
            row_mapper=lambda row: _map_document_row(row),
        )
        raw_items = list(page_result.items)
        items = [DocumentOut.model_validate(item) for item in raw_items]
        last_run_ids = {doc.id: doc.last_run_id for doc in raw_items}
        self._attach_last_runs(workspace_id, items, last_run_ids=last_run_ids)
        for item in items:
            self._apply_derived_fields(item)

        logger.info(
            "document.list.success",
            extra=log_context(
                workspace_id=workspace_id,
                limit=page_result.meta.limit,
                count=len(items),
                has_more=page_result.meta.has_more,
            ),
        )

        rows = [self._build_list_row(item) for item in items]

        return DocumentListPage(items=rows, meta=page_result.meta, facets=facets)

    def list_document_comments(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        limit: int,
        cursor: str | None,
        resolved_sort: ResolvedCursorSort[DocumentComment],
        include_total: bool,
    ) -> DocumentCommentPage:
        self._get_document(workspace_id, document_id)

        stmt = (
            select(DocumentComment)
            .where(DocumentComment.workspace_id == workspace_id)
            .where(DocumentComment.document_id == document_id)
            .options(
                selectinload(DocumentComment.author_user),
                selectinload(DocumentComment.mentions).selectinload(DocumentCommentMention.mentioned_user),
            )
        )

        page_result = paginate_query_cursor(
            self._session,
            stmt,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
            changes_cursor="0",
        )

        items = [DocumentCommentOut.model_validate(item) for item in page_result.items]
        return DocumentCommentPage(items=items, meta=page_result.meta, facets=page_result.facets)

    def create_document_comment(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        body: str,
        mentions: Sequence[UUID] | None,
        actor: User,
    ) -> DocumentCommentOut:
        document = self._get_document(workspace_id, document_id)

        mention_users = self._resolve_comment_mentions(
            workspace_id=workspace_id,
            mentions=mentions,
        )

        comment = DocumentComment(
            workspace_id=workspace_id,
            document_id=document_id,
            author_user_id=actor.id,
            body=body,
        )
        comment.author_user = actor
        comment.mentions = [
            DocumentCommentMention(mentioned_user=user) for user in mention_users
        ]

        document.comment_count = (document.comment_count or 0) + 1
        document.version += 1

        self._session.add(comment)
        self._session.flush()

        return DocumentCommentOut.model_validate(comment)

    def list_document_changes(
        self,
        *,
        workspace_id: UUID,
        cursor_token: str,
        limit: int,
        max_cursor: int | None = None,
        include_rows: bool = False,
    ) -> DocumentChangesPage:
        try:
            cursor = int(cursor_token)
        except (TypeError, ValueError) as exc:
            raise ValueError("cursor must be an integer string") from exc

        page = self._events.list_changes(
            workspace_id=workspace_id,
            cursor=cursor,
            limit=limit,
            max_cursor=max_cursor,
        )
        entries = self.build_change_entries(
            workspace_id=workspace_id,
            events=page.items,
            include_rows=include_rows,
        )
        return DocumentChangesPage(items=entries, next_cursor=str(page.next_cursor))

    def build_change_entries(
        self,
        *,
        workspace_id: UUID,
        events: Sequence[DocumentEvent],
        include_rows: bool,
    ) -> list[DocumentChangeEntry]:
        rows_by_id: dict[UUID, DocumentListRow] = {}
        if include_rows:
            changed_ids = {
                change.document_id
                for change in events
                if change.event_type == DocumentEventType.CHANGED
            }
            if changed_ids:
                stmt = (
                    self._repository.base_query(workspace_id)
                    .where(Document.id.in_(changed_ids))
                    .where(Document.deleted_at.is_(None))
                )
                result = self._session.execute(stmt)
                documents = list(result.scalars())
                payloads = [DocumentOut.model_validate(item) for item in documents]
                self._attach_last_runs(workspace_id, payloads, last_run_ids={doc.id: doc.last_run_id for doc in documents})
                for payload in payloads:
                    self._apply_derived_fields(payload)
                    rows_by_id[payload.id] = self._build_list_row(payload)

        entries: list[DocumentChangeEntry] = []
        for change in events:
            row = None
            if include_rows and change.event_type == DocumentEventType.CHANGED:
                row = rows_by_id.get(change.document_id)
            entries.append(
                DocumentChangeEntry(
                    cursor=str(change.cursor),
                    type=change.event_type.value,
                    document_id=str(change.document_id),
                    occurred_at=change.occurred_at,
                    document_version=change.document_version,
                    row=row,
                )
            )
        return entries

    def get_document(self, *, workspace_id: UUID, document_id: UUID) -> DocumentOut:
        """Return document metadata for ``document_id``."""

        logger.debug(
            "document.get.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )
        document = self._get_document(workspace_id, document_id)
        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload], last_run_ids={document.id: document.last_run_id})
        self._apply_derived_fields(payload)

        logger.info(
            "document.get.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                byte_size=document.byte_size,
            ),
        )
        return payload

    def get_document_list_row(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> DocumentListRow:
        """Return a table-ready row projection for ``document_id``."""

        logger.debug(
            "document.list_row.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )
        document = self._get_document(workspace_id, document_id)
        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload], last_run_ids={document.id: document.last_run_id})
        self._apply_derived_fields(payload)
        row = self._build_list_row(payload)

        logger.info(
            "document.list_row.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                byte_size=document.byte_size,
            ),
        )
        return row

    def update_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        payload: DocumentUpdateRequest,
    ) -> DocumentOut:
        document = self._get_document(workspace_id, document_id)
        changed = False

        if "assignee_user_id" in payload.model_fields_set:
            if document.assignee_user_id != payload.assignee_user_id:
                document.assignee_user_id = payload.assignee_user_id
                changed = True
        if "metadata" in payload.model_fields_set and payload.metadata is not None:
            next_metadata = dict(payload.metadata)
            if document.attributes != next_metadata:
                document.attributes = next_metadata
                changed = True

        if changed:
            document.version += 1

        if changed:
            self._session.flush()
            self._session.refresh(document, attribute_names=["assignee_user"])

        updated = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [updated], last_run_ids={document.id: document.last_run_id})
        self._apply_derived_fields(updated)
        return updated

    def list_document_sheets(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> list[DocumentSheet]:
        """Return worksheet descriptors for ``document_id``."""

        logger.debug(
            "document.sheets.list.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )

        document = self._get_document(workspace_id, document_id)
        storage = self._storage_for(workspace_id)
        path = storage.path_for(document.stored_uri)

        if not path.exists():
            raise DocumentFileMissingError(
                document_id=document_id,
                stored_uri=document.stored_uri,
            )

        suffix = Path(document.original_filename).suffix.lower()
        if suffix == ".xlsx":
            try:
                sheets = _run_with_timeout(
                    self._inspect_workbook,
                    timeout=self._settings.preview_timeout_seconds,
                    path=path,
                )
                logger.info(
                    "document.sheets.list.success",
                    extra=log_context(
                        workspace_id=workspace_id,
                        document_id=document_id,
                        sheet_count=len(sheets),
                        kind="workbook",
                    ),
                )
                return sheets
            except FuturesTimeout as exc:
                raise DocumentWorksheetParseError(
                    document_id=document_id,
                    stored_uri=document.stored_uri,
                    reason="timeout",
                ) from exc
            except Exception as exc:  # pragma: no cover - defensive fallback
                raise DocumentWorksheetParseError(
                    document_id=document_id,
                    stored_uri=document.stored_uri,
                    reason=type(exc).__name__,
                ) from exc

        name = self._default_sheet_name(document.original_filename)
        sheets = [DocumentSheet(name=name, index=0, kind="file", is_active=True)]
        logger.info(
            "document.sheets.list.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                sheet_count=len(sheets),
                kind="fallback",
            ),
        )
        return sheets

    def get_document_preview(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        max_rows: int,
        max_columns: int,
        trim_empty_columns: bool = False,
        trim_empty_rows: bool = False,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
    ) -> WorkbookSheetPreview:
        """Return a table-ready preview for a document workbook."""

        effective_sheet_index = sheet_index
        if sheet_name is None and sheet_index is None:
            effective_sheet_index = 0

        logger.debug(
            "document.preview.start",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                sheet_name=sheet_name,
                sheet_index=effective_sheet_index,
            ),
        )

        document = self._get_document(workspace_id, document_id)
        storage = self._storage_for(workspace_id)
        path = storage.path_for(document.stored_uri)

        if not path.exists():
            raise DocumentFileMissingError(
                document_id=document_id,
                stored_uri=document.stored_uri,
            )

        suffix = Path(document.original_filename).suffix.lower()
        try:
            if suffix == ".xlsx":
                preview = _run_with_timeout(
                    build_workbook_preview_from_xlsx,
                    timeout=self._settings.preview_timeout_seconds,
                    path=path,
                    max_rows=max_rows,
                    max_columns=max_columns,
                    trim_empty_columns=trim_empty_columns,
                    trim_empty_rows=trim_empty_rows,
                    sheet_name=sheet_name,
                    sheet_index=effective_sheet_index,
                )
            elif suffix == ".csv":
                preview = _run_with_timeout(
                    build_workbook_preview_from_csv,
                    timeout=self._settings.preview_timeout_seconds,
                    path=path,
                    max_rows=max_rows,
                    max_columns=max_columns,
                    trim_empty_columns=trim_empty_columns,
                    trim_empty_rows=trim_empty_rows,
                    sheet_name=sheet_name,
                    sheet_index=effective_sheet_index,
                )
            else:
                raise DocumentPreviewUnsupportedError(
                    document_id=document_id,
                    file_type=suffix or "unknown",
                )
        except (KeyError, IndexError) as exc:
            requested = sheet_name if sheet_name is not None else str(effective_sheet_index)
            raise DocumentPreviewSheetNotFoundError(
                document_id=document_id,
                sheet=requested,
            ) from exc
        except FuturesTimeout as exc:
            raise DocumentPreviewParseError(
                document_id=document_id,
                stored_uri=document.stored_uri,
                reason="timeout",
            ) from exc
        except DocumentPreviewUnsupportedError:
            raise
        except Exception as exc:  # pragma: no cover - defensive fallback
            raise DocumentPreviewParseError(
                document_id=document_id,
                stored_uri=document.stored_uri,
                reason=type(exc).__name__,
            ) from exc

        logger.info(
            "document.preview.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                sheet_name=preview.name,
                sheet_index=preview.index,
            ),
        )
        return preview

    def stream_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
    ) -> tuple[DocumentOut, Iterator[bytes]]:
        """Return a document record and iterator for its bytes."""

        logger.debug(
            "document.stream.start",
            extra=log_context(workspace_id=workspace_id, document_id=document_id),
        )

        document = self._get_document(workspace_id, document_id)
        storage = self._storage_for(workspace_id)
        path = storage.path_for(document.stored_uri)
        if not path.exists():
            logger.warning(
                "document.stream.missing_file",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                    stored_uri=document.stored_uri,
                ),
            )
            raise DocumentFileMissingError(
                document_id=document_id,
                stored_uri=document.stored_uri,
            )

        stream = storage.stream(document.stored_uri)

        def _guarded() -> Iterator[bytes]:
            try:
                for chunk in stream:
                    yield chunk
            except FileNotFoundError as exc:
                logger.warning(
                    "document.stream.file_lost_during_stream",
                    extra=log_context(
                        workspace_id=workspace_id,
                        document_id=document_id,
                        stored_uri=document.stored_uri,
                    ),
                )
                raise DocumentFileMissingError(
                    document_id=document_id,
                    stored_uri=document.stored_uri,
                ) from exc

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload], last_run_ids={document.id: document.last_run_id})
        self._apply_derived_fields(payload)

        logger.info(
            "document.stream.ready",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                byte_size=document.byte_size,
                content_type=document.content_type,
            ),
        )
        return payload, _guarded()

    def delete_document(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        actor: User | None = None,
    ) -> None:
        """Soft delete ``document_id`` and remove the stored file."""

        actor_id: UUID | None = actor.id if actor is not None else None
        logger.debug(
            "document.delete.start",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                user_id=actor_id,
            ),
        )

        document = self._get_document(workspace_id, document_id)
        now = datetime.now(tz=UTC)
        document.deleted_at = now
        document.version += 1
        if actor is not None:
            document.deleted_by_user_id = actor_id
        self._session.flush()

        storage = self._storage_for(workspace_id)
        storage.delete(document.stored_uri)

        logger.info(
            "document.delete.success",
            extra=log_context(
                workspace_id=workspace_id,
                document_id=document_id,
                user_id=actor_id,
                stored_uri=document.stored_uri,
            ),
        )

    def delete_documents_batch(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
        actor: User | None = None,
    ) -> list[UUID]:
        """Soft delete multiple documents and remove stored files."""

        ordered_ids = list(dict.fromkeys(document_ids))
        if not ordered_ids:
            return []

        actor_id: UUID | None = actor.id if actor is not None else None
        documents = self._require_documents(
            workspace_id=workspace_id,
            document_ids=ordered_ids,
        )
        document_by_id = {doc.id: doc for doc in documents}

        now = datetime.now(tz=UTC)
        for document in documents:
            document.deleted_at = now
            document.version += 1
            if actor is not None:
                document.deleted_by_user_id = actor_id

        self._session.flush()

        storage = self._storage_for(workspace_id)
        for document in documents:
            storage.delete(document.stored_uri)

        for document_id in ordered_ids:
            document = document_by_id[document_id]

        return ordered_ids

    def replace_document_tags(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        tags: list[str],
    ) -> DocumentOut:
        """Replace tags on a document in a single transaction."""

        try:
            normalized = normalize_tag_list(tags, max_tags=MAX_TAGS_PER_DOCUMENT)
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        document = self._get_document(workspace_id, document_id)
        document.tags = [DocumentTag(document_id=document.id, tag=tag) for tag in normalized]
        document.version += 1
        self._session.flush()

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload], last_run_ids={document.id: document.last_run_id})
        self._apply_derived_fields(payload)
        return payload

    def patch_document_tags(
        self,
        *,
        workspace_id: UUID,
        document_id: UUID,
        add: list[str] | None = None,
        remove: list[str] | None = None,
    ) -> DocumentOut:
        """Add or remove tags on a document."""

        if not add and not remove:
            raise InvalidDocumentTagsError("add or remove is required")

        try:
            normalized_add = normalize_tag_list(add or [])
            normalized_remove = normalize_tag_list(remove or [])
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        document = self._get_document(workspace_id, document_id)
        existing = {tag.tag for tag in document.tags}
        next_tags = (existing | set(normalized_add)) - set(normalized_remove)
        if len(next_tags) > MAX_TAGS_PER_DOCUMENT:
            raise InvalidDocumentTagsError(f"Too many tags; max {MAX_TAGS_PER_DOCUMENT}.")

        to_remove = existing - next_tags
        if to_remove:
            document.tags = [tag for tag in document.tags if tag.tag not in to_remove]
        to_add = next_tags - existing
        for tag in to_add:
            document.tags.append(DocumentTag(document_id=document.id, tag=tag))

        document.version += 1
        self._session.flush()

        payload = DocumentOut.model_validate(document)
        self._attach_last_runs(workspace_id, [payload], last_run_ids={document.id: document.last_run_id})
        self._apply_derived_fields(payload)
        return payload

    def patch_document_tags_batch(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
        add: list[str] | None = None,
        remove: list[str] | None = None,
    ) -> list[DocumentOut]:
        """Add or remove tags on multiple documents."""

        if not add and not remove:
            raise InvalidDocumentTagsError("add or remove is required")

        try:
            normalized_add = normalize_tag_list(add or [])
            normalized_remove = normalize_tag_list(remove or [])
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        ordered_ids = list(dict.fromkeys(document_ids))
        documents = self._require_documents(
            workspace_id=workspace_id,
            document_ids=ordered_ids,
        )
        document_by_id = {doc.id: doc for doc in documents}

        for document in documents:
            existing = {tag.tag for tag in document.tags}
            next_tags = (existing | set(normalized_add)) - set(normalized_remove)
            if len(next_tags) > MAX_TAGS_PER_DOCUMENT:
                raise InvalidDocumentTagsError(f"Too many tags; max {MAX_TAGS_PER_DOCUMENT}.")

            to_remove = existing - next_tags
            if to_remove:
                document.tags = [tag for tag in document.tags if tag.tag not in to_remove]
            to_add = next_tags - existing
            for tag in to_add:
                document.tags.append(DocumentTag(document_id=document.id, tag=tag))

            document.version += 1

        self._session.flush()

        payloads = [DocumentOut.model_validate(document_by_id[doc_id]) for doc_id in ordered_ids]
        last_run_ids = {doc.id: doc.last_run_id for doc in document_by_id.values()}
        self._attach_last_runs(workspace_id, payloads, last_run_ids=last_run_ids)
        for payload in payloads:
            self._apply_derived_fields(payload)
        return payloads

    def list_tag_catalog(
        self,
        *,
        workspace_id: UUID,
        limit: int,
        cursor: str | None,
        include_total: bool,
        sort: list[str],
        q: str | None = None,
    ) -> TagCatalogPage:
        """Return tag catalog entries with counts."""

        try:
            normalized_q = normalize_tag_query(q)
        except TagValidationError as exc:
            raise InvalidDocumentTagsError(str(exc)) from exc

        count_expr = func.count(DocumentTag.document_id).label("document_count")
        stmt = (
            select(
                DocumentTag.tag.label("tag"),
                count_expr,
            )
            .join(Document, DocumentTag.document_id == Document.id)
            .where(
                Document.workspace_id == workspace_id,
                Document.deleted_at.is_(None),
            )
            .group_by(DocumentTag.tag)
        )

        if normalized_q:
            pattern = f"%{normalized_q}%"
            stmt = stmt.where(DocumentTag.tag.like(pattern))

        sort_fields = {
            "name": (DocumentTag.tag.asc(), DocumentTag.tag.desc()),
            "count": (count_expr.asc(), count_expr.desc()),
        }

        cursor_fields = {
            "id": cursor_field(lambda item: item.tag, parse_str),
            "name": cursor_field(lambda item: item.tag, parse_str),
            "count": cursor_field(lambda item: item.document_count, parse_int),
        }
        resolved_sort = resolve_cursor_sort(
            sort,
            allowed=sort_fields,
            cursor_fields=cursor_fields,
            default=["name"],
            id_field=(DocumentTag.tag.asc(), DocumentTag.tag.desc()),
        )
        page_result = paginate_query_cursor(
            self._session,
            stmt,
            resolved_sort=resolved_sort,
            limit=limit,
            cursor=cursor,
            include_total=include_total,
            changes_cursor="0",
            row_mapper=lambda row: TagCatalogItem(
                tag=row["tag"],
                document_count=int(row["document_count"] or 0),
            ),
        )

        return TagCatalogPage(items=page_result.items, meta=page_result.meta, facets=page_result.facets)

    def _get_document(self, workspace_id: UUID, document_id: UUID) -> Document:
        document = self._repository.get_document(
            workspace_id=workspace_id,
            document_id=document_id,
        )
        if document is None:
            logger.warning(
                "document.get.not_found",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=document_id,
                ),
            )
            raise DocumentNotFoundError(document_id)
        return document

    def _resolve_comment_mentions(
        self,
        *,
        workspace_id: UUID,
        mentions: Sequence[UUID] | None,
    ) -> list[User]:
        if not mentions:
            return []
        unique_ids = list(dict.fromkeys(mentions))
        stmt = (
            select(User)
            .join(
                WorkspaceMembership,
                WorkspaceMembership.user_id == User.id,
            )
            .where(WorkspaceMembership.workspace_id == workspace_id)
            .where(User.id.in_(unique_ids))
        )
        users = list(self._session.execute(stmt).scalars())
        found = {user.id for user in users}
        missing = [str(user_id) for user_id in unique_ids if user_id not in found]
        if missing:
            raise InvalidDocumentCommentMentionsError(
                f"Unknown or non-member mentions: {', '.join(missing)}"
            )
        user_by_id = {user.id: user for user in users}
        return [user_by_id[user_id] for user_id in unique_ids]

    def _require_documents(
        self,
        *,
        workspace_id: UUID,
        document_ids: Sequence[UUID],
    ) -> list[Document]:
        if not document_ids:
            return []

        stmt = (
            self._repository.base_query(workspace_id)
            .where(Document.deleted_at.is_(None))
            .where(Document.id.in_(document_ids))
        )
        result = self._session.execute(stmt)
        documents = list(result.scalars())
        found = {doc.id for doc in documents}
        missing = [document_id for document_id in document_ids if document_id not in found]
        if missing:
            logger.warning(
                "document.require_documents.not_found",
                extra=log_context(
                    workspace_id=workspace_id,
                    document_id=str(missing[0]),
                    missing_count=len(missing),
                ),
            )
            raise DocumentNotFoundError(missing[0])

        return documents

    def _attach_last_runs(
        self,
        workspace_id: UUID,
        documents: Sequence[DocumentOut],
        last_run_ids: dict[UUID, UUID | None] | None = None,
    ) -> None:
        """Populate last run summaries on each document."""

        if not documents:
            return

        if last_run_ids is None:
            doc_ids = [doc.id for doc in documents]
            result = self._session.execute(
                select(Document.id, Document.last_run_id).where(Document.id.in_(doc_ids))
            )
            last_run_ids = {}
            for doc_id, run_id in result.all():
                if not isinstance(doc_id, UUID):
                    doc_id = UUID(str(doc_id))
                last_run_ids[doc_id] = run_id

        last_run_ids = {
            doc_id: run_id
            for doc_id, run_id in (last_run_ids or {}).items()
            if run_id is not None
        }
        last_runs = self._last_runs_by_id(
            workspace_id=workspace_id,
            last_run_ids=last_run_ids,
        )
        last_successful_runs = self._last_successful_runs(
            workspace_id=workspace_id,
            documents=documents,
        )
        for document in documents:
            document.last_run = last_runs.get(document.id)
            document.last_successful_run = last_successful_runs.get(document.id)

    def _apply_derived_fields(self, document: DocumentOut) -> None:
        updated_at = document.updated_at
        latest_at = self._last_run_at(document.last_run)
        document.activity_at = latest_at if latest_at and latest_at > updated_at else updated_at
        document.latest_result = self._derive_latest_result(document)
        document.etag = format_weak_etag(build_etag_token(document.id, document.version))

    def _build_list_row(self, document: DocumentOut) -> DocumentListRow:
        activity_at = document.activity_at or document.updated_at
        return DocumentListRow(
            id=document.id,
            workspace_id=document.workspace_id,
            name=document.name,
            file_type=self._derive_file_type(document.name),
            uploader=document.uploader,
            assignee=document.assignee,
            tags=document.tags,
            byte_size=document.byte_size,
            comment_count=document.comment_count,
            created_at=document.created_at,
            updated_at=document.updated_at,
            activity_at=activity_at,
            version=document.version,
            etag=document.etag,
            last_run=document.last_run,
            last_successful_run=document.last_successful_run,
            latest_result=document.latest_result,
        )

    def _build_document_facets(self, stmt: Select) -> dict[str, Any]:
        filtered_ids = stmt.order_by(None).with_only_columns(Document.id).distinct().subquery()

        def coerce_value(value: Any) -> Any:
            if isinstance(value, Enum):
                return value.value
            if isinstance(value, UUID):
                return str(value)
            return value

        def build_buckets(expr) -> list[dict[str, Any]]:
            rows = self._session.execute(
                select(
                    expr.label("value"),
                    func.count().label("count"),
                )
                .select_from(Document)
                .join(filtered_ids, filtered_ids.c.id == Document.id)
                .group_by(expr)
            ).all()
            buckets = [
                {"value": coerce_value(value), "count": int(count or 0)}
                for value, count in rows
            ]
            buckets.sort(key=lambda bucket: str(bucket["value"]))
            return buckets

        lower_name = func.lower(Document.original_filename)
        file_type_expr = case(
            (lower_name.like("%.xlsx"), DocumentFileType.XLSX.value),
            (lower_name.like("%.xls"), DocumentFileType.XLS.value),
            (lower_name.like("%.csv"), DocumentFileType.CSV.value),
            (lower_name.like("%.pdf"), DocumentFileType.PDF.value),
            else_=DocumentFileType.UNKNOWN.value,
        )

        phase_expr = case(
            (Run.id.is_(None), None),
            (Run.status != RunStatus.QUEUED, Run.status),
            (Environment.status == "ready", Run.status),
            else_=literal("building"),
        )
        phase_rows = self._session.execute(
            select(
                phase_expr.label("value"),
                func.count().label("count"),
            )
            .select_from(Document)
            .join(filtered_ids, filtered_ids.c.id == Document.id)
            .outerjoin(Run, Run.id == Document.last_run_id)
            .outerjoin(
                Environment,
                (Environment.workspace_id == Run.workspace_id)
                & (Environment.configuration_id == Run.configuration_id)
                & (Environment.engine_spec == Run.engine_spec)
                & (Environment.deps_digest == Run.deps_digest),
            )
            .group_by(phase_expr)
        ).all()
        phase_buckets = [
            {"value": coerce_value(value), "count": int(count or 0)}
            for value, count in phase_rows
        ]
        phase_buckets.sort(key=lambda bucket: str(bucket["value"]))

        return {
            "lastRunPhase": {"buckets": phase_buckets},
            "fileType": {"buckets": build_buckets(file_type_expr)},
        }

    @staticmethod
    def _derive_file_type(name: str) -> DocumentFileType:
        suffix = Path(name).suffix.lower().lstrip(".")
        if suffix == "xlsx":
            return DocumentFileType.XLSX
        if suffix == "xls":
            return DocumentFileType.XLS
        if suffix == "csv":
            return DocumentFileType.CSV
        if suffix == "pdf":
            return DocumentFileType.PDF
        return DocumentFileType.UNKNOWN

    @staticmethod
    def _coerce_int(value: Any) -> int | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        return None

    @classmethod
    def _derive_latest_result(cls, document: DocumentOut) -> DocumentResultSummary:
        metadata = document.metadata or {}
        candidate = (
            metadata.get("mapping")
            or metadata.get("mapping_health")
            or metadata.get("mapping_quality")
        )
        if isinstance(candidate, Mapping):
            attention = cls._coerce_int(candidate.get("issues"))
            if attention is None:
                attention = cls._coerce_int(candidate.get("attention")) or 0
            unmapped = cls._coerce_int(candidate.get("unmapped")) or 0
            pending = candidate.get("status") == "pending"
            return DocumentResultSummary(
                attention=attention,
                unmapped=unmapped,
                pending=True if pending else None,
            )

        attention = cls._coerce_int(metadata.get("mapping_issues"))
        unmapped = cls._coerce_int(metadata.get("unmapped_columns"))
        if attention is not None or unmapped is not None:
            return DocumentResultSummary(
                attention=attention or 0,
                unmapped=unmapped or 0,
            )

        if document.last_run and document.last_run.status in {RunStatus.QUEUED, RunStatus.RUNNING}:
            return DocumentResultSummary(attention=0, unmapped=0, pending=True)

        return DocumentResultSummary(attention=0, unmapped=0)

    @staticmethod
    def _last_run_at(run: DocumentRunSummary | None) -> datetime | None:
        if run is None:
            return None
        return run.completed_at or run.started_at or run.created_at

    def _last_runs_by_id(
        self,
        *,
        workspace_id: UUID,
        last_run_ids: dict[UUID, UUID],
    ) -> dict[UUID, DocumentRunSummary]:
        if not last_run_ids:
            return {}

        run_ids = list({run_id for run_id in last_run_ids.values() if run_id is not None})
        if not run_ids:
            return {}

        stmt = (
            select(
                Run.id.label("run_id"),
                Run.input_document_id.label("document_id"),
                Run.status.label("status"),
                Run.error_message.label("error_message"),
                Run.completed_at.label("completed_at"),
                Run.started_at.label("started_at"),
                Run.created_at.label("created_at"),
                Environment.status.label("env_status"),
            )
            .outerjoin(
                Environment,
                (Environment.workspace_id == Run.workspace_id)
                & (Environment.configuration_id == Run.configuration_id)
                & (Environment.engine_spec == Run.engine_spec)
                & (Environment.deps_digest == Run.deps_digest),
            )
            .where(
                Run.workspace_id == workspace_id,
                Run.id.in_(run_ids),
                Run.input_document_id.is_not(None),
            )
        )
        result = self._session.execute(stmt)

        runs_by_doc: dict[UUID, DocumentRunSummary] = {}
        for row in result.mappings():
            document_id = row["document_id"]
            if not isinstance(document_id, UUID):
                document_id = UUID(str(document_id))
            status = row["status"]
            if not isinstance(status, RunStatus):
                status = RunStatus(str(status))
            env_status = row.get("env_status")
            phase, phase_reason = self._derive_run_phase_details(
                status=status,
                env_status=env_status,
            )
            runs_by_doc[document_id] = DocumentRunSummary(
                id=row["run_id"],
                status=status,
                phase=phase,
                phase_reason=phase_reason,
                created_at=self._ensure_utc(row.get("created_at")),
                started_at=self._ensure_utc(row.get("started_at")),
                completed_at=self._ensure_utc(row.get("completed_at")),
                error_summary=self._last_run_message(error_message=row.get("error_message")),
            )

        return runs_by_doc

    def _last_successful_runs(
        self,
        *,
        workspace_id: UUID,
        documents: Sequence[DocumentOut],
    ) -> dict[UUID, DocumentRunSummary]:
        doc_ids = [doc.id for doc in documents]
        if not doc_ids:
            return {}

        timestamp = func.coalesce(Run.completed_at, Run.started_at, Run.created_at)
        ranked_runs = (
            select(
                Run.input_document_id.label("document_id"),
                Run.id.label("run_id"),
                Run.status.label("status"),
                Run.error_message.label("error_message"),
                Run.completed_at.label("completed_at"),
                Run.started_at.label("started_at"),
                Run.created_at.label("created_at"),
                func.row_number()
                .over(
                    partition_by=Run.input_document_id,
                    order_by=timestamp.desc(),
                )
                .label("rank"),
            )
            .where(
                Run.workspace_id == workspace_id,
                Run.input_document_id.is_not(None),
                Run.input_document_id.in_(doc_ids),
                Run.status == RunStatus.SUCCEEDED,
            )
            .subquery()
        )

        stmt = select(ranked_runs).where(ranked_runs.c.rank == 1)
        result = self._session.execute(stmt)

        latest: dict[UUID, DocumentRunSummary] = {}
        for row in result.mappings():
            document_id = row["document_id"]
            if not isinstance(document_id, UUID):
                document_id = UUID(str(document_id))
            status = row["status"]
            if not isinstance(status, RunStatus):
                status = RunStatus(str(status))
            latest[document_id] = DocumentRunSummary(
                id=row["run_id"],
                status=status,
                phase=DocumentRunPhase.SUCCEEDED,
                created_at=self._ensure_utc(row.get("created_at")),
                started_at=self._ensure_utc(row.get("started_at")),
                completed_at=self._ensure_utc(row.get("completed_at")),
                error_summary=self._last_run_message(error_message=row.get("error_message")),
            )

        return latest

    @staticmethod
    def _last_run_message(*, error_message: str | None) -> str | None:
        if error_message:
            return error_message
        return None

    @staticmethod
    def _derive_run_phase_details(
        *,
        status: RunStatus,
        env_status: Any | None,
    ) -> tuple[DocumentRunPhase, DocumentRunPhaseReason | None]:
        if status != RunStatus.QUEUED:
            return DocumentRunPhase(status.value), None
        if env_status is None:
            return DocumentRunPhase.BUILDING, DocumentRunPhaseReason.ENVIRONMENT_MISSING
        if isinstance(env_status, Enum):
            env_value = env_status.value
        else:
            env_value = str(env_status)
        if env_value == "ready":
            return DocumentRunPhase.QUEUED, None
        reason = {
            "queued": DocumentRunPhaseReason.ENVIRONMENT_QUEUED,
            "building": DocumentRunPhaseReason.ENVIRONMENT_BUILDING,
            "failed": DocumentRunPhaseReason.ENVIRONMENT_FAILED,
        }.get(env_value)
        return DocumentRunPhase.BUILDING, reason

    @staticmethod
    def _ensure_utc(dt: datetime | None) -> datetime | None:
        if dt is None:
            return None
        if dt.tzinfo is None:
            return dt.replace(tzinfo=UTC)
        return dt

    def _resolve_expiration(self, override: str | None, now: datetime) -> datetime:
        if override is None:
            expires = now + self._settings.storage_document_retention_period
            return expires

        candidate = override.strip()
        if not candidate:
            raise InvalidDocumentExpirationError("expires_at must not be blank")

        if candidate.endswith(("z", "Z")):
            candidate = f"{candidate[:-1]}+00:00"

        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError as exc:
            raise InvalidDocumentExpirationError(
                "expires_at must be a valid ISO 8601 timestamp"
            ) from exc

        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        else:
            parsed = parsed.astimezone(UTC)

        if parsed <= now:
            raise InvalidDocumentExpirationError("expires_at must be in the future")

        return parsed

    def _normalise_filename(self, name: str | None) -> str:
        """Return a safe, display-friendly filename for stored documents."""

        if name is None:
            return _FALLBACK_FILENAME

        candidate = name.strip()
        if not candidate:
            return _FALLBACK_FILENAME

        # Strip control characters (including newlines) to avoid header injection and
        # other control sequence issues when the filename is rendered in responses.
        filtered = "".join(ch for ch in candidate if unicodedata.category(ch)[0] != "C").strip()

        if not filtered:
            return _FALLBACK_FILENAME

        if len(filtered) > _MAX_FILENAME_LENGTH:
            filtered = filtered[:_MAX_FILENAME_LENGTH].rstrip()

        return filtered or _FALLBACK_FILENAME

    def _normalise_content_type(self, content_type: str | None) -> str | None:
        if content_type is None:
            return None
        candidate = content_type.strip()
        return candidate or None

    def _storage_for(self, workspace_id: UUID) -> DocumentStorage:
        base = workspace_documents_root(self._settings, workspace_id)
        return DocumentStorage(base)

    @staticmethod
    def _inspect_workbook(path: Path) -> list[DocumentSheet]:
        with path.open("rb") as raw:
            workbook = openpyxl.load_workbook(
                raw,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            try:
                sheetnames = workbook.sheetnames
                active = workbook.active.title if sheetnames else None
                return [
                    DocumentSheet(
                        name=title,
                        index=index,
                        kind="worksheet",
                        is_active=title == active,
                    )
                    for index, title in enumerate(sheetnames)
                ]
            finally:
                workbook.close()

    @staticmethod
    def _default_sheet_name(name: str | None) -> str:
        stem = Path(name or "Sheet").stem.strip() or "Sheet"
        return stem[:64]


__all__ = ["DocumentsService"]
