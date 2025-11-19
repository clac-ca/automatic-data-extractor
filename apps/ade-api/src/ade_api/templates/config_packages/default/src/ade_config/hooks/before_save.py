"""before_save hook that styles the workbook."""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def run(*, workbook, **_: Any):
    """Convert the active sheet to an Excel table for readability."""

    sheet = workbook.active
    sheet.freeze_panes = "A2"
    right = get_column_letter(sheet.max_column)
    table_ref = f"A1:{right}{sheet.max_row}"
    excel_table = Table(displayName="NormalizedData", ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(excel_table)
    return workbook
