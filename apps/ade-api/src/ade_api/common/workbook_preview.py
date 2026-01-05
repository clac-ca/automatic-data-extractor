from __future__ import annotations

import csv
from collections.abc import Sequence
from datetime import date, datetime
from pathlib import Path

import openpyxl
from pydantic import Field

from ade_api.common.schema import BaseSchema

DEFAULT_PREVIEW_ROWS = 200
DEFAULT_PREVIEW_COLUMNS = 50
MAX_PREVIEW_ROWS = 500
MAX_PREVIEW_COLUMNS = 200


class WorkbookSheetPreview(BaseSchema):
    """Table-ready preview for a single workbook sheet."""

    name: str
    index: int = Field(ge=0)
    headers: list[str]
    rows: list[list[str]]
    total_rows: int = Field(alias="totalRows")
    total_columns: int = Field(alias="totalColumns")
    truncated_rows: bool = Field(alias="truncatedRows")
    truncated_columns: bool = Field(alias="truncatedColumns")


def build_workbook_preview_from_xlsx(
    path: Path,
    *,
    max_rows: int = DEFAULT_PREVIEW_ROWS,
    max_columns: int = DEFAULT_PREVIEW_COLUMNS,
    trim_empty_columns: bool = False,
    trim_empty_rows: bool = False,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> WorkbookSheetPreview:
    with path.open("rb") as handle:
        workbook = openpyxl.load_workbook(
            handle,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            index, sheet = _select_xlsx_sheet(workbook, sheet_name, sheet_index)
            return _preview_xlsx_sheet(
                sheet,
                index=index,
                max_rows=max_rows,
                max_columns=max_columns,
                trim_empty_columns=trim_empty_columns,
                trim_empty_rows=trim_empty_rows,
            )
        finally:
            workbook.close()


def build_workbook_preview_from_csv(
    path: Path,
    *,
    max_rows: int = DEFAULT_PREVIEW_ROWS,
    max_columns: int = DEFAULT_PREVIEW_COLUMNS,
    trim_empty_columns: bool = False,
    trim_empty_rows: bool = False,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> WorkbookSheetPreview:
    name = path.stem or "Sheet1"
    if sheet_name and sheet_name != name:
        raise KeyError(f"Sheet {sheet_name!r} not found")
    if sheet_index not in (None, 0):
        raise IndexError("sheet_index out of range")

    rows: list[list[str]] = []
    total_rows = 0
    total_columns = 0

    with path.open("r", newline="", encoding="utf-8", errors="replace") as handle:
        reader = csv.reader(handle)
        for row in reader:
            total_rows += 1
            total_columns = max(total_columns, len(row))
            if len(rows) < max_rows:
                rows.append([_normalize_cell(cell) for cell in row])

    return _preview_sheet_from_rows(
        name=name,
        index=0,
        rows=rows,
        total_rows=total_rows,
        total_columns=total_columns,
        max_rows=max_rows,
        max_columns=max_columns,
        trim_empty_columns=trim_empty_columns,
        trim_empty_rows=trim_empty_rows,
    )


def _select_xlsx_sheet(workbook, sheet_name: str | None, sheet_index: int | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found")
        index = workbook.sheetnames.index(sheet_name)
        return index, workbook[sheet_name]
    effective_index = sheet_index if sheet_index is not None else 0
    if effective_index < 0 or effective_index >= len(workbook.sheetnames):
        raise IndexError("sheet_index out of range")
    return effective_index, workbook[workbook.sheetnames[effective_index]]


def _preview_xlsx_sheet(
    sheet,
    *,
    index: int,
    max_rows: int,
    max_columns: int,
    trim_empty_columns: bool,
    trim_empty_rows: bool,
) -> WorkbookSheetPreview:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(
        min_row=1,
        max_row=max_rows,
        max_col=max_columns,
        values_only=True,
    ):
        rows.append([_normalize_cell(cell) for cell in row])

    total_rows = sheet.max_row or 0
    total_columns = sheet.max_column or 0
    return _preview_sheet_from_rows(
        name=sheet.title,
        index=index,
        rows=rows,
        total_rows=total_rows,
        total_columns=total_columns,
        max_rows=max_rows,
        max_columns=max_columns,
        trim_empty_columns=trim_empty_columns,
        trim_empty_rows=trim_empty_rows,
    )


def _preview_sheet_from_rows(
    *,
    name: str,
    index: int,
    rows: Sequence[Sequence[str]],
    total_rows: int,
    total_columns: int,
    max_rows: int,
    max_columns: int,
    trim_empty_columns: bool,
    trim_empty_rows: bool,
) -> WorkbookSheetPreview:
    header_row = list(rows[0]) if rows else []
    header_has_names = any(cell.strip() for cell in header_row)
    preview_columns = max(
        1,
        min(
            max_columns,
            max(total_columns, len(header_row)),
        ),
    )
    headers = _build_headers(header_row, preview_columns)
    body_rows = [
        _normalize_row(list(row), len(headers)) for row in rows[1:]
    ]
    if trim_empty_columns:
        headers, body_rows = _trim_empty_columns(
            headers,
            body_rows,
            header_has_names=header_has_names,
        )
    if trim_empty_rows:
        body_rows = _trim_empty_rows(body_rows)

    return WorkbookSheetPreview(
        name=name,
        index=index,
        headers=headers,
        rows=body_rows,
        total_rows=total_rows,
        total_columns=total_columns,
        truncated_rows=total_rows > max_rows,
        truncated_columns=total_columns > max_columns,
    )


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _normalize_row(row: Sequence[str], length: int) -> list[str]:
    return [row[index] if index < len(row) else "" for index in range(length)]


def _build_headers(raw: Sequence[str], total_columns: int) -> list[str]:
    trimmed = [cell.strip() for cell in raw]
    has_named = any(trimmed)
    header_count = max(total_columns, len(trimmed), 1)
    if has_named:
        headers = trimmed
    else:
        headers = [_column_label(index) for index in range(header_count)]
    return _normalize_row(headers, header_count)


def _column_label(index: int) -> str:
    label = ""
    position = index + 1
    while position > 0:
        remainder = (position - 1) % 26
        label = f"{chr(65 + remainder)}{label}"
        position = (position - 1) // 26
    return f"Column {label}"


def _trim_empty_columns(
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    *,
    header_has_names: bool,
) -> tuple[list[str], list[list[str]]]:
    if not headers:
        return list(headers), [list(row) for row in rows]

    keep_indices: list[int] = []
    for index, header in enumerate(headers):
        if header_has_names and header.strip():
            keep_indices.append(index)
            continue
        if any(row[index].strip() for row in rows):
            keep_indices.append(index)

    if not keep_indices:
        keep_indices = [0]

    trimmed_headers = [headers[index] for index in keep_indices]
    trimmed_rows = [
        [row[index] for index in keep_indices] for row in rows
    ]
    return trimmed_headers, trimmed_rows


def _trim_empty_rows(rows: Sequence[Sequence[str]]) -> list[list[str]]:
    return [list(row) for row in rows if any(cell.strip() for cell in row)]


__all__ = [
    "DEFAULT_PREVIEW_COLUMNS",
    "DEFAULT_PREVIEW_ROWS",
    "MAX_PREVIEW_COLUMNS",
    "MAX_PREVIEW_ROWS",
    "WorkbookSheetPreview",
    "build_workbook_preview_from_csv",
    "build_workbook_preview_from_xlsx",
]
