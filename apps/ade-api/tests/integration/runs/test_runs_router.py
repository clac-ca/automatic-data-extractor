"""Integration coverage for the runs API."""

from __future__ import annotations

import asyncio
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import timedelta
from collections.abc import AsyncIterator
from io import BytesIO
from pathlib import Path
from typing import Any
import tomllib

import openpyxl
import pytest
from httpx import AsyncClient

from ade_api.common.encoding import json_dumps
from ade_api.features.builds.fingerprint import compute_build_fingerprint
from ade_api.features.builds import service as builds_service_module
from ade_api.features.builds.builder import (
    BuildArtifacts,
    BuilderArtifactsEvent,
    BuilderEvent,
    BuilderLogEvent,
    BuilderStepEvent,
    BuildStep,
)
from ade_api.core.models import (
    Build,
    BuildStatus,
    Configuration,
    ConfigurationStatus,
    Document,
    Run,
    RunStatus,
    Workspace,
)
from ade_api.features.configs.storage import compute_config_digest
from ade_api.features.documents.storage import DocumentStorage
from ade_api.features.runs.schemas import RunCreateOptions
from ade_api.features.runs.service import RunExecutionContext, RunsService
from ade_api.features.system_settings.service import SafeModeService
from ade_api.settings import Settings, get_settings
from ade_api.common.time import utc_now
from ade_api.infra.db.mixins import generate_uuid7
from ade_api.infra.db.session import get_sessionmaker
from ade_api.infra.storage import (
    build_venv_marker_path,
    build_venv_path,
    workspace_config_root,
    workspace_documents_root,
)
from tests.utils import login
from ade_engine.schemas import AdeEvent

pytestmark = pytest.mark.asyncio

_SETTINGS = get_settings()
_REPO_ROOT = Path(__file__).resolve().parents[5]
_ENGINE_PACKAGE = _REPO_ROOT / "apps" / "ade-engine"
_CONFIG_TEMPLATE = (
    _REPO_ROOT
    / "apps"
    / "ade-api"
    / "src"
    / "ade_api"
    / "templates"
    / "config_packages"
    / "default"
)
_WHEEL_CACHE_DIR: Path | None = None
_ENGINE_WHEEL: Path | None = None
_CONFIG_WHEEL: Path | None = None
_RUNTIME_CACHE_DIR: Path | None = None
_CACHED_VENV: Path | None = None
_REAL_CONFIG_CACHE: dict[str, str] = {}


class StubBuilder:
    """Test double that avoids real venv work during run router tests."""

    events: list[BuilderEvent] = []

    def __init__(self) -> None:
        self._events = [*type(self).events]

    async def build_stream(
        self,
        *,
        build_id: str,
        workspace_id: str,
        configuration_id: str,
        venv_root: Path,
        config_path: Path,
        engine_spec: str,
        pip_cache_dir: Path | None,
        python_bin: str | None,
        timeout: float,
        fingerprint: str | None = None,
    ) -> AsyncIterator[BuilderEvent]:
        json_dumps(
            {
                "build_id": build_id,
                "workspace_id": workspace_id,
                "configuration_id": configuration_id,
            }
        )
        venv_root.mkdir(parents=True, exist_ok=True)
        for event in self._events:
            yield event


@pytest.fixture(autouse=True)
def _stub_builder(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(builds_service_module, "VirtualEnvironmentBuilder", StubBuilder)
    StubBuilder.events = [
        BuilderStepEvent(step=BuildStep.CREATE_VENV, message="create venv"),
        BuilderArtifactsEvent(
            artifacts=BuildArtifacts(python_version="3.14.0", engine_version="1.2.3")
        ),
        BuilderLogEvent(message="stub build complete"),
    ]


def _engine_version_hint(spec: str) -> str | None:
    """Best-effort version detection mirroring the builds service."""

    spec_path = Path(spec)
    if spec_path.exists() and spec_path.is_dir():
        pyproject = spec_path / "pyproject.toml"
        if pyproject.exists():
            try:
                parsed = tomllib.loads(pyproject.read_text(encoding="utf-8"))
                return parsed.get("project", {}).get("version")
            except Exception:
                return None
    if "==" in spec:
        return spec.split("==", 1)[1]
    return None


def _ensure_wheels() -> tuple[Path, Path]:
    """Build engine/config wheels once per test session to avoid per-test rebuilds."""

    global _WHEEL_CACHE_DIR, _ENGINE_WHEEL, _CONFIG_WHEEL
    if _ENGINE_WHEEL and _CONFIG_WHEEL:
        return _ENGINE_WHEEL, _CONFIG_WHEEL

    cache_dir = Path(tempfile.mkdtemp(prefix="ade-wheel-cache-"))
    _WHEEL_CACHE_DIR = cache_dir

    def _build_wheel(source: Path, pattern: str) -> Path:
        subprocess.run(
            [sys.executable, "-m", "pip", "wheel", "-w", str(cache_dir), str(source)],
            check=True,
        )
        wheels = sorted(cache_dir.glob(pattern))
        if not wheels:
            msg = f"Failed to build wheel for {source}"
            raise RuntimeError(msg)
        return wheels[-1]

    _ENGINE_WHEEL = _build_wheel(_ENGINE_PACKAGE, "ade_engine-*.whl")
    _CONFIG_WHEEL = _build_wheel(_CONFIG_TEMPLATE, "ade_config-*.whl")
    return _ENGINE_WHEEL, _CONFIG_WHEEL


def _ensure_cached_runtime(venv_interpreter: str) -> Path:
    """Create a cached venv with engine + config wheels for reuse across streaming tests."""

    global _RUNTIME_CACHE_DIR, _CACHED_VENV
    if _CACHED_VENV and _CACHED_VENV.exists():
        return _CACHED_VENV

    runtime_root = Path(tempfile.mkdtemp(prefix="ade-runtime-cache-"))
    _RUNTIME_CACHE_DIR = runtime_root
    cached_venv = runtime_root / ".venv"
    subprocess.run([venv_interpreter, "-m", "venv", str(cached_venv)], check=True)

    pip_exe = _venv_executable(cached_venv, "pip")
    engine_wheel, config_wheel = _ensure_wheels()
    _pip_install(pip_exe, engine_wheel)
    _pip_install(pip_exe, config_wheel)

    _CACHED_VENV = cached_venv
    return cached_venv


@pytest.fixture(scope="session", autouse=True)
def _warm_streaming_runtime() -> None:
    """Build cached wheels and runtime once to avoid first-test penalty."""

    interpreter = str(Path(_SETTINGS.python_bin).resolve()) if _SETTINGS.python_bin else sys.executable
    _ensure_cached_runtime(interpreter)


async def _prepare_service(
    session,
    tmp_path: Path,
) -> tuple[RunsService, RunExecutionContext, Document]:
    """Create a runnable configuration and document for workspace runs tests."""

    settings = get_settings()
    workspace = Workspace(name="Workspace", slug=f"ws-{generate_uuid7().hex[:8]}")
    session.add(workspace)
    await session.flush()

    configuration_id = generate_uuid7()
    configuration = Configuration(
        id=configuration_id,
        workspace_id=workspace.id,
        display_name="Config",
        status=ConfigurationStatus.ACTIVE,
        content_digest="digest",
    )
    session.add(configuration)
    await session.flush()

    build_id = generate_uuid7()
    venv_path = build_venv_path(
        settings, str(workspace.id), str(configuration.id), str(build_id)
    )
    venv_path.mkdir(parents=True, exist_ok=True)
    config_root = workspace_config_root(
        settings, str(workspace.id), str(configuration.id)
    )
    config_root.mkdir(parents=True, exist_ok=True)
    (config_root / "pyproject.toml").write_text(
        "[project]\nname='demo'\nversion='0.0.1'\n",
        encoding="utf-8",
    )
    engine_version = _engine_version_hint(settings.engine_spec)
    digest = compute_config_digest(config_root)
    fingerprint = compute_build_fingerprint(
        config_digest=digest,
        engine_spec=settings.engine_spec,
        engine_version=engine_version,
        python_version=".".join(map(str, sys.version_info[:3])),
        python_bin=settings.python_bin,
        extra={},
    )

    marker = build_venv_marker_path(
        settings, str(workspace.id), str(configuration.id), str(build_id)
    )
    marker.parent.mkdir(parents=True, exist_ok=True)
    marker.write_text(
        json_dumps({"build_id": build_id, "fingerprint": fingerprint}, indent=2),
        encoding="utf-8",
    )
    bin_dir = venv_path / ("Scripts" if os.name == "nt" else "bin")
    bin_dir.mkdir(parents=True, exist_ok=True)
    python_name = "python.exe" if os.name == "nt" else "python"
    (bin_dir / python_name).write_text("", encoding="utf-8")

    build = Build(
        id=build_id,
        workspace_id=workspace.id,
        configuration_id=configuration.id,
        status=BuildStatus.ACTIVE,
        created_at=utc_now(),
        started_at=utc_now(),
        finished_at=utc_now(),
        exit_code=0,
        fingerprint=fingerprint,
        engine_spec=settings.engine_spec,
        engine_version=engine_version,
        python_version=".".join(map(str, sys.version_info[:3])),
        python_interpreter=settings.python_bin or sys.executable,
        config_digest=digest,
    )
    session.add(build)

    configuration.active_build_id = build_id  # type: ignore[attr-defined]
    configuration.active_build_fingerprint = fingerprint  # type: ignore[attr-defined]
    configuration.content_digest = digest
    await session.flush()

    storage = DocumentStorage(workspace_documents_root(settings, workspace.id))
    document_id = generate_uuid7()
    stored_uri = storage.make_stored_uri(str(document_id))
    source_path = storage.path_for(stored_uri)
    source_path.parent.mkdir(parents=True, exist_ok=True)
    source_path.write_text("name\nAlice", encoding="utf-8")

    document = Document(
        id=document_id,
        workspace_id=workspace.id,
        original_filename="input.csv",
        content_type="text/csv",
        byte_size=source_path.stat().st_size,
        sha256="deadbeef",
        stored_uri=stored_uri,
        attributes={},
        uploaded_by_user_id=None,
        status="processed",
        expires_at=utc_now() + timedelta(days=30),
    )
    session.add(document)
    await session.commit()

    service = RunsService(session=session, settings=settings)
    run, context = await service.prepare_run(
        configuration_id=str(configuration.id),
        options=RunCreateOptions(input_document_id=str(document.id)),
    )
    return service, context, document


async def _seed_configuration(
    *,
    settings: Settings,
    workspace_id: str,
    tmp_path: Path,
) -> str:
    """Insert a configuration and active build used to drive runs tests."""

    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        workspace_id_str = str(workspace_id)
        config_id = generate_uuid7()
        config = Configuration(
            id=config_id,
            workspace_id=workspace_id_str,
            display_name="Test Configuration",
            status=ConfigurationStatus.ACTIVE,
        )
        session.add(config)
        await session.flush()

        config_id_str = str(config.id)
        build_id = generate_uuid7()
        venv_path = build_venv_path(settings, workspace_id_str, config_id_str, str(build_id))
        venv_path.mkdir(parents=True, exist_ok=True)
        config_root = workspace_config_root(settings, workspace_id_str, config_id_str)
        config_root.mkdir(parents=True, exist_ok=True)
        (config_root / "pyproject.toml").write_text(
            "[project]\nname='demo'\nversion='0.0.1'\n",
            encoding="utf-8",
        )
        digest = compute_config_digest(config_root)
        engine_version = _engine_version_hint(settings.engine_spec)
        fingerprint = compute_build_fingerprint(
            config_digest=digest,
            engine_spec=settings.engine_spec,
            engine_version=engine_version,
            python_version=".".join(map(str, sys.version_info[:3])),
            python_bin=settings.python_bin,
            extra={},
        )

        marker = build_venv_marker_path(settings, workspace_id_str, config_id_str, str(build_id))
        marker.parent.mkdir(parents=True, exist_ok=True)
        marker.write_text(
        json_dumps({"build_id": build_id, "fingerprint": fingerprint}, indent=2),
        encoding="utf-8",
    )
        bin_dir = venv_path / ("Scripts" if os.name == "nt" else "bin")
        bin_dir.mkdir(parents=True, exist_ok=True)
        python_name = "python.exe" if os.name == "nt" else "python"
        (bin_dir / python_name).write_text("", encoding="utf-8")

        build = Build(
            id=build_id,
            workspace_id=workspace_id_str,
            configuration_id=config.id,
            status=BuildStatus.ACTIVE,
            created_at=utc_now(),
            started_at=utc_now(),
            finished_at=utc_now(),
            exit_code=0,
            fingerprint=fingerprint,
            engine_spec=settings.engine_spec,
            engine_version=engine_version,
            python_version=".".join(map(str, sys.version_info[:3])),
            python_interpreter=settings.python_bin or sys.executable,
            config_digest=digest,
        )
        session.add(build)

        config.active_build_id = build_id  # type: ignore[attr-defined]
        config.active_build_fingerprint = fingerprint  # type: ignore[attr-defined]
        config.content_digest = digest
        await session.commit()
        return config_id_str


def _venv_executable(venv_path: Path, name: str) -> Path:
    """Return an executable path inside ``venv_path`` honoring platform layout."""

    bin_dir = "Scripts" if os.name == "nt" else "bin"
    suffix = ".exe" if os.name == "nt" else ""
    return venv_path / bin_dir / f"{name}{suffix}"


def _pip_install(pip_executable: Path, package: Path | str) -> None:
    """Install ``package`` into the runtime venv using pip."""

    env = os.environ.copy()
    env.setdefault("PIP_DISABLE_PIP_VERSION_CHECK", "1")
    env.setdefault("PIP_ROOT_USER_ACTION", "ignore")
    subprocess.run([str(pip_executable), "install", str(package)], check=True, env=env)


async def _seed_real_configuration(
    *,
    settings: Settings,
    workspace_id: str,
    tmp_path: Path,
) -> str:
    """Create a configuration with a fully provisioned runtime environment."""

    workspace_id_str = str(workspace_id)

    if workspace_id_str in _REAL_CONFIG_CACHE:
        # Reuse the existing configuration for this workspace to avoid repeated builds.
        cached_config_id = _REAL_CONFIG_CACHE[workspace_id_str]
        session_factory = get_sessionmaker(settings=settings)
        async with session_factory() as session:
            existing = await session.get(Configuration, cached_config_id)
            if existing is not None:
                return cached_config_id

    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        config_id = generate_uuid7()
        config = Configuration(
            id=config_id,
            workspace_id=workspace_id_str,
            display_name="Real Config",
            status=ConfigurationStatus.ACTIVE,
            content_digest="integration-test",
        )
        session.add(config)
        await session.flush()

        config_id_str = str(config.id)
        build_id = generate_uuid7()
        build_id_str = str(build_id)
        venv_path = build_venv_path(settings, workspace_id_str, config_id_str, build_id_str)
        python_bin = settings.python_bin
        resolved_python_bin = str(Path(python_bin).resolve()) if python_bin else None
        venv_interpreter = resolved_python_bin or sys.executable
        config_src = workspace_config_root(settings, workspace_id_str, config_id_str)
        if not config_src.exists():
            shutil.copytree(_CONFIG_TEMPLATE, config_src, dirs_exist_ok=True)
        engine_version = _engine_version_hint(settings.engine_spec)
        try:
            python_version = subprocess.check_output(
                [resolved_python_bin or sys.executable, "-c", "import sys; print('.'.join(map(str, sys.version_info[:3])))"],
                text=True,
            ).strip()
        except Exception:
            python_version = None
        python_interpreter = resolved_python_bin
        digest = compute_config_digest(config_src)
        now = utc_now()
        fingerprint = compute_build_fingerprint(
            config_digest=digest,
            engine_spec=settings.engine_spec,
            engine_version=engine_version,
            python_version=python_version,
            python_bin=resolved_python_bin,
            extra={},
        )
        marker = build_venv_marker_path(settings, workspace_id_str, config_id_str, build_id_str)
        existing_marker = marker if marker.exists() else None
        reuse_fingerprint = None
        reuse_build_id = None
        if existing_marker:
            try:
                parsed = json.loads(existing_marker.read_text(encoding="utf-8"))
                reuse_fingerprint = parsed.get("fingerprint")
                reuse_build_id = parsed.get("build_id")
            except Exception:
                reuse_fingerprint = None
                reuse_build_id = None

        if reuse_fingerprint and reuse_fingerprint == fingerprint and venv_path.exists():
            build_id_str = str(reuse_build_id or build_id_str)
        else:
            if venv_path.exists():
                shutil.rmtree(venv_path)
            venv_path.parent.mkdir(parents=True, exist_ok=True)
            cached_venv = _ensure_cached_runtime(venv_interpreter)
            shutil.copytree(cached_venv, venv_path, dirs_exist_ok=True)

            marker = build_venv_marker_path(
                settings, workspace_id_str, config_id_str, build_id_str
            )
            marker.parent.mkdir(parents=True, exist_ok=True)
            marker.write_text(
                json_dumps({"build_id": build_id_str, "fingerprint": fingerprint}, indent=2),
                    encoding="utf-8",
                )

        python_exe = _venv_executable(venv_path, "python")
        config.content_digest = digest  # type: ignore[attr-defined]
        build = Build(
            id=build_id_str,
            workspace_id=workspace_id_str,
            configuration_id=config.id,
            status=BuildStatus.ACTIVE,
            created_at=now,
            started_at=now,
            finished_at=now,
            exit_code=0,
            fingerprint=fingerprint,
            engine_spec=settings.engine_spec,
            engine_version=engine_version,
            python_version=python_version,
            python_interpreter=python_interpreter,
            config_digest=digest,
        )
        session.add(build)
        config.active_build_id = build_id_str  # type: ignore[attr-defined]
        config.active_build_fingerprint = fingerprint  # type: ignore[attr-defined]
        await session.commit()
        _REAL_CONFIG_CACHE[workspace_id_str] = str(config.id)
        return str(config.id)


async def _seed_document(
    *,
    settings: Settings,
    workspace_id: str,
) -> Document:
    """Persist a CSV document and return its metadata row."""

    csv_bytes = (
        b"member_id,email,first_name,last_name\n"
        b"1,alice@example.com,Alice,Anderson\n"
        b"2,bob@example.com,Bob,Brown\n"
    )
    workspace_id_str = str(workspace_id)
    document_id = generate_uuid7()
    storage = DocumentStorage(workspace_documents_root(settings, workspace_id_str))
    stored_uri = storage.make_stored_uri(str(document_id))
    target_path = storage.path_for(stored_uri)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(csv_bytes)

    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        document = Document(
            id=document_id,
            workspace_id=workspace_id_str,
            original_filename="members.csv",
            content_type="text/csv",
            byte_size=len(csv_bytes),
            sha256=hashlib.sha256(csv_bytes).hexdigest(),
            stored_uri=stored_uri,
            attributes={},
            expires_at=utc_now() + timedelta(days=30),
        )
        session.add(document)
        await session.commit()
        await session.refresh(document)
    return document


async def _seed_workbook_document(
    *,
    settings: Settings,
    workspace_id: str,
) -> Document:
    """Persist a multi-sheet XLSX document and return its metadata row."""

    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Members"
    header = ["member_id", "email", "first_name", "last_name"]
    first.append(header)
    first.append(["101", "first-sheet@example.com", "First", "Sheet"])

    second = workbook.create_sheet("Selected")
    second.append(header)
    second.append(["202", "selected@example.com", "Second", "Worksheet"])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    xlsx_bytes = buffer.getvalue()

    workspace_id_str = str(workspace_id)
    document_id = generate_uuid7()
    storage = DocumentStorage(workspace_documents_root(settings, workspace_id_str))
    stored_uri = storage.make_stored_uri(str(document_id))
    target_path = storage.path_for(stored_uri)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(xlsx_bytes)

    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        document = Document(
            id=document_id,
            workspace_id=workspace_id_str,
            original_filename="members.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            byte_size=len(xlsx_bytes),
            sha256=hashlib.sha256(xlsx_bytes).hexdigest(),
            stored_uri=stored_uri,
            attributes={},
            expires_at=utc_now() + timedelta(days=30),
        )
        session.add(document)
        await session.commit()
        await session.refresh(document)
        return document


async def _auth_headers(client: AsyncClient, *, email: str, password: str) -> dict[str, str]:
    token, _ = await login(client, email=email, password=password)
    return {"Authorization": f"Bearer {token}"}


async def _wait_for_completion(
    client: AsyncClient,
    run_id: str,
    *,
    attempts: int = 10,
    delay: float = 0.05,
    headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Poll the run endpoint until it leaves the queued state."""

    payload: dict[str, Any] = {}
    for _ in range(attempts):
        response = await client.get(f"/api/v1/runs/{run_id}", headers=headers)
        payload = response.json()
        if payload.get("status") != "queued":
            return payload
        await asyncio.sleep(delay)
    return payload


async def _collect_run_events(
    client: AsyncClient,
    run_id: str,
    *,
    headers: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Stream run events via SSE until completion and return the payloads."""

    events: list[dict[str, Any]] = []
    async with client.stream(
        "GET",
        f"/api/v1/runs/{run_id}/events/stream",
        headers=headers,
        params={"after_sequence": 0},
    ) as response:
        assert response.status_code == 200
        async for line in response.aiter_lines():
            if not line or not line.startswith("data: "):
                continue
            event = json.loads(line.removeprefix("data: "))
            events.append(event)
            if event.get("type") == "run.completed":
                break
    return events


async def test_stream_run_safe_mode(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    override_app_settings,
    tmp_path,
) -> None:
    """Streaming a run in safe mode should emit events and persist state."""

    settings = override_app_settings(safe_mode=True)
    configuration_id = await _seed_configuration(
        settings=settings,
        workspace_id=seed_identity["workspace_id"],
        tmp_path=tmp_path,
    )
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    create_response = await async_client.post(
        f"/api/v1/configurations/{configuration_id}/runs",
        headers=headers,
        json={},
    )
    assert create_response.status_code == 201
    run_id = create_response.json()["id"]
    events = await _collect_run_events(async_client, run_id, headers=headers)

    assert events, "expected streaming events"
    assert events[0]["type"] == "run.queued"
    assert events[-1]["type"] == "run.completed"

    run_response = await async_client.get(f"/api/v1/runs/{run_id}", headers=headers)
    assert run_response.status_code == 200
    run_payload = run_response.json()
    assert run_payload["status"] == "succeeded"
    assert run_payload["exit_code"] == 0
    assert run_payload["build_id"]

    console_messages = [
        event["payload"]["message"]
        for event in events
        if event.get("type") == "console.line" and "payload" in event
    ]
    assert any("safe mode" in message.lower() for message in console_messages)


async def test_stream_run_respects_persisted_safe_mode_override(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    override_app_settings,
    tmp_path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Persisted safe mode state should override environment defaults."""

    settings = override_app_settings(safe_mode=True)
    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        service = SafeModeService(session=session, settings=settings)
        await service.update_status(
            enabled=False, detail="Disable safe mode for integration coverage"
        )

    configuration_id = await _seed_configuration(
        settings=settings,
        workspace_id=seed_identity["workspace_id"],
        tmp_path=tmp_path,
    )
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    async def fake_execute_engine(
        self: RunsService,
        *,
        run,
        context,
        options,
        safe_mode_enabled: bool = False,
    ):
        assert not safe_mode_enabled
        completion = await self._complete_run(
            run,
            status=RunStatus.SUCCEEDED,
            exit_code=0,
            summary="Safe mode override respected",
        )
        yield await self._event_dispatcher.emit(
            type="run.completed",
            source="api",
            workspace_id=str(completion.workspace_id),
            configuration_id=str(completion.configuration_id),
            run_id=str(completion.id),
            build_id=str(completion.build_id),
            payload={
                "status": "succeeded",
                "execution": {"exit_code": completion.exit_code},
                "summary": {"run": {"status": "succeeded"}},
            },
        )

    monkeypatch.setattr(RunsService, "_execute_engine", fake_execute_engine)

    create_response = await async_client.post(
        f"/api/v1/configurations/{configuration_id}/runs",
        headers=headers,
        json={},
    )
    assert create_response.status_code == 201
    run_id = create_response.json()["id"]
    events = await _collect_run_events(async_client, run_id, headers=headers)

    assert events, "expected streaming events"
    assert events[-1]["type"] == "run.completed"

    run_response = await async_client.get(f"/api/v1/runs/{run_id}", headers=headers)
    assert run_response.status_code == 200
    run_payload = run_response.json()
    assert run_payload["status"] == "succeeded"
    assert run_payload["exit_code"] == 0
    assert run_payload.get("failure_message") is None
    assert run_payload["build_id"]


async def test_non_stream_run_executes_in_background(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    override_app_settings,
    tmp_path,
) -> None:
    """Non-streaming requests should return immediately and finish in the background."""

    settings = override_app_settings(safe_mode=True)
    configuration_id = await _seed_configuration(
        settings=settings,
        workspace_id=seed_identity["workspace_id"],
        tmp_path=tmp_path,
    )
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    response = await async_client.post(
        f"/api/v1/configurations/{configuration_id}/runs",
        headers=headers,
        json={},
    )
    assert response.status_code == 201
    payload = response.json()
    run_id = payload["id"]

    completed = await _wait_for_completion(async_client, run_id, headers=headers)
    assert completed["status"] == "succeeded"
    assert completed["exit_code"] == 0


async def test_stream_run_processes_real_documents(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    override_app_settings,
    tmp_path,
) -> None:
    """Streaming runs should execute the full engine pipeline against inputs."""

    settings = override_app_settings(safe_mode=False)
    workspace_id = seed_identity["workspace_id"]
    configuration_id = await _seed_real_configuration(
        settings=settings,
        workspace_id=workspace_id,
        tmp_path=tmp_path,
    )
    document = await _seed_document(settings=settings, workspace_id=workspace_id)
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    create_response = await async_client.post(
        f"/api/v1/configurations/{configuration_id}/runs",
        headers=headers,
        json={"options": {"input_document_id": str(document.id)}},
    )
    assert create_response.status_code == 201, create_response.text
    run_id = create_response.json()["id"]
    events = await _collect_run_events(async_client, run_id, headers=headers)

    assert events and events[0]["type"] == "run.queued"
    assert events[-1]["type"] == "run.completed"
    assert events[-1].get("payload", {}).get("status") == "succeeded"

    run_response = await async_client.get(f"/api/v1/runs/{run_id}", headers=headers)
    assert run_response.status_code == 200
    run_payload = run_response.json()
    assert run_payload["status"] == "succeeded"
    assert run_payload["exit_code"] == 0
    assert run_payload["build_id"]
    assert str(document.id) in run_payload["input"]["document_ids"]
    assert run_payload["input"]["input_sheet_names"] == []

    outputs_response = await async_client.get(
        f"/api/v1/runs/{run_id}/outputs", headers=headers
    )
    assert outputs_response.status_code == 200
    outputs_payload = outputs_response.json()
    assert outputs_payload["files"], "expected normalized outputs"
    normalized = next(
        (entry for entry in outputs_payload["files"] if entry["name"].endswith("normalized.xlsx")),
        None,
    )
    assert normalized is not None

    console_events = [event for event in events if event.get("type") == "console.line"]
    assert console_events, "expected ADE engine logs to appear in the event stream"

    session_factory = get_sessionmaker(settings=settings)
    async with session_factory() as session:
        refreshed = await session.get(Document, document.id)
        assert refreshed is not None
        assert refreshed.last_run_at is not None


@pytest.mark.asyncio()
async def test_list_workspace_runs_filters_by_status_and_document(
    session,
    tmp_path: Path,
) -> None:
    service, context, document = await _prepare_service(session, tmp_path)
    configuration = await session.get(Configuration, context.configuration_id)
    assert configuration is not None

    run_success = Run(
        id=generate_uuid7(),
        workspace_id=configuration.workspace_id,
        configuration_id=configuration.id,
        input_document_id=document.id,
        status=RunStatus.SUCCEEDED,
        created_at=utc_now(),
    )
    run_failed = Run(
        id=generate_uuid7(),
        workspace_id=configuration.workspace_id,
        configuration_id=configuration.id,
        status=RunStatus.FAILED,
        created_at=utc_now(),
    )

    session.add_all([run_success, run_failed])
    await session.commit()

    page = await service.list_runs(
        workspace_id=configuration.workspace_id,
        statuses=None,
        input_document_id=None,
        page=1,
        page_size=10,
        include_total=True,
    )

    assert page.total == 3
    assert {str(run.id) for run in page.items} == {
        str(context.run_id),
        str(run_success.id),
        str(run_failed.id),
    }

    filtered = await service.list_runs(
        workspace_id=configuration.workspace_id,
        statuses=[RunStatus.SUCCEEDED],
        input_document_id=str(document.id),
        page=1,
        page_size=5,
        include_total=True,
    )

    assert filtered.total == 1
    assert [str(run.id) for run in filtered.items] == [str(run_success.id)]


async def test_list_configuration_runs_scopes_and_filters(
    async_client: AsyncClient,
    session,
    seed_identity: dict[str, Any],
) -> None:
    """Configuration-scoped run listings should filter within the workspace."""

    workspace_id = seed_identity["workspace_id"]
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    configuration = Configuration(
        workspace_id=workspace_id,
        display_name="Primary",
        status=ConfigurationStatus.ACTIVE,
    )
    other_configuration = Configuration(
        workspace_id=workspace_id,
        display_name="Secondary",
        status=ConfigurationStatus.ACTIVE,
    )
    session.add_all([configuration, other_configuration])
    await session.flush()

    run_succeeded = Run(
        workspace_id=workspace_id,
        configuration_id=configuration.id,
        status=RunStatus.SUCCEEDED,
        created_at=utc_now(),
    )
    run_failed = Run(
        workspace_id=workspace_id,
        configuration_id=configuration.id,
        status=RunStatus.FAILED,
        created_at=utc_now(),
    )
    run_other = Run(
        workspace_id=workspace_id,
        configuration_id=other_configuration.id,
        status=RunStatus.SUCCEEDED,
        created_at=utc_now(),
    )
    session.add_all([run_succeeded, run_failed, run_other])
    await session.commit()

    list_response = await async_client.get(
        f"/api/v1/configurations/{configuration.id}/runs",
        headers=headers,
        params={"include_total": "true"},
    )

    assert list_response.status_code == 200, list_response.text
    payload = list_response.json()
    assert payload["total"] == 2
    assert {item["id"] for item in payload["items"]} == {
        str(run_succeeded.id),
        str(run_failed.id),
    }

    filtered_response = await async_client.get(
        f"/api/v1/configurations/{configuration.id}/runs",
        headers=headers,
        params={"status": RunStatus.SUCCEEDED.value},
    )
    assert filtered_response.status_code == 200, filtered_response.text
    filtered_payload = filtered_response.json()
    assert [item["id"] for item in filtered_payload["items"]] == [
        str(run_succeeded.id)
    ]


async def test_list_configuration_runs_returns_not_found(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
) -> None:
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )
    missing_id = generate_uuid7()

    response = await async_client.get(
        f"/api/v1/configurations/{missing_id}/runs", headers=headers
    )

    assert response.status_code == 404


async def test_stream_run_processes_all_worksheets_when_unspecified(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    override_app_settings,
    tmp_path,
) -> None:
    """Streaming runs should ingest every worksheet when no override is provided."""

    settings = override_app_settings(safe_mode=False)
    workspace_id = seed_identity["workspace_id"]
    configuration_id = await _seed_real_configuration(
        settings=settings,
        workspace_id=workspace_id,
        tmp_path=tmp_path,
    )
    document = await _seed_workbook_document(
        settings=settings, workspace_id=workspace_id
    )
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    create_response = await async_client.post(
        f"/api/v1/configurations/{configuration_id}/runs",
        headers=headers,
        json={
            "options": {
                "input_document_id": str(document.id),
            },
        },
    )
    assert create_response.status_code == 201, create_response.text
    run_id = create_response.json()["id"]
    events = await _collect_run_events(async_client, run_id, headers=headers)

    assert events and events[0]["type"] == "run.queued"
    assert events[-1]["type"] == "run.completed"
    assert events[-1].get("payload", {}).get("status") == "succeeded"

    outputs_response = await async_client.get(
        f"/api/v1/runs/{run_id}/outputs", headers=headers
    )
    assert outputs_response.status_code == 200
    outputs_payload = outputs_response.json()
    normalized = next(
        (entry for entry in outputs_payload["files"] if entry["name"].endswith("normalized.xlsx")),
        None,
    )
    assert normalized is not None, "expected normalized workbook output"

    download_response = await async_client.get(normalized["download_url"], headers=headers)
    assert download_response.status_code == 200
    workbook = openpyxl.load_workbook(BytesIO(download_response.content), read_only=True)
    rows = [list(sheet.iter_rows(values_only=True)) for sheet in workbook]
    workbook.close()

    flattened = [cell for sheet_rows in rows for row in sheet_rows for cell in row if cell]
    assert "first-sheet@example.com" in flattened
    assert "selected@example.com" in flattened


async def test_stream_run_sheet_selection_variants(
    async_client: AsyncClient,
    seed_identity: dict[str, Any],
    override_app_settings,
    tmp_path,
) -> None:
    """Streaming runs should honor sheet selection options without rebuilding venvs."""

    settings = override_app_settings(safe_mode=False)
    workspace_id = seed_identity["workspace_id"]
    configuration_id = await _seed_real_configuration(
        settings=settings,
        workspace_id=workspace_id,
        tmp_path=tmp_path,
    )
    document = await _seed_workbook_document(settings=settings, workspace_id=workspace_id)
    owner = seed_identity["workspace_owner"]
    headers = await _auth_headers(
        async_client, email=owner["email"], password=owner["password"]
    )

    scenarios = [
        {
            "label": "all_sheets",
            "options": {"input_document_id": str(document.id)},
            "expect_selected": True,
            "expect_first_sheet": True,
        },
        {
            "label": "single_name",
            "options": {"input_document_id": str(document.id), "input_sheet_name": "Selected"},
            "expect_selected": True,
            "expect_first_sheet": False,
        },
        {
            "label": "list_names",
            "options": {
                "input_document_id": str(document.id),
                "input_sheet_names": ["Selected"],
            },
            "expect_selected": True,
            "expect_first_sheet": False,
        },
    ]

    for scenario in scenarios:
        create_response = await async_client.post(
            f"/api/v1/configurations/{configuration_id}/runs",
            headers=headers,
            json={"options": scenario["options"]},
        )
        assert create_response.status_code == 201, create_response.text
        run_id = create_response.json()["id"]
        events = await _collect_run_events(async_client, run_id, headers=headers)

        assert events and events[0]["type"] == "run.queued"
        assert events[-1]["type"] == "run.completed"
        assert events[-1].get("payload", {}).get("status") == "succeeded"

        outputs_response = await async_client.get(
            f"/api/v1/runs/{run_id}/outputs", headers=headers
        )
        assert outputs_response.status_code == 200
        outputs_payload = outputs_response.json()
        normalized = next(
            (entry for entry in outputs_payload["files"] if entry["name"].endswith("normalized.xlsx")),
            None,
        )
        assert normalized is not None, f"expected normalized workbook output for {scenario['label']}"

        download_response = await async_client.get(normalized["download_url"], headers=headers)
        assert download_response.status_code == 200
        workbook = openpyxl.load_workbook(BytesIO(download_response.content), read_only=True)
        rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
        workbook.close()

        assert rows[0][:4] == ("member_id", "email", "first_name", "last_name")
        data_rows = rows[1:]
        flattened = [
            [str(cell) for cell in row if cell not in (None, "")]
            for row in data_rows
        ]
        has_selected = any("selected@example.com" in row for row in flattened)
        has_first_sheet = any("first-sheet@example.com" in row for row in flattened)
        assert has_selected is scenario["expect_selected"]
        assert has_first_sheet is scenario["expect_first_sheet"]
