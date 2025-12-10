"""Table detection utilities."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping

from openpyxl.worksheet.worksheet import Worksheet

from ade_engine.config.loader import ConfigRuntime
from ade_engine.config.row_detectors import RowDetector
from ade_engine.exceptions import ConfigError
from ade_engine.runtime import PluginInvoker
from ade_engine.types.contexts import RunContext
from ade_engine.types.origin import TableRegion

HEADER_SCORE_THRESHOLD = 0.6
DATA_SCORE_THRESHOLD = 0.5


@dataclass(frozen=True)
class RowDetectorScore:
    row_index: int
    values: list[Any]
    header_score: float
    data_score: float


def _normalize_row_detector_scores(result: Any, *, detector: RowDetector) -> Mapping[str, float]:
    if isinstance(result, Mapping):
        if "scores" in result:
            raise ConfigError(
                f"{detector.qualified_name} must return a float or dict of label deltas (no 'scores' wrapper')"
            )
        return {str(k): float(v) for k, v in result.items()}

    if result is None:
        return {}

    try:
        delta = float(result)
    except (TypeError, ValueError) as exc:
        raise ConfigError(f"{detector.qualified_name} returned a non-numeric score: {result!r}") from exc

    if detector.default_label is None:
        raise ConfigError(
            f"{detector.qualified_name} returned a bare score but no default label could be inferred; "
            "return a dict like {'header': score} or set a default label on the detector."
        )
    return {detector.default_label: delta}


def _iter_rows(ws: Worksheet) -> Iterator[tuple[int, list[Any]]]:
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        yield idx, list(row)


def _score_rows(
    *,
    rows: Iterable[tuple[int, list[Any]]],
    detectors: tuple[RowDetector, ...],
    invoker: PluginInvoker,
) -> list[RowDetectorScore]:
    scored: list[RowDetectorScore] = []

    for row_index, values in rows:
        totals: dict[str, float] = {}

        for detector in detectors:
            result = invoker.call(
                detector.func,
                row_index=row_index,
                row_values=values,
            )
            deltas = _normalize_row_detector_scores(result, detector=detector)
            for label, delta in deltas.items():
                totals[label] = totals.get(label, 0.0) + delta

        scored.append(
            RowDetectorScore(
                row_index=row_index,
                values=values,
                header_score=totals.get("header", 0.0),
                data_score=totals.get("data", 0.0),
            )
        )

    return scored


def _region_width(header: list[Any], data_rows: list[RowDetectorScore]) -> int:
    return max(len(header), max((len(row.values) for row in data_rows), default=0), 1)


def _detect_regions(
    *,
    scored_rows: list[RowDetectorScore],
    header_threshold: float,
    data_threshold: float,
) -> list[tuple[int, int, int]]:
    """Return a list of ``(min_row, max_row, max_col)`` tuples."""

    regions: list[tuple[int, int, int]] = []
    i = 0

    while i < len(scored_rows):
        header_row = scored_rows[i]
        if header_row.header_score < header_threshold:
            i += 1
            continue

        data_rows: list[RowDetectorScore] = []
        j = i + 1
        while j < len(scored_rows) and scored_rows[j].data_score >= data_threshold:
            data_rows.append(scored_rows[j])
            j += 1

        if data_rows:
            max_col = _region_width(header_row.values, data_rows)
            regions.append((header_row.row_index, data_rows[-1].row_index, max_col))

        i = j if j > i else i + 1

    return regions


class TableDetector:
    """Detect table regions on a worksheet using config-provided row detectors."""

    def __init__(
        self,
        *,
        header_threshold: float = HEADER_SCORE_THRESHOLD,
        data_threshold: float = DATA_SCORE_THRESHOLD,
    ) -> None:
        self.header_threshold = header_threshold
        self.data_threshold = data_threshold

    def detect(
        self,
        *,
        source_path: Path,  # noqa: ARG002 - available via run_ctx.source_path
        worksheet: Worksheet,
        runtime: ConfigRuntime,
        run_ctx: RunContext,  # noqa: ARG002 - available via invoker.base_kwargs()
        invoker: PluginInvoker,
        logger=None,  # noqa: ARG002 - reserved for future visibility
    ) -> list[TableRegion]:
        scored_rows = _score_rows(rows=_iter_rows(worksheet), detectors=runtime.row_detectors, invoker=invoker)
        if logger and logger.isEnabledFor(logging.DEBUG):
            logger.debug(
                "Scored rows for detection",
                extra={
                    "data": {
                        "stage": "detect",
                        "row_count": len(scored_rows),
                        "header_threshold": self.header_threshold,
                        "data_threshold": self.data_threshold,
                        "detector_count": len(runtime.row_detectors),
                        "sheet_name": worksheet.title,
                    },
                },
            )

        regions = [
            TableRegion(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col)
            for min_row, max_row, max_col in _detect_regions(
                scored_rows=scored_rows,
                header_threshold=self.header_threshold,
                data_threshold=self.data_threshold,
            )
        ]
        if logger and logger.isEnabledFor(logging.DEBUG):
            logger.debug(
                "Detected table regions",
                extra={
                    "data": {
                        "stage": "detect",
                        "sheet_name": worksheet.title,
                        "regions": [
                            {"min_row": r.min_row, "max_row": r.max_row, "min_col": r.min_col, "max_col": r.max_col}
                            for r in regions
                        ],
                    },
                },
            )
        return regions


__all__ = ["TableDetector", "HEADER_SCORE_THRESHOLD", "DATA_SCORE_THRESHOLD", "RowDetectorScore"]
