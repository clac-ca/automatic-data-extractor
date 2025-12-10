"""Sheet/Table orchestration for Engine."""

from __future__ import annotations

import logging
from typing import Callable

from openpyxl import Workbook

from ade_engine.config.loader import ConfigRuntime
from ade_engine.hooks.dispatcher import HookDispatcher
from ade_engine.pipeline.detect import TableDetector
from ade_engine.pipeline.extract import TableExtractor
from ade_engine.pipeline.layout import SheetLayout
from ade_engine.pipeline.mapping import ColumnMapper
from ade_engine.pipeline.normalize import TableNormalizer
from ade_engine.pipeline.render import TableRenderer
from ade_engine.runtime import PluginInvoker
from ade_engine.types.contexts import RunContext, TableContext, TableView, WorksheetContext
from ade_engine.types.origin import TableOrigin, TableRegion
from ade_engine.types.issues import Severity


class Pipeline:
    """Coordinates detection → extraction → mapping → normalization → rendering."""

    def __init__(
        self,
        *,
        detector: TableDetector | None = None,
        extractor: TableExtractor | None = None,
        mapper: ColumnMapper | None = None,
        normalizer: TableNormalizer | None = None,
        renderer_factory: Callable[[], TableRenderer] | None = None,
    ) -> None:
        self.detector = detector or TableDetector()
        self.extractor = extractor or TableExtractor()
        self.mapper = mapper or ColumnMapper()
        self.normalizer = normalizer or TableNormalizer()
        self.renderer_factory = renderer_factory or (lambda: TableRenderer(layout=SheetLayout()))

    def process_sheet(
        self,
        *,
        runtime: ConfigRuntime,
        run_ctx: RunContext,
        hook_dispatcher: HookDispatcher,
        invoker: PluginInvoker,
        source_wb: Workbook,
        output_wb: Workbook,
        sheet_name: str,
        sheet_position: int,
        sheet_index_lookup: dict[str, int],
        logger: logging.Logger | logging.LoggerAdapter | None = None,
    ) -> None:
        emitter = logger if hasattr(logger, "event") else invoker.logger

        src_ws = source_wb[sheet_name]
        out_ws = output_wb.create_sheet(title=sheet_name, index=sheet_position)
        sheet_ctx = WorksheetContext(
            run=run_ctx,
            sheet_index=sheet_index_lookup.get(sheet_name, sheet_position),
            source_worksheet=src_ws,
            output_worksheet=out_ws,
        )

        emitter.event(
            "sheet.started",
            message=f"Sheet started: {sheet_name}",
            sheet_name=sheet_name,
            sheet_index=sheet_ctx.sheet_index,
        )
        hook_dispatcher.on_sheet_start(sheet_ctx)

        regions = self.detector.detect(
            source_path=run_ctx.source_path,
            worksheet=src_ws,
            runtime=runtime,
            run_ctx=run_ctx,
            invoker=invoker,
            logger=logger,
        )
        row_count = getattr(src_ws, "max_row", None) or 0
        emitter.event(
            "sheet.tables_detected",
            message=f"Detected {len(regions)} table(s) in {sheet_name}",
            sheet_name=sheet_name,
            sheet_index=sheet_ctx.sheet_index,
            input_file=str(run_ctx.source_path),
            row_count=row_count,
            table_count=len(regions),
            tables=[
                {
                    "table_index": idx,
                    "region": {
                        "min_row": r.min_row,
                        "max_row": r.max_row,
                        "min_col": r.min_col,
                        "max_col": r.max_col,
                    },
                }
                for idx, r in enumerate(regions)
            ],
        )

        renderer = self.renderer_factory()
        for table_index, region in enumerate(regions):
            self.process_table(
                runtime=runtime,
                run_ctx=run_ctx,
                hook_dispatcher=hook_dispatcher,
                invoker=invoker,
                sheet_ctx=sheet_ctx,
                region=region,
                table_index=table_index,
                renderer=renderer,
                logger=logger,
            )

    def process_table(
        self,
        *,
        runtime: ConfigRuntime,
        run_ctx: RunContext,
        hook_dispatcher: HookDispatcher,
        invoker: PluginInvoker,
        sheet_ctx: WorksheetContext,
        region: TableRegion,
        table_index: int,
        renderer: TableRenderer,
        logger: logging.Logger | logging.LoggerAdapter | None = None,
    ) -> TableContext:
        sheet_name = sheet_ctx.source_worksheet.title
        emitter = logger if hasattr(logger, "event") else invoker.logger

        origin = TableOrigin(
            source_path=run_ctx.source_path,
            sheet_name=sheet_name,
            sheet_index=sheet_ctx.sheet_index,
            table_index=table_index,
        )
        table_ctx = TableContext(sheet=sheet_ctx, origin=origin, region=region)

        row_count = region.max_row - region.min_row + 1
        col_count = region.max_col - region.min_col + 1

        emitter.event(
            "table.detected",
            message=f"Table detected in {sheet_name} (#{table_index + 1})",
            sheet_name=sheet_name,
            sheet_index=sheet_ctx.sheet_index,
            table_index=table_index,
            input_file=str(run_ctx.source_path),
            region={
                "min_row": region.min_row,
                "max_row": region.max_row,
                "min_col": region.min_col,
                "max_col": region.max_col,
            },
            row_count=row_count,
            column_count=col_count,
        )

        extracted = self.extractor.extract(sheet_ctx.source_worksheet, origin, region, logger=logger)
        table_ctx.extracted = extracted

        col_count = max(len(extracted.header), max((len(r) for r in extracted.rows), default=0))
        emitter.event(
            "table.extracted",
            message=f"Extracted {len(extracted.rows)} row(s) from {sheet_name} (#{table_index + 1})",
            sheet_name=sheet_name,
            table_index=table_index,
            row_count=len(extracted.rows),
            col_count=col_count,
        )

        hook_dispatcher.on_table_detected(table_ctx)

        mapped = self.mapper.map(extracted, runtime, run_ctx, invoker=invoker, logger=logger)
        table_ctx.mapped = mapped

        mapped_count = sum(1 for f in mapped.mapping.fields if f.source_col is not None)
        emitter.event(
            "table.mapped",
            message=f"Mapped {mapped_count}/{len(mapped.mapping.fields)} fields for {sheet_name} (#{table_index + 1})",
            sheet_name=sheet_name,
            table_index=table_index,
            mapped_fields=mapped_count,
            total_fields=len(mapped.mapping.fields),
            passthrough_fields=len(mapped.mapping.passthrough),
        )

        patch = hook_dispatcher.on_table_mapped(table_ctx)
        if patch:
            table_ctx.mapping_patch = patch
            mapped = self.mapper.apply_patch(mapped, patch, runtime.manifest)
            table_ctx.mapped = mapped
            emitter.event(
                "table.mapping_patched",
                message=f"Applied mapping patch for {sheet_name} (#{table_index + 1})",
                sheet_name=sheet_name,
                table_index=table_index,
            )

        normalized = self.normalizer.normalize(mapped, runtime, run_ctx, invoker=invoker, logger=logger)
        table_ctx.normalized = normalized

        issues = normalized.issues or []
        counts = {sev.value: 0 for sev in Severity}
        for issue in issues:
            counts[issue.severity.value] = counts.get(issue.severity.value, 0) + 1

        emitter.event(
            "table.normalized",
            message=f"Normalized {len(normalized.rows)} row(s) for {sheet_name} (#{table_index + 1})",
            sheet_name=sheet_name,
            table_index=table_index,
            row_count=len(normalized.rows),
            issue_count=len(issues),
            issues_by_severity=counts,
        )

        placement = renderer.write_table(sheet_ctx.output_worksheet, normalized)
        table_ctx.placement = placement
        table_ctx.view = TableView(placement.worksheet, placement.cell_range)

        emitter.event(
            "table.written",
            message=f"Wrote normalized table to output: {sheet_name}!{placement.cell_range.coord}",
            sheet_name=sheet_name,
            table_index=table_index,
            output_range=placement.cell_range.coord,
        )

        hook_dispatcher.on_table_written(table_ctx)
        return table_ctx


__all__ = ["Pipeline"]
