"""Convert detected regions into extracted tables."""

from __future__ import annotations

import logging

from openpyxl.worksheet.worksheet import Worksheet

from ade_engine.types.origin import TableOrigin, TableRegion
from ade_engine.types.tables import ExtractedTable


class TableExtractor:
    """Slice worksheet values into an :class:`~ade_engine.types.tables.ExtractedTable`."""

    def extract(
        self,
        worksheet: Worksheet,
        origin: TableOrigin,
        region: TableRegion,
        *,
        logger: logging.Logger | logging.LoggerAdapter | None = None,
    ) -> ExtractedTable:
        row_iter = worksheet.iter_rows(
            min_row=region.min_row,
            max_row=region.max_row,
            min_col=region.min_col,
            max_col=region.max_col,
            values_only=True,
        )

        try:
            header_row = next(row_iter)
        except StopIteration:
            return ExtractedTable(origin=origin, region=region, header=[], rows=[])

        header = ["" if cell is None else str(cell) for cell in header_row]
        rows = [list(row_values) for row_values in row_iter]

        if logger and logger.isEnabledFor(logging.DEBUG):
            logger.debug(
                "Extracted table region",
                extra={
                    "data": {
                        "stage": "extract",
                        "sheet_name": origin.sheet_name,
                        "table_index": origin.table_index,
                        "header_columns": len(header),
                        "row_count": len(rows),
                        "region": {
                            "min_row": region.min_row,
                            "max_row": region.max_row,
                            "min_col": region.min_col,
                            "max_col": region.max_col,
                        },
                    },
                },
            )

        return ExtractedTable(origin=origin, region=region, header=header, rows=rows)


__all__ = ["TableExtractor"]
